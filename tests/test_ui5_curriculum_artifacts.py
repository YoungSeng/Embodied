from __future__ import annotations

import json
import math
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from eaglevl.train.ui5_curriculum_artifacts import (
    CHECKPOINT_COLUMNS,
    SHEET_ORDER,
    TASKS,
    load_checkpoints_state,
    normalize_scorer_metrics,
    train_curve_rows_from_trainer_state,
    update_curriculum_artifacts,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
UPDATE_SCRIPT = PROJECT_ROOT / "scripts" / "update_ui5_curriculum_artifacts.py"


def prf(tp: int, fp: int, fn: int) -> dict[str, float]:
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = (
        2.0 * precision * recall / (precision + recall)
        if precision + recall
        else 0.0
    )
    return {"precision": precision, "recall": recall, "f1": f1}


def scorer_metrics(
    *,
    image_counts: list[tuple[int, int, int, int]],
    bbox_counts: list[tuple[int, int, int]],
) -> dict:
    scorer_names = (
        "occlusion",
        "cropping",
        "text_overflow",
        "text_ellipsis",
        "content_missing",
    )
    tasks = {}
    for index, task in enumerate(scorer_names):
        image_tp, image_fp, image_fn, image_tn = image_counts[index]
        bbox_tp, bbox_fp, bbox_fn = bbox_counts[index]
        image = {
            **prf(image_tp, image_fp, image_fn),
            "tp": image_tp,
            "fp": image_fp,
            "fn": image_fn,
            "tn": image_tn,
        }
        image["accuracy"] = (image_tp + image_tn) / sum(image_counts[index])
        bbox = {
            **prf(bbox_tp, bbox_fp, bbox_fn),
            "tp": bbox_tp,
            "fp": bbox_fp,
            "fn": bbox_fn,
            "count_accuracy": 0.5 + index * 0.1,
        }
        tasks[task] = {"image": image, "bbox": bbox}
    macro = {
        granularity: {
            name: sum(tasks[task][granularity][name] for task in scorer_names)
            / len(scorer_names)
            for name in ("precision", "recall", "f1")
        }
        for granularity in ("image", "bbox")
    }
    return {"schema_version": 1, "tasks": tasks, "macro": macro}


def uniform_metrics(image_f1: float, bbox_f1: float) -> dict:
    """Build count-consistent metrics for a few useful exact F1 values."""

    count_by_f1 = {
        0.4: (2, 3, 3),
        0.5: (1, 1, 1),
        0.6: (3, 2, 2),
        2 / 3: (2, 1, 1),
        0.8: (4, 1, 1),
    }
    image_key = min(count_by_f1, key=lambda value: abs(value - image_f1))
    bbox_key = min(count_by_f1, key=lambda value: abs(value - bbox_f1))
    if not math.isclose(image_key, image_f1) or not math.isclose(
        bbox_key, bbox_f1
    ):
        raise ValueError("test helper only supports its predefined F1 values")
    image_tuple = (*count_by_f1[image_key], 10)
    return scorer_metrics(
        image_counts=[image_tuple] * 5,
        bbox_counts=[count_by_f1[bbox_key]] * 5,
    )


def make_checkpoint(path: Path, step: int) -> Path:
    path.mkdir(parents=True)
    (path / "config.json").write_text("{}\n", encoding="utf-8")
    (path / "model.safetensors").write_bytes(b"weights")
    (path / "trainer_state.json").write_text(
        json.dumps({"global_step": step}), encoding="utf-8"
    )
    (path / "optimizer.pt").write_bytes(b"optimizer")
    (path / "scheduler.pt").write_bytes(b"scheduler")
    (path / "scaler.pt").write_bytes(b"scaler")
    (path / "rng_state_0.pth").write_bytes(b"rng")
    (path / "rng_state_1.pth").write_bytes(b"rng")
    (path / "training_args.bin").write_bytes(b"training args")
    (path / "dataloader_state_rank0.pt").write_bytes(b"sampler rank 0")
    (path / "dataloader_state_rank1.pt").write_bytes(b"sampler rank 1")
    (path / "continuity_state.json").write_text(
        json.dumps({"global_step": step, "world_size": 2}), encoding="utf-8"
    )
    (path / "checkpoint_complete.json").write_text(
        json.dumps({"global_step": step}), encoding="utf-8"
    )
    nested = path / "global_step" / "rank0"
    nested.mkdir(parents=True)
    (nested / "optim_states.pt").write_bytes(b"nested optimizer")
    return path


class MetricsTests(unittest.TestCase):
    def test_normalizes_five_tasks_and_recomputes_macro_micro_and_joint(self):
        metrics = scorer_metrics(
            image_counts=[
                (1, 0, 1, 8),
                (2, 1, 0, 7),
                (3, 2, 1, 4),
                (0, 1, 2, 7),
                (4, 0, 0, 6),
            ],
            bbox_counts=[
                (1, 1, 1),
                (2, 0, 1),
                (1, 3, 2),
                (0, 1, 2),
                (5, 1, 0),
            ],
        )
        normalized = normalize_scorer_metrics(metrics)
        self.assertEqual(tuple(normalized["tasks"]), TASKS)

        expected_image_macro = sum(
            values["image"]["f1"] for values in metrics["tasks"].values()
        ) / 5
        self.assertAlmostEqual(
            normalized["macro"]["image"]["f1"], expected_image_macro
        )

        image_tp = sum(values["image"]["tp"] for values in metrics["tasks"].values())
        image_fp = sum(values["image"]["fp"] for values in metrics["tasks"].values())
        image_fn = sum(values["image"]["fn"] for values in metrics["tasks"].values())
        expected_image_micro = prf(image_tp, image_fp, image_fn)["f1"]
        self.assertAlmostEqual(
            normalized["micro"]["image"]["f1"], expected_image_micro
        )
        self.assertNotAlmostEqual(expected_image_macro, expected_image_micro)
        self.assertAlmostEqual(
            normalized["overall"]["joint_score"],
            (
                normalized["macro"]["image"]["f1"]
                + normalized["macro"]["bbox"]["f1"]
            )
            / 2,
        )

    def test_rejects_missing_task_and_inconsistent_reported_f1(self):
        metrics = uniform_metrics(0.5, 0.5)
        del metrics["tasks"]["cropping"]
        with self.assertRaisesRegex(ValueError, "ui_cropping"):
            normalize_scorer_metrics(metrics)

        metrics = uniform_metrics(0.5, 0.5)
        metrics["tasks"]["occlusion"]["image"]["f1"] = 0.9
        with self.assertRaisesRegex(ValueError, "disagrees with confusion counts"):
            normalize_scorer_metrics(metrics)

    def test_extracts_and_coalesces_trainer_log_history(self):
        rows = train_curve_rows_from_trainer_state(
            {
                "global_step": 600,
                "log_history": [
                    {"step": 200, "loss": 1.2, "learning_rate": 9e-7},
                    {"step": 200, "grad_norm": 3.5},
                    {"step": 401, "loss_lm": 0.8},
                    {"step": 600, "eval_loss": 1.0},
                ],
            },
            expected_step=600,
        )
        self.assertEqual([row["step"] for row in rows], [200, 401])
        self.assertEqual(rows[0]["phase"], 1)
        self.assertEqual(rows[0]["learning_rate"], 9e-7)
        self.assertEqual(rows[0]["grad_norm"], 3.5)
        self.assertEqual(rows[1]["phase"], 2)
        self.assertEqual(rows[1]["hard_ratio"], 0.45)
        self.assertEqual(rows[1]["learning_rate"], 7e-7)

        baseline = train_curve_rows_from_trainer_state(
            {"global_step": 0, "log_history": []}, expected_step=0
        )
        self.assertEqual(baseline[0]["phase"], "baseline")


class ArtifactTests(unittest.TestCase):
    def update(
        self,
        root: Path,
        *,
        step: int,
        metrics: dict,
        checkpoint: Path,
        **kwargs,
    ) -> dict:
        return update_curriculum_artifacts(
            step=step,
            scorer_metrics=metrics,
            candidate_checkpoint=checkpoint,
            checkpoints_json=root / "checkpoints.json",
            workbook_path=(
                root
                / "diagnostics"
                / "ui5_crop_rollout4_curriculum_evaluation.xlsx"
            ),
            formal_checkpoint_root=root / "checkpoints",
            resume_from=str(kwargs.pop("resume_from", "")),
            evaluation_seconds=float(kwargs.pop("evaluation_seconds", step + 0.5)),
            validate_candidate=False,
            **kwargs,
        )

    def test_best_only_copy_strict_improvement_and_idempotence(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            baseline = make_checkpoint(root / "source_crop", 0)
            rolling = make_checkpoint(root / "resume" / "latest", 200)

            baseline_result = self.update(
                root,
                step=0,
                metrics=uniform_metrics(0.5, 0.4),
                checkpoint=baseline,
                train_curve_rows=[
                    {
                        "step": 0,
                        "phase": "baseline",
                        "hard_ratio": 0.6,
                        "anchor_ratio": 0.25,
                        "global_replay_ratio": 0.15,
                    }
                ],
            )
            self.assertFalse(baseline_result["checkpoint_preserved"])
            self.assertFalse((root / "checkpoints" / "step-000000").exists())
            baseline_row = load_checkpoints_state(root / "checkpoints.json")[
                "evaluations"
            ][0]
            self.assertFalse(baseline_row["improved_image"])
            self.assertFalse(baseline_row["improved_bbox"])
            self.assertFalse(baseline_row["improved_joint"])

            step200 = self.update(
                root,
                step=200,
                metrics=uniform_metrics(0.6, 0.4),
                checkpoint=rolling,
                resume_from=str(baseline),
                hard_transition_rows=[
                    {
                        "task": "ui_occlusion",
                        "group": "0/4",
                        "rollout_id": 0,
                        "samples": 12,
                        "correct": 4,
                        "accuracy": 1 / 3,
                    }
                ],
                anchor_retention_rows=[
                    {
                        "task": "ui_occlusion",
                        "samples": 10,
                        "baseline_score": 0.8,
                        "current_score": 0.8,
                        "delta": 0.0,
                        "retained": True,
                    }
                ],
            )
            self.assertTrue(step200["checkpoint_preserved"])
            self.assertEqual(tuple(step200["metrics"]["tasks"]), TASKS)
            self.assertEqual(step200["overall"]["image_macro_f1"], 0.6)
            self.assertTrue(step200["improved_image"])
            self.assertFalse(step200["improved_bbox"])
            self.assertTrue(step200["improved_joint"])
            preserved200 = root / "checkpoints" / "step-000200"
            self.assertTrue((preserved200 / "scheduler.pt").is_file())
            self.assertTrue((preserved200 / "scaler.pt").is_file())
            self.assertTrue((preserved200 / "dataloader_state_rank0.pt").is_file())
            self.assertTrue((preserved200 / "continuity_state.json").is_file())
            self.assertTrue((preserved200 / "checkpoint_complete.json").is_file())
            self.assertTrue(
                (preserved200 / "global_step" / "rank0" / "optim_states.pt").is_file()
            )

            idempotent = self.update(
                root,
                step=200,
                metrics=uniform_metrics(0.6, 0.4),
                checkpoint=rolling,
                resume_from=str(baseline),
            )
            self.assertTrue(idempotent["idempotent"])
            self.assertEqual(
                [path.name for path in (root / "checkpoints").iterdir()],
                ["step-000200"],
            )

            rolling400 = root / "resume" / "latest-400"
            make_checkpoint(rolling400, 400)
            step400 = self.update(
                root,
                step=400,
                metrics=uniform_metrics(0.5, 0.8),
                checkpoint=rolling400,
                resume_from=str(rolling),
            )
            self.assertTrue(step400["checkpoint_preserved"])

            rolling600 = root / "resume" / "latest-600"
            make_checkpoint(rolling600, 600)
            step600 = self.update(
                root,
                step=600,
                # Image ties the best while BBox and joint are lower.
                metrics=uniform_metrics(0.6, 2 / 3),
                checkpoint=rolling600,
                resume_from=str(rolling400),
            )
            self.assertFalse(step600["checkpoint_preserved"])
            self.assertFalse((root / "checkpoints" / "step-000600").exists())

            for step in (800, 1000, 1200):
                later = make_checkpoint(
                    root / "resume" / f"latest-{step}", step
                )
                result = self.update(
                    root,
                    step=step,
                    metrics=uniform_metrics(0.5, 0.5),
                    checkpoint=later,
                    resume_from=str(rolling600),
                )
                self.assertFalse(result["checkpoint_preserved"])

            state = load_checkpoints_state(root / "checkpoints.json")
            self.assertEqual(
                [row["step"] for row in state["evaluations"]],
                [0, 200, 400, 600, 800, 1000, 1200],
            )
            row200 = state["evaluations"][1]
            self.assertTrue(row200["improved_image"])
            self.assertFalse(row200["improved_bbox"])
            self.assertTrue(row200["improved_joint"])
            self.assertEqual(state["best_image"]["step"], 200)
            self.assertEqual(state["best_bbox"]["step"], 400)
            self.assertEqual(state["best_joint"]["step"], 400)
            self.assertEqual(
                state["best_bbox"]["checkpoint_path"],
                state["best_joint"]["checkpoint_path"],
            )
            self.assertEqual(
                {path.name for path in (root / "checkpoints").iterdir()},
                {"step-000200", "step-000400"},
            )

            workbook_path = (
                root
                / "diagnostics"
                / "ui5_crop_rollout4_curriculum_evaluation.xlsx"
            )
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            try:
                self.assertEqual(tuple(workbook.sheetnames), SHEET_ORDER)
                checkpoint_sheet = workbook["checkpoints"]
                header = tuple(cell.value for cell in checkpoint_sheet[1])
                self.assertEqual(header, CHECKPOINT_COLUMNS)
                self.assertEqual(checkpoint_sheet.max_row - 1, 7)
                self.assertEqual(workbook["ui5_overall"].max_row - 1, 7)
                self.assertEqual(workbook["ui5_by_task"].max_row - 1, 70)
                self.assertEqual(workbook["hard_transition"].max_row - 1, 1)
                self.assertEqual(workbook["anchor_retention"].max_row - 1, 1)
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            self.assertNotEqual(cell.data_type, "f")
                            self.assertNotEqual(cell.data_type, "e")
            finally:
                workbook.close()

    def test_idempotent_registration_rejects_a_damaged_preserved_checkpoint(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            baseline = make_checkpoint(root / "source_crop", 0)
            rolling = make_checkpoint(root / "resume" / "latest", 200)
            self.update(
                root,
                step=0,
                metrics=uniform_metrics(0.5, 0.4),
                checkpoint=baseline,
            )
            self.update(
                root,
                step=200,
                metrics=uniform_metrics(0.6, 0.5),
                checkpoint=rolling,
                resume_from=str(baseline),
            )
            damaged = root / "checkpoints" / "step-000200" / "scheduler.pt"
            damaged.unlink()
            with self.assertRaisesRegex(RuntimeError, "inventory differs"):
                self.update(
                    root,
                    step=200,
                    metrics=uniform_metrics(0.6, 0.5),
                    checkpoint=rolling,
                    resume_from=str(baseline),
                )

    def test_conflicting_same_step_is_rejected_without_state_change(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            baseline = make_checkpoint(root / "source_crop", 0)
            self.update(
                root,
                step=0,
                metrics=uniform_metrics(0.5, 0.4),
                checkpoint=baseline,
            )
            before = (root / "checkpoints.json").read_bytes()
            with self.assertRaisesRegex(ValueError, "different metrics"):
                self.update(
                    root,
                    step=0,
                    metrics=uniform_metrics(0.6, 0.4),
                    checkpoint=baseline,
                )
            self.assertEqual((root / "checkpoints.json").read_bytes(), before)

    def test_rerun_repairs_workbook_after_json_first_crash(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            baseline = make_checkpoint(root / "source_crop", 0)
            rolling = make_checkpoint(root / "resume" / "latest", 200)
            self.update(
                root,
                step=0,
                metrics=uniform_metrics(0.5, 0.4),
                checkpoint=baseline,
            )
            workbook_path = (
                root
                / "diagnostics"
                / "ui5_crop_rollout4_curriculum_evaluation.xlsx"
            )
            with patch(
                "eaglevl.train.ui5_curriculum_artifacts.write_curriculum_workbook",
                side_effect=RuntimeError("simulated workbook crash"),
            ):
                with self.assertRaisesRegex(RuntimeError, "simulated workbook crash"):
                    self.update(
                        root,
                        step=200,
                        metrics=uniform_metrics(0.6, 0.4),
                        checkpoint=rolling,
                    )

            state = load_checkpoints_state(root / "checkpoints.json")
            self.assertEqual(
                [row["step"] for row in state["evaluations"]], [0, 200]
            )
            stale = load_workbook(workbook_path, read_only=True)
            try:
                self.assertEqual(stale["checkpoints"].max_row - 1, 1)
            finally:
                stale.close()

            repaired = self.update(
                root,
                step=200,
                metrics=uniform_metrics(0.6, 0.4),
                checkpoint=rolling,
            )
            self.assertTrue(repaired["idempotent"])
            self.assertEqual(
                [path.name for path in (root / "checkpoints").iterdir()],
                ["step-000200"],
            )
            workbook = load_workbook(workbook_path, read_only=True)
            try:
                self.assertEqual(
                    [
                        int(row[0].value)
                        for row in workbook["checkpoints"].iter_rows(min_row=2)
                    ],
                    [0, 200],
                )
            finally:
                workbook.close()

    def test_strict_candidate_validation_rejects_step_mismatch(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            candidate = make_checkpoint(root / "resume" / "latest", 200)
            validation = {
                "valid": True,
                "errors": [],
                "details": {"global_step": 201},
            }
            with patch(
                "eaglevl.train.ui5_checkpoint_utils.validate_checkpoint",
                return_value=validation,
            ) as validator:
                with self.assertRaisesRegex(ValueError, "does not match"):
                    update_curriculum_artifacts(
                        step=200,
                        scorer_metrics=uniform_metrics(0.5, 0.4),
                        candidate_checkpoint=candidate,
                        checkpoints_json=root / "checkpoints.json",
                        workbook_path=root / "diagnostics.xlsx",
                        formal_checkpoint_root=root / "checkpoints",
                        expected_ranks=2,
                    )
            validator.assert_called_once_with(
                candidate.resolve(),
                mode="resume",
                expected_ranks=2,
                strict=True,
                require_completion_marker=True,
            )

    def test_cli_uses_required_default_paths(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            baseline = make_checkpoint(root / "source_crop", 0)
            (baseline / "trainer_state.json").write_text(
                json.dumps(
                    {
                        "global_step": 0,
                        "log_history": [
                            {"step": 0, "loss": 2.5, "learning_rate": 1e-6}
                        ],
                    }
                ),
                encoding="utf-8",
            )
            metric_path = root / "metrics.json"
            metric_path.write_text(
                json.dumps(uniform_metrics(0.5, 0.4)), encoding="utf-8"
            )
            completed = subprocess.run(
                [
                    sys.executable,
                    "-B",
                    str(UPDATE_SCRIPT),
                    "--run-dir",
                    str(root),
                    "--step",
                    "0",
                    "--metrics-json",
                    str(metric_path),
                    "--candidate-checkpoint",
                    str(baseline),
                    "--evaluation-seconds",
                    "12.5",
                ],
                cwd=PROJECT_ROOT,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(completed.returncode, 0, completed.stderr)
            self.assertTrue((root / "checkpoints.json").is_file())
            self.assertTrue(
                (
                    root
                    / "diagnostics"
                    / "ui5_crop_rollout4_curriculum_evaluation.xlsx"
                ).is_file()
            )
            state = load_checkpoints_state(root / "checkpoints.json")
            self.assertEqual(state["train_curve"][0]["loss_total"], 2.5)
            self.assertEqual(state["train_curve"][0]["phase"], 1)


if __name__ == "__main__":
    unittest.main()
