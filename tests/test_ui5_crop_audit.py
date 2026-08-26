from __future__ import annotations

import json
import sys
import tempfile
import unittest
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from PIL import Image


SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import run_ui5_crop_audit as crop_audit_module  # noqa: E402
from analyze_ui5_source_overlap import TASK_NAMES, analyze  # noqa: E402
from run_ui5_crop_audit import (  # noqa: E402
    CONFIGS,
    TASK_AWARE_CANDIDATES,
    AuditPaths,
    ProgressReporter,
    aggregate_scope,
    audit_state_digest,
    atomic_write_json,
    atomic_write_jsonl,
    build_task_aware_manifest,
    build_preview_rows,
    completed_shard_valid,
    detection_worker_command,
    detector_config,
    ensure_detector_config,
    digest_ids,
    initialize_crop_audit_v2,
    initialize_geometry_worker,
    normalize_gt_in_crop,
    proposal_crops,
    prepared_manifest_valid,
    preflight_icon_runtime,
    resolve_required_directory,
    resolve_python_executable,
    run_crop_audit,
    run_detection_stage,
    geometry_worker,
    evaluate_candidate_gate,
    materialize_image_record,
    uses_task_whole_image_policy,
    validate_training_ready_marker,
    write_excel_report,
    write_statistics_csv,
)


def source_record(image: Path, boxes: list[list[int]]) -> dict:
    return {
        "images": [str(image)],
        "objects": {"bbox": boxes, "bbox_type": "real"},
    }


def locany_record(image: Path, boxes: list[list[int]], label: str = "defect") -> dict:
    if boxes:
        answer = f"<ref>{label}</ref>" + "".join(
            f"<box><{x1}><{y1}><{x2}><{y2}></box>" for x1, y1, x2, y2 in boxes
        )
    else:
        answer = "<box>none</box>"
    return {
        "image": str(image),
        "conversations": [
            {"from": "human", "value": "locate"},
            {"from": "gpt", "value": answer},
        ],
    }


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


class OverlapAuditTest(unittest.TestCase):
    def test_same_basename_different_directories_are_not_overwritten(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source_dir = root / "source"
            locany_dir = root / "locany"
            output_dir = root / "audit"
            images = []
            for index, task in enumerate(TASK_NAMES):
                directory = root / ("left" if index == 0 else "right" if index == 1 else f"d{index}")
                directory.mkdir(parents=True)
                image = directory / ("same.png" if index < 2 else f"image_{index}.png")
                Image.new("RGB", (100, 80), (index * 20, 0, 0)).save(image)
                images.append(image)
            source_names = (
                "train_ui_occlusion_wcnt.jsonl",
                "train_ui_cropping_wcnt.jsonl",
                "train_ui_text_overflow_wcnt.jsonl",
                "train_ui_text_ellipsis_wcnt.jsonl",
                "train_ui_content_missing_wcnt.jsonl",
            )
            for task, source_name, image in zip(TASK_NAMES, source_names, images):
                write_jsonl(source_dir / source_name, [source_record(image, [[1, 1, 20, 20]])])
                write_jsonl(locany_dir / f"{task}_train.jsonl", [locany_record(image, [[10, 10, 200, 200]])])
            result = analyze(source_dir, locany_dir, output_dir)
            actual = result["actual_training_data"]
            self.assertEqual(actual["basename_conflicts"]["count"], 1)
            self.assertEqual(
                actual["path_overlap"]["counts"][TASK_NAMES[0]][TASK_NAMES[1]], 0
            )
            records = [json.loads(line) for line in (output_dir / "training_records.jsonl").read_text(encoding="utf-8").splitlines()]
            self.assertEqual(len({row["canonical_path"] for row in records}), 5)

    def test_same_content_different_paths_has_one_detector_manifest_entry(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source_dir = root / "source"
            locany_dir = root / "locany"
            output = root / "output"
            source_dir.mkdir()
            locany_dir.mkdir()
            images = []
            for index in range(5):
                image = root / f"alias_{index}" / "screen.png"
                image.parent.mkdir()
                color = "white" if index < 2 else (index * 30, 0, 0)
                Image.new("RGB", (100, 100), color).save(image)
                images.append(image)
            source_names = (
                "train_ui_occlusion_wcnt.jsonl",
                "train_ui_cropping_wcnt.jsonl",
                "train_ui_text_overflow_wcnt.jsonl",
                "train_ui_text_ellipsis_wcnt.jsonl",
                "train_ui_content_missing_wcnt.jsonl",
            )
            for task, source_name, image in zip(TASK_NAMES, source_names, images):
                write_jsonl(source_dir / source_name, [source_record(image, [[1, 1, 20, 20]])])
                write_jsonl(
                    locany_dir / f"{task}_train.jsonl",
                    [locany_record(image, [[100, 100, 200, 200]])],
                )
            args = SimpleNamespace(
                source_dir=source_dir,
                locany_data_dir=locany_dir,
                parser_root=root / "parser",
                output_dir=output,
                shard_size=500,
                resume=False,
                max_unique_images=0,
                progress_interval_seconds=60,
                progress_every_images=25,
                text_model_dir=None,
                icon_model=root / "model.pt",
                text_long_side=1920,
                text_box_threshold=0.3,
                icon_long_side=1920,
                icon_confidence=0.05,
                enable_mkldnn=False,
            )
            with mock.patch("run_ui5_crop_audit.print_preflight"):
                unique, task_samples = build_task_aware_manifest(args)
            self.assertEqual(len(unique), 4)
            self.assertEqual(len(task_samples), 5)
            shared = [row for row in unique if len(row["canonical_paths"]) == 2]
            self.assertEqual(len(shared), 1)
            self.assertEqual(
                shared[0]["tasks"], ["ui_cropping", "ui_occlusion"]
            )


def sample(task: str, gt_boxes: list[list[int]], image_id: str = "img_shared") -> dict:
    return {
        "sample_id": f"sample_{task}",
        "image_id": image_id,
        "task": task,
        "canonical_path": "/tmp/source.png",
        "width": 100,
        "height": 100,
        "positive": bool(gt_boxes),
        "gt_boxes": gt_boxes,
        "gt_count": len(gt_boxes),
        "gt_boxes_1000": [
            [round(value / 100 * 1000) for value in box] for box in gt_boxes
        ],
    }


class TaskAwarePreviewTest(unittest.TestCase):
    def test_shared_detection_does_not_merge_different_task_gt(self):
        crops = [[0, 0, 50, 50], [50, 50, 100, 100]]
        paths = [Path("crop_1.png"), Path("crop_2.png")]
        first, _ = build_preview_rows(sample("ui_occlusion", [[5, 5, 20, 20]]), crops, paths, config_name="A")
        second, _ = build_preview_rows(sample("ui_cropping", [[70, 70, 90, 90]]), crops, paths, config_name="A")
        self.assertEqual(first[0]["contained_gt_indices"], [0])
        self.assertEqual(second[-1]["contained_gt_indices"], [0])
        self.assertNotEqual(first[0]["image"], second[-1]["image"])

    def test_positive_and_negative_tasks_for_same_image_remain_independent(self):
        crop = [[0, 0, 100, 100]]
        paths = [Path("shared.png")]
        positive, _ = build_preview_rows(sample("ui_occlusion", [[10, 10, 20, 20]]), crop, paths, config_name="A")
        negative, _ = build_preview_rows(sample("ui_cropping", []), crop, paths, config_name="A")
        self.assertTrue(positive[0]["positive"])
        self.assertFalse(negative[0]["positive"])
        self.assertEqual(positive[0]["image"], negative[0]["image"])

    def test_partial_intersection_is_ineligible_not_negative(self):
        rows, failures = build_preview_rows(
            sample("ui_occlusion", [[40, 40, 80, 80]]),
            [[0, 0, 60, 60]],
            [Path("crop.png")],
            config_name="A",
        )
        self.assertEqual(len(rows), 1)
        self.assertIsNone(rows[0]["positive"])
        self.assertFalse(rows[0]["training_eligible"])
        self.assertEqual(rows[0]["partial_gt_indices"], [0])
        self.assertEqual(failures[0]["failure_type"], "partial_intersection")

    def test_content_missing_policy_is_task_aware(self):
        self.assertTrue(uses_task_whole_image_policy("ui_content_missing"))

    def test_content_missing_reuses_full_original_and_normalized_gt_verbatim(self):
        content = sample("ui_content_missing", [[10, 20, 30, 40]])
        content["gt_boxes_1000"] = [[101, 199, 303, 402]]
        rows, failures = build_preview_rows(
            content,
            [[0, 0, 100, 100]],
            [Path("original.png")],
            config_name="v3",
        )
        self.assertEqual(failures, [])
        self.assertEqual(rows[0]["crop_bbox"], [0, 0, 100, 100])
        self.assertFalse(rows[0]["label_transform_applied"])
        self.assertTrue(rows[0]["roundtrip_gate_excluded"])
        self.assertEqual(rows[0]["original_gt_boxes_1000"], [[101, 199, 303, 402]])
        self.assertEqual(rows[0]["output_gt_boxes_1000"], [[101, 199, 303, 402]])

    def test_one_image_across_five_tasks_keeps_supervision_independent(self):
        paths = [Path("same_physical_crop.png")]
        outputs = {}
        for index, task in enumerate(TASK_NAMES):
            boxes = [] if index % 2 else [[10 + index, 10, 20 + index, 20]]
            rows, _ = build_preview_rows(
                sample(task, boxes), [[0, 0, 100, 100]], paths, config_name="v3"
            )
            outputs[task] = rows[0]
        self.assertEqual({row["image"] for row in outputs.values()}, {str(paths[0])})
        self.assertEqual(
            [outputs[task]["positive"] for task in TASK_NAMES],
            [True, False, True, False, True],
        )
        self.assertEqual(
            [outputs[task]["gt_count"] for task in TASK_NAMES], [1, 0, 1, 0, 1]
        )
        self.assertFalse(uses_task_whole_image_policy("ui_occlusion"))
        self.assertTrue(uses_task_whole_image_policy("ui_content_missing"))

    def test_coordinate_roundtrip_is_within_one_pixel(self):
        result = normalize_gt_in_crop([121, 75, 451, 309], [20, 30, 820, 630])
        self.assertLessEqual(result["roundtrip_max_error_px"], 1)


class FakeCropper:
    @dataclass(frozen=True)
    class DetectionBox:
        index: int
        source: str
        source_index: int
        bbox: tuple[int, int, int, int]
        score: float | None

    @dataclass
    class Group:
        group_id: int
        member_indices: set[int]
        bbox: tuple[int, int, int, int]

    @staticmethod
    def union_bbox(boxes):
        boxes = list(boxes)
        return (
            min(box[0] for box in boxes), min(box[1] for box in boxes),
            max(box[2] for box in boxes), max(box[3] for box in boxes),
        )

    @classmethod
    def build_connected_components(cls, detections, width, height, horizontal_ratio, vertical_ratio):
        return ([list(range(len(detections)))] if detections else []), max(0, len(detections) - 1)

    @classmethod
    def merge_groups_to_limit(cls, components, detections, max_groups):
        members = set(components[0])
        return [cls.Group(0, members, cls.union_bbox(detections[index].bbox for index in members))], []

    @staticmethod
    def merge_overlapping_group_envelopes(groups):
        return groups, []

    @staticmethod
    def make_non_overlapping_context_crops(groups, width, height, context_ratio):
        box = groups[0].bbox
        pad_x = round((box[2] - box[0]) * context_ratio)
        pad_y = round((box[3] - box[1]) * context_ratio)
        return [(max(0, box[0] - pad_x), max(0, box[1] - pad_y), min(width, box[2] + pad_x), min(height, box[3] + pad_y))], []

    @classmethod
    def make_boundary_safe_crop(cls, crop, boxes, width, height, margin_ratio):
        current = tuple(crop)
        changed = True
        while changed:
            changed = False
            for box in boxes:
                intersects = not (current[2] <= box[0] or box[2] <= current[0] or current[3] <= box[1] or box[3] <= current[1])
                contains = current[0] <= box[0] and current[1] <= box[1] and current[2] >= box[2] and current[3] >= box[3]
                if intersects and not contains:
                    current = cls.union_bbox([current, box])
                    changed = True
        return current, 0

    @classmethod
    def make_lightly_trimmed_whole_image_crop(
        cls, boxes, width, height, max_trim_ratio, detection_margin_ratio
    ):
        if not boxes:
            return 0, 0, width, height
        content = cls.union_bbox(boxes)
        max_x = int(width * max_trim_ratio)
        max_y = int(height * max_trim_ratio)
        return (
            min(max_x, content[0]),
            min(max_y, content[1]),
            max(width - max_x, content[2]),
            max(height - max_y, content[3]),
        )


class GeometryTest(unittest.TestCase):
    def test_dense_page_is_one_near_full_crop_without_forced_split(self):
        boxes = [[2, 5, 22, 35], [20, 5, 45, 35], [43, 5, 70, 35], [68, 5, 98, 35]]
        record = {
            "width": 100,
            "height": 40,
            "text_detections": [{"bbox": box, "score": 0.9} for box in boxes],
            "icon_detections": [],
        }
        proposal = proposal_crops(FakeCropper, record, CONFIGS["A"], max_crops=10, boundary_margin_ratio=0.01)
        self.assertEqual(len(proposal["crop_boxes"]), 1)
        self.assertGreater(
            (proposal["crop_boxes"][0][2] - proposal["crop_boxes"][0][0]) *
            (proposal["crop_boxes"][0][3] - proposal["crop_boxes"][0][1]) / 4000,
            0.8,
        )

    def test_no_detector_box_is_cut_by_crop_boundary(self):
        record = {
            "width": 100,
            "height": 100,
            "text_detections": [{"bbox": [10, 10, 90, 90], "score": 0.9}],
            "icon_detections": [{"bbox": [85, 85, 98, 98], "score": 0.8}],
        }
        proposal = proposal_crops(FakeCropper, record, CONFIGS["C"], max_crops=10, boundary_margin_ratio=0.01)
        self.assertEqual(proposal["detector_boundary_cut_count"], 0)
        for detector in proposal["detection_boxes"]:
            self.assertTrue(any(crop[0] <= detector[0] and crop[1] <= detector[1] and crop[2] >= detector[2] and crop[3] >= detector[3] for crop in proposal["crop_boxes"]))

    def test_identical_task_aware_bbox_writes_one_physical_crop(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.png"
            Image.new("RGB", (100, 100), "white").save(source)
            shared_box = [0, 0, 50, 50]
            geometry = {
                "unique_region_boxes": [shared_box],
                "sample_results": [
                    {
                        "sample_id": "occlusion",
                        "task": "ui_occlusion",
                        "crop_kind": "region",
                        "crop_boxes": [shared_box],
                        "gt_boxes": [],
                    },
                    {
                        "sample_id": "cropping",
                        "task": "ui_cropping",
                        "crop_kind": "region",
                        "crop_boxes": [shared_box],
                        "gt_boxes": [],
                    },
                    {
                        "sample_id": "whole",
                        "task": "ui_content_missing",
                        "crop_kind": "whole",
                        "crop_boxes": [[0, 0, 100, 100]],
                        "gt_boxes": [],
                    },
                ],
            }
            result = materialize_image_record(
                manifest={"image_id": "img", "image_path": str(source)},
                geometry=geometry,
                config_root=root / "candidate",
                overview_sample_ids=set(),
            )
            self.assertEqual(len(result["region_paths"]), 1)
            self.assertEqual(
                result["sample_paths"]["occlusion"],
                result["sample_paths"]["cropping"],
            )
            self.assertEqual(result["sample_paths"]["whole"], [str(source.resolve())])
            self.assertEqual(result["whole_paths"], [])
            self.assertEqual(result["region_reference_count"], 2)

    def test_region_roundtrip_gate_is_checked_while_whole_image_is_excluded(self):
        def detail(task: str, roundtrip_errors: int = 0) -> dict:
            return {
                "config": "X", "sample_id": task, "image_id": "img", "task": task,
                "positive": True, "gt_count": 1, "gt_contained_count": 1,
                "partial_only_gt_count": 0, "all_gt_contained": True, "crop_count": 1,
                "original_area": 10000, "union_crop_area": 2500,
                "union_area_ratio": 0.25, "gt_enlargement_gains": [2.0],
                "empty_detection_fallback": False, "forced_merge": False,
                "detector_boundary_cut_count": 0,
                "roundtrip_error_over_1_count": roundtrip_errors,
                "partial_training_eligible_count": 0, "hard_negative_count": 0,
            }

        region_rows = [detail(task) for task in (
            "ui_occlusion", "ui_cropping", "ui_text_overflow", "ui_text_ellipsis"
        )]
        whole = detail("ui_content_missing", roundtrip_errors=99)
        rows = [*region_rows, whole]
        by_scope = {"ALL": aggregate_scope(rows)}
        for task in TASK_NAMES:
            by_scope[task] = aggregate_scope([row for row in rows if row["task"] == task])
        summary = {"candidates": {"X": {"by_scope": by_scope}}}
        self.assertTrue(evaluate_candidate_gate("X", summary, rows)["passes"])
        region_rows[0]["roundtrip_error_over_1_count"] = 1
        rows = [*region_rows, whole]
        self.assertFalse(evaluate_candidate_gate("X", summary, rows)["passes"])


class ResumeAndExcelTest(unittest.TestCase):
    def test_geometry_process_pool_loads_fixed_parser_without_images(self):
        parser_root = Path(__file__).resolve().parents[2] / "ui-region-parser"
        if not (parser_root / "ui_region_cropper.py").is_file():
            self.skipTest("independent fixed parser checkout is not present")
        with tempfile.TemporaryDirectory() as temporary:
            payload = {
                "manifest": {"image_id": "img", "width": 100, "height": 100},
                "detection": {
                    "width": 100,
                    "height": 100,
                    "text_detections": [{"bbox": [10, 10, 30, 30], "score": 0.9}],
                    "icon_detections": [],
                },
                "image_samples": [sample("ui_occlusion", [[15, 15, 20, 20]], "img")],
                "config_name": "C",
                "candidate": TASK_AWARE_CANDIDATES["C"],
                "config_root": temporary,
                "max_crops": 10,
                "boundary_margin_ratio": 0.01,
            }
            with ProcessPoolExecutor(
                max_workers=2,
                initializer=initialize_geometry_worker,
                initargs=(str(parser_root),),
            ) as executor:
                result = list(executor.map(geometry_worker, [payload]))
            self.assertEqual(result[0]["image_id"], "img")
            self.assertEqual(result[0]["config"], "C")

    def test_named_audit_refuses_legacy_output_and_preserves_other_audits(self):
        with tempfile.TemporaryDirectory() as temporary:
            paths = AuditPaths(Path(temporary) / "audit")
            legacy = paths.crop_audit / "config_A" / "crops" / "img"
            legacy.mkdir(parents=True)
            (legacy / "crop_01.png").write_bytes(b"partial")
            args = SimpleNamespace(
                max_crops=10,
                boundary_margin_ratio=0.01,
                overview_samples_per_task=50,
                overview_anomalies_per_category=50,
            )
            with self.assertRaisesRegex(RuntimeError, "choose a new --crop-audit-name"):
                initialize_crop_audit_v2(args, paths, [{"image_id": "img"}])
            self.assertTrue((legacy / "crop_01.png").is_file())
            fresh = AuditPaths(paths.output, "crop_audit_v3_retry")
            initialize_crop_audit_v2(args, fresh, [{"image_id": "img"}])
            self.assertTrue((fresh.crop_audit / "audit_state.json").is_file())
            self.assertTrue((legacy / "crop_01.png").is_file())

    def test_detection_workers_can_use_separate_python_environments(self):
        args = SimpleNamespace(
            text_python="/envs/paddle/bin/python",
            icon_python="/envs/locateanything/bin/python",
            source_dir=Path("/source"),
            locany_data_dir=Path("/data"),
            parser_root=Path("/parser"),
            output_dir=Path("/output"),
            gpus="0,1,2,3",
            workers_per_gpu=1,
            image_loader_threads=4,
            shard_size=750,
            max_unique_images=0,
            progress_interval_seconds=10,
            progress_every_images=25,
            text_long_side=1920,
            text_box_threshold=0.3,
            icon_long_side=1920,
            icon_confidence=0.05,
            text_model_dir=None,
            icon_model=None,
            resume=True,
            enable_mkldnn=False,
        )
        text_command = detection_worker_command(args, "text", 0, 4)
        icon_command = detection_worker_command(args, "icon", 0, 4)
        self.assertEqual(text_command[0], args.text_python)
        self.assertEqual(icon_command[0], args.icon_python)

    def test_icon_runtime_preflight_preserves_original_import_error(self):
        failed = __import__("subprocess").CompletedProcess(
            args=[],
            returncode=1,
            stdout="",
            stderr="RuntimeError: operator torchvision::nms does not exist\n",
        )
        with mock.patch("run_ui5_crop_audit.subprocess.run", return_value=failed):
            with self.assertRaisesRegex(
                RuntimeError,
                r"(?s)尚未启动 GPU worker.*torchvision::nms does not exist",
            ):
                preflight_icon_runtime(
                    "/envs/icon/bin/python",
                    gpu="0",
                    model_path=Path("/parser/weights/icon_detect_v3/model.pt"),
                )

    def test_python_executable_defaults_to_current_interpreter(self):
        self.assertEqual(
            Path(resolve_python_executable(None, "--icon-python")),
            Path(sys.executable).absolute(),
        )

    def test_missing_cli_directory_reports_option_and_resolved_path(self):
        with tempfile.TemporaryDirectory() as temporary:
            missing = Path(temporary) / "does-not-exist"
            with self.assertRaisesRegex(
                FileNotFoundError,
                r"--locany-data-dir directory does not exist:.*do not create an empty placeholder",
            ):
                resolve_required_directory(missing, "--locany-data-dir")

    def test_progress_reporter_writes_atomic_status_with_eta_fields(self):
        with tempfile.TemporaryDirectory() as temporary:
            reporter = ProgressReporter(
                stage="text",
                total=100,
                output_dir=Path(temporary),
                interval_seconds=60,
            )
            payload = reporter.update(25, force=True)
            stored = json.loads(
                (Path(temporary) / "run_status.json").read_text(encoding="utf-8")
            )
            self.assertEqual(stored["stage"], "text")
            self.assertEqual(stored["completed"], 25)
            self.assertEqual(stored["total"], 100)
            self.assertEqual(stored["percent"], 0.25)
            self.assertIn("eta_seconds", payload)

    def test_ocr_defaults_to_paddle_auto_download_but_icon_is_explicit(self):
        args = SimpleNamespace(
            text_model_dir=None,
            icon_model=None,
            parser_root=Path("/tmp/ui-region-parser"),
            text_long_side=1920,
            text_box_threshold=0.3,
            icon_long_side=1920,
            icon_confidence=0.05,
        )
        config = detector_config(args)
        self.assertIsNone(config["text"]["model_dir"])
        self.assertTrue(config["text"]["auto_download"])
        self.assertFalse(config["text"]["enable_mkldnn"])
        self.assertTrue(
            Path(config["icon"]["model"]).as_posix().endswith(
                "weights/icon_detect_v3/model.pt"
            )
        )

    def test_old_config_migrates_to_mkldnn_disabled_before_any_text_shard(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            args = SimpleNamespace(
                text_model_dir=None,
                icon_model=None,
                parser_root=root / "parser",
                text_long_side=1920,
                text_box_threshold=0.3,
                icon_long_side=1920,
                icon_confidence=0.05,
                enable_mkldnn=False,
            )
            desired = detector_config(args)
            old = json.loads(json.dumps(desired))
            old["text"].pop("enable_mkldnn")
            config_path = root / "detections" / "detector_config.json"
            atomic_write_json(config_path, old)
            ensure_detector_config(config_path, desired)
            migrated = json.loads(config_path.read_text(encoding="utf-8"))
            self.assertFalse(migrated["text"]["enable_mkldnn"])

    def test_complete_prepare_manifest_is_resume_eligible(self):
        with tempfile.TemporaryDirectory() as temporary:
            paths = AuditPaths(Path(temporary))
            unique = [{"image_id": "img_a"}, {"image_id": "img_b"}]
            samples = [
                {"sample_id": "sample_a", "image_id": "img_a"},
                {"sample_id": "sample_b", "image_id": "img_b"},
            ]
            atomic_write_jsonl(paths.unique_images, unique)
            atomic_write_jsonl(paths.task_samples, samples)
            atomic_write_jsonl(paths.shards / "shard_00000.jsonl", unique)
            atomic_write_json(
                paths.manifest / "prepare_summary.json",
                {"unique_images": 2, "task_samples": 2, "shards": 1},
            )
            self.assertTrue(prepared_manifest_valid(paths))

    def test_interrupted_shard_is_not_skipped_but_complete_shard_is(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            input_path = root / "shard_00000.jsonl"
            output_path = root / "output.jsonl"
            done_path = root / "output.done.json"
            inputs = [{"image_id": "a"}, {"image_id": "b"}]
            outputs = [{"image_id": "a"}, {"image_id": "b"}]
            atomic_write_jsonl(input_path, inputs)
            atomic_write_jsonl(output_path, outputs)
            atomic_write_json(done_path, {"stage": "text", "count": 2, "image_id_digest": digest_ids(["a", "b"])})
            self.assertTrue(completed_shard_valid(input_path, output_path, done_path, "text"))
            atomic_write_jsonl(output_path, outputs[:1])
            self.assertFalse(completed_shard_valid(input_path, output_path, done_path, "text"))

    def test_fully_completed_resume_does_not_load_model_or_start_workers(self):
        with tempfile.TemporaryDirectory() as temporary:
            paths = AuditPaths(Path(temporary))
            rows = [{"image_id": "a"}, {"image_id": "b"}]
            shard = paths.shards / "shard_00000.jsonl"
            output = paths.stage_dir("text") / shard.name
            marker = paths.stage_dir("text") / "shard_00000.done.json"
            atomic_write_jsonl(paths.unique_images, rows)
            atomic_write_jsonl(shard, rows)
            atomic_write_jsonl(output, rows)
            atomic_write_json(
                marker,
                {
                    "stage": "text",
                    "count": 2,
                    "image_id_digest": digest_ids(["a", "b"]),
                },
            )
            args = SimpleNamespace(
                output_dir=Path(temporary),
                resume=True,
                progress_interval_seconds=60,
            )
            with (
                mock.patch("run_ui5_crop_audit.print_preflight"),
                mock.patch("run_ui5_crop_audit.subprocess.Popen") as popen,
                mock.patch("run_ui5_crop_audit.preflight_icon_runtime") as runtime_probe,
            ):
                run_detection_stage(args, "text")
            popen.assert_not_called()
            runtime_probe.assert_not_called()
            status = json.loads(
                (Path(temporary) / "run_status.json").read_text(encoding="utf-8")
            )
            self.assertEqual(status["completed"], 2)
            self.assertEqual(status["status"], "completed")

    def test_excel_summary_matches_json_and_statistics_csv(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            self.fail(f"openpyxl is required by the project: {exc}")
        detail = {
            "config": "A", "sample_id": "s", "image_id": "i", "task": TASK_NAMES[0],
            "positive": True, "gt_count": 1, "gt_contained_count": 1,
            "partial_only_gt_count": 0, "all_gt_contained": True, "crop_count": 1,
            "original_area": 10000, "union_crop_area": 2500, "union_area_ratio": 0.25,
            "pixel_reduction_ratio": 0.75, "gt_enlargement_gains": [2.0],
            "empty_detection_fallback": False, "forced_merge": False,
            "detector_boundary_cut_count": 0, "roundtrip_error_over_1_count": 0,
            "overview": "overview.png", "source_image": "source.png",
        }
        metric = aggregate_scope([detail])
        empty = aggregate_scope([])
        summary = {"configs": {}}
        for config in CONFIGS:
            by_scope = {"ALL": metric, **{task: (metric if task == TASK_NAMES[0] else empty) for task in TASK_NAMES}}
            summary["configs"][config] = {"by_scope": by_scope}
        matrix = {task: {other: int(task == other) for other in TASK_NAMES} for task in TASK_NAMES}
        jaccard = {task: {other: float(task == other) for other in TASK_NAMES} for task in TASK_NAMES}
        overlap_dataset = {"path_overlap": {"counts": matrix, "jaccard": jaccard}, "content_overlap": {"counts": matrix, "jaccard": jaccard}}
        overlap = {"source_data": overlap_dataset, "actual_training_data": overlap_dataset}
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "ui5_crop_audit.xlsx"
            csv_path = root / "statistics.csv"
            write_statistics_csv(csv_path, [detail])
            write_excel_report(workbook_path, summary, overlap, [detail], [])
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                headers = [cell.value for cell in workbook["summary"][1]]
                values = next(workbook["summary"].iter_rows(min_row=2, values_only=True))
                excel_row = dict(zip(headers, values))
                self.assertEqual(excel_row["gt_contained_count"], metric["gt_contained_count"])
                self.assertAlmostEqual(excel_row["pixel_reduction_ratio"], metric["pixel_reduction_ratio"])
                self.assertEqual(tuple(workbook.sheetnames), ("summary", "task_overlap", "image_detail", "gt_failures", "config_compare"))
            finally:
                workbook.close()
            with csv_path.open("r", encoding="utf-8", newline="") as handle:
                csv_row = next(__import__("csv").DictReader(handle))
            self.assertEqual(int(csv_row["gt_contained_count"]), metric["gt_contained_count"])
            self.assertAlmostEqual(float(csv_row["pixel_reduction_ratio"]), metric["pixel_reduction_ratio"])

    def test_cpu_crop_audit_writes_all_machine_and_excel_outputs(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            output = root / "audit"
            paths = AuditPaths(output)
            image_path = root / "shared.png"
            Image.new("RGB", (100, 100), "white").save(image_path)
            unique = {
                "image_id": "img_shared",
                "content_id": "content_shared",
                "image_path": str(image_path),
                "width": 100,
                "height": 100,
                "tasks": ["ui_occlusion", "ui_content_missing"],
            }
            samples = [
                {
                    **sample("ui_occlusion", [[15, 15, 25, 25]]),
                    "canonical_path": str(image_path),
                },
                {
                    **sample("ui_content_missing", []),
                    "canonical_path": str(image_path),
                },
            ]
            merged = {
                "image_id": "img_shared",
                "content_id": "content_shared",
                "image": str(image_path),
                "width": 100,
                "height": 100,
                "text_detections": [{"bbox": [10, 10, 30, 30], "score": 0.9}],
                "icon_detections": [],
            }
            atomic_write_jsonl(paths.unique_images, [unique])
            atomic_write_jsonl(paths.task_samples, samples)
            atomic_write_jsonl(paths.shards / "shard_00000.jsonl", [unique])
            atomic_write_jsonl(paths.merged, [merged])
            atomic_write_json(paths.detector_config, {"fixed": True})
            matrix = {task: {other: 0 for other in TASK_NAMES} for task in TASK_NAMES}
            jaccard = {task: {other: 0.0 for other in TASK_NAMES} for task in TASK_NAMES}
            overlap_dataset = {
                "path_overlap": {"counts": matrix, "jaccard": jaccard},
                "content_overlap": {"counts": matrix, "jaccard": jaccard},
                "basename_conflicts": {"details": []},
            }
            atomic_write_json(
                paths.manifest / "overlap" / "source_overlap.json",
                {"source_data": overlap_dataset, "actual_training_data": overlap_dataset},
            )
            args = SimpleNamespace(
                output_dir=output,
                parser_root=root,
                max_crops=10,
                boundary_margin_ratio=0.01,
                progress_interval_seconds=60,
                resume=True,
                crop_workers=1,
                crop_audit_name="crop_audit_v3",
                overview_samples_per_task=50,
                overview_anomalies_per_category=50,
            )
            with mock.patch("run_ui5_crop_audit.load_parser_module", return_value=FakeCropper):
                result = run_crop_audit(args)
            self.assertFalse(result["next_stage_gate"]["training_started"])
            self.assertTrue((paths.crop_audit / "summary.json").is_file())
            self.assertTrue((paths.crop_audit / "statistics.csv").is_file())
            self.assertTrue((paths.crop_audit / "ui5_crop_audit.xlsx").is_file())
            selected = result["materialized_candidate"]
            selected_root = paths.crop_audit / f"candidate_{selected}"
            self.assertTrue((selected_root / "preview" / "ui_occlusion.jsonl").is_file())
            self.assertEqual(
                len(list((selected_root / "crops" / "img_shared").glob("region_*.png"))),
                1,
            )
            self.assertFalse(any(selected_root.rglob("whole_*.png")))
            for name in TASK_AWARE_CANDIDATES:
                self.assertTrue(
                    (
                        paths.crop_audit
                        / f"candidate_{name}"
                        / "geometry"
                        / "shard_00000.done.json"
                    ).is_file()
                )
            self.assertEqual(result["materialization"]["candidate"], selected)
            self.assertEqual(result["materialization"]["whole_generated_file_count"], 0)
            self.assertTrue(result["next_stage_gate"]["training_ready"])
            marker_path = paths.crop_audit / "training_ready.json"
            self.assertTrue(marker_path.is_file())
            validated = validate_training_ready_marker(paths.crop_audit)
            self.assertEqual(validated["recommended_config"], selected)
            marker = json.loads(marker_path.read_text(encoding="utf-8"))
            summary = json.loads(
                (paths.crop_audit / "summary.json").read_text(encoding="utf-8")
            )
            state = json.loads(
                (paths.crop_audit / "audit_state.json").read_text(encoding="utf-8")
            )
            self.assertTrue(marker["created_after_all_checks"])
            self.assertEqual(marker["audit_state_digest"], audit_state_digest(state))
            self.assertEqual(
                marker["input_snapshot_digest"],
                audit_state_digest(summary["input_snapshot_after"]),
            )
            self.assertEqual(
                marker["summary_file_digest"],
                crop_audit_module.content_fingerprint(paths.crop_audit / "summary.json"),
            )
            summary_path = paths.crop_audit / "summary.json"
            summary_bytes = summary_path.read_bytes()
            summary_path.write_bytes(summary_bytes + b"\n")
            with self.assertRaisesRegex(RuntimeError, "summary digest mismatch"):
                validate_training_ready_marker(paths.crop_audit)
            summary_path.write_bytes(summary_bytes)

            state_path = paths.crop_audit / "audit_state.json"
            state_bytes = state_path.read_bytes()
            changed_state = dict(state)
            changed_state["test_mutation"] = True
            atomic_write_json(state_path, changed_state)
            with self.assertRaisesRegex(RuntimeError, "audit state digest mismatch"):
                validate_training_ready_marker(paths.crop_audit)
            state_path.write_bytes(state_bytes)

            merged_bytes = paths.merged.read_bytes()
            paths.merged.write_bytes(merged_bytes + b"\n")
            with self.assertRaisesRegex(RuntimeError, "live manifest/detections"):
                validate_training_ready_marker(paths.crop_audit)
            paths.merged.write_bytes(merged_bytes)
            region_path = next(
                (selected_root / "crops" / "img_shared").glob("region_*.png")
            )
            region_before = (region_path.read_bytes(), region_path.stat().st_mtime_ns)

            with (
                mock.patch(
                    "run_ui5_crop_audit.load_parser_module", return_value=FakeCropper
                ) as parser_loader,
                mock.patch(
                    "run_ui5_crop_audit.compute_geometry_record",
                    side_effect=AssertionError("resume must not recompute geometry"),
                ),
                mock.patch(
                    "run_ui5_crop_audit.open_raw_image",
                    side_effect=AssertionError("resume must not decode completed images"),
                ),
            ):
                resumed = run_crop_audit(args)
            parser_loader.assert_not_called()
            self.assertEqual(resumed["materialized_candidate"], selected)
            self.assertEqual(
                region_before,
                (region_path.read_bytes(), region_path.stat().st_mtime_ns),
            )

            original_snapshot = crop_audit_module.audit_input_snapshot
            snapshot_calls = 0

            def changed_after_snapshot(*call_args, **call_kwargs):
                nonlocal snapshot_calls
                snapshot_calls += 1
                value = original_snapshot(*call_args, **call_kwargs)
                if snapshot_calls == 2:
                    value = dict(value)
                    value["merged_detections_file_digest"] = "changed-during-run"
                return value

            with (
                mock.patch(
                    "run_ui5_crop_audit.audit_input_snapshot",
                    side_effect=changed_after_snapshot,
                ),
                mock.patch(
                    "run_ui5_crop_audit.compute_geometry_record",
                    side_effect=AssertionError("resume must not recompute geometry"),
                ),
                mock.patch(
                    "run_ui5_crop_audit.open_raw_image",
                    side_effect=AssertionError("resume must not rewrite crop PNG"),
                ),
            ):
                changed = run_crop_audit(args)
            self.assertFalse(changed["next_stage_gate"]["training_ready"])
            self.assertFalse(
                changed["next_stage_gate"]["conditions"]["input_snapshot_unchanged"]
            )
            self.assertFalse(marker_path.exists())

            with mock.patch(
                "run_ui5_crop_audit.load_parser_module", return_value=FakeCropper
            ):
                restored = run_crop_audit(args)
            self.assertTrue(restored["next_stage_gate"]["training_ready"])
            self.assertTrue(marker_path.is_file())

            with mock.patch(
                "run_ui5_crop_audit.write_excel_report",
                side_effect=RuntimeError("simulated report interruption"),
            ):
                with self.assertRaisesRegex(RuntimeError, "simulated report interruption"):
                    run_crop_audit(args)
            self.assertFalse(marker_path.exists())
            interrupted_summary = json.loads(
                (paths.crop_audit / "summary.json").read_text(encoding="utf-8")
            )
            self.assertFalse(interrupted_summary["training_ready"])
            self.assertFalse(
                interrupted_summary["next_stage_gate"]["conditions"][
                    "all_reports_written_successfully"
                ]
            )

    def test_failed_gate_uses_best_candidate_and_never_touches_detector_shards(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            output = root / "audit"
            paths = AuditPaths(output, "crop_audit_v3_fail")
            image_path = root / "shared.png"
            Image.new("RGB", (100, 100), "white").save(image_path)
            unique = {
                "image_id": "img", "content_id": "content", "image_path": str(image_path),
                "width": 100, "height": 100, "tasks": ["ui_occlusion"],
            }
            failing_sample = {
                **sample("ui_occlusion", [[80, 80, 95, 95]], "img"),
                "canonical_path": str(image_path),
            }
            merged = {
                "image_id": "img", "content_id": "content", "image": str(image_path),
                "width": 100, "height": 100,
                "text_detections": [{"bbox": [5, 5, 15, 15], "score": 0.9}],
                "icon_detections": [],
            }
            atomic_write_jsonl(paths.unique_images, [unique])
            atomic_write_jsonl(paths.task_samples, [failing_sample])
            atomic_write_jsonl(paths.shards / "shard_00000.jsonl", [unique])
            atomic_write_jsonl(paths.merged, [merged])
            atomic_write_json(paths.detector_config, {"fixed": True})
            text_sentinel = paths.stage_dir("text") / "shard_00000.jsonl"
            icon_sentinel = paths.stage_dir("icon") / "shard_00000.jsonl"
            text_sentinel.parent.mkdir(parents=True, exist_ok=True)
            icon_sentinel.parent.mkdir(parents=True, exist_ok=True)
            text_sentinel.write_bytes(b"immutable text detector output\n")
            icon_sentinel.write_bytes(b"immutable icon detector output\n")
            before = (text_sentinel.read_bytes(), icon_sentinel.read_bytes())
            matrix = {task: {other: 0 for other in TASK_NAMES} for task in TASK_NAMES}
            jaccard = {task: {other: 0.0 for other in TASK_NAMES} for task in TASK_NAMES}
            overlap_dataset = {
                "path_overlap": {"counts": matrix, "jaccard": jaccard},
                "content_overlap": {"counts": matrix, "jaccard": jaccard},
                "basename_conflicts": {"details": []},
            }
            atomic_write_json(
                paths.manifest / "overlap" / "source_overlap.json",
                {
                    "source_data": overlap_dataset,
                    "actual_training_data": overlap_dataset,
                    "same_content_cross_train_val": {"count": 0, "details": []},
                },
            )
            args = SimpleNamespace(
                output_dir=output, parser_root=root, max_crops=10,
                boundary_margin_ratio=0.01, progress_interval_seconds=60,
                resume=True, crop_workers=1, crop_audit_name="crop_audit_v3_fail",
                overview_samples_per_task=0, overview_anomalies_per_category=0,
            )
            with mock.patch("run_ui5_crop_audit.load_parser_module", return_value=FakeCropper):
                result = run_crop_audit(args)
            self.assertNotIn("recommended_config", result)
            self.assertIn("best_candidate_config", result)
            self.assertFalse(result["next_stage_gate"]["training_ready"])
            self.assertFalse((paths.crop_audit / "training_ready.json").exists())
            self.assertEqual(result["detector_stages_executed"], [])
            self.assertTrue(result["input_snapshot_unchanged"])
            self.assertEqual(before, (text_sentinel.read_bytes(), icon_sentinel.read_bytes()))


if __name__ == "__main__":
    unittest.main()
