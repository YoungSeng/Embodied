import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from eaglevl.train.cpt_excel import build_cpt_workbook


class CPTExcelTest(unittest.TestCase):
    def test_exactly_three_sheets_and_missing_values_remain_blank(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "cpt_train_metrics.jsonl").write_text(
                json.dumps(
                    {
                        "step": 100,
                        "scope": "lifetime",
                        "task": "vqa",
                        "train_main_token_ce": 1.5,
                    }
                )
                + "\n",
                encoding="utf-8",
            )
            defect_metrics = {
                "per_class": {
                    "cropping": {
                        "display_label": "元素裁切",
                        "image": {
                            "tp": 2,
                            "fp": 1,
                            "fn": 0,
                            "tn": 7,
                            "precision": 2 / 3,
                            "recall": 1.0,
                            "f1": 0.8,
                            "accuracy": 0.9,
                            "images": 10,
                        },
                        "bbox": {
                            "tp": 3,
                            "fp": 1,
                            "fn": 2,
                            "precision": 0.75,
                            "recall": 0.6,
                            "f1": 2 / 3,
                        },
                    }
                },
                "image_macro": {"precision": 2 / 3, "recall": 1.0, "f1": 0.8},
                "bbox_macro": {"precision": 0.75, "recall": 0.6, "f1": 2 / 3},
                "image_micro": {"tp": 2, "fp": 1, "fn": 0, "tn": 7, "f1": 0.8},
                "bbox_micro": {"tp": 3, "fp": 1, "fn": 2, "f1": 2 / 3},
            }
            (root / "cpt_eval_metrics.jsonl").write_text(
                "\n".join(
                    json.dumps(row)
                    for row in (
                        {
                            "checkpoint": "checkpoint-100",
                            "step": 100,
                            "split": "heldout",
                            "task": "vqa",
                            "primary_metric": 0.9,
                            "eval_token_ce": 1.2,
                        },
                        {
                            "checkpoint": "checkpoint-200",
                            "step": 200,
                            "split": "heldout",
                            "task": "vqa",
                            "primary_metric": 0.8,
                            "eval_token_ce": 1.0,
                        },
                        {
                            "checkpoint": "checkpoint-200",
                            "step": 200,
                            "split": "heldout",
                            "task": "ui_defect",
                            "primary_metric": 2 / 3,
                            "metrics": defect_metrics,
                            "base_metrics": defect_metrics,
                        },
                        {
                            "checkpoint": "checkpoint-200",
                            "step": 200,
                            "split": "heldout",
                            "task": "__task_macro__",
                            "primary_metric": 0.8,
                        },
                    )
                )
                + "\n",
                encoding="utf-8",
            )
            output = root / "metrics.xlsx"
            self.assertTrue(build_cpt_workbook(root, output))
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["TrainMetrics", "EvalMetrics", "UIDefectMetrics"],
            )
            train = workbook["TrainMetrics"]
            train_headers = {cell.value: cell.column for cell in train[1]}
            self.assertEqual(train.cell(2, train_headers["task"]).value, "vqa")
            defect = workbook["UIDefectMetrics"]
            defect_headers = {cell.value: cell.column for cell in defect[1]}
            rows = [
                {
                    name: defect.cell(row_index, column).value
                    for name, column in defect_headers.items()
                }
                for row_index in range(2, defect.max_row + 1)
            ]
            cropping_bbox = next(
                row
                for row in rows
                if row["model"] == "checkpoint"
                and row["class"] == "cropping"
                and row["granularity"] == "bbox"
            )
            self.assertAlmostEqual(cropping_bbox["f1"], 2 / 3)
            self.assertIsNone(cropping_bbox["tn"])
            for sheet in workbook.worksheets:
                self.assertEqual(sheet.freeze_panes, "A2")
                self.assertIsNotNone(sheet.auto_filter.ref)

    def test_excel_import_failure_is_only_a_warning(self):
        with tempfile.TemporaryDirectory() as directory:
            original_import = __import__

            def rejecting_import(name, *args, **kwargs):
                if name == "openpyxl" or name.startswith("openpyxl."):
                    raise ImportError("intentional test failure")
                return original_import(name, *args, **kwargs)

            with mock.patch("builtins.__import__", side_effect=rejecting_import):
                with self.assertWarns(UserWarning):
                    self.assertFalse(build_cpt_workbook(directory))


if __name__ == "__main__":
    unittest.main()
