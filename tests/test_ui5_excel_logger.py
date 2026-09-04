from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from eaglevl.train.ui5_excel_logger import (
    EVAL_COLUMNS,
    EXPECTED_SHEETS,
    TRAIN_COLUMNS,
    TRAIN_TASKS,
    UI5ExcelLogger,
    build_eval_rows,
)


def training_metrics(step: int) -> dict:
    return {
        "segment_epoch": step / 1000,
        "global_epoch": step / 1000,
        "gpu_num": 4,
        "max_num_tokens": 12800,
        "learning_rate": 2e-5,
        "base_learning_rate": 1e-5,
        "ui_relation_learning_rate": 2e-5,
        "init_checkpoint": "/models/cpt/checkpoint-3000",
        "init_cpt_step": 3000,
        "sft_step": step,
        "loss_total": 1.0,
        "pbd_delta_norm": 0.25,
        "pbd_active_positions": 6,
        "box_anchor_count": 1,
        "loss_box_l1": 0.4,
        "loss_box_giou": 0.3,
        "loss_attn_kl": 0.2,
        "loss_coverage": 0.1,
        "coarse_iou_mean": 0.5,
        "coarse_recall_03": 0.7,
        "coarse_recall_05": 0.6,
        "matched_slots": 2,
        "unmatched_slots": 6,
        "unique_slot_count": 2,
        "duplicate_slot_rate": 0.0,
        "tasks": {
            task: {
                "detail_weight_l5": 0.2,
                "detail_weight_l15": 0.3,
                "detail_weight_l26": 0.5,
                "scale_entropy": 1.03,
                "scale_batch_std_l5": 0.01,
                "scale_batch_std_l15": 0.02,
                "scale_batch_std_l26": 0.03,
            }
            for task in TRAIN_TASKS
        },
    }


def scorer_metrics(value: float = 0.5) -> dict:
    tasks = {}
    for task in (
        "text_overflow",
        "text_ellipsis",
        "occlusion",
        "cropping",
        "content_missing",
    ):
        tasks[task] = {
            "image": {
                "precision": value,
                "recall": value,
                "f1": value,
                "tp": 2,
                "fp": 1,
                "fn": 1,
                "tn": 10,
                "accuracy": 0.8,
            },
            "bbox": {
                "precision": value,
                "recall": value,
                "f1": value,
                "tp": 2,
                "fp": 1,
                "fn": 1,
                "count_accuracy": 0.7,
            },
        }
    return {
        "tasks": tasks,
        "macro": {
            "image": {"precision": value, "recall": value, "f1": value},
            "bbox": {"precision": value, "recall": value, "f1": value},
        },
    }


class UI5ExcelLoggerTest(unittest.TestCase):
    def test_legacy_noncontiguous_train_header_migrates_by_name(self):
        """Reproduce the production workbook that failed after Gate splitting."""

        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "ui5_training_evaluation.xlsx"
            workbook = Workbook()
            train = workbook.active
            train.title = "train_100steps"
            evaluation = workbook.create_sheet("eval_1000steps")
            # This intentionally is not a prefix of TRAIN_COLUMNS.  In the old
            # schema gate_grad_norm followed relation_grad_norm and represented
            # both Gate branches as one value.
            legacy_header = (
                "step",
                "epoch",
                "text_overflow_detail_weight_l5",
                "relation_grad_norm",
                "gate_grad_norm",
                "pbd_grad_norm",
            )
            train.append(list(legacy_header))
            train.append([100, 0.1, 0.2, 0.5, 0.75, 0.25])
            evaluation.append(list(EVAL_COLUMNS))
            workbook.save(path)
            workbook.close()

            logger = UI5ExcelLogger(path)
            rows = build_eval_rows(
                step=0,
                checkpoint="checkpoint-0",
                metrics=scorer_metrics(0.4),
            )
            self.assertTrue(logger.append_eval(0, rows))

            migrated = load_workbook(path, read_only=True)
            try:
                self.assertEqual(tuple(migrated.sheetnames), EXPECTED_SHEETS)
                train_sheet = migrated["train_100steps"]
                header = tuple(cell.value for cell in train_sheet[1])
                self.assertEqual(header, TRAIN_COLUMNS)
                values = next(train_sheet.iter_rows(min_row=2, values_only=True))
                row = dict(zip(header, values))
                self.assertEqual(row["step"], 100)
                self.assertEqual(row["segment_epoch"], 0.1)
                self.assertEqual(row["global_epoch"], 0.1)
                self.assertEqual(row["gate_grad_norm"], 0.75)
                self.assertIsNone(row["image_gate_grad_norm"])
                self.assertIsNone(row["slot_gate_grad_norm"])
                self.assertEqual(
                    migrated["eval_1000steps"].max_row - 1,
                    14,
                )
            finally:
                migrated.close()

    def test_train_resume_deduplicates_and_keeps_exactly_two_sheets(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "diagnostics" / "ui5_training_evaluation.xlsx"
            logger = UI5ExcelLogger(path)
            self.assertTrue(logger.update_train(100, training_metrics(100)))
            self.assertTrue(logger.update_train(200, training_metrics(200)))
            self.assertFalse(logger.update_train(100, training_metrics(100)))
            self.assertEqual(logger.latest_train_global_epoch(), 0.2)
            workbook = load_workbook(path, read_only=True)
            try:
                steps_at_250 = [
                    row[0]
                    for row in workbook["train_100steps"].iter_rows(
                        min_row=2, values_only=True
                    )
                ]
                self.assertEqual(steps_at_250, [100, 200])
            finally:
                workbook.close()
            # A run ending at 250 writes no partial window; resume adds only 300.
            self.assertTrue(logger.update_train(300, training_metrics(300)))
            workbook = load_workbook(path, read_only=True)
            try:
                self.assertEqual(tuple(workbook.sheetnames), EXPECTED_SHEETS)
                steps = [
                    row[0]
                    for row in workbook["train_100steps"].iter_rows(
                        min_row=2, values_only=True
                    )
                ]
                self.assertEqual(steps, [100, 200, 300])
                self.assertFalse(
                    any(
                        isinstance(cell.value, str) and cell.value.startswith("=")
                        for sheet in workbook.worksheets
                        for row in sheet.iter_rows()
                        for cell in row
                    )
                )
            finally:
                workbook.close()

    def test_retention_metadata_updates_both_existing_sheets(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "ui5_training_evaluation.xlsx"
            logger = UI5ExcelLogger(path)
            logger.update_train(1000, training_metrics(1000))
            rows = build_eval_rows(
                step=1000,
                checkpoint="checkpoint-1000",
                metrics=scorer_metrics(0.5),
                metadata={
                    "init_checkpoint": "/models/cpt/checkpoint-3000",
                    "init_cpt_step": 3000,
                },
            )
            logger.append_eval(1000, rows)
            self.assertTrue(
                logger.update_checkpoint_status(
                    1000,
                    is_best_image=True,
                    is_best_bbox=False,
                    is_4000_milestone=False,
                    checkpoint_kept=True,
                )
            )
            workbook = load_workbook(path, read_only=True)
            try:
                for sheet_name in EXPECTED_SHEETS:
                    sheet = workbook[sheet_name]
                    header = [cell.value for cell in sheet[1]]
                    for values in sheet.iter_rows(min_row=2, values_only=True):
                        row = dict(zip(header, values))
                        self.assertEqual(row["sft_step"], 1000)
                        self.assertTrue(row["is_best_image"])
                        self.assertFalse(row["is_best_bbox"])
                        self.assertTrue(row["checkpoint_kept"])
            finally:
                workbook.close()

    def test_five_task_detail_weights_and_active_pbd_are_written(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "ui5_training_evaluation.xlsx"
            logger = UI5ExcelLogger(path)
            logger.update_train(100, training_metrics(100))
            workbook = load_workbook(path, read_only=True)
            try:
                self.assertEqual(tuple(workbook.sheetnames), EXPECTED_SHEETS)
                sheet = workbook["train_100steps"]
                headers = [cell.value for cell in sheet[1]]
                values = next(sheet.iter_rows(min_row=2, values_only=True))
                row = dict(zip(headers, values))
                for task in TRAIN_TASKS:
                    self.assertIn(f"{task}_lm_loss", headers)
                    self.assertIn(f"{task}_slot_objectness_loss", headers)
                    weights = [
                        row[f"{task}_detail_weight_l5"],
                        row[f"{task}_detail_weight_l15"],
                        row[f"{task}_detail_weight_l26"],
                    ]
                    self.assertAlmostEqual(sum(weights), 1.0, places=7)
                self.assertEqual(row["pbd_active_positions"], 6)
                self.assertEqual(row["box_anchor_count"], 1)
                self.assertAlmostEqual(row["pbd_delta_norm"], 0.25)
                self.assertAlmostEqual(row["loss_box_l1"], 0.4)
                self.assertAlmostEqual(row["coarse_recall_05"], 0.6)
            finally:
                workbook.close()

    def test_eval_has_five_tasks_and_separate_macro_micro_rows_per_step(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "ui5_training_evaluation.xlsx"
            logger = UI5ExcelLogger(path)
            for step, value in ((0, 0.4), (1000, 0.5)):
                rows = build_eval_rows(
                    step=step,
                    checkpoint=f"checkpoint-{step}",
                    metrics=scorer_metrics(value),
                )
                self.assertTrue(logger.append_eval(step, rows))
            self.assertFalse(
                logger.append_eval(
                    1000,
                    build_eval_rows(
                        step=1000,
                        checkpoint="checkpoint-1000",
                        metrics=scorer_metrics(0.5),
                    ),
                )
            )
            workbook = load_workbook(path, read_only=True)
            try:
                rows = list(
                    workbook["eval_1000steps"].iter_rows(
                        min_row=2, values_only=True
                    )
                )
                self.assertEqual(len(rows), 28)
                self.assertEqual(sum(row[0] == 0 for row in rows), 14)
                self.assertEqual(sum(row[0] == 1000 for row in rows), 14)
                f1_change_index = EVAL_COLUMNS.index("f1_change_from_previous")
                task_index = EVAL_COLUMNS.index("task")
                self.assertTrue(
                    all(
                        abs(row[f1_change_index] - 0.1) < 1e-9
                        for row in rows
                        if row[0] == 1000
                        and row[task_index] != "five_task_micro"
                    )
                )
                self.assertTrue(
                    all(
                        abs(row[f1_change_index]) < 1e-9
                        for row in rows
                        if row[0] == 1000
                        and row[task_index] == "five_task_micro"
                    )
                )
                granularity_index = EVAL_COLUMNS.index("granularity")
                tp_index = EVAL_COLUMNS.index("tp")
                for granularity in ("image", "bbox"):
                    macro = next(
                        row
                        for row in rows
                        if row[0] == 1000
                        and row[task_index] == "five_task_macro"
                        and row[granularity_index] == granularity
                    )
                    micro = next(
                        row
                        for row in rows
                        if row[0] == 1000
                        and row[task_index] == "five_task_micro"
                        and row[granularity_index] == granularity
                    )
                    self.assertIsNone(macro[tp_index])
                    self.assertEqual(micro[tp_index], 10)
            finally:
                workbook.close()

    def test_eval_replaces_same_step_when_cache_identity_changes(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "ui5_training_evaluation.xlsx"
            logger = UI5ExcelLogger(path)
            first = build_eval_rows(
                step=0,
                checkpoint="checkpoint-0",
                metrics=scorer_metrics(0.4),
                metadata={
                    "git_commit": "old",
                    "config_hash": "old-config",
                    "model_signature": "old-model",
                },
            )
            second = build_eval_rows(
                step=0,
                checkpoint="checkpoint-0",
                metrics=scorer_metrics(0.5),
                metadata={
                    "git_commit": "new",
                    "config_hash": "new-config",
                    "model_signature": "new-model",
                },
            )
            self.assertTrue(logger.append_eval(0, first))
            self.assertTrue(logger.append_eval(0, second))
            workbook = load_workbook(path, read_only=True)
            try:
                sheet = workbook["eval_1000steps"]
                header = [cell.value for cell in sheet[1]]
                rows = [
                    dict(zip(header, values))
                    for values in sheet.iter_rows(min_row=2, values_only=True)
                ]
                self.assertEqual(len(rows), 14)
                self.assertEqual({row["git_commit"] for row in rows}, {"new"})
                self.assertEqual(
                    {row["model_signature"] for row in rows}, {"new-model"}
                )
            finally:
                workbook.close()

    def test_soft_eval_keeps_separate_observe_raw_metrics(self):
        gate_metrics = {
            task: {
                "relation_gate_mode": "soft",
                "selected_gate_threshold": 0.3,
                "gated_precision": 0.7,
                "gated_recall": 0.6,
                "gated_f1": 0.646,
                "gated_predicted_positive": 3,
                "gate_filter_rate": 0.2,
                "selected_slot_iou": 0.41,
                "oracle_8slot_iou": 0.58,
                "route_top1_match_accuracy": 0.7,
                "pbd_enabled": True,
                "coordinate_bridge_enabled": True,
                "slot_routing_enabled": True,
            }
            for task in (
                "text_overflow",
                "text_ellipsis",
                "occlusion",
                "cropping",
                "content_missing",
            )
        }
        rows = build_eval_rows(
            step=1000,
            checkpoint="checkpoint-1000",
            metrics=scorer_metrics(0.6),
            raw_metrics=scorer_metrics(0.4),
            gate_metrics=gate_metrics,
        )
        image = next(
            row
            for row in rows
            if row["task"] == "text_overflow" and row["granularity"] == "image"
        )
        bbox = next(
            row
            for row in rows
            if row["task"] == "text_overflow" and row["granularity"] == "bbox"
        )
        self.assertEqual(image["raw_f1"], 0.4)
        self.assertEqual(image["soft_f1"], 0.6)
        self.assertEqual(image["selected_slot_iou"], 0.41)
        self.assertEqual(image["oracle_8slot_iou"], 0.58)
        self.assertLessEqual(
            image["selected_slot_iou"], image["oracle_8slot_iou"]
        )
        self.assertTrue(image["pbd_enabled"])
        self.assertTrue(image["coordinate_bridge_enabled"])
        self.assertTrue(image["slot_routing_enabled"])
        self.assertEqual(bbox["raw_f1"], 0.4)
        self.assertEqual(bbox["soft_f1"], 0.6)
        self.assertIsNone(bbox["gated_f1"])
        self.assertIsNone(bbox["diagnostic_upper_bound_f1"])

    def test_failed_atomic_replace_leaves_original_workbook_readable(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "ui5_training_evaluation.xlsx"
            logger = UI5ExcelLogger(path)
            logger.update_train(100, training_metrics(100))
            with mock.patch(
                "eaglevl.train.ui5_excel_logger.os.replace",
                side_effect=OSError("simulated interruption"),
            ):
                with self.assertRaises(OSError):
                    logger.update_train(200, training_metrics(200))
            workbook = load_workbook(path, read_only=True)
            try:
                steps = [
                    row[0]
                    for row in workbook["train_100steps"].iter_rows(
                        min_row=2, values_only=True
                    )
                ]
                self.assertEqual(steps, [100])
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
