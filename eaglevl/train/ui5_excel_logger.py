"""Atomic two-sheet Excel diagnostics for LocateAnything UI5 training/eval."""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
import math
import os
import json
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from eaglevl.ui_task_registry import UI_TASKS, UI5_TASKS, UI9_TASKS, load_registry


LEGACY_TRAIN_TASKS = (
    "text_overflow",
    "text_ellipsis",
    "element_overlap",
    "element_cropping",
    "content_missing",
)

TRAIN_TASKS = (*LEGACY_TRAIN_TASKS, *(t.diagnostic_name for t in UI9_TASKS))

TRAIN_BASE_COLUMNS = (
    "step",
    "task_id",
    "tc_msed_stage",
    "epoch",
    "segment_epoch",
    "global_epoch",
    "gpu_num",
    "max_num_tokens",
    "learning_rate",
    "gate_loss_weight",
    "slot_gate_loss_weight",
    "slot_objectness_loss_weight",
    "attention_loss_weight",
    "gate_threshold",
    "focal_beta",
    "focal_gamma",
    "relation_num_slots",
    "loss_total",
    "loss_total_min",
    "loss_total_max",
    "loss_lm",
    "loss_gate",
    "loss_image_gate",
    "loss_slot_gate",
    "loss_slot_objectness",
    "loss_attention",
    "weighted_gate_loss",
    "weighted_slot_gate_loss",
    "weighted_slot_objectness_loss",
    "weighted_attention_loss",
    "loss_lm_contribution",
    "loss_image_gate_contribution",
    "loss_slot_gate_contribution",
    "loss_attention_contribution",
    "loss_box_l1_contribution",
    "loss_box_giou_contribution",
    "loss_coverage_contribution",
    "loss_coordinate_bridge_contribution",
    "relation_aux_budget_scale",
    "relation_aux_raw_contribution",
    "relation_aux_scaled_contribution",
    "loss_reconstructed",
    "loss_reconstruction_error",
    "attention_active_batch_rate",
    "loss_box_l1",
    "loss_box_giou",
    "loss_attn_kl",
    "loss_coverage",
    "grad_norm",
    "grad_norm_max",
    "samples",
    "positive_samples",
    "negative_samples",
    "peak_gpu_memory_mb",
)
TRAIN_TASK_SUFFIXES = (
    "samples",
    "positive",
    "negative",
    "lm_loss",
    "gate_loss",
    "attention_loss",
    "p_defect_pos",
    "p_defect_neg",
    "gate_precision",
    "gate_recall",
    "gate_f1",
    "gate_pr_auc",
    "slot_gate_loss",
    "slot_objectness_loss",
    "box_l1_loss",
    "box_giou_loss",
    "slot_positive",
    "slot_negative",
    "detail_weight_l5",
    "detail_weight_l15",
    "detail_weight_l26",
    "scale_w_l5",
    "scale_w_l15",
    "scale_w_l26",
    "scale_entropy",
    "scale_batch_std_l5",
    "scale_batch_std_l15",
    "scale_batch_std_l26",
    "coord_prior_lambda",
    "soft_gate_beta",
    "expert_grad_norm",
)
TRAIN_MODULE_COLUMNS = (
    "detail_layer5_norm",
    "detail_layer15_norm",
    "detail_layer26_norm",
    "detail_layer5_abs_max",
    "detail_layer15_abs_max",
    "detail_layer26_abs_max",
    "detail_layer5_saturation_fraction",
    "detail_layer15_saturation_fraction",
    "detail_layer26_saturation_fraction",
    "detail_norm_ratio",
    "detail_fused_norm",
    "relation_context_norm",
    "relation_gate_prob_mean",
    "pbd_delta_norm",
    "pbd_active_positions",
    "coarse_iou_mean",
    "coarse_recall_03",
    "coarse_recall_05",
    "selected_slot_iou",
    "oracle_8slot_iou",
    "route_top1_match_accuracy",
    "pre_mask_route_top1_accuracy",
    "oracle_slot_hit_rate",
    "selected_oracle_iou_ratio",
    "per_slot_usage_histogram",
    "predicted_center_diversity",
    "attention_diversity",
    "matched_slots",
    "unmatched_slots",
    "slot_usage_entropy",
    "box_anchor_count",
    "unique_slot_count",
    "duplicate_slot_rate",
    "pbd_enabled",
    "coordinate_bridge_enabled",
    "slot_routing_enabled",
    "pbd_to_hidden_ratio",
    "pbd_delta_norm_active",
    "relation_grad_norm",
    # Kept as the backwards-compatible aggregate of image/slot Gate gradients.
    # Older workbooks contain this column, while newer runs also expose the two
    # components below.  Retaining it lets name-based schema migration preserve
    # completed evaluations instead of rejecting the legacy header.
    "gate_grad_norm",
    "image_gate_grad_norm",
    "slot_gate_grad_norm",
    "pbd_grad_norm",
    "relation_grad_seen_steps",
    "image_gate_grad_seen_steps",
    "slot_gate_grad_seen_steps",
    "pbd_grad_seen_steps",
    "relation_absolute_update_norm",
    "relation_relative_update_norm",
    "relation_changed_element_count",
    "image_gate_absolute_update_norm",
    "image_gate_relative_update_norm",
    "image_gate_changed_element_count",
    "slot_gate_absolute_update_norm",
    "slot_gate_relative_update_norm",
    "slot_gate_changed_element_count",
    "pbd_absolute_update_norm",
    "pbd_relative_update_norm",
    "pbd_changed_element_count",
    "coarse_box_grad_norm",
    "coord_bridge_grad_norm",
    "coarse_box_grad_seen_steps",
    "coord_bridge_grad_seen_steps",
    "coarse_box_relative_update_norm",
    "coord_bridge_relative_update_norm",
    # Task-book spellings kept as explicit aliases so downstream analysis does
    # not need to know the legacy diagnostic column names above.
    "grad_relation",
    "grad_coarse_box",
    "grad_pbd",
    "grad_coord_bridge",
    "update_ratio_relation",
    "update_ratio_pbd",
    "update_ratio_coord_bridge",
    "cross_task_shared_gradient_cosine",
)
TRAIN_COLUMNS = (
    *TRAIN_BASE_COLUMNS,
    *(
        f"{task}_{suffix}"
        for task in TRAIN_TASKS
        for suffix in TRAIN_TASK_SUFFIXES
    ),
    *TRAIN_MODULE_COLUMNS,
    "crop_train_mode",
    "ui_sampling_mode",
    "eval_inference_crop_mode",
    "scan_name",
    "recipe_digest",
    "code_digest",
    "git_commit",
    "run_name",
    "config_hash",
    "base_learning_rate",
    "ui_relation_learning_rate",
    "task_registry",
    "init_checkpoint",
    "init_cpt_step",
    "sft_step",
    "is_best_image",
    "is_best_bbox",
    "is_4000_milestone",
    "checkpoint_kept",
)

EVAL_COLUMNS = (
    "step",
    "task_id", "task_key", "source_dataset", "source_version", "view_policy", "positive_count", "negative_count",
    "checkpoint",
    "evaluation_split",
    "cache_scope",
    "development_test_reuse",
    "git_sha",
    "git_dirty",
    "recipe_digest",
    "cache_digest",
    "code_digest",
    "crop_train_mode",
    "ui_sampling_mode",
    "eval_inference_crop_mode",
    "scan_name",
    "task",
    "granularity",
    "precision",
    "recall",
    "f1",
    "tp",
    "fp",
    "fn",
    "tn",
    "accuracy",
    "predicted_positive",
    "gate_positive",
    "gate_filtered",
    "p_defect_pos",
    "p_defect_neg",
    "parse_error",
    "f1_change_from_previous",
    "best_f1_so_far",
    "best_step_so_far",
    "raw_precision",
    "raw_recall",
    "raw_f1",
    "raw_predicted_positive",
    "selected_gate_threshold",
    "gated_precision",
    "gated_recall",
    "gated_f1",
    "gated_predicted_positive",
    "gate_filter_rate",
    "bbox_metrics_genuinely_rescored",
    "gate_metric_status",
    "soft_precision",
    "soft_recall",
    "soft_f1",
    "diagnostic_upper_bound_f1",
    "gt_average_box_count",
    "pred_average_box_count",
    "count_mae",
    "coarse_recall_03",
    "coarse_recall_05",
    "selected_slot_iou",
    "oracle_8slot_iou",
    "route_top1_match_accuracy",
    "pre_mask_route_top1_accuracy",
    "oracle_slot_hit_rate",
    "selected_oracle_iou_ratio",
    "per_slot_usage_histogram",
    "predicted_center_diversity",
    "attention_diversity",
    "duplicate_slot_rate",
    "pbd_enabled",
    "coordinate_bridge_enabled",
    "slot_routing_enabled",
    "raw_best_f1_so_far",
    "raw_best_step_so_far",
    "git_commit",
    "run_name",
    "tc_msed_stage",
    "config_hash",
    "model_signature",
    "task_name",
    "defect_type",
    "init_checkpoint",
    "init_cpt_step",
    "sft_step",
    "is_best_image",
    "is_best_bbox",
    "is_4000_milestone",
    "checkpoint_kept",
)

SHEET_TRAIN = "train_100steps"
SHEET_EVAL = "eval_1000steps"
EXPECTED_SHEETS = (SHEET_TRAIN, SHEET_EVAL)


def _openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "UI5 Excel diagnostics require openpyxl>=3.1 in the training environment"
        ) from exc
    return openpyxl, Alignment, Font, PatternFill


def _value(value: Any) -> Any:
    """Convert tensor/numpy scalar-like values without importing those packages."""

    if isinstance(value, float) and not math.isfinite(value):
        return None
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple, dict)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    detach = getattr(value, "detach", None)
    if callable(detach):
        value = detach()
    cpu = getattr(value, "cpu", None)
    if callable(cpu):
        value = cpu()
    item = getattr(value, "item", None)
    if callable(item):
        try:
            scalar = item()
            if isinstance(scalar, float) and not math.isfinite(scalar):
                return None
            return scalar
        except (ValueError, RuntimeError):
            pass
    return value


def _flatten_train_metrics(metrics: Mapping[str, Any]) -> dict[str, Any]:
    row = {key: _value(value) for key, value in metrics.items() if key != "tasks"}
    tasks = metrics.get("tasks", {})
    if isinstance(tasks, Mapping):
        for task, values in tasks.items():
            if task not in TRAIN_TASKS or not isinstance(values, Mapping):
                continue
            for suffix, value in values.items():
                row[f"{task}_{suffix}"] = _value(value)
    return row


def _first_environment_value(*names: str) -> str | None:
    for name in names:
        value = os.environ.get(name)
        if value:
            return value
    return None


def _training_audit_context() -> dict[str, str | None]:
    """Capture the immutable crop recipe/code identity on every train row."""

    recipe_digest = _first_environment_value(
        "UI5_RECIPE_DIGEST",
        "UI5_CROP_RECIPE_DIGEST",
        "RECIPE_DIGEST",
    )
    if recipe_digest is None:
        recipe_path = _first_environment_value("UI5_CROP_META_PATH")
        if recipe_path:
            try:
                digest = hashlib.sha256()
                with Path(recipe_path).expanduser().open("rb") as handle:
                    for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                        digest.update(chunk)
                recipe_digest = digest.hexdigest()
            except OSError:
                # Diagnostics must not abort training solely because an optional
                # audit path disappeared after dataset construction.
                recipe_digest = None
    return {
        "crop_train_mode": _first_environment_value("UI5_CROP_TRAIN_MODE"),
        "ui_sampling_mode": _first_environment_value("UI5_UI_SAMPLING_MODE"),
        "eval_inference_crop_mode": _first_environment_value(
            "EVAL_INFERENCE_CROP_MODE"
        ),
        "scan_name": _first_environment_value("EVAL_SCAN_NAME"),
        "task_registry": Path(os.environ["UI_TASK_REGISTRY"]).read_text(encoding="utf-8") if os.environ.get("UI_TASK_REGISTRY") else None,
        "recipe_digest": recipe_digest,
        "code_digest": _first_environment_value(
            "UI5_CODE_DIGEST", "CODE_DIGEST", "GIT_COMMIT"
        ),
    }


@dataclass
class UI5ExcelLogger:
    """Append-only, resume-safe writer for the requested two diagnostic sheets."""

    path: Path | str
    task_keys: Sequence[str] | None = None

    def __post_init__(self) -> None:
        self.path = Path(self.path).expanduser().resolve()
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def _expected_eval_pairs(self):
        keys = self.task_keys
        if keys is None:
            registry_path = os.environ.get("UI_TASK_REGISTRY")
            keys = [r["task_key"] for r in load_registry(registry_path)] if registry_path else [t.task_key for t in UI5_TASKS]
        names = [t.diagnostic_name for t in UI_TASKS if t.task_key in keys]
        aggregates = ["five_task_macro", "five_task_micro"]
        if any(t.task_key in keys for t in UI9_TASKS): aggregates += ["ui9_macro", "ui9_micro"]
        return {(task, kind) for task in [*names, *aggregates] for kind in ("image", "bbox")}

    def _new_workbook(self):
        openpyxl, Alignment, Font, PatternFill = _openpyxl()
        workbook = openpyxl.Workbook()
        train = workbook.active
        train.title = SHEET_TRAIN
        evaluation = workbook.create_sheet(SHEET_EVAL)
        train.append(list(TRAIN_COLUMNS))
        evaluation.append(list(EVAL_COLUMNS))
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in (train, evaluation):
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.row_dimensions[1].height = 28
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for index, heading in enumerate(sheet[1], start=1):
                sheet.column_dimensions[heading.column_letter].width = min(
                    max(12, len(str(heading.value)) + 2), 30
                )
        return workbook

    def _load_or_create(self):
        openpyxl, _, _, _ = _openpyxl()
        if not self.path.is_file():
            return self._new_workbook()
        workbook = openpyxl.load_workbook(self.path)
        if tuple(workbook.sheetnames) != EXPECTED_SHEETS:
            workbook.close()
            raise ValueError(
                f"Expected exactly sheets {EXPECTED_SHEETS}, found {workbook.sheetnames}"
            )
        for sheet_name, expected in (
            (SHEET_TRAIN, TRAIN_COLUMNS),
            (SHEET_EVAL, EVAL_COLUMNS),
        ):
            sheet = workbook[sheet_name]
            current = tuple(cell.value for cell in sheet[1])
            if current != expected:
                legacy_columns = (
                    {"epoch"}
                    if sheet_name == SHEET_TRAIN
                    else {"inference_crop_mode"}
                )
                if not set(current).issubset(set(expected) | legacy_columns):
                    workbook.close()
                    raise ValueError(
                        f"Cannot migrate {sheet_name} header: current={current}"
                    )
                rows = [
                    dict(zip(current, values))
                    for values in sheet.iter_rows(min_row=2, values_only=True)
                ]
                if sheet_name == SHEET_TRAIN and "epoch" in current:
                    for row in rows:
                        legacy_epoch = row.get("epoch")
                        row.setdefault("segment_epoch", legacy_epoch)
                        row.setdefault("global_epoch", legacy_epoch)
                if sheet_name == SHEET_EVAL and "inference_crop_mode" in current:
                    for row in rows:
                        row.setdefault(
                            "eval_inference_crop_mode",
                            row.get("inference_crop_mode"),
                        )
                sheet.delete_rows(1, sheet.max_row)
                sheet.append(list(expected))
                for row in rows:
                    sheet.append([row.get(column) for column in expected])
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
        return workbook

    def _atomic_save(self, workbook) -> None:
        openpyxl, _, _, _ = _openpyxl()
        temporary = self.path.with_name(
            f".{self.path.stem}.tmp-{os.getpid()}{self.path.suffix}"
        )
        try:
            workbook.save(temporary)
            workbook.close()
            verification = openpyxl.load_workbook(temporary, read_only=True)
            try:
                if tuple(verification.sheetnames) != EXPECTED_SHEETS:
                    raise ValueError(
                        f"Temporary workbook has unexpected sheets: {verification.sheetnames}"
                    )
                if tuple(
                    cell.value for cell in next(verification[SHEET_TRAIN].iter_rows())
                ) != TRAIN_COLUMNS:
                    raise ValueError("Temporary workbook train header verification failed")
                if tuple(
                    cell.value for cell in next(verification[SHEET_EVAL].iter_rows())
                ) != EVAL_COLUMNS:
                    raise ValueError("Temporary workbook eval header verification failed")
            finally:
                verification.close()
            os.replace(temporary, self.path)
        finally:
            try:
                workbook.close()
            except Exception:
                pass
            temporary.unlink(missing_ok=True)

    @staticmethod
    def _steps(sheet) -> set[int]:
        return {
            int(value)
            for (value,) in sheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if value is not None
        }

    def has_train_step(self, step: int) -> bool:
        if not self.path.is_file():
            return False
        workbook = self._load_or_create()
        try:
            return int(step) in self._steps(workbook[SHEET_TRAIN])
        finally:
            workbook.close()

    def has_eval_step(self, step: int) -> bool:
        if not self.path.is_file():
            return False
        workbook = self._load_or_create()
        try:
            sheet = workbook[SHEET_EVAL]
            headers = [cell.value for cell in sheet[1]]
            rows = [
                dict(zip(headers, row))
                for row in workbook[SHEET_EVAL].iter_rows(
                    min_row=2, values_only=True
                )
                if row[0] is not None and int(row[0]) == int(step)
            ]
            expected_pairs = self._expected_eval_pairs()
            return len(rows) == len(expected_pairs) and {
                (row.get("task"), row.get("granularity")) for row in rows
            } == expected_pairs
        finally:
            workbook.close()

    def latest_train_global_epoch(self) -> float:
        """Return the last cumulative epoch recorded before a new segment starts."""
        if not self.path.is_file():
            return 0.0
        workbook = self._load_or_create()
        try:
            sheet = workbook[SHEET_TRAIN]
            headers = [cell.value for cell in sheet[1]]
            rows = [
                dict(zip(headers, values))
                for values in sheet.iter_rows(min_row=2, values_only=True)
            ]
            valid = [
                row
                for row in rows
                if row.get("step") is not None and row.get("global_epoch") is not None
            ]
            if not valid:
                return 0.0
            latest = max(valid, key=lambda row: int(row["step"]))
            return float(latest["global_epoch"])
        finally:
            workbook.close()

    def update_train(self, step: int, window_metrics: Mapping[str, Any]) -> bool:
        step = int(step)
        if step <= 0 or step % 100:
            raise ValueError(f"train_100steps accepts positive multiples of 100, got {step}")
        workbook = self._load_or_create()
        sheet = workbook[SHEET_TRAIN]
        if step in self._steps(sheet):
            workbook.close()
            return False
        values = _flatten_train_metrics(window_metrics)
        for name, value in _training_audit_context().items():
            values.setdefault(name, value)
        values["step"] = step
        sheet.append([values.get(column) for column in TRAIN_COLUMNS])
        sheet.auto_filter.ref = sheet.dimensions
        self._atomic_save(workbook)
        return True

    @staticmethod
    def _decorate_eval_rows(existing_rows: list[dict[str, Any]], rows: list[dict[str, Any]]) -> None:
        for row in rows:
            history = [
                old
                for old in existing_rows
                if old.get("task") == row.get("task")
                and old.get("granularity") == row.get("granularity")
                and int(old.get("step", -1)) < int(row["step"])
                and old.get("f1") is not None
            ]
            previous = max(history, key=lambda item: int(item["step"]), default=None)
            row["f1_change_from_previous"] = (
                float(row["f1"]) - float(previous["f1"])
                if previous is not None and row.get("f1") is not None
                else None
            )
            candidates = [*history, row]
            valid = [item for item in candidates if item.get("f1") is not None]
            best = max(valid, key=lambda item: float(item["f1"]), default=None)
            row["best_f1_so_far"] = best.get("f1") if best else None
            row["best_step_so_far"] = best.get("step") if best else None
            raw_history = [item for item in history if item.get("raw_f1") is not None]
            raw_candidates = [*raw_history, row]
            raw_valid = [item for item in raw_candidates if item.get("raw_f1") is not None]
            raw_best = max(raw_valid, key=lambda item: float(item["raw_f1"]), default=None)
            row["raw_best_f1_so_far"] = raw_best.get("raw_f1") if raw_best else None
            row["raw_best_step_so_far"] = raw_best.get("step") if raw_best else None

    def append_eval(self, step: int, task_metrics: Sequence[Mapping[str, Any]]) -> bool:
        step = int(step)
        rows = [dict(row) for row in task_metrics]
        for row in rows:
            row["step"] = step
        expected_pairs = self._expected_eval_pairs()
        actual_pairs = {(row.get("task"), row.get("granularity")) for row in rows}
        if len(rows) != len(expected_pairs) or actual_pairs != expected_pairs:
            raise ValueError(
                "eval_1000steps requires the complete registered task set plus its group macro/micro, each at image/bbox; "
                f"found={sorted(actual_pairs)}"
            )

        workbook = self._load_or_create()
        sheet = workbook[SHEET_EVAL]
        headers = [cell.value for cell in sheet[1]]
        existing = [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]
        existing_for_step = [
            row for row in existing if int(row.get("step", -1)) == step
        ]
        identity_columns = (
            "git_sha",
            "git_commit",
            "config_hash",
            "model_signature",
            "evaluation_split",
            "cache_scope",
            "recipe_digest",
            "cache_digest",
            "code_digest",
            "crop_train_mode",
            "ui_sampling_mode",
            "eval_inference_crop_mode",
            "scan_name",
        )
        new_identity = tuple(rows[0].get(column) for column in identity_columns)
        if any(
            tuple(row.get(column) for column in identity_columns) != new_identity
            for row in rows[1:]
        ):
            workbook.close()
            raise ValueError(
                "All eval rows for one step must share the same audit identity"
            )
        if len(existing_for_step) == len(expected_pairs):
            existing_pairs = {
                (row.get("task"), row.get("granularity"))
                for row in existing_for_step
            }
            if existing_pairs == expected_pairs and all(
                tuple(row.get(column) for column in identity_columns) == new_identity
                for row in existing_for_step
            ):
                workbook.close()
                return False

        retained = [row for row in existing if int(row.get("step", -1)) != step]
        if len(retained) != len(existing):
            sheet.delete_rows(2, sheet.max_row - 1)
            for row in retained:
                sheet.append([row.get(column) for column in EVAL_COLUMNS])

        self._decorate_eval_rows(retained, rows)
        for row in rows:
            sheet.append([_value(row.get(column)) for column in EVAL_COLUMNS])
        sheet.auto_filter.ref = sheet.dimensions
        self._atomic_save(workbook)
        return True

    def update_checkpoint_status(
        self,
        step: int,
        *,
        is_best_image: bool,
        is_best_bbox: bool,
        is_4000_milestone: bool,
        checkpoint_kept: bool,
    ) -> bool:
        """Apply scorer-derived retention flags to both sheets for one SFT step."""

        if not self.path.is_file():
            return False
        step = int(step)
        updates = {
            "sft_step": step,
            "is_best_image": bool(is_best_image),
            "is_best_bbox": bool(is_best_bbox),
            "is_4000_milestone": bool(is_4000_milestone),
            "checkpoint_kept": bool(checkpoint_kept),
        }
        workbook = self._load_or_create()
        changed = False
        for sheet_name in EXPECTED_SHEETS:
            sheet = workbook[sheet_name]
            headers = {
                cell.value: cell.column for cell in sheet[1] if cell.value is not None
            }
            for row_index in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row_index, column=headers["step"]).value
                if value is None or int(value) != step:
                    continue
                for name, new_value in updates.items():
                    cell = sheet.cell(row=row_index, column=headers[name])
                    if cell.value != new_value:
                        cell.value = new_value
                        changed = True
        if changed:
            self._atomic_save(workbook)
        else:
            workbook.close()
        return changed


def _build_group_eval_rows(
    *,
    step: int,
    checkpoint: str,
    metrics: Mapping[str, Any],
    gate_metrics: Mapping[str, Mapping[str, Any]] | None = None,
    raw_metrics: Mapping[str, Any] | None = None,
    metadata: Mapping[str, Any] | None = None,
    audit_context: Mapping[str, Any] | None = None,
    task_specs=UI5_TASKS,
    group_prefix="five_task",
) -> list[dict[str, Any]]:
    """Convert scorer and Gate summaries into task, macro, and micro rows.

    Macro rows contain only task-averaged metrics.  Micro rows contain summed
    confusion counts and metrics recomputed from those counts.  BBox Gate
    metrics never fall back to image-Gate metrics.
    """

    scorer_to_diagnostic = {t.task_key: t.diagnostic_name for t in task_specs}
    gate_metrics = gate_metrics or {}
    metadata = dict(metadata or {})
    if raw_metrics is None:
        raw_metrics = metrics
    audit_context = dict(audit_context or {})
    context_columns = {
        key: audit_context.get(key)
        for key in (
            "evaluation_split",
            "cache_scope",
            "development_test_reuse",
            "git_sha",
            "git_dirty",
            "recipe_digest",
            "cache_digest",
            "code_digest",
            "crop_train_mode",
            "ui_sampling_mode",
        )
    }
    context_columns["eval_inference_crop_mode"] = audit_context.get(
        "eval_inference_crop_mode", audit_context.get("inference_crop_mode")
    )
    context_columns["scan_name"] = audit_context.get(
        "scan_name", audit_context.get("eval_scan_name")
    )

    def mean(values: Sequence[Any]) -> float | None:
        numeric = [float(value) for value in values if value is not None]
        return sum(numeric) / len(numeric) if numeric else None

    def sum_present(source_rows: Sequence[Mapping[str, Any]], name: str) -> int | None:
        values = [row.get(name) for row in source_rows]
        present = [int(value) for value in values if value is not None]
        return sum(present) if present else None

    def metrics_from_counts(
        *,
        tp: int,
        fp: int,
        fn: int,
        tn: int | None,
        granularity: str,
    ) -> dict[str, Any]:
        precision = tp / (tp + fp) if tp + fp else 0.0
        recall = tp / (tp + fn) if tp + fn else 0.0
        f1 = (
            2.0 * precision * recall / (precision + recall)
            if precision + recall
            else 0.0
        )
        if granularity == "image":
            denominator = tp + fp + fn + int(tn or 0)
            accuracy = (tp + int(tn or 0)) / denominator if denominator else 0.0
        else:
            denominator = tp + fp + fn
            accuracy = tp / denominator if denominator else 0.0
        return {
            "precision": precision,
            "recall": recall,
            "f1": f1,
            "accuracy": accuracy,
        }

    def aggregate_gate_status(source_rows: Sequence[Mapping[str, Any]]) -> str:
        statuses = {str(row.get("gate_metric_status")) for row in source_rows}
        if "not_rescored" in statuses:
            return "not_rescored"
        if len(statuses) == 1:
            return next(iter(statuses))
        return "mixed"

    def aggregate_diagnostics(
        source_rows: Sequence[Mapping[str, Any]],
    ) -> dict[str, Any]:
        mean_columns = (
            "soft_precision",
            "soft_recall",
            "soft_f1",
            "diagnostic_upper_bound_f1",
            "gt_average_box_count",
            "pred_average_box_count",
            "count_mae",
            "coarse_recall_03",
            "coarse_recall_05",
            "selected_slot_iou",
            "oracle_8slot_iou",
            "route_top1_match_accuracy",
            "pre_mask_route_top1_accuracy",
            "oracle_slot_hit_rate",
            "selected_oracle_iou_ratio",
            "predicted_center_diversity",
            "attention_diversity",
            "duplicate_slot_rate",
        )
        result = {
            name: mean([row.get(name) for row in source_rows])
            for name in mean_columns
        }
        histograms = [
            row.get("per_slot_usage_histogram")
            for row in source_rows
            if isinstance(row.get("per_slot_usage_histogram"), (list, tuple))
        ]
        result["per_slot_usage_histogram"] = (
            [
                sum(
                    int(histogram[slot]) if slot < len(histogram) else 0
                    for histogram in histograms
                )
                for slot in range(8)
            ]
            if histograms
            else None
        )
        for name in (
            "pbd_enabled",
            "coordinate_bridge_enabled",
            "slot_routing_enabled",
        ):
            result[name] = next(
                (row.get(name) for row in source_rows if row.get(name) is not None),
                None,
            )
        return result

    row_metadata = {
        "git_commit": metadata.get("git_commit"),
        "run_name": metadata.get("run_name"),
        "tc_msed_stage": metadata.get("tc_msed_stage"),
        "config_hash": metadata.get("config_hash"),
        "model_signature": metadata.get("model_signature"),
        "init_checkpoint": metadata.get("init_checkpoint"),
        "init_cpt_step": metadata.get("init_cpt_step"),
        "sft_step": step,
        "is_best_image": metadata.get("is_best_image", False),
        "is_best_bbox": metadata.get("is_best_bbox", False),
        "is_4000_milestone": metadata.get("is_4000_milestone", False),
        "checkpoint_kept": metadata.get("checkpoint_kept", False),
    }

    rows: list[dict[str, Any]] = []
    for scorer_task, diagnostic_task in scorer_to_diagnostic.items():
        task_values = metrics.get("tasks", {}).get(scorer_task, {})
        gate = gate_metrics.get(scorer_task, {})
        for granularity in ("image", "bbox"):
            values = task_values.get(granularity, {})
            raw_values = (
                raw_metrics.get("tasks", {})
                .get(scorer_task, {})
                .get(granularity, {})
            )
            rescored_by_granularity = gate.get(
                "gated_metrics_by_granularity", {}
            )
            rescored_candidate = (
                rescored_by_granularity.get(granularity, {})
                if isinstance(rescored_by_granularity, Mapping)
                else {}
            )
            rescored = (
                rescored_candidate
                if isinstance(rescored_candidate, Mapping)
                and any(
                    rescored_candidate.get(name) is not None
                    for name in (
                        "precision",
                        "recall",
                        "f1",
                        "tp",
                        "fp",
                        "fn",
                        "tn",
                        "predicted_positive",
                    )
                )
                and not (
                    granularity == "bbox"
                    and gate.get("bbox_metrics_genuinely_rescored") is False
                )
                else {}
            )
            if rescored:
                gated = dict(rescored)
                gate_metric_status = "genuinely_rescored"
            elif granularity == "image" and any(
                gate.get(name) is not None
                for name in (
                    "gated_precision",
                    "gated_recall",
                    "gated_f1",
                    "gated_tp",
                    "gated_fp",
                    "gated_fn",
                    "gated_tn",
                )
            ):
                # The image-Gate sidecar sweep is a real image-level rescore.
                # It is not a localization scorer and must never populate BBox.
                gated = {
                    "precision": gate.get("gated_precision"),
                    "recall": gate.get("gated_recall"),
                    "f1": gate.get("gated_f1"),
                    "tp": gate.get("gated_tp"),
                    "fp": gate.get("gated_fp"),
                    "fn": gate.get("gated_fn"),
                    "tn": gate.get("gated_tn"),
                    "predicted_positive": gate.get("gated_predicted_positive"),
                }
                gate_metric_status = "image_gate_sidecar_rescored"
            else:
                gated = {}
                gate_metric_status = "not_rescored"
            tp = values.get("tp")
            fp = values.get("fp")
            primary_predicted_positive = (
                int(tp) + int(fp)
                if tp is not None and fp is not None
                else None
            )
            raw_tp = raw_values.get("tp")
            raw_fp = raw_values.get("fp")
            raw_predicted_positive = (
                int(raw_tp) + int(raw_fp)
                if raw_tp is not None and raw_fp is not None
                else None
            )
            gated_tp = gated.get("tp")
            gated_fp = gated.get("fp")
            gated_fn = gated.get("fn")
            gated_tn = gated.get("tn") if granularity == "image" else None
            gated_predicted_positive = (
                int(gated_tp) + int(gated_fp)
                if gated_tp is not None and gated_fp is not None
                else gated.get("predicted_positive")
            )
            rows.append(
                {
                    **context_columns,
                    "step": step,
                    "git_commit": metadata.get("git_commit"),
                    "run_name": metadata.get("run_name"),
                    "tc_msed_stage": metadata.get("tc_msed_stage"),
                    "config_hash": metadata.get("config_hash"),
                    "model_signature": metadata.get("model_signature"),
                    "init_checkpoint": metadata.get("init_checkpoint"),
                    "init_cpt_step": metadata.get("init_cpt_step"),
                    "sft_step": step,
                    "is_best_image": metadata.get("is_best_image", False),
                    "is_best_bbox": metadata.get("is_best_bbox", False),
                    "is_4000_milestone": metadata.get(
                        "is_4000_milestone", False
                    ),
                    "checkpoint_kept": metadata.get("checkpoint_kept", False),
                    "checkpoint": checkpoint,
                    "task": diagnostic_task,
                    "task_name": diagnostic_task,
                    "defect_type": next(t.task_id for t in task_specs if t.diagnostic_name == diagnostic_task),
                    "task_id": next(t.task_id for t in task_specs if t.diagnostic_name == diagnostic_task),
                    "task_key": scorer_task,
                    "source_dataset": task_values.get("source_dataset", next(t.source_dataset for t in task_specs if t.task_key == scorer_task)),
                    "source_version": task_values.get("source_version"),
                    "view_policy": next(t.view_policy for t in task_specs if t.task_key == scorer_task),
                    "positive_count": task_values.get("positive_count", gate.get("positive_count")),
                    "negative_count": task_values.get("negative_count", gate.get("negative_count")),
                    "granularity": granularity,
                    "precision": values.get("precision"),
                    "recall": values.get("recall"),
                    "f1": values.get("f1"),
                    "tp": tp,
                    "fp": fp,
                    "fn": values.get("fn"),
                    "tn": values.get("tn") if granularity == "image" else None,
                    "accuracy": values.get(
                        "accuracy" if granularity == "image" else "count_accuracy"
                    ),
                    "predicted_positive": primary_predicted_positive,
                    "gate_positive": gate.get("gate_positive"),
                    "gate_filtered": gate.get("gate_filtered"),
                    "p_defect_pos": gate.get("p_defect_pos"),
                    "p_defect_neg": gate.get("p_defect_neg"),
                    "parse_error": gate.get("parse_error"),
                    "raw_precision": raw_values.get("precision"),
                    "raw_recall": raw_values.get("recall"),
                    "raw_f1": raw_values.get("f1"),
                    "raw_predicted_positive": raw_predicted_positive,
                    "selected_gate_threshold": (
                        gate.get("selected_gate_threshold") if gated else None
                    ),
                    "gated_precision": gated.get("precision"),
                    "gated_recall": gated.get("recall"),
                    "gated_f1": gated.get("f1"),
                    "gated_predicted_positive": gated_predicted_positive,
                    "gate_filter_rate": (
                        1.0
                        - int(gated_predicted_positive)
                        / max(1, int(raw_predicted_positive))
                        if gated_predicted_positive is not None
                        and raw_predicted_positive is not None
                        else None
                    ),
                    "bbox_metrics_genuinely_rescored": (
                        bool(rescored) if granularity == "bbox" else None
                    ),
                    "gate_metric_status": gate_metric_status,
                    "soft_precision": (
                        values.get("precision") if gate.get("relation_gate_mode") == "soft" else None
                    ),
                    "soft_recall": (
                        values.get("recall") if gate.get("relation_gate_mode") == "soft" else None
                    ),
                    "soft_f1": (
                        values.get("f1") if gate.get("relation_gate_mode") == "soft" else None
                    ),
                    "diagnostic_upper_bound_f1": (
                        gate.get("gated_f1") if granularity == "image" else None
                    ),
                    "gt_average_box_count": gate.get("gt_average_box_count"),
                    "pred_average_box_count": gate.get("pred_average_box_count"),
                    "count_mae": gate.get("count_mae"),
                    "coarse_recall_03": gate.get("coarse_recall_03"),
                    "coarse_recall_05": gate.get("coarse_recall_05"),
                    "selected_slot_iou": gate.get("selected_slot_iou"),
                    "oracle_8slot_iou": gate.get("oracle_8slot_iou"),
                    "route_top1_match_accuracy": gate.get("route_top1_match_accuracy"),
                    "pre_mask_route_top1_accuracy": gate.get("pre_mask_route_top1_accuracy"),
                    "oracle_slot_hit_rate": gate.get("oracle_slot_hit_rate"),
                    "selected_oracle_iou_ratio": gate.get("selected_oracle_iou_ratio"),
                    "per_slot_usage_histogram": gate.get("per_slot_usage_histogram"),
                    "predicted_center_diversity": gate.get("predicted_center_diversity"),
                    "attention_diversity": gate.get("attention_diversity"),
                    "duplicate_slot_rate": gate.get("duplicate_slot_rate"),
                    "pbd_enabled": gate.get("pbd_enabled"),
                    "coordinate_bridge_enabled": gate.get("coordinate_bridge_enabled"),
                    "slot_routing_enabled": gate.get("slot_routing_enabled"),
                    "_primary_tp": values.get("tp"),
                    "_primary_fp": values.get("fp"),
                    "_primary_fn": values.get("fn"),
                    "_primary_tn": values.get("tn") if granularity == "image" else None,
                    "_raw_tp": raw_values.get("tp"),
                    "_raw_fp": raw_values.get("fp"),
                    "_raw_fn": raw_values.get("fn"),
                    "_raw_tn": raw_values.get("tn") if granularity == "image" else None,
                    "_gated_tp": gated_tp,
                    "_gated_fp": gated_fp,
                    "_gated_fn": gated_fn,
                    "_gated_tn": gated_tn,
                }
            )

    for granularity in ("image", "bbox"):
        values = metrics.get("macro", {}).get(granularity, {})
        source_rows = [
            row
            for row in rows
            if row["granularity"] == granularity
        ]
        positive_count = sum(
            int(gate_metrics.get(task, {}).get("positive_count", 0))
            for task in scorer_to_diagnostic
        )
        negative_count = sum(
            int(gate_metrics.get(task, {}).get("negative_count", 0))
            for task in scorer_to_diagnostic
        )
        p_defect_pos = (
            sum(
                float(gate_metrics.get(task, {}).get("p_defect_pos") or 0.0)
                * int(gate_metrics.get(task, {}).get("positive_count", 0))
                for task in scorer_to_diagnostic
            )
            / positive_count
            if positive_count
            else None
        )
        p_defect_neg = (
            sum(
                float(gate_metrics.get(task, {}).get("p_defect_neg") or 0.0)
                * int(gate_metrics.get(task, {}).get("negative_count", 0))
                for task in scorer_to_diagnostic
            )
            / negative_count
            if negative_count
            else None
        )
        all_gated = all(row.get("gated_f1") is not None for row in source_rows)
        macro_row = {
            **context_columns,
            **row_metadata,
            "step": step,
            "checkpoint": checkpoint,
            "task": f"{group_prefix}_macro",
            "task_name": f"{group_prefix}_macro",
            "defect_type": None,
            "granularity": granularity,
            "precision": values.get("precision", mean([row.get("precision") for row in source_rows])),
            "recall": values.get("recall", mean([row.get("recall") for row in source_rows])),
            "f1": values.get("f1", mean([row.get("f1") for row in source_rows])),
            # Macro is a mean-of-tasks row.  Counts intentionally remain empty.
            "tp": None,
            "fp": None,
            "fn": None,
            "tn": None,
            "accuracy": mean([row.get("accuracy") for row in source_rows]),
            "predicted_positive": None,
            "gate_positive": None,
            "gate_filtered": None,
            "p_defect_pos": p_defect_pos,
            "p_defect_neg": p_defect_neg,
            "parse_error": None,
            "raw_precision": mean([row.get("raw_precision") for row in source_rows]),
            "raw_recall": mean([row.get("raw_recall") for row in source_rows]),
            "raw_f1": mean([row.get("raw_f1") for row in source_rows]),
            "raw_predicted_positive": None,
            "selected_gate_threshold": (
                mean([row.get("selected_gate_threshold") for row in source_rows])
                if all_gated
                else None
            ),
            "gated_precision": mean([row.get("gated_precision") for row in source_rows]) if all_gated else None,
            "gated_recall": mean([row.get("gated_recall") for row in source_rows]) if all_gated else None,
            "gated_f1": mean([row.get("gated_f1") for row in source_rows]) if all_gated else None,
            "gated_predicted_positive": None,
            "gate_filter_rate": mean([row.get("gate_filter_rate") for row in source_rows]) if all_gated else None,
            "bbox_metrics_genuinely_rescored": (
                all(bool(row.get("bbox_metrics_genuinely_rescored")) for row in source_rows)
                if granularity == "bbox"
                else None
            ),
            "gate_metric_status": aggregate_gate_status(source_rows),
            **aggregate_diagnostics(source_rows),
        }
        rows.append(macro_row)

        primary_tp = int(sum_present(source_rows, "_primary_tp") or 0)
        primary_fp = int(sum_present(source_rows, "_primary_fp") or 0)
        primary_fn = int(sum_present(source_rows, "_primary_fn") or 0)
        primary_tn = (
            int(sum_present(source_rows, "_primary_tn") or 0)
            if granularity == "image"
            else None
        )
        primary_micro = metrics_from_counts(
            tp=primary_tp,
            fp=primary_fp,
            fn=primary_fn,
            tn=primary_tn,
            granularity=granularity,
        )
        raw_tp = int(sum_present(source_rows, "_raw_tp") or 0)
        raw_fp = int(sum_present(source_rows, "_raw_fp") or 0)
        raw_fn = int(sum_present(source_rows, "_raw_fn") or 0)
        raw_tn = (
            int(sum_present(source_rows, "_raw_tn") or 0)
            if granularity == "image"
            else None
        )
        raw_micro = metrics_from_counts(
            tp=raw_tp, fp=raw_fp, fn=raw_fn, tn=raw_tn, granularity=granularity
        )
        all_gated_counts = all(
            row.get("_gated_tp") is not None
            and row.get("_gated_fp") is not None
            and row.get("_gated_fn") is not None
            for row in source_rows
        )
        gated_micro: dict[str, Any] | None = None
        gated_tp = gated_fp = gated_fn = gated_tn = None
        if all_gated_counts:
            gated_tp = int(sum_present(source_rows, "_gated_tp") or 0)
            gated_fp = int(sum_present(source_rows, "_gated_fp") or 0)
            gated_fn = int(sum_present(source_rows, "_gated_fn") or 0)
            gated_tn = (
                int(sum_present(source_rows, "_gated_tn") or 0)
                if granularity == "image"
                else None
            )
            gated_micro = metrics_from_counts(
                tp=gated_tp,
                fp=gated_fp,
                fn=gated_fn,
                tn=gated_tn,
                granularity=granularity,
            )
        rows.append(
            {
                **context_columns,
                **row_metadata,
                "step": step,
                "checkpoint": checkpoint,
                "task": f"{group_prefix}_micro",
                "task_name": f"{group_prefix}_micro",
                "defect_type": None,
                "granularity": granularity,
                **primary_micro,
                "tp": primary_tp,
                "fp": primary_fp,
                "fn": primary_fn,
                "tn": primary_tn,
                "predicted_positive": primary_tp + primary_fp,
                "gate_positive": sum(
                    int(gate_metrics.get(task, {}).get("gate_positive", 0))
                    for task in scorer_to_diagnostic
                ),
                "gate_filtered": sum(
                    int(gate_metrics.get(task, {}).get("gate_filtered", 0))
                    for task in scorer_to_diagnostic
                ),
                "p_defect_pos": p_defect_pos,
                "p_defect_neg": p_defect_neg,
                "parse_error": sum(
                    int(gate_metrics.get(task, {}).get("parse_error", 0))
                    for task in scorer_to_diagnostic
                ),
                "raw_precision": raw_micro["precision"],
                "raw_recall": raw_micro["recall"],
                "raw_f1": raw_micro["f1"],
                "raw_predicted_positive": raw_tp + raw_fp,
                "selected_gate_threshold": mean([row.get("selected_gate_threshold") for row in source_rows]),
                "gated_precision": gated_micro["precision"] if gated_micro else None,
                "gated_recall": gated_micro["recall"] if gated_micro else None,
                "gated_f1": gated_micro["f1"] if gated_micro else None,
                "gated_predicted_positive": (
                    int(gated_tp) + int(gated_fp) if gated_micro else None
                ),
                "gate_filter_rate": (
                    1.0
                    - (int(gated_tp) + int(gated_fp))
                    / max(1, raw_tp + raw_fp)
                    if gated_micro
                    else None
                ),
                "bbox_metrics_genuinely_rescored": (
                    all(
                        bool(row.get("bbox_metrics_genuinely_rescored"))
                        for row in source_rows
                    )
                    if granularity == "bbox"
                    else None
                ),
                "gate_metric_status": aggregate_gate_status(source_rows),
                "soft_precision": (
                    sum(float(row.get("soft_precision") or 0.0) for row in source_rows)
                    / len(source_rows)
                    if any(row.get("soft_precision") is not None for row in source_rows)
                    else None
                ),
                "soft_recall": (
                    sum(float(row.get("soft_recall") or 0.0) for row in source_rows)
                    / len(source_rows)
                    if any(row.get("soft_recall") is not None for row in source_rows)
                    else None
                ),
                "soft_f1": (
                    sum(float(row.get("soft_f1") or 0.0) for row in source_rows)
                    / len(source_rows)
                    if any(row.get("soft_f1") is not None for row in source_rows)
                    else None
                ),
                "diagnostic_upper_bound_f1": (
                    sum(float(row.get("diagnostic_upper_bound_f1") or 0.0) for row in source_rows)
                    / len(source_rows)
                    if any(row.get("diagnostic_upper_bound_f1") is not None for row in source_rows)
                    else None
                ),
                "gt_average_box_count": (
                    sum(float(row.get("gt_average_box_count") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "pred_average_box_count": (
                    sum(float(row.get("pred_average_box_count") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "count_mae": (
                    sum(float(row.get("count_mae") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "coarse_recall_03": (
                    sum(float(row.get("coarse_recall_03") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "coarse_recall_05": (
                    sum(float(row.get("coarse_recall_05") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "selected_slot_iou": (
                    sum(float(row.get("selected_slot_iou") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "oracle_8slot_iou": (
                    sum(float(row.get("oracle_8slot_iou") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "route_top1_match_accuracy": (
                    sum(
                        float(row.get("route_top1_match_accuracy") or 0.0)
                        for row in source_rows
                    )
                    / len(source_rows)
                ),
                "pre_mask_route_top1_accuracy": (
                    sum(float(row.get("pre_mask_route_top1_accuracy") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "oracle_slot_hit_rate": (
                    sum(float(row.get("oracle_slot_hit_rate") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "selected_oracle_iou_ratio": (
                    sum(float(row.get("selected_oracle_iou_ratio") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "per_slot_usage_histogram": [
                    sum(
                        int((row.get("per_slot_usage_histogram") or [0] * 8)[slot])
                        if slot < len(row.get("per_slot_usage_histogram") or []) else 0
                        for row in source_rows
                    )
                    for slot in range(8)
                ],
                "predicted_center_diversity": (
                    sum(float(row.get("predicted_center_diversity") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "attention_diversity": (
                    sum(float(row.get("attention_diversity") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "duplicate_slot_rate": (
                    sum(float(row.get("duplicate_slot_rate") or 0.0) for row in source_rows)
                    / len(source_rows)
                ),
                "pbd_enabled": source_rows[0].get("pbd_enabled"),
                "coordinate_bridge_enabled": source_rows[0].get(
                    "coordinate_bridge_enabled"
                ),
                "slot_routing_enabled": source_rows[0].get(
                    "slot_routing_enabled"
                ),
            }
        )
    return rows


def build_eval_rows(*, step, checkpoint, metrics, gate_metrics=None, raw_metrics=None, metadata=None, audit_context=None):
    common = dict(step=step, checkpoint=checkpoint, gate_metrics=gate_metrics, metadata=metadata, audit_context=audit_context)
    ui5 = _build_group_eval_rows(metrics=metrics, raw_metrics=raw_metrics, **common)
    new_keys = {t.task_key for t in UI9_TASKS}
    if new_keys & set(metrics.get("tasks", {})):
        if not new_keys.issubset(metrics["tasks"]): raise ValueError("Incomplete UI9 evaluation")
        subset = {"tasks": {k: metrics["tasks"][k] for k in new_keys}}
        raw_subset = {"tasks": {k: raw_metrics["tasks"][k] for k in new_keys}} if raw_metrics else subset
        ui5 += _build_group_eval_rows(metrics=subset, raw_metrics=raw_subset, task_specs=UI9_TASKS, group_prefix="ui9", **common)
    return ui5
