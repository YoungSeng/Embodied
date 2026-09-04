"""UI5 curriculum evaluation metrics, best checkpoints, and Excel artifacts.

This module deliberately has no dependency on the training or inference entry
points.  The scheduler calls :func:`update_curriculum_artifacts` after the five
task scorer has completed.  ``checkpoints.json`` is the authoritative state;
the workbook is rebuilt from that state on every update so an interrupted Excel
write can be repaired by safely re-running the same step.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
import json
import math
import os
from pathlib import Path
import shutil
import uuid
from typing import Any


TASK_ALIASES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("ui_occlusion", ("ui_occlusion", "occlusion", "element_overlap")),
    ("ui_cropping", ("ui_cropping", "cropping", "element_cropping")),
    (
        "ui_text_overflow",
        ("ui_text_overflow", "text_overflow"),
    ),
    (
        "ui_text_ellipsis",
        ("ui_text_ellipsis", "text_ellipsis"),
    ),
    (
        "ui_content_missing",
        ("ui_content_missing", "content_missing"),
    ),
)
TASKS = tuple(task for task, _ in TASK_ALIASES)
GRANULARITIES = ("image", "bbox")

SHEET_TRAIN_CURVE = "train_curve"
SHEET_UI5_OVERALL = "ui5_overall"
SHEET_UI5_BY_TASK = "ui5_by_task"
SHEET_HARD_TRANSITION = "hard_transition"
SHEET_ANCHOR_RETENTION = "anchor_retention"
SHEET_CHECKPOINTS = "checkpoints"
SHEET_ORDER = (
    SHEET_TRAIN_CURVE,
    SHEET_UI5_OVERALL,
    SHEET_UI5_BY_TASK,
    SHEET_HARD_TRANSITION,
    SHEET_ANCHOR_RETENTION,
    SHEET_CHECKPOINTS,
)

TRAIN_CURVE_COLUMNS = (
    "step",
    "phase",
    "learning_rate",
    "loss_total",
    "loss_lm",
    "hard_ratio",
    "anchor_ratio",
    "global_replay_ratio",
    "hard_samples",
    "anchor_samples",
    "global_replay_samples",
)
UI5_OVERALL_COLUMNS = (
    "step",
    "image_macro_precision",
    "image_macro_recall",
    "image_macro_f1",
    "image_micro_precision",
    "image_micro_recall",
    "image_micro_f1",
    "bbox_macro_precision",
    "bbox_macro_recall",
    "bbox_macro_f1",
    "bbox_micro_precision",
    "bbox_micro_recall",
    "bbox_micro_f1",
    "joint_score",
    "candidate_checkpoint",
    "evaluation_seconds",
)
UI5_BY_TASK_COLUMNS = (
    "step",
    "task",
    "granularity",
    "precision",
    "recall",
    "f1",
    "tp",
    "fp",
    "fn",
    "tn",
    "accuracy",
    "predicted_positive",
    "candidate_checkpoint",
)
HARD_TRANSITION_COLUMNS = (
    "step",
    "task",
    "group",
    "rollout_id",
    "samples",
    "correct",
    "accuracy",
    "transition",
)
ANCHOR_RETENTION_COLUMNS = (
    "step",
    "task",
    "samples",
    "baseline_score",
    "current_score",
    "delta",
    "retained",
)
CHECKPOINT_COLUMNS = (
    "step",
    "image_macro_f1",
    "bbox_macro_f1",
    "joint_score",
    "improved_image",
    "improved_bbox",
    "improved_joint",
    "checkpoint_preserved",
    "checkpoint_path",
    "resume_from",
    "evaluation_seconds",
)

_PRESERVED_MARKER = "ui5_preserved_checkpoint.json"
_FLOAT_TOLERANCE = 1e-8
_CURRICULUM_PHASES = (
    (0.60, 0.25, 0.15, 1.0e-6),
    (0.45, 0.35, 0.20, 7.0e-7),
    (0.30, 0.30, 0.40, 5.0e-7),
)


def _finite_float(value: Any, label: str, *, bounded: bool = False) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} must be a finite number, got boolean")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be a finite number, got {value!r}") from exc
    if not math.isfinite(result):
        raise ValueError(f"{label} must be finite, got {result!r}")
    if bounded and not 0.0 <= result <= 1.0:
        raise ValueError(f"{label} must be in [0, 1], got {result!r}")
    return result


def _curriculum_phase(
    step: int, *, total_steps: int
) -> tuple[int, tuple[float, float, float, float]]:
    if total_steps <= 0 or total_steps % len(_CURRICULUM_PHASES):
        raise ValueError("total_steps must be positive and divisible by three")
    if step < 0 or step > total_steps:
        raise ValueError(f"training log step must be in [0, {total_steps}], got {step}")
    phase_width = total_steps // len(_CURRICULUM_PHASES)
    optimizer_step = max(1, step)
    phase_index = min(
        (optimizer_step - 1) // phase_width, len(_CURRICULUM_PHASES) - 1
    )
    return phase_index, _CURRICULUM_PHASES[phase_index]


def train_curve_rows_from_trainer_state(
    trainer_state: Path | str | Mapping[str, Any],
    *,
    expected_step: int | None = None,
    total_steps: int = 1200,
) -> list[dict[str, Any]]:
    """Extract upsert-ready train-curve rows from HF ``trainer_state``.

    ``trainer_state`` may be the JSON object, the JSON file, or a checkpoint
    directory containing ``trainer_state.json``. Ratios and the phase LR are
    filled from the fixed three-stage schedule when a log entry does not carry
    them; actual logged values always take precedence.
    """

    source_label = "trainer_state"
    if isinstance(trainer_state, Mapping):
        state = dict(trainer_state)
    else:
        source = Path(trainer_state).expanduser().resolve(strict=True)
        if source.is_dir():
            source = source / "trainer_state.json"
        if not source.is_file():
            raise FileNotFoundError(f"trainer_state.json is missing: {source}")
        source_label = str(source)
        state = json.loads(source.read_text(encoding="utf-8"))
    if not isinstance(state, dict):
        raise ValueError(f"{source_label} must contain an object")

    if expected_step is not None:
        global_step = _count(
            state.get("global_step"), f"{source_label}.global_step"
        )
        if global_step != int(expected_step):
            raise ValueError(
                "trainer_state global_step does not match the evaluation step: "
                f"{global_step} != {expected_step}"
            )
    raw_history = state.get("log_history", [])
    if not isinstance(raw_history, list):
        raise ValueError(f"{source_label}.log_history must be a list")

    by_step: dict[int, dict[str, Any]] = {}
    training_fields = {
        "loss",
        "loss_total",
        "train_loss",
        "loss_lm",
        "lm_loss",
        "learning_rate",
        "grad_norm",
        "curriculum_hard_samples",
        "curriculum_anchor_samples",
        "curriculum_global_replay_samples",
        "hard_samples",
        "anchor_samples",
        "global_replay_samples",
    }
    for index, raw in enumerate(raw_history):
        if not isinstance(raw, Mapping):
            raise ValueError(
                f"{source_label}.log_history[{index}] must be an object"
            )
        if raw.get("step") is None or not any(name in raw for name in training_fields):
            continue
        row_step = _count(
            raw["step"], f"{source_label}.log_history[{index}].step"
        )
        if expected_step is not None and row_step > expected_step:
            raise ValueError(
                f"trainer log step {row_step} exceeds evaluation step {expected_step}"
            )
        phase_index, schedule = _curriculum_phase(
            row_step, total_steps=total_steps
        )
        row = by_step.setdefault(
            row_step,
            {
                "step": row_step,
                "phase": phase_index + 1,
                "learning_rate": schedule[3],
                "loss_total": None,
                "loss_lm": None,
                "hard_ratio": schedule[0],
                "anchor_ratio": schedule[1],
                "global_replay_ratio": schedule[2],
                "hard_samples": None,
                "anchor_samples": None,
                "global_replay_samples": None,
            },
        )
        aliases = {
            "learning_rate": ("learning_rate",),
            "loss_total": ("loss_total", "loss", "train_loss"),
            "loss_lm": ("loss_lm", "lm_loss"),
            "hard_ratio": ("hard_ratio",),
            "anchor_ratio": ("anchor_ratio",),
            "global_replay_ratio": ("global_replay_ratio",),
            "hard_samples": ("hard_samples", "curriculum_hard_samples"),
            "anchor_samples": (
                "anchor_samples",
                "curriculum_anchor_samples",
            ),
            "global_replay_samples": (
                "global_replay_samples",
                "curriculum_global_replay_samples",
            ),
        }
        for destination, names in aliases.items():
            for name in names:
                if raw.get(name) is not None:
                    row[destination] = raw[name]
                    break
        for name in ("epoch", "grad_norm"):
            if raw.get(name) is not None:
                row[name] = raw[name]

    if expected_step == 0 and not by_step:
        _, schedule = _curriculum_phase(0, total_steps=total_steps)
        by_step[0] = {
            "step": 0,
            "phase": "baseline",
            "learning_rate": schedule[3],
            "loss_total": None,
            "loss_lm": None,
            "hard_ratio": schedule[0],
            "anchor_ratio": schedule[1],
            "global_replay_ratio": schedule[2],
            "hard_samples": None,
            "anchor_samples": None,
            "global_replay_samples": None,
        }
    return [by_step[step] for step in sorted(by_step)]


def _count(value: Any, label: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label} must be a non-negative integer, got boolean")
    try:
        result = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{label} must be a non-negative integer, got {value!r}"
        ) from exc
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        numeric = float(result)
    if result < 0 or numeric != result:
        raise ValueError(
            f"{label} must be a non-negative integer, got {value!r}"
        )
    return result


def _prf(tp: int, fp: int, fn: int) -> dict[str, float]:
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = (
        2.0 * precision * recall / (precision + recall)
        if precision + recall
        else 0.0
    )
    return {"precision": precision, "recall": recall, "f1": f1}


def _validate_reported_metric(
    raw: Mapping[str, Any], name: str, expected: float, label: str
) -> None:
    if raw.get(name) is None:
        return
    actual = _finite_float(raw[name], f"{label}.{name}", bounded=True)
    if not math.isclose(
        actual, expected, rel_tol=_FLOAT_TOLERANCE, abs_tol=_FLOAT_TOLERANCE
    ):
        raise ValueError(
            f"{label}.{name} disagrees with confusion counts: "
            f"reported={actual}, recomputed={expected}"
        )


def _normalize_metric_group(
    raw: Any, *, task: str, granularity: str
) -> dict[str, Any]:
    label = f"tasks.{task}.{granularity}"
    if not isinstance(raw, Mapping):
        raise ValueError(f"{label} must be an object")
    tp = _count(raw.get("tp"), f"{label}.tp")
    fp = _count(raw.get("fp"), f"{label}.fp")
    fn = _count(raw.get("fn"), f"{label}.fn")
    calculated = _prf(tp, fp, fn)
    for name, expected in calculated.items():
        _validate_reported_metric(raw, name, expected, label)

    output: dict[str, Any] = {
        **calculated,
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "predicted_positive": tp + fp,
    }
    if granularity == "image":
        tn = _count(raw.get("tn"), f"{label}.tn")
        total = tp + fp + fn + tn
        accuracy = (tp + tn) / total if total else 0.0
        _validate_reported_metric(raw, "accuracy", accuracy, label)
        output.update({"tn": tn, "accuracy": accuracy, "samples": total})
    else:
        accuracy_value = raw.get("count_accuracy", raw.get("accuracy"))
        output["tn"] = None
        output["accuracy"] = (
            _finite_float(
                accuracy_value, f"{label}.count_accuracy", bounded=True
            )
            if accuracy_value is not None
            else None
        )

    for name in ("mean_iou", "matched_iou_sum"):
        if raw.get(name) is not None:
            output[name] = _finite_float(raw[name], f"{label}.{name}")
    if raw.get("matched_iou_count") is not None:
        output["matched_iou_count"] = _count(
            raw["matched_iou_count"], f"{label}.matched_iou_count"
        )
    return output


def _mean(values: Sequence[Any]) -> float | None:
    present = [float(value) for value in values if value is not None]
    return sum(present) / len(present) if present else None


def _validate_scorer_macro(
    raw_macro: Any, calculated: Mapping[str, Mapping[str, Any]]
) -> None:
    if raw_macro is None:
        return
    if not isinstance(raw_macro, Mapping):
        raise ValueError("macro must be an object")
    for granularity in GRANULARITIES:
        raw_group = raw_macro.get(granularity)
        if raw_group is None:
            continue
        if not isinstance(raw_group, Mapping):
            raise ValueError(f"macro.{granularity} must be an object")
        for name in ("precision", "recall", "f1"):
            if raw_group.get(name) is None:
                continue
            actual = _finite_float(
                raw_group[name], f"macro.{granularity}.{name}", bounded=True
            )
            expected = float(calculated[granularity][name])
            if not math.isclose(
                actual,
                expected,
                rel_tol=_FLOAT_TOLERANCE,
                abs_tol=_FLOAT_TOLERANCE,
            ):
                raise ValueError(
                    f"macro.{granularity}.{name} disagrees with the five tasks: "
                    f"reported={actual}, recomputed={expected}"
                )


def normalize_scorer_metrics(metrics: Mapping[str, Any]) -> dict[str, Any]:
    """Validate and normalize one complete five-task scorer result.

    Task-level precision/recall/F1 are recomputed from confusion counts.  Macro
    values are equal-weight means over the five tasks; micro values are computed
    after summing counts.  This keeps checkpoint selection independent from any
    spreadsheet formulas or pre-rounded display values.
    """

    if not isinstance(metrics, Mapping):
        raise ValueError("scorer metrics must be an object")
    raw_tasks = metrics.get("tasks")
    if not isinstance(raw_tasks, Mapping):
        raise ValueError("scorer metrics must contain a tasks object")

    consumed: set[str] = set()
    tasks: dict[str, dict[str, Any]] = {}
    for canonical, aliases in TASK_ALIASES:
        matches = [name for name in aliases if name in raw_tasks]
        if len(matches) != 1:
            raise ValueError(
                f"expected exactly one scorer entry for {canonical}; found={matches}"
            )
        source_name = matches[0]
        consumed.add(source_name)
        raw_task = raw_tasks[source_name]
        if not isinstance(raw_task, Mapping):
            raise ValueError(f"tasks.{source_name} must be an object")
        tasks[canonical] = {
            granularity: _normalize_metric_group(
                raw_task.get(granularity),
                task=canonical,
                granularity=granularity,
            )
            for granularity in GRANULARITIES
        }
    extras = sorted(str(name) for name in raw_tasks if name not in consumed)
    if extras:
        raise ValueError(f"unexpected scorer task entries: {extras}")

    macro: dict[str, dict[str, Any]] = {}
    micro: dict[str, dict[str, Any]] = {}
    for granularity in GRANULARITIES:
        groups = [tasks[task][granularity] for task in TASKS]
        macro[granularity] = {
            name: _mean([group.get(name) for group in groups])
            for name in ("precision", "recall", "f1", "accuracy")
        }

        tp = sum(int(group["tp"]) for group in groups)
        fp = sum(int(group["fp"]) for group in groups)
        fn = sum(int(group["fn"]) for group in groups)
        micro_group: dict[str, Any] = {
            **_prf(tp, fp, fn),
            "tp": tp,
            "fp": fp,
            "fn": fn,
            "predicted_positive": tp + fp,
        }
        if granularity == "image":
            tn = sum(int(group["tn"]) for group in groups)
            total = tp + fp + fn + tn
            micro_group.update(
                {
                    "tn": tn,
                    "accuracy": (tp + tn) / total if total else 0.0,
                    "samples": total,
                }
            )
        else:
            weights = [int(tasks[task]["image"]["samples"]) for task in TASKS]
            weighted = [
                (tasks[task]["bbox"].get("accuracy"), weights[index])
                for index, task in enumerate(TASKS)
            ]
            accuracy_terms = [
                (float(value), weight)
                for value, weight in weighted
                if value is not None
            ]
            weight_total = sum(weight for _, weight in accuracy_terms)
            micro_group.update(
                {
                    "tn": None,
                    "accuracy": (
                        sum(value * weight for value, weight in accuracy_terms)
                        / weight_total
                        if weight_total
                        else None
                    ),
                }
            )
        matched_count = sum(
            int(group.get("matched_iou_count", 0)) for group in groups
        )
        if matched_count:
            matched_sum = sum(
                float(group.get("matched_iou_sum", 0.0)) for group in groups
            )
            micro_group.update(
                {
                    "matched_iou_sum": matched_sum,
                    "matched_iou_count": matched_count,
                    "mean_iou": matched_sum / matched_count,
                }
            )
        micro[granularity] = micro_group

    _validate_scorer_macro(metrics.get("macro"), macro)
    image_macro_f1 = float(macro["image"]["f1"])
    bbox_macro_f1 = float(macro["bbox"]["f1"])
    joint_score = (image_macro_f1 + bbox_macro_f1) / 2.0
    return {
        "schema_version": 1,
        "tasks": tasks,
        "macro": macro,
        "micro": micro,
        "overall": {
            "image_macro_f1": image_macro_f1,
            "bbox_macro_f1": bbox_macro_f1,
            "joint_score": joint_score,
        },
    }


def _default_state() -> dict[str, Any]:
    return {
        "schema_version": 1,
        "metric_definition": {
            "macro": "equal-weight mean over the five UI5 tasks",
            "micro": "precision/recall/F1 recomputed from summed confusion counts",
            "joint_score": "(image_macro_f1 + bbox_macro_f1) / 2",
            "improvement": "strictly greater than the previous best",
        },
        "evaluations": [],
        "best_image": None,
        "best_bbox": None,
        "best_joint": None,
        "train_curve": [],
        "hard_transition": [],
        "anchor_retention": [],
    }


def load_checkpoints_state(path: Path | str) -> dict[str, Any]:
    path = Path(path).expanduser().resolve()
    if not path.is_file():
        return _default_state()
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"invalid checkpoints state: {path}")
    evaluations = value.get("evaluations")
    if not isinstance(evaluations, list) or not all(
        isinstance(row, dict) for row in evaluations
    ):
        raise ValueError(f"checkpoints.json evaluations must be a list: {path}")
    state = _default_state()
    state.update(value)
    for name in ("train_curve", "hard_transition", "anchor_retention"):
        rows = state.get(name, [])
        if not isinstance(rows, list) or not all(isinstance(row, dict) for row in rows):
            raise ValueError(f"checkpoints.json {name} must be a list: {path}")
    steps = [int(row["step"]) for row in evaluations]
    if steps != sorted(steps) or len(steps) != len(set(steps)):
        raise ValueError("checkpoints.json evaluation steps must be unique and sorted")
    return state


def _atomic_write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{uuid.uuid4().hex}")
    try:
        payload = (
            json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
        )
        with temporary.open("w", encoding="utf-8", newline="") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _tree_inventory(root: Path, *, ignore_marker: bool = False) -> dict[str, int]:
    inventory: dict[str, int] = {}
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        relative = path.relative_to(root).as_posix()
        if ignore_marker and relative == _PRESERVED_MARKER:
            continue
        inventory[relative] = path.stat().st_size
    return inventory


def _checkpoint_marker(
    *, step: int, source: Path, overall: Mapping[str, Any]
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "step": step,
        "source_checkpoint": str(source),
        "image_macro_f1": overall["image_macro_f1"],
        "bbox_macro_f1": overall["bbox_macro_f1"],
        "joint_score": overall["joint_score"],
    }


def preserve_checkpoint(
    candidate_checkpoint: Path | str,
    formal_checkpoint_root: Path | str,
    *,
    step: int,
    overall: Mapping[str, Any],
    expected_ranks: int | None = None,
    validate_resume: bool = True,
) -> Path:
    """Copy one complete candidate directory and publish it atomically."""

    source = Path(candidate_checkpoint).expanduser().resolve(strict=True)
    if not source.is_dir():
        raise ValueError(f"candidate checkpoint is not a directory: {source}")
    source_inventory = _tree_inventory(source)
    if not source_inventory:
        raise ValueError(f"candidate checkpoint is empty: {source}")

    root = Path(formal_checkpoint_root).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    destination = root / f"step-{step:06d}"
    expected_marker = _checkpoint_marker(
        step=step, source=source, overall=overall
    )
    if destination.exists():
        marker_path = destination / _PRESERVED_MARKER
        if not marker_path.is_file():
            raise FileExistsError(
                f"formal checkpoint exists without a preservation marker: {destination}"
            )
        marker = json.loads(marker_path.read_text(encoding="utf-8"))
        if marker != expected_marker:
            raise FileExistsError(
                f"formal checkpoint conflicts with this evaluation: {destination}"
            )
        destination_inventory = _tree_inventory(destination, ignore_marker=True)
        if destination_inventory != source_inventory:
            raise RuntimeError(
                "existing formal checkpoint inventory differs from its evaluated "
                f"source: {destination}"
            )
        if validate_resume:
            from eaglevl.train.ui5_checkpoint_utils import validate_checkpoint

            report = validate_checkpoint(
                destination,
                mode="resume",
                expected_ranks=expected_ranks,
                strict=True,
                require_completion_marker=True,
            )
            if not report.get("valid") or int(
                report.get("details", {}).get("global_step", -1)
            ) != int(step):
                raise RuntimeError(
                    "existing formal checkpoint is not strictly resumable at the "
                    f"recorded step {step}: {destination}; "
                    f"errors={report.get('errors', [])}"
                )
        return destination.resolve()

    temporary = root / f".{destination.name}.tmp-{uuid.uuid4().hex}"
    try:
        shutil.copytree(source, temporary)
        copied_inventory = _tree_inventory(temporary)
        if copied_inventory != source_inventory:
            raise RuntimeError(
                f"checkpoint copy verification failed: {source} -> {temporary}"
            )
        _atomic_write_json(temporary / _PRESERVED_MARKER, expected_marker)
        os.replace(temporary, destination)
    finally:
        if temporary.exists():
            shutil.rmtree(temporary)
    return destination.resolve()


def _validate_candidate_checkpoint(
    candidate: Path, *, step: int, expected_ranks: int | None
) -> dict[str, Any]:
    """Use the shared checkpoint contract before registering/preserving."""

    from eaglevl.train.ui5_checkpoint_utils import validate_checkpoint

    mode = "eval" if step == 0 else "resume"
    report = validate_checkpoint(
        candidate,
        mode=mode,
        expected_ranks=expected_ranks if step > 0 else None,
        strict=step > 0,
        require_completion_marker=step > 0,
    )
    if not report.get("valid"):
        errors = report.get("errors") or ["unknown checkpoint validation error"]
        raise ValueError(
            f"candidate checkpoint is not valid for step {step}: "
            + "; ".join(str(error) for error in errors)
        )
    if step > 0:
        checkpoint_step = int(report.get("details", {}).get("global_step", -1))
        if checkpoint_step != step:
            raise ValueError(
                "candidate checkpoint global_step does not match evaluation step: "
                f"{checkpoint_step} != {step}"
            )
    return report


def _same_metrics(
    existing: Mapping[str, Any], normalized: Mapping[str, Any]
) -> bool:
    previous = existing.get("metrics")
    if previous is None:
        return all(
            math.isclose(
                float(existing[name]),
                float(normalized["overall"][name]),
                rel_tol=_FLOAT_TOLERANCE,
                abs_tol=_FLOAT_TOLERANCE,
            )
            for name in ("image_macro_f1", "bbox_macro_f1", "joint_score")
        )
    return previous == normalized


def _best_record(
    evaluations: Sequence[Mapping[str, Any]], metric: str
) -> dict[str, Any] | None:
    if not evaluations:
        return None
    selected = max(evaluations, key=lambda row: (float(row[metric]), -int(row["step"])))
    return {
        "step": int(selected["step"]),
        "score": float(selected[metric]),
        "checkpoint_preserved": bool(selected["checkpoint_preserved"]),
        "checkpoint_path": str(selected.get("checkpoint_path") or ""),
    }


def _normalize_auxiliary_rows(
    rows: Sequence[Mapping[str, Any]] | None, *, default_step: int
) -> list[dict[str, Any]] | None:
    if rows is None:
        return None
    normalized: list[dict[str, Any]] = []
    for raw in rows:
        if not isinstance(raw, Mapping):
            raise ValueError("auxiliary workbook rows must be objects")
        row = dict(raw)
        row["step"] = int(row.get("step", default_step))
        normalized.append(row)
    return normalized


def _merge_auxiliary_rows(
    existing: Sequence[Mapping[str, Any]], incoming: list[dict[str, Any]] | None
) -> list[dict[str, Any]]:
    if incoming is None:
        return [dict(row) for row in existing]
    replaced_steps = {int(row["step"]) for row in incoming}
    retained = [
        dict(row)
        for row in existing
        if int(row.get("step", -1)) not in replaced_steps
    ]
    retained.extend(incoming)
    return sorted(
        retained,
        key=lambda row: (
            int(row.get("step", -1)),
            str(row.get("task", "")),
            str(row.get("group", "")),
            int(row.get("rollout_id", -1)),
        ),
    )


def _excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, str):
        if not value:
            return None
        return "'" + value if value.startswith("=") else value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _same_excel_value(actual: Any, expected: Any) -> bool:
    if isinstance(actual, (int, float)) and not isinstance(actual, bool):
        if isinstance(expected, (int, float)) and not isinstance(expected, bool):
            return math.isclose(
                float(actual),
                float(expected),
                rel_tol=_FLOAT_TOLERANCE,
                abs_tol=_FLOAT_TOLERANCE,
            )
    return actual == expected


def _ordered_columns(
    rows: Sequence[Mapping[str, Any]], preferred: Sequence[str]
) -> list[str]:
    output = list(preferred)
    seen = set(output)
    extras = sorted(
        {
            str(key)
            for row in rows
            for key in row
            if str(key) not in seen
        }
    )
    output.extend(extras)
    return output


def _overall_rows(state: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for evaluation in state["evaluations"]:
        metrics = evaluation["metrics"]
        rows.append(
            {
                "step": evaluation["step"],
                **{
                    f"{granularity}_{aggregation}_{name}": metrics[aggregation][
                        granularity
                    ][name]
                    for granularity in GRANULARITIES
                    for aggregation in ("macro", "micro")
                    for name in ("precision", "recall", "f1")
                },
                "joint_score": evaluation["joint_score"],
                "candidate_checkpoint": evaluation["candidate_checkpoint"],
                "evaluation_seconds": evaluation["evaluation_seconds"],
            }
        )
    return rows


def _by_task_rows(state: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for evaluation in state["evaluations"]:
        for task in TASKS:
            for granularity in GRANULARITIES:
                values = evaluation["metrics"]["tasks"][task][granularity]
                rows.append(
                    {
                        "step": evaluation["step"],
                        "task": task,
                        "granularity": granularity,
                        **{
                            name: values.get(name)
                            for name in (
                                "precision",
                                "recall",
                                "f1",
                                "tp",
                                "fp",
                                "fn",
                                "tn",
                                "accuracy",
                                "predicted_positive",
                            )
                        },
                        "candidate_checkpoint": evaluation[
                            "candidate_checkpoint"
                        ],
                    }
                )
    return rows


def _checkpoint_rows(state: Mapping[str, Any]) -> list[dict[str, Any]]:
    return [
        {column: evaluation.get(column) for column in CHECKPOINT_COLUMNS}
        for evaluation in state["evaluations"]
    ]


def write_curriculum_workbook(
    path: Path | str, state: Mapping[str, Any]
) -> Path:
    """Atomically rebuild and verify the six-sheet curriculum workbook."""

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl>=3.1 is required for UI5 diagnostics") from exc

    path = Path(path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    tables: dict[str, tuple[list[dict[str, Any]], list[str]]] = {
        SHEET_TRAIN_CURVE: (
            [dict(row) for row in state.get("train_curve", [])],
            _ordered_columns(state.get("train_curve", []), TRAIN_CURVE_COLUMNS),
        ),
        SHEET_UI5_OVERALL: (_overall_rows(state), list(UI5_OVERALL_COLUMNS)),
        SHEET_UI5_BY_TASK: (_by_task_rows(state), list(UI5_BY_TASK_COLUMNS)),
        SHEET_HARD_TRANSITION: (
            [dict(row) for row in state.get("hard_transition", [])],
            _ordered_columns(
                state.get("hard_transition", []), HARD_TRANSITION_COLUMNS
            ),
        ),
        SHEET_ANCHOR_RETENTION: (
            [dict(row) for row in state.get("anchor_retention", [])],
            _ordered_columns(
                state.get("anchor_retention", []), ANCHOR_RETENTION_COLUMNS
            ),
        ),
        SHEET_CHECKPOINTS: (_checkpoint_rows(state), list(CHECKPOINT_COLUMNS)),
    }

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    decimal_names = {
        "precision",
        "recall",
        "f1",
        "accuracy",
        "image_macro_f1",
        "bbox_macro_f1",
        "joint_score",
        "baseline_score",
        "current_score",
        "delta",
        "hard_ratio",
        "anchor_ratio",
        "global_replay_ratio",
    }
    for sheet_name in SHEET_ORDER:
        rows, columns = tables[sheet_name]
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(columns)
        for row in rows:
            sheet.append([_excel_value(row.get(column)) for column in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, column in enumerate(columns, start=1):
            if sheet.max_row >= 2 and (
                column in decimal_names
                or column.endswith("_precision")
                or column.endswith("_recall")
                or column.endswith("_f1")
            ):
                for cell in sheet.iter_cols(
                    min_col=index,
                    max_col=index,
                    min_row=2,
                    max_row=sheet.max_row,
                ):
                    for item in cell:
                        item.number_format = "0.000000"
            if sheet.max_row >= 2 and column == "evaluation_seconds":
                for cell in sheet.iter_cols(
                    min_col=index,
                    max_col=index,
                    min_row=2,
                    max_row=sheet.max_row,
                ):
                    for item in cell:
                        item.number_format = "0.0"
        for column_cells in sheet.columns:
            width = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width + 2, 10), 48
            )

    temporary = path.with_name(f".{path.stem}.tmp-{uuid.uuid4().hex}.xlsx")
    try:
        workbook.save(temporary)
        workbook.close()
        verification = openpyxl.load_workbook(
            temporary, read_only=True, data_only=False
        )
        try:
            if tuple(verification.sheetnames) != SHEET_ORDER:
                raise RuntimeError(
                    f"workbook sheet verification failed: {verification.sheetnames}"
                )
            for sheet_name in SHEET_ORDER:
                expected_rows, expected_columns = tables[sheet_name]
                sheet = verification[sheet_name]
                actual_rows = list(sheet.iter_rows())
                actual_header = tuple(cell.value for cell in actual_rows[0])
                if actual_header != tuple(expected_columns):
                    raise RuntimeError(
                        f"{sheet_name} header verification failed: {actual_header}"
                    )
                for row_number, (actual_row, expected_row) in enumerate(
                    zip(actual_rows[1:], expected_rows, strict=False), start=2
                ):
                    expected_values = tuple(
                        _excel_value(expected_row.get(column))
                        for column in expected_columns
                    )
                    actual_values = tuple(cell.value for cell in actual_row)
                    if len(actual_values) != len(expected_values) or any(
                        not _same_excel_value(actual, expected)
                        for actual, expected in zip(
                            actual_values, expected_values, strict=True
                        )
                    ):
                        raise RuntimeError(
                            f"{sheet_name} data verification failed at row {row_number}"
                        )
                if sheet.max_row - 1 != len(expected_rows):
                    raise RuntimeError(
                        f"{sheet_name} row count verification failed"
                    )
                for row in actual_rows:
                    for cell in row:
                        if cell.data_type == "f":
                            raise RuntimeError(
                                f"unexpected formula in {sheet_name}!{cell.coordinate}"
                            )
                        if cell.data_type == "e":
                            raise RuntimeError(
                                f"spreadsheet error in {sheet_name}!{cell.coordinate}"
                            )
        finally:
            verification.close()
        os.replace(temporary, path)
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        temporary.unlink(missing_ok=True)
    return path


def update_curriculum_artifacts(
    *,
    step: int,
    scorer_metrics: Mapping[str, Any],
    candidate_checkpoint: Path | str,
    checkpoints_json: Path | str,
    workbook_path: Path | str,
    formal_checkpoint_root: Path | str,
    resume_from: str = "",
    evaluation_seconds: float = 0.0,
    train_curve_rows: Sequence[Mapping[str, Any]] | None = None,
    hard_transition_rows: Sequence[Mapping[str, Any]] | None = None,
    anchor_retention_rows: Sequence[Mapping[str, Any]] | None = None,
    validate_candidate: bool = True,
    expected_ranks: int | None = None,
) -> dict[str, Any]:
    """Register one evaluation and update all durable artifacts.

    Step 0 establishes the three baselines but is intentionally not copied into
    the formal checkpoint directory.  Repeating an already-recorded step with
    identical normalized metrics is idempotent; different metrics for that step
    are rejected so best history cannot silently change.
    """

    step = int(step)
    if step < 0:
        raise ValueError("step must be non-negative")
    duration = _finite_float(evaluation_seconds, "evaluation_seconds")
    if duration < 0:
        raise ValueError("evaluation_seconds must be non-negative")
    candidate = Path(candidate_checkpoint).expanduser().resolve(strict=True)
    if not candidate.is_dir():
        raise ValueError(f"candidate checkpoint is not a directory: {candidate}")
    if expected_ranks is not None and int(expected_ranks) <= 0:
        raise ValueError("expected_ranks must be positive")
    if validate_candidate:
        _validate_candidate_checkpoint(
            candidate,
            step=step,
            expected_ranks=expected_ranks,
        )
    normalized = normalize_scorer_metrics(scorer_metrics)
    state_path = Path(checkpoints_json).expanduser().resolve()
    state = load_checkpoints_state(state_path)
    evaluations = [dict(row) for row in state["evaluations"]]

    existing = next(
        (row for row in evaluations if int(row["step"]) == step), None
    )
    if existing is not None:
        if not _same_metrics(existing, normalized):
            raise ValueError(
                f"step {step} is already recorded with different metrics"
            )
        evaluation = existing
        if evaluation.get("checkpoint_preserved"):
            # Re-enter the preservation transaction even when the directory is
            # present.  Its existing branch verifies the marker, complete file
            # inventory, strict resume contract, and recorded global step.
            repaired = preserve_checkpoint(
                candidate,
                formal_checkpoint_root,
                step=step,
                overall=normalized["overall"],
                expected_ranks=expected_ranks,
                validate_resume=validate_candidate,
            )
            evaluation["checkpoint_path"] = str(repaired)
    else:
        if not evaluations and step != 0:
            raise ValueError("step 0 baseline must be recorded before training steps")
        if evaluations and step <= int(evaluations[-1]["step"]):
            raise ValueError(
                f"new evaluation step must exceed {evaluations[-1]['step']}, got {step}"
            )
        overall = normalized["overall"]
        previous_image = _best_record(evaluations, "image_macro_f1")
        previous_bbox = _best_record(evaluations, "bbox_macro_f1")
        previous_joint = _best_record(evaluations, "joint_score")
        # Step 0 establishes the comparison baseline; it does not "improve"
        # over a prior measured checkpoint and therefore is never preserved as
        # a formal best-only copy.
        improved_image = previous_image is not None and float(
            overall["image_macro_f1"]
        ) > float(previous_image["score"])
        improved_bbox = previous_bbox is not None and float(
            overall["bbox_macro_f1"]
        ) > float(previous_bbox["score"])
        improved_joint = previous_joint is not None and float(
            overall["joint_score"]
        ) > float(previous_joint["score"])
        preserve = step > 0 and any(
            (improved_image, improved_bbox, improved_joint)
        )
        checkpoint_path = ""
        if preserve:
            checkpoint_path = str(
                preserve_checkpoint(
                    candidate,
                    formal_checkpoint_root,
                    step=step,
                    overall=overall,
                    expected_ranks=expected_ranks,
                    validate_resume=validate_candidate,
                )
            )
        elif step == 0:
            # The source crop checkpoint remains the recoverable baseline, but
            # no duplicate is retained under the formal results.
            checkpoint_path = str(candidate)
        evaluation = {
            "step": step,
            "candidate_checkpoint": str(candidate),
            "metrics": normalized,
            "image_macro_f1": float(overall["image_macro_f1"]),
            "bbox_macro_f1": float(overall["bbox_macro_f1"]),
            "joint_score": float(overall["joint_score"]),
            "improved_image": improved_image,
            "improved_bbox": improved_bbox,
            "improved_joint": improved_joint,
            "checkpoint_preserved": preserve,
            "checkpoint_path": checkpoint_path,
            "resume_from": str(resume_from),
            "evaluation_seconds": duration,
        }
        evaluations.append(evaluation)

    evaluations.sort(key=lambda row: int(row["step"]))
    state["evaluations"] = evaluations
    state["best_image"] = _best_record(evaluations, "image_macro_f1")
    state["best_bbox"] = _best_record(evaluations, "bbox_macro_f1")
    state["best_joint"] = _best_record(evaluations, "joint_score")
    for name, rows in (
        ("train_curve", train_curve_rows),
        ("hard_transition", hard_transition_rows),
        ("anchor_retention", anchor_retention_rows),
    ):
        incoming = _normalize_auxiliary_rows(rows, default_step=step)
        state[name] = _merge_auxiliary_rows(state.get(name, []), incoming)

    _atomic_write_json(state_path, state)
    write_curriculum_workbook(workbook_path, state)
    return {
        "step": step,
        "idempotent": existing is not None,
        "candidate_checkpoint": str(evaluation["candidate_checkpoint"]),
        "resume_from": str(evaluation.get("resume_from") or ""),
        "evaluation_seconds": float(evaluation["evaluation_seconds"]),
        "metrics": normalized,
        "overall": dict(normalized["overall"]),
        "improved_image": bool(evaluation["improved_image"]),
        "improved_bbox": bool(evaluation["improved_bbox"]),
        "improved_joint": bool(evaluation["improved_joint"]),
        "checkpoint_preserved": bool(evaluation["checkpoint_preserved"]),
        "checkpoint_path": str(evaluation.get("checkpoint_path") or ""),
        "checkpoints_json": str(state_path),
        "workbook": str(Path(workbook_path).expanduser().resolve()),
        "best_image": state["best_image"],
        "best_bbox": state["best_bbox"],
        "best_joint": state["best_joint"],
    }
