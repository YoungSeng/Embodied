"""Best-effort CPT JSON/JSONL to three-sheet Excel projection.

The JSON files are authoritative.  ``openpyxl`` is imported inside the public
function so training never gains a hard spreadsheet dependency.
"""

from __future__ import annotations

import json
import os
import tempfile
import warnings
from pathlib import Path
from typing import Any, Iterable, Mapping


SHEETS = ("TrainMetrics", "EvalMetrics", "UIDefectMetrics")


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        value = json.load(handle)
    return value if isinstance(value, dict) else {}


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"{path}:{line_number}: expected a JSON object")
            rows.append(value)
    return rows


def _flatten(value: Mapping[str, Any], prefix: str = "") -> dict[str, Any]:
    output = {}
    for key, item in value.items():
        name = f"{prefix}_{key}" if prefix else str(key)
        if isinstance(item, Mapping):
            output.update(_flatten(item, name))
        elif isinstance(item, (list, tuple, set)):
            output[name] = json.dumps(list(item), ensure_ascii=False, sort_keys=True)
        else:
            output[name] = item
    return output


def _ordered_columns(rows: Iterable[Mapping[str, Any]], preferred: list[str]) -> list[str]:
    keys = set()
    for row in rows:
        keys.update(row)
    return [key for key in preferred if key in keys] + sorted(keys.difference(preferred))


def _ui_defect_metric_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Project nested five-class image/bbox metrics into an analysis table."""

    output: list[dict[str, Any]] = []
    for row in rows:
        if row.get("task") not in {"ui_defect", "ui_defect_external"}:
            continue
        common = {
            key: row.get(key)
            for key in (
                "checkpoint",
                "step",
                "split",
                "task",
                "evaluation_kind",
                "manifest_id",
                "evaluation_protocol_id",
                "samples_per_task",
            )
        }
        for model, metrics in (
            ("checkpoint", row.get("metrics")),
            ("base", row.get("base_metrics")),
        ):
            if not isinstance(metrics, Mapping):
                continue
            bbox_iou_threshold = metrics.get(
                "iou_threshold", row.get("iou_threshold", 0.5)
            )
            for class_name, class_metrics in metrics.get("per_class", {}).items():
                if not isinstance(class_metrics, Mapping):
                    continue
                for granularity in ("image", "bbox"):
                    values = class_metrics.get(granularity, {})
                    if not isinstance(values, Mapping):
                        continue
                    output.append(
                        {
                            **common,
                            "model": model,
                            "aggregate": "class",
                            "class": class_name,
                            "class_label": class_metrics.get(
                                "display_label", class_name
                            ),
                            "granularity": granularity,
                            "iou_threshold": (
                                bbox_iou_threshold if granularity == "bbox" else None
                            ),
                            **{
                                key: values.get(key)
                                for key in (
                                    "tp",
                                    "fp",
                                    "fn",
                                    "tn",
                                    "precision",
                                    "recall",
                                    "f1",
                                    "accuracy",
                                    "images",
                                )
                            },
                        }
                    )
            for aggregate in ("macro", "micro"):
                for granularity in ("image", "bbox"):
                    values = metrics.get(f"{granularity}_{aggregate}", {})
                    if not isinstance(values, Mapping):
                        continue
                    output.append(
                        {
                            **common,
                            "model": model,
                            "aggregate": aggregate,
                            "class": f"__{aggregate}__",
                            "class_label": f"five-class {aggregate}",
                            "granularity": granularity,
                            "iou_threshold": (
                                bbox_iou_threshold if granularity == "bbox" else None
                            ),
                            **{
                                key: values.get(key)
                                for key in (
                                    "tp",
                                    "fp",
                                    "fn",
                                    "tn",
                                    "precision",
                                    "recall",
                                    "f1",
                                    "accuracy",
                                    "images",
                                )
                            },
                        }
                    )
    return output


def _latest_by_task(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    output = {}
    for row in rows:
        task = row.get("task")
        if task is None:
            continue
        if task not in output or int(row.get("step") or -1) >= int(output[task].get("step") or -1):
            output[str(task)] = row
    return output


def _best_by_task(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Select each task's best held-out row by primary metric, then main CE."""
    output: dict[str, dict[str, Any]] = {}
    for row in rows:
        task = row.get("task")
        primary = row.get("primary_metric")
        if task is None or task == "__task_macro__" or not isinstance(primary, (int, float)):
            continue
        current = output.get(str(task))
        candidate_key = (
            float(primary),
            -float(row["eval_token_ce"])
            if isinstance(row.get("eval_token_ce"), (int, float))
            else float("-inf"),
            int(row.get("step") or -1),
        )
        current_key = (
            float(current["primary_metric"]),
            -float(current["eval_token_ce"])
            if current and isinstance(current.get("eval_token_ce"), (int, float))
            else float("-inf"),
            int(current.get("step") or -1) if current else -1,
        ) if current else None
        if current_key is None or candidate_key > current_key:
            output[str(task)] = row
    return output


def _overview_rows(
    run_config: Mapping[str, Any],
    split_summary: Mapping[str, Any],
    data_stats: Mapping[str, Any],
    train_rows: list[dict[str, Any]],
    eval_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    latest_train = _latest_by_task(train_rows)
    heldout_eval = [
        row
        for row in eval_rows
        if row.get("split") == "heldout" and row.get("task") != "__task_macro__"
    ]
    latest_eval = _latest_by_task(heldout_eval)
    best_eval = _best_by_task(heldout_eval)
    task_names = set(split_summary.get("tasks", {}))
    task_names.update(data_stats.get("tasks", {}))
    task_names.update(latest_train)
    task_names.update(latest_eval)
    task_names.update(
        str(item.get("task"))
        for item in run_config.get("datasets", [])
        if item.get("task")
    )
    sampling = {
        str(item.get("name")): item
        for item in (run_config.get("sampling") or {}).get("tasks", [])
    }
    dataset_config = {
        str(item.get("task")): item
        for item in run_config.get("datasets", [])
        if item.get("task")
    }
    output = []
    for task in sorted(task_names):
        split = split_summary.get("tasks", {}).get(task, {})
        lengths = data_stats.get("tasks", {}).get(task, {})
        length_values = lengths.get("lengths", {})
        train = latest_train.get(task, {})
        evaluation = latest_eval.get(task, {})
        best = best_eval.get(task, {})
        output.append(
            {
                "run_name": run_config.get("run_name"),
                "seed": run_config.get("seed"),
                "world_size": run_config.get("world_size"),
                "sampling_mode": (run_config.get("sampling") or {}).get("mode"),
                "task": task,
                "train_rows": split.get("train_rows"),
                "val_rows": split.get("val_rows"),
                "val_fast_rows": split.get("val_fast_rows"),
                "train_groups": split.get("train_groups"),
                "val_groups": split.get("val_groups"),
                "group_leakage": split_summary.get("group_intersection"),
                "pre_mtp_p50": length_values.get("pre_mtp_seq_len", {}).get("p50"),
                "pre_mtp_p95": length_values.get("pre_mtp_seq_len", {}).get("p95"),
                "pre_mtp_p99": length_values.get("pre_mtp_seq_len", {}).get("p99"),
                "post_mtp_p50": length_values.get("post_mtp_seq_len", {}).get("p50"),
                "post_mtp_p95": length_values.get("post_mtp_seq_len", {}).get("p95"),
                "post_mtp_p99": length_values.get("post_mtp_seq_len", {}).get("p99"),
                "post_mtp_max": length_values.get("post_mtp_seq_len", {}).get("max"),
                "static_oversize_rate": lengths.get("oversize_rate"),
                "sampling_probability": sampling.get(task, {}).get(
                    "probability", dataset_config.get(task, {}).get("probability")
                ),
                "sample_share": train.get("sample_share"),
                "total_token_share": train.get("total_token_share"),
                "oversize_skip_rate": train.get("oversize_skip_rate"),
                "row_coverage": train.get("row_coverage"),
                "group_coverage": train.get("group_coverage"),
                "effective_epoch": train.get("effective_epoch"),
                "repeat_factor": train.get("repeat_factor"),
                "train_total_token_ce": train.get("train_total_token_ce"),
                "best_step": best.get("step"),
                "best_primary": best.get("primary_metric"),
                "base_primary": evaluation.get("base_primary"),
                "primary_metric": evaluation.get("primary_metric"),
                "delta_vs_base": evaluation.get("delta_vs_base"),
            }
        )
    return output


def _write_table(sheet, rows: list[dict[str, Any]], preferred: list[str], table_name: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    columns = _ordered_columns(rows, preferred)
    if not columns:
        columns = preferred or ["status"]
    values = [columns]
    values.extend([[row.get(column) for column in columns] for row in rows])
    for row in values:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    header = sheet[1]
    for cell in header:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    header[0].parent.row_dimensions[1].height = 26
    for column_cells in sheet.columns:
        sample = list(column_cells[: min(len(column_cells), 200)])
        width = max((len(str(cell.value)) if cell.value is not None else 0 for cell in sample), default=0)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 36)
    rate_columns = {
        "static_oversize_rate", "sampling_probability", "sample_share",
        "main_token_share", "mtp_token_share", "total_token_share",
        "oversize_skip_rate", "row_coverage", "group_coverage",
        "packing_efficiency", "task_conditional_packing_efficiency",
    }
    count_columns = {
        "step", "train_rows", "val_rows", "val_fast_rows", "train_groups",
        "val_groups", "attempted_samples", "accepted_samples", "trained_samples",
        "oversize_skipped_samples", "main_supervised_tokens",
        "mtp_supervised_tokens", "total_supervised_tokens", "main_loss_tokens",
        "mtp_loss_tokens", "eval_loss_tokens", "unique_record_count",
        "unique_group_count", "best_step",
    }
    for column_index, name in enumerate(columns, start=1):
        number_format = None
        if name in rate_columns or name.endswith("_rate") or name.endswith("_share"):
            number_format = "0.00%"
        elif name in count_columns or name.endswith("_count") or name.endswith("_tokens"):
            number_format = "#,##0"
        elif name.endswith("_ce") or name.endswith("_gap") or name in {
            "primary_metric", "base_primary", "best_primary", "delta_vs_base",
            "effective_epoch", "repeat_factor", "token_dominance_ratio",
        }:
            number_format = "0.0000"
        if number_format:
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).number_format = number_format
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def build_cpt_workbook(
    diagnostics_dir: str | os.PathLike[str],
    output_path: str | os.PathLike[str] | None = None,
) -> bool:
    """Build the workbook; warn and return False on any optional Excel failure."""
    diagnostics = Path(diagnostics_dir)
    output = Path(output_path) if output_path else diagnostics / "cpt_training_evaluation.xlsx"
    try:
        from openpyxl import Workbook
        raw_train_rows = _read_jsonl(diagnostics / "cpt_train_metrics.jsonl")
        raw_eval_rows = _read_jsonl(diagnostics / "cpt_eval_metrics.jsonl")
        train_rows = [_flatten(row) for row in raw_train_rows]
        eval_rows = [_flatten(row) for row in raw_eval_rows]
        ui_defect_rows = _ui_defect_metric_rows(raw_eval_rows)

        workbook = Workbook()
        workbook.remove(workbook.active)
        specs = (
            ("TrainMetrics", train_rows, ["step", "epoch", "scope", "task", "learning_rate", "global_loss", "train_main_token_ce", "train_mtp_token_ce", "train_total_token_ce", "main_loss_tokens", "mtp_loss_tokens", "attempted_samples", "accepted_samples", "trained_samples", "oversize_skipped_samples", "oversize_skip_rate", "sample_share", "main_supervised_tokens", "mtp_supervised_tokens", "total_supervised_tokens", "total_token_share", "avg_post_mtp_length", "p95_post_mtp_length", "packing_efficiency", "row_coverage", "group_coverage", "effective_epoch", "repeat_factor"], "CPTTrainMetrics"),
            ("EvalMetrics", eval_rows, ["checkpoint", "step", "split", "task", "manifest_id", "evaluation_protocol_id", "subset_strategy", "samples_per_task", "ce_kind", "train_main_token_ce", "eval_token_ce", "train_val_main_ce_gap", "train_val_ce_gap", "eval_loss_tokens", "primary_name", "primary_metric", "base_primary", "delta_vs_base", "is_best_overall", "complete_ten_task_heldout", "eval_wall_time_seconds"], "CPTEvalMetrics"),
            ("UIDefectMetrics", ui_defect_rows, ["checkpoint", "step", "split", "task", "evaluation_kind", "model", "aggregate", "class", "class_label", "granularity", "iou_threshold", "tp", "fp", "fn", "tn", "precision", "recall", "f1", "accuracy", "images", "manifest_id", "evaluation_protocol_id", "samples_per_task"], "CPTUIDefectMetrics"),
        )
        for name, rows, preferred, table_name in specs:
            sheet = workbook.create_sheet(name)
            _write_table(sheet, rows, preferred, table_name)
        if tuple(sheet.title for sheet in workbook.worksheets) != SHEETS:
            raise RuntimeError("CPT workbook must contain exactly three sheets")
        output.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=output.parent, prefix=output.name + ".", delete=False
        ) as handle:
            temporary = Path(handle.name)
        workbook.save(temporary)
        os.replace(temporary, output)
        return True
    except Exception as exc:
        warnings.warn(f"CPT Excel export skipped; JSON/JSONL remain authoritative: {exc}")
        return False
