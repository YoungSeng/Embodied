#!/usr/bin/env python3
"""Aggregate the eight UI5 train rollouts without re-running inference."""
from __future__ import annotations

import argparse
import json
import math
import os
import statistics
import sys
from collections import Counter, defaultdict
from contextlib import ExitStack
from datetime import datetime, timezone
from itertools import zip_longest
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

from run_ui5_train_rollout_worker import (
    MAX_NUM_TOKENS_PER_SAMPLE,
    MAX_SEQ_LENGTH,
    PROCESSOR_IN_TOKEN_LIMIT,
    ROLLOUT_MAX_NEW_TOKENS,
    TRAINING_MAX_NUM_TOKENS,
    fixed_interleaved_samples,
    last_jsonl_row,
    load_module,
    score_prediction,
)
from snapshot_ui5_train_rollouts import (
    DIFFICULTIES,
    build_difficulty_records,
    execution_architecture,
    write_difficulty_exports,
)


SCHEMA_VERSION = 6
MODELS = ("m31", "crop")
TASKS = ("occlusion", "cropping", "text_overflow", "text_ellipsis", "content_missing")
THRESHOLDS = (0.1, 0.3, 0.5)
ERROR_TYPES = (
    "TN",
    "FP_ONLY",
    "FN_NO_PRED",
    "LOC_WRONG",
    "PARTIAL_MISS",
    "PARTIAL_EXTRA",
    "PARTIAL_BOTH",
    "EXACT_TP",
    "PARSE_ERROR",
)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--bundle-root", type=Path, required=True)
    parser.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parents[1])
    return parser.parse_args(argv)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def atomic_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    try:
        temporary.write_text(
            json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def read_json_if_present(path: Path) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError(f"expected JSON object: {path}")
    return value


def raw_iterator(root: Path, model: str, rollout: int) -> Iterator[dict[str, Any]]:
    directory = root / "raw" / model / f"rollout_{rollout}"
    parts = sorted(directory.glob("part-*.jsonl"))
    if not parts:
        raise FileNotFoundError(f"no raw rollout parts: {directory}")
    for part in parts:
        with part.open("r", encoding="utf-8") as handle:
            for line_no, line in enumerate(handle, 1):
                if not line.strip():
                    continue
                value = json.loads(line)
                if not isinstance(value, dict):
                    raise ValueError(f"expected object at {part}:{line_no}")
                yield value


def grouped_rollouts(root: Path, model: str) -> Iterator[list[dict[str, Any]]]:
    iterators = [raw_iterator(root, model, rollout) for rollout in range(4)]
    for row_index, values in enumerate(zip_longest(*iterators), 1):
        if any(value is None for value in values):
            raise RuntimeError(f"{model} rollout lengths differ at row {row_index}")
        rows = [value for value in values if value is not None]
        ids = {str(row["record_id"]) for row in rows}
        rollout_ids = {int(row["rollout_id"]) for row in rows}
        if len(ids) != 1 or rollout_ids != {0, 1, 2, 3}:
            raise RuntimeError(
                f"unaligned {model} rollouts at row {row_index}: ids={ids}, rollouts={rollout_ids}"
            )
        yield sorted(rows, key=lambda row: int(row["rollout_id"]))


def bundle_sample_keys(bundle: Path) -> tuple[list[tuple[str, str]], int]:
    manifest = json.loads((bundle / "bundle_manifest.json").read_text(encoding="utf-8"))
    expected_total = int(manifest["rollout_samples"])
    rows: list[dict[str, Any]] = []
    path = bundle / "manifest" / "task_samples.jsonl"
    with path.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, 1):
            if not line.strip():
                continue
            rows.append(json.loads(line))
    keys = [
        (str(row["record_id"]), str(row["sample_id"]))
        for row in fixed_interleaved_samples(rows)
    ]
    if len(keys) != expected_total:
        raise RuntimeError(
            "bundle expected_total disagrees with task_samples: "
            f"manifest={expected_total} rows={len(keys)}"
        )
    if len(set(keys)) != len(keys):
        raise RuntimeError("bundle task_samples has duplicate record_id/sample_id keys")
    return keys, expected_total


def validate_raw_alignment(
    root: Path, bundle: Path
) -> tuple[list[dict[str, Any]], int]:
    expected_keys, expected_total = bundle_sample_keys(bundle)
    reports: list[dict[str, Any]] = []
    for model in MODELS:
        for rollout in range(4):
            actual_count = 0
            for row_index, row in enumerate(raw_iterator(root, model, rollout)):
                if row_index >= expected_total:
                    raise RuntimeError(
                        f"{model} rollout {rollout} exceeds expected_total={expected_total}"
                    )
                actual = (str(row["record_id"]), str(row["sample_id"]))
                expected = expected_keys[row_index]
                if actual != expected:
                    raise RuntimeError(
                        f"{model} rollout {rollout} sample alignment mismatch at "
                        f"row={row_index + 1}: actual={actual} expected={expected}"
                    )
                if int(row["rollout_id"]) != rollout:
                    raise RuntimeError(
                        f"{model} rollout directory contains rollout_id={row['rollout_id']}"
                    )
                actual_count += 1
            if actual_count != expected_total:
                raise RuntimeError(
                    f"{model} rollout {rollout} count={actual_count}, "
                    f"expected_total={expected_total}"
                )
            reports.append(
                {
                    "model_id": model,
                    "rollout_id": rollout,
                    "expected_total": expected_total,
                    "actual_total": actual_count,
                    "sample_ids_aligned": True,
                    "complete": True,
                }
            )
    return reports, expected_total


def safe_div(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


class Accumulator:
    def __init__(self) -> None:
        self.total = 0
        self.image = Counter()
        self.tp_box = 0
        self.fp_box = 0
        self.fn_box = 0
        self.parse_errors = 0
        self.exact = 0
        self.pair_count = 0
        self.iou_sum = 0.0
        self.center_px_sum = 0.0
        self.center_norm_sum = 0.0
        self.area_ratio_sum = 0.0

    def add(self, score: Mapping[str, Any]) -> None:
        self.total += 1
        self.image[str(score["image_confusion"])] += 1
        self.tp_box += int(score["TP_box"])
        self.fp_box += int(score["FP_box"])
        self.fn_box += int(score["FN_box"])
        self.parse_errors += int(score.get("error_type") == "PARSE_ERROR")
        self.exact += int(bool(score.get("exact_correct")))
        for pair in score.get("matched_pairs", []):
            if not pair.get("is_tp"):
                continue
            self.pair_count += 1
            self.iou_sum += float(pair["iou"])
            self.center_px_sum += float(pair["center_distance_px"])
            self.center_norm_sum += float(pair["center_distance_normalized"])
            self.area_ratio_sum += float(pair["pred_gt_area_ratio"])

    def metrics(self) -> dict[str, Any]:
        tp, tn = self.image["TP"], self.image["TN"]
        fp, fn = self.image["FP"], self.image["FN"]
        precision = safe_div(tp, tp + fp)
        recall = safe_div(tp, tp + fn)
        specificity = safe_div(tn, tn + fp)
        fpr = safe_div(fp, fp + tn)
        fnr = safe_div(fn, fn + tp)
        npv = safe_div(tn, tn + fn)
        accuracy = safe_div(tp + tn, self.total)
        balanced = (recall + specificity) / 2
        f1 = safe_div(2 * precision * recall, precision + recall)
        bbox_precision = safe_div(self.tp_box, self.tp_box + self.fp_box)
        bbox_recall = safe_div(self.tp_box, self.tp_box + self.fn_box)
        bbox_f1 = safe_div(
            2 * bbox_precision * bbox_recall, bbox_precision + bbox_recall
        )
        return {
            "total_samples": self.total,
            "image_TP": tp,
            "image_TN": tn,
            "image_FP": fp,
            "image_FN": fn,
            "image_TP_ratio": safe_div(tp, self.total),
            "image_TN_ratio": safe_div(tn, self.total),
            "image_FP_ratio": safe_div(fp, self.total),
            "image_FN_ratio": safe_div(fn, self.total),
            "precision": precision,
            "recall": recall,
            "specificity": specificity,
            "FPR": fpr,
            "FNR": fnr,
            "NPV": npv,
            "accuracy": accuracy,
            "balanced_accuracy": balanced,
            "F1": f1,
            "bbox_TP": self.tp_box,
            "bbox_FP": self.fp_box,
            "bbox_FN": self.fn_box,
            "bbox_precision": bbox_precision,
            "bbox_recall": bbox_recall,
            "bbox_F1": bbox_f1,
            "matched_pair_count": self.pair_count,
            "mean_matched_iou": safe_div(self.iou_sum, self.pair_count),
            "mean_center_distance_px": safe_div(self.center_px_sum, self.pair_count),
            "mean_center_distance_normalized": safe_div(
                self.center_norm_sum, self.pair_count
            ),
            "mean_pred_gt_area_ratio": safe_div(self.area_ratio_sum, self.pair_count),
            "parse_errors": self.parse_errors,
            "exact_correct": self.exact,
            "exact_correct_ratio": safe_div(self.exact, self.total),
        }


class JsonlWriter:
    def __init__(self, path: Path):
        self.path = path
        path.parent.mkdir(parents=True, exist_ok=True)
        self.temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
        self.handle = self.temporary.open("w", encoding="utf-8")
        self.count = 0

    def write(self, row: Mapping[str, Any]) -> None:
        self.handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
        self.count += 1

    def close(self) -> None:
        self.handle.flush()
        os.fsync(self.handle.fileno())
        self.handle.close()
        os.replace(self.temporary, self.path)


def compact_rollout(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "model_id": row["model_id"],
        "rollout_id": row["rollout_id"],
        "seed": row["seed"],
        "checkpoint": row["checkpoint"],
        "git_commit": row["git_commit"],
        "baseline_git_commit": row["baseline_git_commit"],
        "generation_config": row["generation_config"],
        "parse_status": row["parse_status"],
        "parse_warnings": row.get("parse_warnings", []),
        "contains_crop_parse_error": bool(row.get("contains_crop_parse_error")),
        "error_type": row["error_type"],
        "reward": row["exact_correct"],
        "exact_correct": row["exact_correct"],
        "raw_output": row.get("raw_output"),
        "pred_local": row.get("pred_local"),
        "pred_global": row.get("pred_global"),
        "gt_local": row.get("gt_local"),
        "gt_global": row.get("gt_global"),
        "matched_pairs": row.get("matched_pairs", []),
        "TP_box": row["TP_box"],
        "FP_box": row["FP_box"],
        "FN_box": row["FN_box"],
        "image_confusion": row.get("image_confusion"),
        "crop_outputs": row.get("crop_outputs", []),
        "latency_seconds": row["latency_seconds"],
        "inference_success": row.get("inference_success", True),
        "runtime_error": row.get("runtime_error"),
        "token_usage": row.get("token_usage"),
        "oom_recovered": bool(row.get("oom_recovered")),
        "oom_events": int(row.get("oom_events", 0)),
        "oom_final_failure": bool(row.get("oom_final_failure")),
        "oom_retry": row.get("oom_retry"),
    }


def selection_base(rows: Sequence[Mapping[str, Any]], correct_count: int) -> dict[str, Any]:
    first = rows[0]
    return {
        "record_id": first["record_id"],
        "sample_id": first["sample_id"],
        "source_image_id": first["source_image_id"],
        "task": first["task"],
        "image_relpath": first["image_relpath"],
        "prompt": first["prompt"],
        "gt_global": first["gt_global"],
        "source_records": first.get("source_records", []),
        "original_training_record": first.get("original_training_record"),
        "correct_count": correct_count,
        "rollouts": [compact_rollout(row) for row in rows],
        "pipeline_coverage_failure": bool(first.get("pipeline_coverage_failure")),
        "annotation_anomaly": bool(first.get("annotation_anomaly")),
        "coordinate_transform_anomaly": bool(
            first.get("coordinate_transform_anomaly")
        ),
    }


def metric_fields() -> tuple[str, ...]:
    return (
        "precision",
        "recall",
        "specificity",
        "FPR",
        "FNR",
        "NPV",
        "accuracy",
        "balanced_accuracy",
        "F1",
        "bbox_precision",
        "bbox_recall",
        "bbox_F1",
        "mean_matched_iou",
        "mean_center_distance_normalized",
        "mean_pred_gt_area_ratio",
        "exact_correct_ratio",
    )


def mean_std(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for field in metric_fields():
        values = [float(row.get(field, 0.0)) for row in rows]
        output[f"{field}_mean"] = statistics.fmean(values) if values else 0.0
        output[f"{field}_std"] = statistics.pstdev(values) if len(values) > 1 else 0.0
    return output


def macro_row(task_rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    return {
        field: statistics.fmean(float(row.get(field, 0.0)) for row in task_rows)
        for field in metric_fields()
    }


def workbook(
    path: Path,
    overview: Mapping[str, Any],
    tables: Mapping[str, Sequence[Mapping[str, Any]]],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl>=3.1 is required for the formal report") from exc
    book = openpyxl.Workbook()
    book.remove(book.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in tables.items():
        sheet = book.create_sheet(sheet_name[:31])
        if sheet_name == "overview":
            sheet.append(["metric", "value"])
            for key, value in overview.items():
                sheet.append(
                    [key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value]
                )
        else:
            columns = sorted({key for row in rows for key in row})
            if not columns:
                columns = ["note"]
                rows = [{"note": "no rows"}]
            sheet.append(columns)
            for row in rows:
                sheet.append(
                    [
                        json.dumps(value, ensure_ascii=False, sort_keys=True)
                        if isinstance(value, (dict, list))
                        else value
                        for value in (row.get(column) for column in columns)
                    ]
                )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells[:200]]
            width = min(48, max(10, max(map(len, values), default=8) + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
                cell.alignment = Alignment(vertical="top", wrap_text=False)
        if sheet.max_row > 2:
            for cell in sheet[2]:
                cell.fill = section_fill
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}.xlsx")
    book.save(temporary)
    verified = openpyxl.load_workbook(temporary, read_only=True, data_only=False)
    required = set(tables)
    if not required.issubset(set(verified.sheetnames)):
        raise RuntimeError("workbook verification missed required sheets")
    for sheet in verified.worksheets:
        if sheet.max_row < 2 or sheet.max_column < 1:
            raise RuntimeError(f"workbook sheet is unexpectedly empty: {sheet.title}")
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    raise RuntimeError(f"formula/error token in {sheet.title}!{cell.coordinate}")
    verified.close()
    os.replace(temporary, path)


def run(args: argparse.Namespace) -> dict[str, Any]:
    root = args.output_root.expanduser().resolve(strict=True)
    bundle = args.bundle_root.expanduser().resolve(strict=True)
    repo = args.repo_root.expanduser().resolve(strict=True)
    reports = root / "reports"
    selection_dir = root / "selection"
    reports.mkdir(parents=True, exist_ok=True)
    selection_dir.mkdir(parents=True, exist_ok=True)
    formal_run_status = read_json_if_present(
        root / "diagnostics" / "formal_run_valid.json"
    )
    model_load_root = root / "diagnostics" / "model_load"
    model_load_statuses = {
        path.stem: read_json_if_present(path)
        for path in sorted(model_load_root.glob("*.json"))
    } if model_load_root.is_dir() else {}
    if model_load_statuses and not bool((formal_run_status or {}).get("valid")):
        raise RuntimeError(
            "model-load diagnostics exist but formal_run_valid.json does not "
            "confirm MODEL_LOAD_OK for all eight current physical workers"
        )
    oom_summary = read_json_if_present(root / "reports" / "oom_summary.json")
    raw_alignment, expected_total = validate_raw_alignment(root, bundle)
    scorer = load_module(
        repo / "qwen3vl_merge_and_score_fixed_5tasks.py",
        f"ui5_aggregate_formal_scorer_{os.getpid()}",
    )
    accumulators: dict[tuple[str, int, str, float], Accumulator] = defaultdict(Accumulator)
    for model in MODELS:
        for rollout in range(4):
            for scope in (*TASKS, "micro"):
                for threshold in THRESHOLDS:
                    accumulators[(model, rollout, scope, threshold)]
    error_counts = Counter()
    consistency: dict[str, dict[str, dict[str, Any]]] = {model: {} for model in MODELS}
    correct_distributions = Counter()
    selection_counts = Counter()
    execution_counts = Counter()
    runtime_error_types = Counter()
    runtime_excluded_groups = Counter()
    parse_excluded_groups = Counter()
    run_metadata: dict[str, dict[str, Any]] = {}
    for model in MODELS:
        for group in grouped_rollouts(root, model):
            first = group[0]
            if model not in run_metadata:
                run_metadata[model] = {
                    "checkpoint": first["checkpoint"],
                    "git_commit": first["git_commit"],
                    "baseline_git_commit": first["baseline_git_commit"],
                    "generation_config": first["generation_config"],
                    "seeds": [int(row["seed"]) for row in group],
                }
            elif [int(row["seed"]) for row in group] != run_metadata[model]["seeds"]:
                raise RuntimeError(f"{model} rollout seeds changed between records")
            if any(
                row["checkpoint"] != run_metadata[model]["checkpoint"]
                or row["git_commit"] != run_metadata[model]["git_commit"]
                or row["generation_config"] != run_metadata[model]["generation_config"]
                for row in group
            ):
                raise RuntimeError(f"{model} checkpoint/code/generation identity changed")
            task = str(first["task"])
            if task not in TASKS:
                raise ValueError(f"unknown task in raw rollout: {task}")
            threshold_scores: dict[tuple[int, float], dict[str, Any]] = {}
            width = int(first["image_size"]["width"])
            height = int(first["image_size"]["height"])
            for row in group:
                rollout_id = int(row["rollout_id"])
                for scope in (task, "micro"):
                    execution_counts[(model, rollout_id, scope, "attempted")] += 1
                runtime_error = row.get("runtime_error")
                if runtime_error:
                    runtime_type = str(runtime_error.get("type") or "UNKNOWN")
                    for scope in (task, "micro"):
                        execution_counts[(model, rollout_id, scope, "runtime_error")] += 1
                        runtime_error_types[
                            (model, rollout_id, scope, runtime_type)
                        ] += 1
                    continue
                for scope in (task, "micro"):
                    execution_counts[(model, rollout_id, scope, "inference_success")] += 1
                has_parse_error = bool(
                    row.get("parse_status") == "parse_error"
                    or row.get("contains_crop_parse_error")
                )
                if has_parse_error:
                    for scope in (task, "micro"):
                        execution_counts[(model, rollout_id, scope, "parse_error")] += 1
                    # Parsing failures are technical failures in v6.  Preserve
                    # them in execution/error reports, but never manufacture a
                    # TP/FP/FN score or count them as a wrong model answer.
                    continue
                for threshold in THRESHOLDS:
                    score = score_prediction(
                        scorer,
                        row["gt_global"],
                        row["pred_global"],
                        row["parse_status"],
                        threshold,
                        (width, height),
                    )
                    threshold_scores[(rollout_id, threshold)] = score
                    accumulators[(model, rollout_id, task, threshold)].add(score)
                    accumulators[(model, rollout_id, "micro", threshold)].add(score)
                    if threshold == 0.1:
                        error_counts[(model, rollout_id, task, score["error_type"])] += 1
            group_has_runtime_error = any(row.get("runtime_error") for row in group)
            group_has_parse_error = any(
                row.get("parse_status") == "parse_error"
                or row.get("contains_crop_parse_error")
                for row in group
                if not row.get("runtime_error")
            )
            if group_has_runtime_error:
                runtime_excluded_groups[(model, task)] += 1
                runtime_excluded_groups[(model, "micro")] += 1
            if group_has_parse_error:
                parse_excluded_groups[(model, task)] += 1
                parse_excluded_groups[(model, "micro")] += 1
            if group_has_runtime_error or group_has_parse_error:
                continue
            correct_count = sum(
                bool(threshold_scores[(int(row["rollout_id"]), 0.1)]["exact_correct"])
                for row in group
            )
            correct_distributions[(model, task, correct_count)] += 1
            correct_distributions[(model, "micro", correct_count)] += 1
            base = selection_base(group, correct_count)
            base["model_id"] = model
            compact_gallery = []
            for row in group:
                compact_gallery.append(
                    {
                        "rollout_id": row["rollout_id"],
                        "pred_global": row["pred_global"],
                        "matched_pairs": row["matched_pairs"],
                        "error_type": row["error_type"],
                        "exact_correct": row["exact_correct"],
                        "crop_boundaries": [
                            item["crop_xyxy"] for item in row.get("crop_outputs", [])
                        ],
                    }
                )
            consistency[model][str(first["record_id"])] = {
                **base,
                "model_id": model,
                "error_types": sorted({str(row["error_type"]) for row in group}),
                "gallery_rollouts": compact_gallery,
            }

    common_ids = sorted(set(consistency["m31"]) & set(consistency["crop"]))
    joint_0to8 = Counter()
    joint_5x5 = Counter()
    cross_writers = {
        category: JsonlWriter(selection_dir / f"{category}.jsonl")
        for category in ("both_hard", "m31_better", "crop_better")
    }
    for record_id in common_ids:
        m31 = consistency["m31"][record_id]
        crop = consistency["crop"][record_id]
        m31_count, crop_count = int(m31["correct_count"]), int(crop["correct_count"])
        joint_0to8[m31_count + crop_count] += 1
        joint_5x5[(m31_count, crop_count)] += 1
        category = None
        if m31_count == 0 and crop_count == 0:
            category = "both_hard"
        elif m31_count > crop_count:
            category = "m31_better"
        elif crop_count > m31_count:
            category = "crop_better"
        if category:
            row = {
                "record_id": record_id,
                "sample_id": m31["sample_id"],
                "source_image_id": m31["source_image_id"],
                "task": m31["task"],
                "image_relpath": m31["image_relpath"],
                "prompt": m31["prompt"],
                "gt_global": m31["gt_global"],
                "source_records": m31["source_records"],
                "m31_correct_count": m31_count,
                "crop_correct_count": crop_count,
                "category": category,
            }
            cross_writers[category].write(row)
            selection_counts[("cross", category)] += 1
    for writer in cross_writers.values():
        writer.close()

    difficulty_records, _ = build_difficulty_records(root, bundle)
    difficulty_file_counts = write_difficulty_exports(
        selection_dir, difficulty_records
    )
    complete_difficulty_records = [
        row
        for row in difficulty_records
        if row.get("cross_model_complete8") is True
    ]
    difficulty_counts = Counter(
        str(row["difficulty"]) for row in complete_difficulty_records
    )
    difficulty_task_counts = Counter(
        (str(row["task"]), str(row["difficulty"]))
        for row in complete_difficulty_records
    )
    for difficulty in DIFFICULTIES:
        selection_counts[("cross_model_8", difficulty)] = difficulty_file_counts[
            difficulty
        ]
    selection_counts[("cross_model_8", "grpo_m31_ready")] = (
        difficulty_file_counts["grpo_m31_ready"]
    )
    selection_counts[("cross_model_8", "grpo_crop_ready")] = (
        difficulty_file_counts["grpo_crop_ready"]
    )

    per_task: list[dict[str, Any]] = []
    per_rollout: list[dict[str, Any]] = []
    for (model, rollout, scope, threshold), accumulator in sorted(accumulators.items()):
        row = {
            "model_id": model,
            "rollout_id": rollout,
            "scope": scope,
            "iou_threshold": threshold,
            **accumulator.metrics(),
        }
        if scope == "micro":
            per_rollout.append(row)
        else:
            row["task"] = scope
            per_task.append(row)

    execution_rows = []
    for model in MODELS:
        for rollout in range(4):
            for scope in (*TASKS, "micro"):
                attempted = execution_counts[(model, rollout, scope, "attempted")]
                inference_success = execution_counts[
                    (model, rollout, scope, "inference_success")
                ]
                runtime_error = execution_counts[
                    (model, rollout, scope, "runtime_error")
                ]
                parse_error = execution_counts[(model, rollout, scope, "parse_error")]
                execution_rows.append(
                    {
                        "model_id": model,
                        "rollout_id": rollout,
                        "scope": scope,
                        "attempted": attempted,
                        "inference_success": inference_success,
                        "runtime_error": runtime_error,
                        "parse_error": parse_error,
                        "inference_success_ratio": safe_div(
                            inference_success, attempted
                        ),
                        "runtime_error_ratio": safe_div(runtime_error, attempted),
                        "parse_error_ratio_of_inference_success": safe_div(
                            parse_error, inference_success
                        ),
                    }
                )
    runtime_error_rows = [
        {
            "model_id": model,
            "rollout_id": rollout,
            "scope": scope,
            "runtime_error_type": runtime_type,
            "samples": count,
            "ratio_of_attempted": safe_div(
                count,
                execution_counts[(model, rollout, scope, "attempted")],
            ),
        }
        for (model, rollout, scope, runtime_type), count in sorted(
            runtime_error_types.items()
        )
    ]

    four_rollout_summary: list[dict[str, Any]] = []
    for model in MODELS:
        for threshold in THRESHOLDS:
            micro_rows = [
                row
                for row in per_rollout
                if row["model_id"] == model and row["iou_threshold"] == threshold
            ]
            macro_rows = []
            for rollout in range(4):
                task_rows = [
                    row
                    for row in per_task
                    if row["model_id"] == model
                    and row["rollout_id"] == rollout
                    and row["iou_threshold"] == threshold
                ]
                macro_rows.append(macro_row(task_rows))
            four_rollout_summary.extend(
                [
                    {
                        "model_id": model,
                        "iou_threshold": threshold,
                        "aggregation": "micro",
                        **mean_std(micro_rows),
                    },
                    {
                        "model_id": model,
                        "iou_threshold": threshold,
                        "aggregation": "macro_task",
                        **mean_std(macro_rows),
                    },
                ]
            )

    correct_rows = [
        {
            "model_id": model,
            "task": task,
            "correct_count": count,
            "samples": correct_distributions[(model, task, count)],
            "proportion": safe_div(
                correct_distributions[(model, task, count)],
                sum(correct_distributions[(model, task, item)] for item in range(5)),
            ),
            "runtime_excluded_samples": runtime_excluded_groups[(model, task)],
            "parse_excluded_samples": parse_excluded_groups[(model, task)],
        }
        for model in MODELS
        for task in (*TASKS, "micro")
        for count in range(5)
    ]
    difficulty_rows = []
    for task in (*TASKS, "micro"):
        denominator = (
            len(complete_difficulty_records)
            if task == "micro"
            else sum(
                1
                for row in complete_difficulty_records
                if str(row["task"]) == task
            )
        )
        for difficulty in DIFFICULTIES:
            samples = (
                difficulty_counts[difficulty]
                if task == "micro"
                else difficulty_task_counts[(task, difficulty)]
            )
            difficulty_rows.append(
                {
                    "task": task,
                    "difficulty": difficulty,
                    "samples": samples,
                    "proportion": safe_div(samples, denominator),
                }
            )
    joint_rows = [
        {
            "table": "joint_0to8",
            "m31_correct_count": None,
            "crop_correct_count": None,
            "joint_correct_count": count,
            "samples": joint_0to8[count],
            "proportion": safe_div(joint_0to8[count], len(common_ids)),
        }
        for count in range(9)
    ] + [
        {
            "table": "m31_x_crop_5x5",
            "m31_correct_count": m31_count,
            "crop_correct_count": crop_count,
            "joint_correct_count": m31_count + crop_count,
            "samples": joint_5x5[(m31_count, crop_count)],
            "proportion": safe_div(joint_5x5[(m31_count, crop_count)], len(common_ids)),
        }
        for m31_count in range(5)
        for crop_count in range(5)
    ]
    error_rows = [
        {
            "model_id": model,
            "rollout_id": rollout,
            "task": task,
            "error_type": error_type,
            "samples": error_counts[(model, rollout, task, error_type)],
        }
        for model in MODELS
        for rollout in range(4)
        for task in TASKS
        for error_type in ERROR_TYPES
    ]

    gallery_writer = JsonlWriter(reports / "gallery_selection.jsonl")
    gallery_counts = Counter()
    for model in MODELS:
        for record_id in sorted(consistency[model]):
            row = consistency[model][record_id]
            count = int(row["correct_count"])
            categories = [
                "4_of_4" if count == 4 else ("0_of_4" if count == 0 else f"{count}_of_4")
            ]
            for error_type in row["error_types"]:
                if error_type == "FP_ONLY":
                    categories.append("FP")
                elif error_type in {"FN_NO_PRED", "LOC_WRONG"}:
                    categories.append(error_type)
                elif error_type.startswith("PARTIAL_"):
                    categories.append("PARTIAL")
            other = consistency["crop" if model == "m31" else "m31"].get(record_id)
            if other:
                own, peer = count, int(other["correct_count"])
                if own == 0 and peer == 0:
                    categories.append("both_hard")
                elif model == "m31" and own > peer or model == "crop" and peer > own:
                    categories.append("m31_better")
                elif model == "crop" and own > peer or model == "m31" and peer > own:
                    categories.append("crop_better")
            for category in sorted(set(categories)):
                key = (model, row["task"], category)
                if gallery_counts[key] >= 10:
                    continue
                gallery_writer.write(
                    {
                        "model_id": model,
                        "task": row["task"],
                        "category": category,
                        "record_id": record_id,
                        "source_image_id": row["source_image_id"],
                        "image_relpath": row["image_relpath"],
                        "gt_global": row["gt_global"],
                        "correct_count": count,
                        "rollouts": row["gallery_rollouts"],
                    }
                )
                gallery_counts[key] += 1
    gallery_writer.close()

    runtime_rows = []
    for model in MODELS:
        for rollout in range(4):
            progress_path = root / "progress" / model / f"rollout_{rollout}.jsonl"
            latest = last_jsonl_row(progress_path) or {}
            runtime_rows.append(
                {
                    "model_id": model,
                    "rollout_id": rollout,
                    "status": latest.get("status"),
                    "attempted": latest.get("attempted", latest.get("completed")),
                    "inference_success": latest.get("inference_success"),
                    "runtime_error": latest.get("runtime_error"),
                    "parse_error": latest.get("parse_error"),
                    "total": latest.get("total"),
                    "elapsed_seconds": latest.get("elapsed_seconds"),
                    "throughput_samples_per_second": latest.get(
                        "throughput_samples_per_second"
                    ),
                    "remaining_seconds": latest.get("remaining_seconds"),
                    "estimated_completion": latest.get("estimated_completion"),
                    "gpu_memory": latest.get("gpu_memory"),
                }
            )
    total_eta = last_jsonl_row(root / "progress" / "total_eta.jsonl")
    selection_count_rows = [
        {"scope": scope, "category": category, "samples": count}
        for (scope, category), count in sorted(selection_counts.items())
    ]
    analysis = {
        "schema_version": SCHEMA_VERSION,
        "created_at": utc_now(),
        "semantics": {
            "rollout_4_plus_4": "cross-model consistency only",
            "not_pass_at_8": True,
            "not_one_policy_grpo_group": True,
            "grpo_group_rule": "same model + same image/crop + same prompt, four answers",
            "official_main_iou_threshold": 0.1,
            "rescored_thresholds_without_inference": [0.3, 0.5],
            "image_confusion": "presence-level GT/pred emptiness; wrong location remains Image TP",
            "bbox_tn_defined": False,
            "difficulty_rule": (
                "only eight complete runtime/parse-error-free records enter "
                "easy/medium/hard"
            ),
            "grpo_ready_rule": (
                "medium sample; same-model complete4; source and parse eligible; "
                "four exact rewards contain both true and false"
            ),
        },
        "paths": {"output_root": str(root), "bundle_root": str(bundle)},
        "per_rollout": per_rollout,
        "per_task": per_task,
        "four_rollout_summary": four_rollout_summary,
        "correct_0to4": correct_rows,
        "difficulty_8route": difficulty_rows,
        "difficulty_file_counts": difficulty_file_counts,
        "joint_0to8_and_5x5": joint_rows,
        "common_image_task_intersection": len(common_ids),
        "error_subtypes": error_rows,
        "execution_counts": execution_rows,
        "runtime_errors": runtime_error_rows,
        "oom_summary_table": (
            [
                {
                    "metric": key,
                    "value": value,
                }
                for key, value in (oom_summary or {}).items()
                if key not in {"unique_oom_samples", "by_model_rollout_oom_events"}
            ]
        ),
        "oom_by_model_rollout": list(
            (oom_summary or {}).get("by_model_rollout_oom_events", [])
        ),
        "oom_sample_index": list((oom_summary or {}).get("unique_oom_samples", [])),
        "model_load_status": [
            row for row in model_load_statuses.values() if row is not None
        ],
        "raw_alignment": raw_alignment,
        "expected_total_per_rollout": expected_total,
        "selection_counts": selection_count_rows,
        "runtime_eta": runtime_rows,
        "total_eta_last_snapshot": total_eta,
        "formal_run_status": formal_run_status,
        "model_load_statuses": model_load_statuses,
        "oom_summary": oom_summary,
    }
    analysis_path = reports / "ui5_train_rollout_analysis.json"
    atomic_json(analysis_path, analysis)
    run_config = {
        "schema_version": SCHEMA_VERSION,
        "created_at": utc_now(),
        "models": {
            model: {
                "checkpoint": run_metadata.get(model, {}).get("checkpoint"),
                "git_commit": run_metadata.get(model, {}).get("git_commit"),
                "baseline_git_commit": run_metadata.get(model, {}).get(
                    "baseline_git_commit"
                ),
                "generation_config": run_metadata.get(model, {}).get(
                    "generation_config"
                ),
                "rollouts": 4,
            }
            for model in MODELS
        },
        "rollout_seeds": {
            model: run_metadata.get(model, {}).get("seeds", [])
            for model in MODELS
        },
        "generation": {
            "dtype": "bf16",
            "text_attention": "sdpa",
            "vision_attention": "flash_attention_2",
            "vision_blocks": "27/27",
            "mode": "hybrid",
            "generation_mode": "hybrid",
            "sampling": True,
            "temperature": 0.7,
            "top_p": 0.9,
            "top_k": 0,
            "repetition_penalty": 1.1,
            "max_seq_length": MAX_SEQ_LENGTH,
            "max_num_tokens_per_sample": MAX_NUM_TOKENS_PER_SAMPLE,
            "training_max_num_tokens": TRAINING_MAX_NUM_TOKENS,
            "training_max_num_tokens_record_only": TRAINING_MAX_NUM_TOKENS,
            "processor_in_token_limit": PROCESSOR_IN_TOKEN_LIMIT,
            "max_new_tokens": ROLLOUT_MAX_NEW_TOKENS,
            "effective_max_new_tokens_rule": (
                "min(512, 7268 - input_tokens)"
            ),
            "n_future_tokens": 6,
            "main_iou": 0.1,
            "rescore_iou": [0.3, 0.5],
        },
        "execution_architecture": execution_architecture(),
        "sample_order": {
            "policy": "sample_major_fixed_task_polarity_round_robin_v1",
            "task_order": list(TASKS),
            "polarity_order": ["positive", "negative"],
            "within_bucket_order": ["record_id", "sample_id"],
        },
        "bundle_root": str(bundle),
        "output_root": str(root),
        "formal_run_status": formal_run_status,
    }
    atomic_json(root / "run_config.snapshot.json", run_config)
    overview = {
        "title": "UI5 train dual-model 4+4 rollout analysis",
        "created_at": analysis["created_at"],
        "main_iou_threshold": 0.1,
        "rollout_semantics": "cross-model consistency; not pass@8",
        "grpo_semantics": "four answers from one model and one image/crop prompt only",
        "common_image_task_intersection": len(common_ids),
        "raw_results": str(root / "raw"),
        "analysis_json": str(analysis_path),
    }
    tables = {
        "overview": [],
        "per_rollout": per_rollout,
        "per_task": per_task,
        "image_presence_confusion": [
            {
                key: row[key]
                for key in (
                    "model_id",
                    "rollout_id",
                    "task",
                    "iou_threshold",
                    "total_samples",
                    "image_TP",
                    "image_TN",
                    "image_FP",
                    "image_FN",
                    "image_TP_ratio",
                    "image_TN_ratio",
                    "image_FP_ratio",
                    "image_FN_ratio",
                    "precision",
                    "recall",
                    "specificity",
                    "FPR",
                    "FNR",
                    "NPV",
                    "accuracy",
                    "balanced_accuracy",
                    "F1",
                )
            }
            for row in per_task
        ],
        "bbox_metrics": [
            {
                key: row[key]
                for key in (
                    "model_id",
                    "rollout_id",
                    "task",
                    "iou_threshold",
                    "bbox_TP",
                    "bbox_FP",
                    "bbox_FN",
                    "bbox_precision",
                    "bbox_recall",
                    "bbox_F1",
                    "mean_matched_iou",
                    "mean_center_distance_px",
                    "mean_center_distance_normalized",
                    "mean_pred_gt_area_ratio",
                )
            }
            for row in per_task
        ],
        "correct_0to4": correct_rows,
        "difficulty_8route": difficulty_rows,
        "joint_0to8": joint_rows,
        "error_subtypes": error_rows,
        "execution_counts": execution_rows,
        "runtime_errors": runtime_error_rows,
        "oom_summary": analysis["oom_summary_table"],
        "oom_by_model_rollout": analysis["oom_by_model_rollout"],
        "oom_sample_index": analysis["oom_sample_index"],
        "model_load_status": analysis["model_load_status"],
        "raw_alignment": raw_alignment,
        "selection_counts": selection_count_rows,
        "runtime_eta": runtime_rows,
        "four_rollout_summary": four_rollout_summary,
        "sample_index": [
            {
                "model_id": model,
                "task": task,
                "category": category,
                "samples": count,
            }
            for (model, task, category), count in sorted(gallery_counts.items())
        ],
    }
    workbook(reports / "ui5_train_rollout_analysis.xlsx", overview, tables)
    atomic_json(
        root / "summary.json",
        {
            "schema_version": SCHEMA_VERSION,
            "created_at": analysis["created_at"],
            "status": "completed",
            "expected_total_per_rollout": expected_total,
            "raw_alignment": raw_alignment,
            "execution_counts": execution_rows,
            "runtime_errors": runtime_error_rows,
            "formal_run_status": formal_run_status,
            "model_load_statuses": model_load_statuses,
            "oom_summary": oom_summary,
            "correct_0to4": correct_rows,
            "difficulty_8route": difficulty_rows,
            "difficulty_file_counts": difficulty_file_counts,
            "analysis_json": str(analysis_path),
            "analysis_excel": str(
                reports / "ui5_train_rollout_analysis.xlsx"
            ),
        },
    )
    print(json.dumps({"analysis": str(analysis_path), "common": len(common_ids)}, ensure_ascii=False))
    return analysis


def main(argv: Sequence[str] | None = None) -> int:
    run(parse_args(argv))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
