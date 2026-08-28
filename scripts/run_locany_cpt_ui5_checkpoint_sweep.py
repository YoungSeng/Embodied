#!/usr/bin/env python3
"""Evaluate many legacy LocateAnything CPT checkpoints on the five UI5 tasks.

The unit of GPU scheduling is a checkpoint, not a task.  Each worker loads one
checkpoint once and runs all five tasks sequentially.  With two GPUs this keeps
two checkpoints in flight without loading the same 3.5B model five times.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone
import json
import math
import os
from pathlib import Path
import queue
import re
import shlex
import subprocess
import sys
import threading
import time
from typing import Any, Iterable, Mapping, Sequence

from locany_ui5_common import TASK_JSONL, TASKS, parse_gpu_devices


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INFERENCE_SCRIPT = PROJECT_ROOT / "scripts" / "inference_ui_defect_locany.py"
DEFAULT_SCORER_SCRIPT = PROJECT_ROOT / "qwen3vl_merge_and_score_fixed_5tasks.py"
CHECKPOINT_RE = re.compile(r"^checkpoint-(\d+)$")


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def atomic_write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}-{threading.get_ident()}")
    try:
        temporary.write_text(value, encoding="utf-8", newline="")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def atomic_write_json(path: Path, value: Any) -> None:
    atomic_write_text(
        path,
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )


def threshold_tag(value: float) -> str:
    return f"iou-{value:g}".replace(".", "p")


def safe_prf(tp: int, fp: int, fn: int) -> dict[str, float]:
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = (
        2.0 * precision * recall / (precision + recall)
        if precision + recall
        else 0.0
    )
    return {"precision": precision, "recall": recall, "f1": f1}


def finite_or_none(value: Any) -> Any:
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def discover_checkpoints(run_dir: Path, requested_steps: Sequence[int] | None) -> list[tuple[int, Path]]:
    found: dict[int, Path] = {}
    for path in run_dir.glob("checkpoint-*"):
        match = CHECKPOINT_RE.fullmatch(path.name)
        if match and path.is_dir():
            found[int(match.group(1))] = path.resolve()
    if requested_steps:
        missing = sorted(set(int(step) for step in requested_steps) - set(found))
        if missing:
            raise FileNotFoundError(
                f"run directory does not contain requested checkpoints: {missing}; run_dir={run_dir}"
            )
        steps = sorted(set(int(step) for step in requested_steps))
    else:
        steps = sorted(found)
    if not steps:
        raise FileNotFoundError(f"no checkpoint-N directories found under {run_dir}")
    return [(step, found[step]) for step in steps]


def has_all_ui5_test_files(path: Path) -> bool:
    return path.is_dir() and all((path / filename).is_file() for filename in TASK_JSONL.values())


def resolve_ui5_input_dir(raw_value: str | None, run_dir: Path) -> tuple[Path, str]:
    """Resolve an explicit path or infer WORKSPACE/data from the CPT run path.

    ``--input-dir "$UI5_TEST_DIR"`` becomes an empty string when the shell
    variable was not set.  Treating that as ``Path('.')`` hid the real problem
    and produced a misleading error under the repository root.
    """

    raw = str(raw_value or "").strip()
    if raw:
        requested = Path(raw).expanduser().resolve()
        if has_all_ui5_test_files(requested):
            return requested, "explicit"
        expected = ", ".join(TASK_JSONL.values())
        raise FileNotFoundError(
            f"explicit --input-dir does not contain all five UI5 JSONL files: {requested}; "
            f"expected: {expected}"
        )

    candidates: list[tuple[str, Path]] = []
    for env_name in ("UI5_TEST_DIR", "EVAL_INPUT_DIR"):
        value = os.environ.get(env_name, "").strip()
        if value:
            candidates.append((f"environment:{env_name}", Path(value).expanduser().resolve()))
    # Expected layout: WORKSPACE/gui_models/RUN_NAME -> WORKSPACE/data.
    candidates.append(("inferred-from-run-dir", run_dir.parent.parent / "data"))
    for source, candidate in candidates:
        candidate = candidate.resolve()
        if has_all_ui5_test_files(candidate):
            print(f"[AUTO] UI5 test directory ({source}): {candidate}", flush=True)
            return candidate, source
    rendered = ", ".join(f"{source}={path}" for source, path in candidates)
    raise FileNotFoundError(
        "--input-dir was empty and no complete UI5 test directory could be inferred; "
        f"checked: {rendered}"
    )


def parse_external_metrics(values: Sequence[str]) -> list[tuple[str, Path]]:
    result: list[tuple[str, Path]] = []
    labels: set[str] = set()
    for value in values:
        if "=" not in value:
            raise ValueError(
                f"--external-metrics must be LABEL=/path/to/all_tasks_evaluation.json, got {value!r}"
            )
        label, raw_path = value.split("=", 1)
        label = label.strip()
        if not label:
            raise ValueError(f"external model label is empty: {value!r}")
        if label in labels:
            raise ValueError(f"duplicate external model label: {label}")
        path = Path(raw_path).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"external metric JSON does not exist: {path}")
        labels.add(label)
        result.append((label, path))
    return result


def load_metric_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"cannot read metric JSON {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise ValueError(f"metric JSON must contain an object: {path}")
    tasks = value.get("tasks")
    if not isinstance(tasks, dict) or set(tasks) != set(TASKS):
        raise ValueError(
            f"metric JSON must contain exactly the five UI5 tasks; "
            f"found={sorted(tasks) if isinstance(tasks, dict) else type(tasks).__name__}: {path}"
        )
    for task in TASKS:
        for granularity in ("image", "bbox"):
            group = tasks[task].get(granularity)
            if not isinstance(group, dict):
                raise ValueError(f"missing tasks.{task}.{granularity} in {path}")
            for key in ("precision", "recall", "f1", "tp", "fp", "fn"):
                if group.get(key) is None:
                    raise ValueError(f"missing tasks.{task}.{granularity}.{key} in {path}")
    return value


def read_log_tail(path: Path, lines: int) -> str:
    if not path.is_file() or lines <= 0:
        return ""
    max_bytes = 256 * 1024
    with path.open("rb") as handle:
        handle.seek(0, os.SEEK_END)
        size = handle.tell()
        handle.seek(max(0, size - max_bytes), os.SEEK_SET)
        payload = handle.read()
    text = payload.decode("utf-8", errors="replace")
    if size > max_bytes:
        text = text.split("\n", 1)[-1]
    return "\n".join(text.splitlines()[-lines:])


def run_logged(command: Sequence[str], *, env: Mapping[str, str], log_path: Path, dry_run: bool) -> int:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    rendered = shlex.join(str(item) for item in command)
    print(f"[COMMAND] {rendered}", flush=True)
    if dry_run:
        return 0
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(f"\n===== {utc_now()} =====\n{rendered}\n")
        handle.flush()
        completed = subprocess.run(
            list(command),
            env=dict(env),
            cwd=PROJECT_ROOT,
            stdout=handle,
            stderr=subprocess.STDOUT,
            check=False,
        )
    return int(completed.returncode)


def build_inference_command(args: argparse.Namespace, checkpoint: Path, prediction_dir: Path, gpu: str) -> list[str]:
    command = [
        args.python,
        str(args.inference_script),
        "--checkpoint",
        str(checkpoint),
        "--processor-path",
        str(args.processor_path),
        "--input-dir",
        str(args.input_dir),
        "--output-dir",
        str(prediction_dir),
        "--summary-path",
        str(prediction_dir / "_summary.json"),
        "--cuda-visible-devices",
        gpu,
        "--device",
        "cuda:0",
        "--dtype",
        args.dtype,
        "--attn-implementation",
        args.attn_implementation,
        "--vision-attn-implementation",
        args.vision_attn_implementation,
        "--generation-mode",
        args.generation_mode,
        "--max-new-tokens",
        str(args.max_new_tokens),
        "--n-future-tokens",
        str(args.n_future_tokens),
        "--seed",
        str(args.seed),
        "--tasks",
        "all",
        "--skip-figma",
        "--fail-fast",
        "--relation-gate-mode",
        "observe",
        "--enable-ui-relation" if args.enable_ui_relation else "--no-enable-ui-relation",
        "--enable-pbd" if args.enable_pbd else "--no-enable-pbd",
    ]
    if args.save_raw_answer:
        command.append("--save-raw-answer")
    if args.save_visualization:
        command.append("--save-visualization")
    if args.greedy:
        command.append("--greedy")
    if args.max_images_per_task:
        command.extend(["--max-images-per-task", str(args.max_images_per_task)])
    if args.overwrite:
        command.append("--overwrite")
    return command


def build_score_command(
    args: argparse.Namespace,
    *,
    prediction_dir: Path,
    output_root: Path,
    run_name: str,
) -> list[str]:
    return [
        args.python,
        str(args.scorer_script),
        "--all_tasks",
        "--input_mode",
        "yolo_dir",
        "--gt_dir",
        str(args.input_dir),
        "--pred_root",
        str(prediction_dir),
        "--output_root",
        str(output_root),
        "--run_name",
        run_name,
        "--yolo_bbox_format",
        "xyxy",
        "--iou_thresh",
        str(args.iou_threshold),
    ]


def totals_for(metrics: Mapping[str, Any], granularity: str) -> dict[str, Any]:
    groups = [metrics["tasks"][task][granularity] for task in TASKS]
    tp = sum(int(group.get("tp", 0)) for group in groups)
    fp = sum(int(group.get("fp", 0)) for group in groups)
    fn = sum(int(group.get("fn", 0)) for group in groups)
    result: dict[str, Any] = {**safe_prf(tp, fp, fn), "tp": tp, "fp": fp, "fn": fn}
    if granularity == "image":
        tn = sum(int(group.get("tn", 0)) for group in groups)
        total = tp + fp + fn + tn
        result.update({"tn": tn, "accuracy": (tp + tn) / total if total else 0.0})
    else:
        sample_weights = [
            int(metrics["tasks"][task]["image"].get("tp", 0))
            + int(metrics["tasks"][task]["image"].get("fp", 0))
            + int(metrics["tasks"][task]["image"].get("fn", 0))
            + int(metrics["tasks"][task]["image"].get("tn", 0))
            for task in TASKS
        ]
        weight = sum(sample_weights)
        result["count_accuracy"] = (
            sum(
                float(metrics["tasks"][task]["bbox"].get("count_accuracy", 0.0))
                * sample_weights[index]
                for index, task in enumerate(TASKS)
            )
            / weight
            if weight
            else 0.0
        )
    return result


def normalized_model_result(
    *,
    label: str,
    kind: str,
    step: int | None,
    checkpoint: Path | None,
    prediction_dir: Path | None,
    metrics_path: Path,
    metrics: Mapping[str, Any],
    iou_threshold: float,
) -> dict[str, Any]:
    macro = metrics.get("macro", {})
    return {
        "label": label,
        "kind": kind,
        "step": step,
        "checkpoint": str(checkpoint) if checkpoint is not None else None,
        "prediction_dir": str(prediction_dir) if prediction_dir is not None else None,
        "metrics_json": str(metrics_path),
        "iou_threshold": iou_threshold,
        "tasks": {task: metrics["tasks"][task] for task in TASKS},
        "macro": {
            granularity: {
                key: finite_or_none(value)
                for key, value in (macro.get(granularity) or {}).items()
            }
            for granularity in ("image", "bbox")
        },
        "micro": {
            granularity: totals_for(metrics, granularity)
            for granularity in ("image", "bbox")
        },
    }


def metric_rows(models: Sequence[Mapping[str, Any]], granularity: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for model in models:
        for task in TASKS:
            values = model["tasks"][task][granularity]
            rows.append(
                {
                    "label": model["label"],
                    "kind": model["kind"],
                    "step": model["step"],
                    "iou_threshold": model["iou_threshold"],
                    "task": task,
                    "aggregation": "task",
                    **{key: finite_or_none(value) for key, value in values.items()},
                }
            )
        for aggregation in ("macro", "micro"):
            values = model[aggregation][granularity]
            rows.append(
                {
                    "label": model["label"],
                    "kind": model["kind"],
                    "step": model["step"],
                    "iou_threshold": model["iou_threshold"],
                    "task": f"__{aggregation}__",
                    "aggregation": aggregation,
                    **{key: finite_or_none(value) for key, value in values.items()},
                }
            )
    return rows


def write_csv(path: Path, rows: Sequence[Mapping[str, Any]], columns: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    try:
        with temporary.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(columns), extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({column: "" if row.get(column) is None else row.get(column) for column in columns})
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def write_comparison_excel(
    path: Path,
    overview_rows: Sequence[Mapping[str, Any]],
    image_rows: Sequence[Mapping[str, Any]],
    bbox_rows: Sequence[Mapping[str, Any]],
) -> str | None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        return f"openpyxl unavailable: {exc}"

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    specs = (
        ("Overview", overview_rows),
        ("ImageMetrics", image_rows),
        ("BBoxMetrics", bbox_rows),
    )
    try:
        for name, rows in specs:
            sheet = workbook.create_sheet(name)
            columns = list(rows[0]) if rows else ["label"]
            sheet.append(columns)
            for row in rows:
                sheet.append([finite_or_none(row.get(column)) for column in columns])
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column_cells in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_name(f".{path.stem}.tmp-{os.getpid()}{path.suffix}")
        try:
            workbook.save(temporary)
            verification = openpyxl.load_workbook(temporary, read_only=True)
            try:
                if verification.sheetnames != ["Overview", "ImageMetrics", "BBoxMetrics"]:
                    raise ValueError(f"unexpected workbook sheets: {verification.sheetnames}")
            finally:
                verification.close()
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)
    except Exception as exc:  # Excel is a convenience artifact, JSON/CSV remain authoritative.
        return f"{type(exc).__name__}: {exc}"
    finally:
        workbook.close()
    return None


def write_comparison(
    args: argparse.Namespace,
    checkpoints: Sequence[tuple[int, Path]],
    external: Sequence[tuple[str, Path]],
) -> dict[str, Any]:
    tag = threshold_tag(args.iou_threshold)
    models: list[dict[str, Any]] = []
    for step, checkpoint in checkpoints:
        metrics_path = args.output_root / "metrics" / f"checkpoint-{step}" / f"{tag}.json"
        if not metrics_path.is_file():
            continue
        prediction_dir = args.output_root / "predictions" / f"checkpoint-{step}"
        models.append(
            normalized_model_result(
                label=f"cpt-step-{step}",
                kind="cpt_checkpoint",
                step=step,
                checkpoint=checkpoint,
                prediction_dir=prediction_dir,
                metrics_path=metrics_path,
                metrics=load_metric_json(metrics_path),
                iou_threshold=args.iou_threshold,
            )
        )
    for label, metrics_path in external:
        models.append(
            normalized_model_result(
                label=label,
                kind="external_model",
                step=None,
                checkpoint=None,
                prediction_dir=None,
                metrics_path=metrics_path,
                metrics=load_metric_json(metrics_path),
                iou_threshold=args.iou_threshold,
            )
        )
    if not models:
        return {"models": [], "artifacts": {}}

    comparison_dir = args.output_root / "comparison"
    base_name = f"checkpoint_comparison_{tag}"
    image_rows = metric_rows(models, "image")
    bbox_rows = metric_rows(models, "bbox")
    overview_rows: list[dict[str, Any]] = []
    for model in models:
        overview_rows.append(
            {
                "label": model["label"],
                "kind": model["kind"],
                "step": model["step"],
                "iou_threshold": model["iou_threshold"],
                "image_macro_precision": model["macro"]["image"].get("precision"),
                "image_macro_recall": model["macro"]["image"].get("recall"),
                "image_macro_f1": model["macro"]["image"].get("f1"),
                "image_micro_precision": model["micro"]["image"].get("precision"),
                "image_micro_recall": model["micro"]["image"].get("recall"),
                "image_micro_f1": model["micro"]["image"].get("f1"),
                "bbox_macro_precision": model["macro"]["bbox"].get("precision"),
                "bbox_macro_recall": model["macro"]["bbox"].get("recall"),
                "bbox_macro_f1": model["macro"]["bbox"].get("f1"),
                "bbox_micro_precision": model["micro"]["bbox"].get("precision"),
                "bbox_micro_recall": model["micro"]["bbox"].get("recall"),
                "bbox_micro_f1": model["micro"]["bbox"].get("f1"),
                "checkpoint": model["checkpoint"],
                "prediction_dir": model["prediction_dir"],
                "metrics_json": model["metrics_json"],
            }
        )

    json_path = comparison_dir / f"{base_name}.json"
    overview_csv = comparison_dir / f"{base_name}_overview.csv"
    image_csv = comparison_dir / f"{base_name}_image.csv"
    bbox_csv = comparison_dir / f"{base_name}_bbox.csv"
    excel_path = comparison_dir / f"{base_name}.xlsx"
    payload = {
        "schema_version": 1,
        "created_at": utc_now(),
        "iou_threshold": args.iou_threshold,
        "tasks": list(TASKS),
        "models": models,
    }
    atomic_write_json(json_path, payload)
    write_csv(overview_csv, overview_rows, list(overview_rows[0]))
    metric_columns = (
        "label",
        "kind",
        "step",
        "iou_threshold",
        "task",
        "aggregation",
        "precision",
        "recall",
        "f1",
        "tp",
        "fp",
        "fn",
        "tn",
        "accuracy",
        "count_accuracy",
    )
    write_csv(image_csv, image_rows, metric_columns)
    write_csv(bbox_csv, bbox_rows, metric_columns)
    excel_warning = write_comparison_excel(excel_path, overview_rows, image_rows, bbox_rows)
    if excel_warning:
        print(f"[WARN] comparison Excel was not written: {excel_warning}", file=sys.stderr)
    return {
        "models": models,
        "artifacts": {
            "json": str(json_path),
            "overview_csv": str(overview_csv),
            "image_csv": str(image_csv),
            "bbox_csv": str(bbox_csv),
            "xlsx": None if excel_warning else str(excel_path),
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run legacy CPT checkpoint sweep on five UI defect test sets",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--run-dir", required=True, help="directory containing checkpoint-N")
    parser.add_argument("--processor-path", required=True, help="base LocateAnything processor/model directory")
    parser.add_argument(
        "--input-dir",
        default=None,
        help=(
            "directory containing five test_ui_* JSONL files; an empty/omitted value "
            "uses UI5_TEST_DIR, EVAL_INPUT_DIR, then WORKSPACE/data inferred from --run-dir"
        ),
    )
    parser.add_argument("--output-root", type=Path, default=None)
    parser.add_argument("--steps", nargs="*", type=int, default=None, help="checkpoint steps; omit to discover all")
    parser.add_argument("--gpu-devices", default="0,1", help="physical GPU IDs; one checkpoint per GPU")
    parser.add_argument(
        "--stage",
        choices=("infer-score", "infer-only", "score-only", "aggregate-only"),
        default="infer-score",
    )
    parser.add_argument("--python", default=sys.executable)
    parser.add_argument("--inference-script", type=Path, default=DEFAULT_INFERENCE_SCRIPT)
    parser.add_argument("--scorer-script", type=Path, default=DEFAULT_SCORER_SCRIPT)
    parser.add_argument("--attn-implementation", choices=("sdpa", "flash_attention_2", "eager", "magi", "auto"), default="sdpa")
    parser.add_argument("--vision-attn-implementation", choices=("sdpa", "flash_attention_2", "eager"), default="flash_attention_2")
    parser.add_argument("--dtype", choices=("bf16", "fp16", "fp32"), default="bf16")
    parser.add_argument("--generation-mode", choices=("fast", "slow", "hybrid"), default="hybrid")
    parser.add_argument("--max-new-tokens", type=int, default=4096)
    parser.add_argument("--n-future-tokens", type=int, default=6)
    parser.add_argument("--seed", type=int, default=20260817)
    parser.add_argument("--max-images-per-task", type=int, default=0, help="0 means full test sets")
    parser.add_argument("--iou-threshold", type=float, default=0.1, help="keep 0.1 to match the existing UI5/Qwen scorer; use 0.5 for stricter reports")
    parser.add_argument("--greedy", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--save-raw-answer", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--save-visualization", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--enable-ui-relation", action=argparse.BooleanOptionalAction, default=False, help="legacy CPT checkpoints should keep this disabled")
    parser.add_argument("--enable-pbd", action=argparse.BooleanOptionalAction, default=False, help="legacy CPT checkpoints should keep relation-to-PBD disabled")
    parser.add_argument("--external-metrics", action="append", default=[], metavar="LABEL=JSON", help="append another model's canonical all_tasks_evaluation.json to comparison outputs")
    parser.add_argument("--failure-log-lines", type=int, default=120)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    if args.max_new_tokens <= 0 or args.n_future_tokens <= 0:
        parser.error("--max-new-tokens and --n-future-tokens must be positive")
    if args.max_images_per_task < 0:
        parser.error("--max-images-per-task cannot be negative")
    if not 0.0 < args.iou_threshold <= 1.0:
        parser.error("--iou-threshold must be in (0, 1]")
    if args.failure_log_lines < 0:
        parser.error("--failure-log-lines cannot be negative")
    return args


def main() -> int:
    args = parse_args()
    raw_run_dir = str(args.run_dir or "").strip()
    raw_processor_path = str(args.processor_path or "").strip()
    if not raw_run_dir:
        raise ValueError(
            "--run-dir expanded to an empty string; define RUN_DIR or pass the absolute path"
        )
    if not raw_processor_path:
        raise ValueError(
            "--processor-path expanded to an empty string; define BASE_MODEL or pass the absolute path"
        )
    args.run_dir = Path(raw_run_dir).expanduser().resolve()
    args.processor_path = Path(raw_processor_path).expanduser().resolve()
    args.input_dir, input_dir_source = resolve_ui5_input_dir(args.input_dir, args.run_dir)
    args.output_root = (
        args.output_root.expanduser().resolve()
        if args.output_root is not None
        else args.run_dir / "ui5_checkpoint_sweep"
    )
    args.inference_script = args.inference_script.expanduser().resolve()
    args.scorer_script = args.scorer_script.expanduser().resolve()
    for path, label in (
        (args.run_dir, "run directory"),
        (args.processor_path, "processor path"),
        (args.input_dir, "test input directory"),
    ):
        if not path.is_dir():
            raise FileNotFoundError(f"{label} does not exist: {path}")
    if args.stage in {"infer-score", "infer-only"} and not args.inference_script.is_file():
        raise FileNotFoundError(f"inference script does not exist: {args.inference_script}")
    if args.stage in {"infer-score", "score-only"} and not args.scorer_script.is_file():
        raise FileNotFoundError(f"scorer script does not exist: {args.scorer_script}")
    for task, filename in TASK_JSONL.items():
        path = args.input_dir / filename
        if not path.is_file():
            raise FileNotFoundError(f"missing {task} test JSONL: {path}")

    checkpoints = discover_checkpoints(args.run_dir, args.steps)
    external = parse_external_metrics(args.external_metrics)
    gpus = parse_gpu_devices(args.gpu_devices)
    args.output_root.mkdir(parents=True, exist_ok=True)
    tag = threshold_tag(args.iou_threshold)
    status_path = args.output_root / "sweep_status.json"
    status_lock = threading.Lock()
    status: dict[str, Any] = {
        "schema_version": 1,
        "run_dir": str(args.run_dir),
        "output_root": str(args.output_root),
        "input_dir": str(args.input_dir),
        "input_dir_source": input_dir_source,
        "processor_path": str(args.processor_path),
        "steps": [step for step, _ in checkpoints],
        "gpu_devices": gpus,
        "one_checkpoint_per_gpu": True,
        "stage": args.stage,
        "iou_threshold": args.iou_threshold,
        "legacy_cpt_overrides": {
            "enable_ui_relation": bool(args.enable_ui_relation),
            "enable_pbd": bool(args.enable_pbd),
        },
        "checkpoints": {},
        "started_at": utc_now(),
        "success": False,
    }
    atomic_write_json(status_path, status)

    print("===== Legacy CPT -> UI5 checkpoint sweep =====")
    print(f"run dir              : {args.run_dir}")
    print(f"steps                : {', '.join(str(step) for step, _ in checkpoints)}")
    print(f"physical GPUs        : {', '.join(gpus)}")
    print("scheduler            : one checkpoint per GPU; five tasks per model load")
    print(f"IoU threshold        : {args.iou_threshold:g}")
    print(f"output               : {args.output_root}")

    def set_checkpoint_status(step: int, values: Mapping[str, Any]) -> None:
        with status_lock:
            status["checkpoints"][str(step)] = dict(values)
            atomic_write_json(status_path, status)

    work: queue.Queue[tuple[int, Path]] = queue.Queue()
    if args.stage != "aggregate-only":
        for item in checkpoints:
            work.put(item)
    stop = threading.Event()

    def worker(gpu: str) -> None:
        while not stop.is_set():
            try:
                step, checkpoint = work.get_nowait()
            except queue.Empty:
                return
            started = time.time()
            prediction_dir = args.output_root / "predictions" / f"checkpoint-{step}"
            metrics_path = args.output_root / "metrics" / f"checkpoint-{step}" / f"{tag}.json"
            log_path = args.output_root / "logs" / f"checkpoint-{step}.log"
            details: dict[str, Any] = {
                "step": step,
                "checkpoint": str(checkpoint),
                "physical_gpu": gpu,
                "prediction_dir": str(prediction_dir),
                "metrics_json": str(metrics_path),
                "status": "running",
                "started_at": utc_now(),
            }
            set_checkpoint_status(step, details)
            print(f"[START] step={step} physical_gpu={gpu}", flush=True)
            try:
                if args.stage in {"infer-score", "infer-only"}:
                    command = build_inference_command(args, checkpoint, prediction_dir, gpu)
                    child_env = dict(os.environ)
                    child_env["CUDA_VISIBLE_DEVICES"] = gpu
                    child_env["PYTHONUNBUFFERED"] = "1"
                    child_env.setdefault("TOKENIZERS_PARALLELISM", "false")
                    child_env.setdefault("PYTORCH_CUDA_ALLOC_CONF", "expandable_segments:True")
                    return_code = run_logged(command, env=child_env, log_path=log_path, dry_run=args.dry_run)
                    details["inference_return_code"] = return_code
                    if return_code:
                        raise RuntimeError(f"inference exited with code {return_code}")

                if args.stage in {"infer-score", "score-only"}:
                    if not prediction_dir.is_dir() and not args.dry_run:
                        raise FileNotFoundError(f"prediction directory does not exist: {prediction_dir}")
                    if metrics_path.is_file() and not args.overwrite:
                        load_metric_json(metrics_path)
                        details["score_cache_hit"] = True
                        print(f"[SCORE CACHE] step={step} metrics={metrics_path}", flush=True)
                    else:
                        attempt = datetime.now().strftime("attempt-%Y%m%d-%H%M%S-%f")
                        score_root = args.output_root / "score_runs" / tag / f"checkpoint-{step}"
                        score_command = build_score_command(
                            args,
                            prediction_dir=prediction_dir,
                            output_root=score_root,
                            run_name=attempt,
                        )
                        score_env = dict(os.environ)
                        score_env["PYTHONUNBUFFERED"] = "1"
                        score_code = run_logged(
                            score_command,
                            env=score_env,
                            log_path=log_path,
                            dry_run=args.dry_run,
                        )
                        details["score_return_code"] = score_code
                        if score_code:
                            raise RuntimeError(f"scorer exited with code {score_code}")
                        if not args.dry_run:
                            produced = score_root / attempt / "all_tasks_evaluation.json"
                            metrics = load_metric_json(produced)
                            metrics["evaluation_metadata"] = {
                                "iou_threshold": args.iou_threshold,
                                "checkpoint": str(checkpoint),
                                "step": step,
                                "prediction_dir": str(prediction_dir),
                                "scorer_output": str(produced),
                                "created_at": utc_now(),
                            }
                            atomic_write_json(metrics_path, metrics)
                            details["score_cache_hit"] = False
                            details["scorer_output"] = str(produced)

                details.update(
                    {
                        "status": "dry_run" if args.dry_run else "success",
                        "finished_at": utc_now(),
                        "elapsed_seconds": round(time.time() - started, 6),
                    }
                )
                set_checkpoint_status(step, details)
                print(
                    f"[DONE] step={step} gpu={gpu} elapsed={details['elapsed_seconds']:.1f}s",
                    flush=True,
                )
            except Exception as exc:
                tail = read_log_tail(log_path, args.failure_log_lines)
                details.update(
                    {
                        "status": "failed",
                        "error": f"{type(exc).__name__}: {exc}",
                        "log_tail": tail,
                        "finished_at": utc_now(),
                        "elapsed_seconds": round(time.time() - started, 6),
                    }
                )
                set_checkpoint_status(step, details)
                print(
                    f"[FAILED] step={step} gpu={gpu}: {details['error']}\n{tail}",
                    file=sys.stderr,
                    flush=True,
                )
                stop.set()
            finally:
                work.task_done()

    threads = [threading.Thread(target=worker, args=(gpu,), name=f"gpu-{gpu}") for gpu in gpus]
    try:
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()
    except KeyboardInterrupt:
        stop.set()
        print("[STOP] interrupted; completed per-image predictions are resumable", file=sys.stderr)
        for thread in threads:
            thread.join(timeout=5)
        raise

    comparison = write_comparison(args, checkpoints, external)
    failed = [row for row in status["checkpoints"].values() if row.get("status") == "failed"]
    if args.stage == "aggregate-only":
        run_success = bool(comparison["models"])
    elif args.dry_run:
        run_success = not failed
    else:
        expected_status = "success"
        run_success = not failed and len(status["checkpoints"]) == len(checkpoints) and all(
            row.get("status") == expected_status for row in status["checkpoints"].values()
        )
    status.update(
        {
            "success": run_success,
            "finished_at": utc_now(),
            "comparison_artifacts": comparison["artifacts"],
            "comparison_model_count": len(comparison["models"]),
        }
    )
    atomic_write_json(status_path, status)
    print("===== Sweep result =====")
    print(f"success              : {run_success}")
    print(f"status               : {status_path}")
    for name, path in comparison["artifacts"].items():
        if path:
            print(f"comparison {name:8s}: {path}")
    return 0 if run_success else 1


if __name__ == "__main__":
    raise SystemExit(main())
