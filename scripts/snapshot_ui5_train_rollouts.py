#!/usr/bin/env python3
"""Create cumulative time snapshots from technically valid UI5 rollout8 groups."""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import shutil
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from run_ui5_train_rollout_worker import TASKS, fixed_interleaved_samples


SCHEMA_VERSION = 4
MODELS = ("m31", "crop")
ROLLOUT_IDS = (0, 1, 2, 3)
FORMAL_SEEDS = (20260903, 20260917, 20260931, 20260947)
DIFFICULTIES = ("easy", "medium", "hard")
CLASSIFICATION_FILES = {
    "easy": "easy.jsonl",
    "medium": "medium.jsonl",
    "hard": "hard.jsonl",
}
GRPO_FILES = {
    "m31": "grpo_m31_ready.jsonl",
    "crop": "grpo_crop_ready.jsonl",
}


def execution_architecture() -> dict[str, Any]:
    """Return the immutable v6 physical-process topology as JSON metadata."""
    return {
        "physical_processes_total": 8,
        "physical_processes_per_gpu": 4,
        "logical_rollouts_total": 8,
        "rollouts_per_physical_process": 1,
        "checkpoint_loads_per_physical_process": 1,
        "fixed_sample_order_shared_by_all_processes": True,
        "rollout_execution": "one_rollout_per_physical_process_fixed_sample_order",
        "global_eta_reduction": "maximum_estimated_completion_across_8_processes",
        "gpu_0": {
            "model_id": "m31",
            "processes": [
                {
                    "process_index": rollout_id + 1,
                    "rollout_id": rollout_id,
                    "rollout_ids": [rollout_id],
                    "seed": FORMAL_SEEDS[rollout_id],
                }
                for rollout_id in ROLLOUT_IDS
            ],
        },
        "gpu_1": {
            "model_id": "crop",
            "processes": [
                {
                    "process_index": rollout_id + 1,
                    "rollout_id": rollout_id,
                    "rollout_ids": [rollout_id],
                    "seed": FORMAL_SEEDS[rollout_id],
                }
                for rollout_id in ROLLOUT_IDS
            ],
        },
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--bundle-root", type=Path, required=True)
    parser.add_argument("--kind", choices=("hourly", "final"), required=True)
    parser.add_argument("--scheduled-hour", type=int)
    parser.add_argument("--started-at-epoch", type=float, required=True)
    parser.add_argument("--export-selection-dir", type=Path)
    return parser.parse_args(argv)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def timestamp_token(epoch: float | None = None) -> str:
    value = time.time() if epoch is None else epoch
    return datetime.fromtimestamp(value, timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def atomic_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    try:
        temporary.write_text(
            json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def atomic_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    try:
        temporary.write_text(value, encoding="utf-8", newline="\n")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def atomic_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    count = 0
    try:
        with temporary.open("w", encoding="utf-8", newline="\n") as handle:
            for row in rows:
                handle.write(
                    json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n"
                )
                count += 1
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
    return count


def atomic_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}")
    columns = list(rows[0]) if rows else ["record_id"]
    try:
        with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        column: (
                            json.dumps(row.get(column), ensure_ascii=False, sort_keys=True)
                            if isinstance(row.get(column), (dict, list))
                            else row.get(column)
                        )
                        for column in columns
                    }
                )
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
    return len(rows)


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, 1):
            if not line.strip():
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise ValueError(f"expected object at {path}:{line_no}")
            rows.append(value)
    return rows


def visible_jsonl_rows(path: Path) -> list[dict[str, Any]]:
    """Read only newline-terminated records while a worker may still append."""
    payload = path.read_bytes()
    final_newline = payload.rfind(b"\n")
    if final_newline < 0:
        return []
    payload = payload[: final_newline + 1]
    rows: list[dict[str, Any]] = []
    for line_no, raw_line in enumerate(payload.splitlines(), 1):
        if not raw_line.strip():
            continue
        try:
            value = json.loads(raw_line.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValueError(f"invalid completed JSONL record at {path}:{line_no}") from exc
        if not isinstance(value, dict):
            raise ValueError(f"expected object at {path}:{line_no}")
        rows.append(value)
    return rows


def bundle_samples(bundle_root: Path) -> list[dict[str, Any]]:
    manifest = json.loads(
        (bundle_root / "bundle_manifest.json").read_text(encoding="utf-8")
    )
    samples = fixed_interleaved_samples(
        read_jsonl(bundle_root / "manifest" / "task_samples.jsonl")
    )
    expected_total = int(manifest["rollout_samples"])
    if len(samples) != expected_total:
        raise RuntimeError(
            f"bundle sample count mismatch: manifest={expected_total} rows={len(samples)}"
        )
    record_ids = [str(row["record_id"]) for row in samples]
    if len(set(record_ids)) != len(record_ids):
        raise RuntimeError("bundle contains duplicate record_id values")
    return samples


def load_visible_stream(
    output_root: Path,
    model: str,
    rollout_id: int,
    expected_ids: set[str],
) -> dict[str, dict[str, Any]]:
    directory = output_root / "raw" / model / f"rollout_{rollout_id}"
    rows: dict[str, dict[str, Any]] = {}
    for part in sorted(directory.glob("part-*.jsonl")) if directory.is_dir() else []:
        for row in visible_jsonl_rows(part):
            record_id = str(row.get("record_id"))
            if record_id not in expected_ids:
                raise RuntimeError(
                    f"{model} rollout {rollout_id} has unknown record_id={record_id}"
                )
            if str(row.get("model_id")) != model or int(row.get("rollout_id", -1)) != rollout_id:
                raise RuntimeError(
                    f"raw stream identity mismatch: {part} record_id={record_id}"
                )
            if record_id in rows:
                raise RuntimeError(
                    f"duplicate record_id in {model} rollout {rollout_id}: {record_id}"
                )
            rows[record_id] = row
    return rows


def compact_runtime_error(row: Mapping[str, Any]) -> dict[str, Any] | None:
    runtime_error = row.get("runtime_error")
    if not isinstance(runtime_error, Mapping):
        return None
    return {
        "type": runtime_error.get("type"),
        "python_type": runtime_error.get("python_type"),
        "message": runtime_error.get("message"),
    }


def snapshot_rollout_payload(
    model: str,
    rollout_id: int,
    row: Mapping[str, Any] | None,
) -> dict[str, Any]:
    if row is None:
        return {
            "model_id": model,
            "rollout_id": rollout_id,
            "status": "missing",
            "reward": None,
            "exact_correct": None,
            "raw_output": None,
            "pred_local": None,
            "pred_global": None,
            "parse_status": "not_available",
            "runtime_error": None,
            "oom_recovered": False,
            "oom_events": 0,
        }
    return {
        "model_id": model,
        "rollout_id": int(row["rollout_id"]),
        "seed": row.get("seed"),
        "status": "runtime_error" if row.get("runtime_error") else "completed",
        "reward": row.get("exact_correct"),
        "exact_correct": row.get("exact_correct"),
        "raw_output": row.get("raw_output"),
        "pred_local": row.get("pred_local"),
        "pred_global": row.get("pred_global"),
        "gt_local": row.get("gt_local"),
        "gt_global": row.get("gt_global"),
        "matched_pairs": row.get("matched_pairs") or [],
        "TP_box": row.get("TP_box"),
        "FP_box": row.get("FP_box"),
        "FN_box": row.get("FN_box"),
        "image_confusion": row.get("image_confusion"),
        "error_type": row.get("error_type"),
        "parse_status": row.get("parse_status"),
        "contains_crop_parse_error": bool(row.get("contains_crop_parse_error")),
        "parse_warnings": row.get("parse_warnings") or [],
        "runtime_error": row.get("runtime_error"),
        "oom_recovered": bool(row.get("oom_recovered")),
        "oom_events": int(row.get("oom_events", 0)),
        "oom_final_failure": bool(row.get("oom_final_failure")),
        "oom_retry": row.get("oom_retry"),
        "token_usage": row.get("token_usage"),
        "crop_outputs": row.get("crop_outputs") or [],
        "latency_seconds": row.get("latency_seconds"),
    }


def model_group_payload(
    model: str,
    rows: Sequence[Mapping[str, Any]],
    base: Mapping[str, Any],
) -> dict[str, Any]:
    ordered = sorted(rows, key=lambda row: int(row["rollout_id"]))
    return {
        **base,
        "group_model_id": model,
        "group_key": str(base["record_id"]),
        "answers": [row.get("raw_output") for row in ordered],
        "rewards_exact": [bool(row.get("exact_correct")) for row in ordered],
        "rollouts": [
            snapshot_rollout_payload(model, int(row["rollout_id"]), row)
            for row in ordered
        ],
        "rollout_ids": [int(row["rollout_id"]) for row in ordered],
        "seeds": [int(row["seed"]) for row in ordered],
        "group_size": 4,
        "cross_model_group": False,
    }


def cumulative_metrics(
    streams: Mapping[tuple[str, int], Mapping[str, Mapping[str, Any]]],
    complete8_record_ids: set[str],
) -> list[dict[str, Any]]:
    """Score only routes belonging to technically valid, complete-eight samples."""
    counters: dict[tuple[str, str, str, str], Counter[str]] = defaultdict(Counter)
    sample_ids: dict[tuple[str, str, str, str], set[str]] = defaultdict(set)

    def keys_for(model: str, rollout_id: int, task: str):
        rollout = str(rollout_id)
        yield ("overall", "all", "all", "all")
        yield ("model", model, "all", "all")
        yield ("task", "all", "all", task)
        yield ("rollout", "all", rollout, "all")
        yield ("model_task", model, "all", task)
        yield ("model_rollout", model, rollout, "all")
        yield ("model_rollout_task", model, rollout, task)

    for (model, rollout_id), stream in streams.items():
        for record_id in complete8_record_ids:
            row = stream[record_id]
            task = str(row.get("task"))
            for key in keys_for(model, rollout_id, task):
                counter = counters[key]
                sample_ids[key].add(record_id)
                counter["scored_predictions"] += 1
                confusion = row.get("image_confusion")
                if confusion in {"TP", "TN", "FP", "FN"}:
                    counter[confusion] += 1
                counter["bbox_TP"] += int(row.get("TP_box") or 0)
                counter["bbox_FP"] += int(row.get("FP_box") or 0)
                counter["bbox_FN"] += int(row.get("FN_box") or 0)
                counter["exact_correct"] += int(row.get("exact_correct") is True)

    expected_keys = {("overall", "all", "all", "all")}
    expected_keys.update(("model", model, "all", "all") for model in MODELS)
    expected_keys.update(("task", "all", "all", task) for task in TASKS)
    expected_keys.update(
        ("rollout", "all", str(rollout_id), "all")
        for rollout_id in ROLLOUT_IDS
    )
    expected_keys.update(
        ("model_task", model, "all", task)
        for model in MODELS
        for task in TASKS
    )
    expected_keys.update(
        ("model_rollout", model, str(rollout_id), "all")
        for model in MODELS
        for rollout_id in ROLLOUT_IDS
    )
    expected_keys.update(
        ("model_rollout_task", model, str(rollout_id), task)
        for model in MODELS
        for rollout_id in ROLLOUT_IDS
        for task in TASKS
    )
    for key in expected_keys:
        counters[key]

    rows: list[dict[str, Any]] = []
    for key in sorted(counters):
        level, model, rollout, task = key
        counter = counters[key]
        scored = counter["TP"] + counter["TN"] + counter["FP"] + counter["FN"]
        tp, tn, fp, fn = (
            counter["TP"],
            counter["TN"],
            counter["FP"],
            counter["FN"],
        )
        precision = tp / (tp + fp) if tp + fp else 0.0
        recall = tp / (tp + fn) if tp + fn else 0.0
        specificity = tn / (tn + fp) if tn + fp else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
        rows.append(
            {
                "level": level,
                "model_id": model,
                "rollout_id": rollout,
                "task": task,
                "scope": task if task != "all" else "micro",
                "complete8_samples": len(sample_ids[key]),
                "scored_predictions": counter["scored_predictions"],
                "TP": tp,
                "TN": tn,
                "FP": fp,
                "FN": fn,
                "image_TP": tp,
                "image_TN": tn,
                "image_FP": fp,
                "image_FN": fn,
                "TP_ratio": tp / scored if scored else 0.0,
                "TN_ratio": tn / scored if scored else 0.0,
                "FP_ratio": fp / scored if scored else 0.0,
                "FN_ratio": fn / scored if scored else 0.0,
                "image_TP_ratio": tp / scored if scored else 0.0,
                "image_TN_ratio": tn / scored if scored else 0.0,
                "image_FP_ratio": fp / scored if scored else 0.0,
                "image_FN_ratio": fn / scored if scored else 0.0,
                "bbox_TP": counter["bbox_TP"],
                "bbox_FP": counter["bbox_FP"],
                "bbox_FN": counter["bbox_FN"],
                "precision": precision,
                "recall": recall,
                "F1": f1,
                "f1": f1,
                "accuracy": (tp + tn) / scored if scored else 0.0,
                "specificity": specificity,
                "exact_correct": counter["exact_correct"],
                "exact_correct_ratio": (
                    counter["exact_correct"] / scored if scored else 0.0
                ),
            }
        )
    return rows


def rollout_technical_issues(row: Mapping[str, Any]) -> list[str]:
    issues: list[str] = []
    if row.get("runtime_error"):
        issues.append("runtime_error")
    if row.get("inference_success") is not True:
        issues.append("inference_not_successful")
    parse_status = row.get("parse_status")
    if parse_status == "parse_error":
        issues.append("parse_error")
    elif not isinstance(parse_status, str) or not parse_status:
        issues.append("missing_parse_status")
    if row.get("contains_crop_parse_error"):
        issues.append("crop_parse_error")
    if row.get("oom_final_failure"):
        issues.append("oom_final_failure")
    if not isinstance(row.get("exact_correct"), bool):
        issues.append("missing_reward")
    if row.get("image_confusion") not in {"TP", "TN", "FP", "FN"}:
        issues.append("missing_image_confusion")
    return issues


def build_difficulty_records(
    output_root: Path, bundle_root: Path
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    samples = bundle_samples(bundle_root)
    expected_ids = {str(row["record_id"]) for row in samples}
    streams = {
        (model, rollout_id): load_visible_stream(
            output_root, model, rollout_id, expected_ids
        )
        for model in MODELS
        for rollout_id in ROLLOUT_IDS
    }
    records: list[dict[str, Any]] = []
    for sample in samples:
        record_id = str(sample["record_id"])
        by_model = {
            model: [
                streams[(model, rollout_id)][record_id]
                for rollout_id in ROLLOUT_IDS
                if record_id in streams[(model, rollout_id)]
            ]
            for model in MODELS
        }
        for model in MODELS:
            for row in by_model[model]:
                if str(row.get("sample_id")) != str(sample.get("sample_id")):
                    raise RuntimeError(
                        f"{model} raw sample_id mismatch for record_id={record_id}"
                    )
                if str(row.get("task")) != str(sample.get("task")):
                    raise RuntimeError(
                        f"{model} raw task mismatch for record_id={record_id}"
                    )
                expected_image = sample.get("source_image_id", sample.get("image_id"))
                if str(row.get("source_image_id")) != str(expected_image):
                    raise RuntimeError(
                        f"{model} raw source_image_id mismatch for record_id={record_id}"
                    )
        runtime_errors = [
            {
                "model_id": model,
                "rollout_id": int(row["rollout_id"]),
                **(compact_runtime_error(row) or {}),
            }
            for model in MODELS
            for row in by_model[model]
            if row.get("runtime_error")
        ]
        completed = {model: len(by_model[model]) for model in MODELS}
        correct = {
            model: sum(row.get("exact_correct") is True for row in by_model[model])
            for model in MODELS
        }
        technical_issues = {
            model: {
                int(row["rollout_id"]): rollout_technical_issues(row)
                for row in by_model[model]
                if rollout_technical_issues(row)
            }
            for model in MODELS
        }
        complete4 = {
            model: bool(
                completed[model] == 4
                and all(
                    not rollout_technical_issues(row)
                    for row in by_model[model]
                )
            )
            for model in MODELS
        }
        complete8 = bool(complete4["m31"] and complete4["crop"])
        total_correct = correct["m31"] + correct["crop"]
        if not complete8:
            difficulty = None
        elif total_correct == 8:
            difficulty = "easy"
        elif total_correct == 0:
            difficulty = "hard"
        else:
            difficulty = "medium"
        first_raw = next(
            (row for model in MODELS for row in by_model[model]), None
        )
        anomaly_flags = {
            "pipeline_coverage_failure": bool(
                (first_raw or sample).get("pipeline_coverage_failure")
            ),
            "annotation_anomaly": bool(
                (first_raw or sample).get("annotation_anomaly")
            ),
            "coordinate_transform_anomaly": bool(
                (first_raw or sample).get("coordinate_transform_anomaly")
            ),
        }
        sample_grpo_eligible = bool(
            sample.get("grpo_eligible", not any(anomaly_flags.values()))
            and not any(anomaly_flags.values())
        )
        parse_clean = {
            model: all(
                row.get("parse_status") != "parse_error"
                and not row.get("contains_crop_parse_error")
                for row in by_model[model]
            )
            for model in MODELS
        }
        grpo_m31 = bool(
            difficulty == "medium"
            and complete4["m31"]
            and 1 <= correct["m31"] <= 3
            and sample_grpo_eligible
            and parse_clean["m31"]
        )
        grpo_crop = bool(
            difficulty == "medium"
            and complete4["crop"]
            and 1 <= correct["crop"] <= 3
            and sample_grpo_eligible
            and parse_clean["crop"]
        )
        base = {
            "record_id": record_id,
            "sample_id": str(sample.get("sample_id", record_id)),
            "source_image_id": sample.get("source_image_id", sample.get("image_id")),
            "task": sample.get("task"),
            "image_relpath": sample.get("image_relpath"),
            "prompt": sample.get("prompt"),
            "gt_global": sample.get("gt_global"),
            "source_records": sample.get("source_records", []),
            "original_training_record": sample.get("original_training_record"),
            "m31_correct_count": correct["m31"],
            "crop_correct_count": correct["crop"],
            "total_correct_count": total_correct,
            "success_rate": total_correct / 8.0 if complete8 else None,
            "difficulty": difficulty,
            "grpo_ready_m31": grpo_m31,
            "grpo_ready_crop": grpo_crop,
            "grpo_source_eligible": sample_grpo_eligible,
            "grpo_parse_clean_m31": parse_clean["m31"],
            "grpo_parse_clean_crop": parse_clean["crop"],
            "m31_complete4": complete4["m31"],
            "crop_complete4": complete4["crop"],
            "m31_completed_rollout_count": completed["m31"],
            "crop_completed_rollout_count": completed["crop"],
            "completed_rollout_count": completed["m31"] + completed["crop"],
            "runtime_error_count": len(runtime_errors),
            "runtime_errors": runtime_errors,
            "parse_error_count": sum(
                bool(
                    row.get("parse_status") == "parse_error"
                    or row.get("contains_crop_parse_error")
                )
                for model in MODELS
                for row in by_model[model]
                if not row.get("runtime_error")
            ),
            "cross_model_complete8": complete8,
            "technical_error_free": complete8,
            "technical_issues": technical_issues,
            "exclusion_reason": (
                None
                if complete8
                else (
                    "technical_error"
                    if any(technical_issues[model] for model in MODELS)
                    else "incomplete_eight_routes"
                )
            ),
            "rollouts": {
                model: [
                    snapshot_rollout_payload(
                        model,
                        rollout_id,
                        streams[(model, rollout_id)].get(record_id),
                    )
                    for rollout_id in ROLLOUT_IDS
                ]
                for model in MODELS
            },
            **anomaly_flags,
            "visualization_rollouts": {
                model: [
                    {
                        "model_id": model,
                        "rollout_id": int(row["rollout_id"]),
                        "pred_global": row.get("pred_global") or [],
                        "matched_pairs": row.get("matched_pairs") or [],
                        "error_type": (
                            f"{model}:{row.get('error_type') or 'INCOMPLETE'}"
                        ),
                        "exact_correct": row.get("exact_correct") is True,
                        "image_confusion": row.get("image_confusion"),
                        "TP_box": row.get("TP_box"),
                        "FP_box": row.get("FP_box"),
                        "FN_box": row.get("FN_box"),
                        "runtime_error": row.get("runtime_error"),
                        "oom_recovered": bool(row.get("oom_recovered")),
                        "oom_events": int(row.get("oom_events", 0)),
                        "crop_boundaries": [
                            item["crop_xyxy"]
                            for item in row.get("crop_outputs", [])
                            if item.get("crop_xyxy")
                        ],
                    }
                    for row in sorted(
                        by_model[model], key=lambda item: int(item["rollout_id"])
                    )
                ]
                for model in MODELS
            },
        }
        group_base = {
            key: value for key, value in base.items() if key != "visualization_rollouts"
        }
        if grpo_m31:
            base["grpo_m31_group"] = model_group_payload(
                "m31", by_model["m31"], group_base
            )
        if grpo_crop:
            base["grpo_crop_group"] = model_group_payload(
                "crop", by_model["crop"], group_base
            )
        records.append(base)
    stream_rows = []
    for model in MODELS:
        for rollout_id in ROLLOUT_IDS:
            stream = streams[(model, rollout_id)]
            stream_rows.append(
                {
                    "model_id": model,
                    "rollout_id": rollout_id,
                    "visible_records": len(stream),
                    "expected_records": len(samples),
                    "complete": len(stream) == len(samples),
                }
            )
    complete8_ids = {
        str(row["record_id"])
        for row in records
        if row.get("cross_model_complete8") is True
    }
    raw_record_ids = {
        record_id
        for stream in streams.values()
        for record_id in stream
    }
    return records, {
        "expected_total": len(samples),
        "raw_streams": stream_rows,
        "raw_records_visible": sum(len(stream) for stream in streams.values()),
        "expected_raw_records": len(samples) * len(MODELS) * len(ROLLOUT_IDS),
        "attempted_unique_samples": len(raw_record_ids),
        "eight_raw_records_samples": sum(
            all(str(row["record_id"]) in stream for stream in streams.values())
            for row in records
        ),
        "complete8_samples": len(complete8_ids),
        "incomplete_or_technical_error_samples": len(records) - len(complete8_ids),
        "cumulative_metrics": cumulative_metrics(streams, complete8_ids),
        "correct_count_4": correct_count_4_rows(records),
        "correct_count_8": correct_count_8_rows(records),
    }


def correct_count_4_rows(records: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for model in MODELS:
        complete_field = f"{model}_complete4"
        count_field = f"{model}_correct_count"
        for scope in (*TASKS, "micro"):
            scoped = [
                row
                for row in records
                if row.get("cross_model_complete8") is True
                and (scope == "micro" or row.get("task") == scope)
            ]
            complete = [row for row in scoped if row.get(complete_field)]
            for correct_count in range(5):
                count = sum(
                    int(row.get(count_field, -1)) == correct_count for row in complete
                )
                result.append(
                    {
                        "model_id": model,
                        "scope": scope,
                        "correct_count_4": correct_count,
                        "samples": count,
                        "complete4_samples": len(complete),
                        "incomplete_samples": len(scoped) - len(complete),
                        "proportion_of_complete4": (
                            count / len(complete) if complete else 0.0
                        ),
                    }
                )
    return result


def correct_count_8_rows(records: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for scope in (*TASKS, "micro"):
        scoped = [row for row in records if scope == "micro" or row.get("task") == scope]
        complete = [row for row in scoped if row.get("cross_model_complete8")]
        for correct_count in range(9):
            count = sum(
                int(row.get("total_correct_count", -1)) == correct_count
                for row in complete
            )
            result.append(
                {
                    "scope": scope,
                    "correct_count_8": correct_count,
                    "samples": count,
                    "complete8_samples": len(complete),
                    "incomplete_samples": len(scoped) - len(complete),
                    "proportion_of_complete8": count / len(complete) if complete else 0.0,
                }
            )
    return result


def write_error_exports(
    destination: Path,
    output_root: Path,
    records: Sequence[Mapping[str, Any]],
) -> dict[str, int]:
    runtime_rows: list[dict[str, Any]] = []
    parse_rows: list[dict[str, Any]] = []
    oom_rows: list[dict[str, Any]] = []
    for record in records:
        identity = {
            key: record.get(key)
            for key in ("record_id", "sample_id", "source_image_id", "task", "image_relpath")
        }
        for model in MODELS:
            for rollout in record.get("rollouts", {}).get(model, []):
                row = {**identity, **rollout}
                if rollout.get("runtime_error"):
                    runtime_rows.append(row)
                if rollout.get("parse_status") == "parse_error" or rollout.get(
                    "contains_crop_parse_error"
                ):
                    parse_rows.append(row)
                if int(rollout.get("oom_events", 0)) > 0 or rollout.get("oom_recovered"):
                    oom_rows.append(row)
    model_load_rows = []
    model_load_root = output_root / "diagnostics" / "model_load"
    for path in sorted(model_load_root.glob("*.json")) if model_load_root.is_dir() else []:
        row = json.loads(path.read_text(encoding="utf-8"))
        if row.get("status") != "MODEL_LOAD_OK":
            model_load_rows.append(row)
    error_root = destination / "errors"
    return {
        "runtime_errors": atomic_jsonl(error_root / "runtime_errors.jsonl", runtime_rows),
        "parse_errors": atomic_jsonl(error_root / "parse_errors.jsonl", parse_rows),
        "oom_events": atomic_jsonl(error_root / "oom_events.jsonl", oom_rows),
        "model_load_errors": atomic_jsonl(
            error_root / "model_load_errors.jsonl", model_load_rows
        ),
    }


def classification_projection(row: Mapping[str, Any]) -> dict[str, Any]:
    excluded = {"grpo_m31_group", "grpo_crop_group", "visualization_rollouts"}
    return {key: value for key, value in row.items() if key not in excluded}


def difficulty_index_projection(row: Mapping[str, Any]) -> dict[str, Any]:
    return {
        key: row.get(key)
        for key in (
            "record_id",
            "sample_id",
            "source_image_id",
            "task",
            "image_relpath",
            "m31_correct_count",
            "crop_correct_count",
            "total_correct_count",
            "success_rate",
            "difficulty",
            "m31_complete4",
            "crop_complete4",
            "cross_model_complete8",
            "grpo_ready_m31",
            "grpo_ready_crop",
        )
    }


def write_difficulty_exports(
    destination: Path,
    records: Sequence[Mapping[str, Any]],
    *,
    previous_records: Mapping[str, Mapping[str, Any]] | None = None,
    write_delta: bool = False,
) -> dict[str, int]:
    destination.mkdir(parents=True, exist_ok=True)
    counts: dict[str, int] = {}
    complete = [row for row in records if row.get("cross_model_complete8") is True]
    incomplete = [row for row in records if row.get("cross_model_complete8") is not True]
    projected_complete = [classification_projection(row) for row in complete]
    difficulty_rows = [difficulty_index_projection(row) for row in complete]
    counts["complete8"] = atomic_jsonl(
        destination / "complete8.jsonl", projected_complete
    )
    counts["sample_difficulty_jsonl"] = atomic_jsonl(
        destination / "sample_difficulty.jsonl", difficulty_rows
    )
    counts["sample_difficulty_csv"] = atomic_csv(
        destination / "sample_difficulty.csv", difficulty_rows
    )
    # Keep the generic sample stream as a complete8-only compatibility alias.
    counts["snapshot_samples"] = atomic_jsonl(
        destination / "samples.jsonl",
        projected_complete,
    )
    counts["incomplete_or_technical_error"] = atomic_jsonl(
        destination / "incomplete_or_technical_error.jsonl",
        (classification_projection(row) for row in incomplete),
    )
    for difficulty, filename in CLASSIFICATION_FILES.items():
        selected = [
            classification_projection(row)
            for row in complete
            if row["difficulty"] == difficulty
        ]
        counts[difficulty] = atomic_jsonl(destination / filename, selected)
    for model, filename in GRPO_FILES.items():
        flag = f"grpo_ready_{model}"
        group = f"grpo_{model}_group"
        selected = [row[group] for row in records if row.get(flag)]
        counts[f"grpo_{model}_ready"] = atomic_jsonl(destination / filename, selected)
    forbidden = destination / "cross_model_complete8.jsonl"
    if forbidden.exists():
        raise RuntimeError(
            f"obsolete cross_model_complete8.jsonl must not exist: {forbidden}"
        )
    if write_delta:
        previous = previous_records or {}
        delta_rows = []
        for row in complete:
            projected = classification_projection(row)
            old = previous.get(str(row["record_id"]))
            if old == projected:
                continue
            delta_rows.append(
                {
                    **projected,
                    "delta_change": "new" if old is None else "updated",
                    "previous_state": (
                        None
                        if old is None
                        else {
                            key: old.get(key)
                            for key in (
                                "m31_correct_count",
                                "crop_correct_count",
                                "total_correct_count",
                                "success_rate",
                                "difficulty",
                                "grpo_ready_m31",
                                "grpo_ready_crop",
                                "completed_rollout_count",
                                "runtime_error_count",
                            )
                        }
                    ),
                }
            )
        counts["delta_since_previous"] = atomic_jsonl(
            destination / "delta_since_previous.jsonl", delta_rows
        )
    return counts


def read_snapshot_records(snapshot: Path) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    path = snapshot / "complete8.jsonl"
    if not path.is_file():
        path = snapshot / "samples.jsonl"
    if path.is_file():
        for row in read_jsonl(path):
            records[str(row["record_id"])] = row
    return records


def previous_snapshot(snapshots_root: Path) -> Path | None:
    candidates = [
        path
        for path in snapshots_root.iterdir()
        if path.is_dir()
        and not path.name.startswith(".")
        and (path / "summary.json").is_file()
        and (path / "manifest.json").is_file()
        and (path / "_SUCCESS").is_file()
    ] if snapshots_root.is_dir() else []
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime_ns)


def current_processing_status(
    output_root: Path,
    cumulative: Mapping[str, Any],
    now_epoch: float,
) -> dict[str, Any]:
    eta_history = output_root / "progress" / "total_eta.jsonl"
    latest_eta = None
    if eta_history.is_file():
        rows = visible_jsonl_rows(eta_history)
        latest_eta = rows[-1] if rows else None

    logical_progress: list[dict[str, Any]] = []
    for model in MODELS:
        for rollout_id in ROLLOUT_IDS:
            path = output_root / "progress" / model / f"rollout_{rollout_id}.jsonl"
            rows = visible_jsonl_rows(path) if path.is_file() else []
            latest = rows[-1] if rows else None
            logical_progress.append(
                {
                    "model_id": model,
                    "rollout_id": rollout_id,
                    "status": (latest or {}).get("status", "missing"),
                    "attempted": int((latest or {}).get("attempted", 0)),
                    "total": int((latest or {}).get("total", cumulative["expected_total"])),
                    "throughput_attempted_per_second": float(
                        (latest or {}).get(
                            "throughput_attempted_per_second",
                            (latest or {}).get("throughput_samples_per_second", 0.0),
                        )
                    ),
                    "remaining_seconds": (latest or {}).get("remaining_seconds"),
                    "estimated_completion": (latest or {}).get("estimated_completion"),
                    "progress_history_rows": len(rows),
                }
            )

    remaining_seconds = (
        latest_eta.get("total_remaining_seconds")
        if isinstance(latest_eta, Mapping)
        else None
    )
    estimated_completion = (
        latest_eta.get("total_estimated_completion")
        if isinstance(latest_eta, Mapping)
        else None
    )
    route_eta_fallback = False
    if remaining_seconds is None:
        route_remaining = [
            float(row["remaining_seconds"])
            for row in logical_progress
            if isinstance(row.get("remaining_seconds"), (int, float))
        ]
        if len(route_remaining) == len(MODELS) * len(ROLLOUT_IDS):
            route_eta_fallback = True
            remaining_seconds = max(route_remaining, default=0.0)
            estimated_completion = datetime.fromtimestamp(
                now_epoch + remaining_seconds, timezone.utc
            ).isoformat()

    expected_total = int(cumulative["expected_total"])
    complete8 = int(cumulative["complete8_samples"])
    return {
        "expected_samples": expected_total,
        "attempted_unique_samples": int(cumulative["attempted_unique_samples"]),
        "raw_records_visible": int(cumulative["raw_records_visible"]),
        "expected_raw_records": int(cumulative["expected_raw_records"]),
        "eight_raw_records_samples": int(cumulative["eight_raw_records_samples"]),
        "complete8_samples": complete8,
        "remaining_samples_until_complete8": max(0, expected_total - complete8),
        "remaining_seconds": remaining_seconds,
        "estimated_completion": estimated_completion,
        "eta_valid": bool(
            remaining_seconds is not None
            and (
                route_eta_fallback
                or (
                    isinstance(latest_eta, Mapping)
                    and latest_eta.get("eta_valid") is True
                )
            )
        ),
        "eta_source": (
            "progress/total_eta.jsonl"
            if latest_eta is not None and not route_eta_fallback
            else "logical_rollout_progress_max"
            if route_eta_fallback
            else None
        ),
        "eta_snapshot": latest_eta,
        "logical_rollouts": logical_progress,
    }


def write_snapshot_workbook(
    path: Path,
    summary: Mapping[str, Any],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for incremental snapshot tables") from exc
    book = openpyxl.Workbook()
    book.remove(book.active)
    tables: dict[str, list[dict[str, Any]]] = {
        "overview": [
            {"metric": "snapshot_kind", "value": summary["snapshot_kind"]},
            {"metric": "scheduled_hour", "value": summary["scheduled_hour"]},
            {"metric": "started_at", "value": summary["started_at"]},
            {"metric": "created_at", "value": summary["created_at"]},
            {"metric": "elapsed_seconds", "value": summary["elapsed_seconds"]},
            {"metric": "expected_total", "value": summary["expected_total"]},
            {"metric": "complete8_samples", "value": summary["complete8_samples"]},
            {
                "metric": "remaining_seconds",
                "value": summary["processing_status"]["remaining_seconds"],
            },
            {
                "metric": "estimated_completion",
                "value": summary["processing_status"]["estimated_completion"],
            },
            {"metric": "previous_snapshot", "value": summary["previous_snapshot"]},
        ],
        "difficulty": [
            {
                "difficulty": difficulty,
                "samples": summary["difficulty_counts"][difficulty],
                "proportion": summary["difficulty_ratios"][difficulty],
            }
            for difficulty in DIFFICULTIES
        ],
        "per_task_difficulty": list(summary["per_task_difficulty"]),
        "raw_streams": list(summary["raw_streams"]),
        "processing_status": list(summary["processing_status"]["logical_rollouts"]),
        "cumulative_metrics": list(summary["cumulative_metrics"]),
        "correct_count_4": list(summary["correct_count_4"]),
        "correct_count_8": list(summary["correct_count_8"]),
        "error_counts": [
            {"error_file": name, "records": count}
            for name, count in summary.get("error_counts", {}).items()
        ],
        "selection_files": [
            {"file": filename, "records": count}
            for filename, count in sorted(summary["file_counts"].items())
        ],
    }
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in tables.items():
        sheet = book.create_sheet(sheet_name[:31])
        columns = list(rows[0]) if rows else ["note"]
        sheet.append(columns)
        for row in rows or [{"note": "no rows"}]:
            sheet.append(
                [
                    json.dumps(row.get(column), ensure_ascii=False, sort_keys=True)
                    if isinstance(row.get(column), (dict, list))
                    else row.get(column)
                    for column in columns
                ]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
        for cells in sheet.columns:
            width = min(48, max(10, max(len(str(cell.value or "")) for cell in cells) + 2))
            sheet.column_dimensions[cells[0].column_letter].width = width
    temporary = path.with_name(f".{path.name}.tmp-{os.getpid()}.xlsx")
    book.save(temporary)
    verified = openpyxl.load_workbook(temporary, read_only=True, data_only=False)
    if set(tables) != set(verified.sheetnames):
        raise RuntimeError("incremental snapshot workbook verification failed")
    verified.close()
    os.replace(temporary, path)


def render_snapshot_gallery(
    snapshot_root: Path,
    bundle_root: Path,
    records: Sequence[Mapping[str, Any]],
    *,
    panel_long_side: int = 420,
) -> dict[str, Any]:
    from PIL import Image
    from render_ui5_train_rollout_gallery import render_panel

    visual_root = snapshot_root / "visualizations"
    counts: Counter[str] = Counter()
    seen_images: dict[str, set[str]] = defaultdict(set)
    cards: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    for record in records:
        categories = [str(record["difficulty"])]
        visualization_map = record.get("visualization_rollouts", {})
        flattened_rollouts = [
            rollout
            for model in MODELS
            for rollout in visualization_map.get(model, [])
        ]
        if any(
            rollout.get("image_confusion") == "FP"
            or int(rollout.get("FP_box") or 0) > 0
            for rollout in flattened_rollouts
        ):
            categories.append("FP")
        if any(
            rollout.get("image_confusion") == "FN"
            or int(rollout.get("FN_box") or 0) > 0
            for rollout in flattened_rollouts
        ):
            categories.append("FN")
        for category in categories:
            task = str(record["task"])
            image_id = str(record["source_image_id"])
            if len(seen_images[category]) >= 10 or image_id in seen_images[category]:
                continue
            seen_images[category].add(image_id)
            rollout_map = visualization_map
            rollouts = [
                rollout
                for model in MODELS
                for rollout in rollout_map.get(model, [])
                if category != "FP"
                or rollout.get("image_confusion") == "FP"
                or int(rollout.get("FP_box") or 0) > 0
                if category != "FN"
                or rollout.get("image_confusion") == "FN"
                or int(rollout.get("FN_box") or 0) > 0
            ]
            source_path = bundle_root / str(record["image_relpath"])
            try:
                with Image.open(source_path) as opened:
                    source = opened.convert("RGB")
                if not rollouts:
                    rollouts = [
                        {
                            "rollout_id": "none",
                            "pred_global": [],
                            "matched_pairs": [],
                            "error_type": "INCOMPLETE",
                            "exact_correct": False,
                            "crop_boundaries": [],
                        }
                    ]
                panels = [
                    render_panel(
                        source,
                        record.get("gt_global") or [],
                        rollout,
                        panel_long_side,
                    )
                    for rollout in rollouts
                ]
                source.close()
                columns = min(4, len(panels))
                rows = (len(panels) + columns - 1) // columns
                gap = 10
                cell_width = max(panel.width for panel in panels)
                cell_height = max(panel.height for panel in panels)
                composite = Image.new(
                    "RGB",
                    (
                        columns * cell_width + (columns - 1) * gap,
                        rows * cell_height + (rows - 1) * gap,
                    ),
                    (238, 242, 247),
                )
                for index, panel in enumerate(panels):
                    x = (index % columns) * (cell_width + gap)
                    y = (index // columns) * (cell_height + gap)
                    composite.paste(panel, (x, y))
                    panel.close()
                relative = (
                    Path("assets") / category / task / f"{record['record_id']}.jpg"
                )
                destination = visual_root / relative
                destination.parent.mkdir(parents=True, exist_ok=True)
                temporary = destination.with_name(
                    f".{destination.name}.tmp-{os.getpid()}"
                )
                composite.save(temporary, format="JPEG", quality=88, subsampling=0)
                composite.close()
                os.replace(temporary, destination)
                cards.append(
                    {
                        "category": category,
                        "task": task,
                        "record_id": record["record_id"],
                        "source_image_id": image_id,
                        "difficulty": record["difficulty"],
                        "m31_correct_count": record["m31_correct_count"],
                        "crop_correct_count": record["crop_correct_count"],
                        "total_correct_count": record["total_correct_count"],
                        "visual_relpath": relative.as_posix(),
                    }
                )
                counts[category] += 1
            except Exception as exc:
                failures.append(
                    {
                        "record_id": record.get("record_id"),
                        "category": category,
                        "error": f"{type(exc).__name__}: {exc}",
                    }
                )
    sections = []
    for category in sorted(counts):
        selected = [
            card
            for card in cards
            if card["category"] == category
        ]
        cards_html = "".join(
            "<article><img loading='lazy' src='"
            + html.escape(str(card["visual_relpath"]))
            + "'><div><code>"
            + html.escape(str(card["record_id"]))
            + "</code> task="
            + html.escape(str(card["task"]))
            + " m31="
            + str(card["m31_correct_count"])
            + "/4 crop="
            + str(card["crop_correct_count"])
            + "/4 total="
            + str(card["total_correct_count"])
            + "/8</div></article>"
            for card in selected
        )
        sections.append(
            f"<section><h2>{html.escape(category)} "
            f"({len(selected)})</h2><div class='grid'>{cards_html}</div></section>"
        )
    document = f"""<!doctype html><html><head><meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<title>UI5 incremental rollout snapshot</title><style>
body{{font-family:system-ui,sans-serif;margin:24px;background:#f6f8fb;color:#172033}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(420px,1fr));gap:14px}}
article{{background:white;border:1px solid #dbe2ea;border-radius:8px;padding:9px}}
img{{width:100%;height:auto}} code{{font-size:12px}}
</style></head><body><h1>UI5 incremental rollout snapshot</h1>
<p>Green=GT, blue=matched prediction, red=unmatched prediction, yellow=crop.</p>
{''.join(sections)}</body></html>"""
    visual_root.mkdir(parents=True, exist_ok=True)
    (visual_root / "index.html").write_text(document, encoding="utf-8")
    result = {
        "index": "visualizations/index.html",
        "rendered": len(cards),
        "failures": failures,
        "counts": dict(sorted(counts.items())),
    }
    atomic_json(visual_root / "gallery_summary.json", result)
    return result


def build_snapshot_manifest(
    snapshot_root: Path,
    summary: Mapping[str, Any],
) -> dict[str, Any]:
    files = []
    for path in sorted(snapshot_root.rglob("*")):
        if not path.is_file() or path.name in {"manifest.json", "_SUCCESS"}:
            continue
        relative = path.relative_to(snapshot_root).as_posix()
        row_count = None
        if path.suffix == ".jsonl":
            with path.open("rb") as handle:
                row_count = sum(1 for line in handle if line.strip())
        files.append(
            {
                "path": relative,
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "jsonl_records": row_count,
            }
        )
    return {
        "schema_version": SCHEMA_VERSION,
        "snapshot_kind": summary["snapshot_kind"],
        "scheduled_hour": summary["scheduled_hour"],
        "created_at": summary["created_at"],
        "previous_snapshot": summary["previous_snapshot"],
        "append_only": True,
        "atomic_publish": True,
        "success_marker": "_SUCCESS",
        "execution_architecture": execution_architecture(),
        "files": files,
    }


def create_snapshot(
    output_root: Path,
    bundle_root: Path,
    *,
    kind: str,
    scheduled_hour: int | None,
    started_at_epoch: float,
    export_selection_dir: Path | None = None,
    created_at_epoch: float | None = None,
) -> tuple[Path, dict[str, Any]]:
    if kind == "hourly":
        if scheduled_hour is None or scheduled_hour < 3 or scheduled_hour % 3:
            raise ValueError("hourly snapshots require --scheduled-hour 3, 6, 9, ...")
    elif kind != "final":
        raise ValueError(f"unsupported snapshot kind: {kind}")
    now_epoch = time.time() if created_at_epoch is None else created_at_epoch
    snapshots_root = output_root / "snapshots"
    snapshots_root.mkdir(parents=True, exist_ok=True)
    previous = previous_snapshot(snapshots_root)
    previous_records = read_snapshot_records(previous) if previous else {}
    records, cumulative = build_difficulty_records(output_root, bundle_root)
    complete_records = [
        row for row in records if row.get("cross_model_complete8") is True
    ]
    prefix = f"hour_{scheduled_hour:03d}" if kind == "hourly" else "final"
    base_name = f"{prefix}_{timestamp_token(now_epoch)}"
    destination = snapshots_root / base_name
    suffix = 1
    while destination.exists():
        destination = snapshots_root / f"{base_name}_{suffix:02d}"
        suffix += 1
    temporary = snapshots_root / f".{destination.name}.tmp-{os.getpid()}"
    if temporary.exists():
        shutil.rmtree(temporary)
    temporary.mkdir(parents=True)
    try:
        file_counts = write_difficulty_exports(
            temporary,
            records,
            previous_records=previous_records,
            write_delta=True,
        )
        error_counts = write_error_exports(temporary, output_root, records)
        difficulty_counts = Counter(str(row["difficulty"]) for row in complete_records)
        task_counts = Counter(
            (str(row["task"]), str(row["difficulty"])) for row in complete_records
        )
        task_totals = Counter(str(row["task"]) for row in complete_records)
        exclusion_counts = Counter(
            str(row.get("exclusion_reason"))
            for row in records
            if row.get("cross_model_complete8") is not True
        )
        expected_total = int(cumulative["expected_total"])
        complete8_total = len(complete_records)
        processing_status = current_processing_status(
            output_root, cumulative, now_epoch
        )
        file_counts["confusion_metrics_jsonl"] = atomic_jsonl(
            temporary / "confusion_metrics.jsonl",
            cumulative["cumulative_metrics"],
        )
        file_counts["confusion_metrics_csv"] = atomic_csv(
            temporary / "confusion_metrics.csv",
            cumulative["cumulative_metrics"],
        )
        file_counts["correct_count_8_jsonl"] = atomic_jsonl(
            temporary / "correct_count_8.jsonl",
            cumulative["correct_count_8"],
        )
        file_counts["correct_count_8_csv"] = atomic_csv(
            temporary / "correct_count_8.csv",
            cumulative["correct_count_8"],
        )
        atomic_json(temporary / "processing_status.json", processing_status)
        correct_count_distribution = {
            str(row["correct_count_8"]): int(row["samples"])
            for row in cumulative["correct_count_8"]
            if row["scope"] == "micro"
        }
        summary = {
            "schema_version": SCHEMA_VERSION,
            "snapshot_kind": kind,
            "scheduled_hour": scheduled_hour,
            "started_at_epoch": started_at_epoch,
            "started_at": datetime.fromtimestamp(
                started_at_epoch, timezone.utc
            ).isoformat(),
            "created_at_epoch": now_epoch,
            "created_at": datetime.fromtimestamp(now_epoch, timezone.utc).isoformat(),
            "elapsed_seconds": max(0.0, now_epoch - started_at_epoch),
            "previous_snapshot": previous.name if previous else None,
            "expected_total": expected_total,
            "complete8_samples": complete8_total,
            "current_processed_samples": processing_status["attempted_unique_samples"],
            "remaining_seconds": processing_status["remaining_seconds"],
            "estimated_completion": processing_status["estimated_completion"],
            "processing_status": processing_status,
            "execution_architecture": execution_architecture(),
            "classification_policy": {
                "complete8_required": True,
                "technical_error_free_required": True,
                "difficulty_ratio_denominator": "complete8_samples",
                "easy": "total_correct_count == 8",
                "medium": "1 <= total_correct_count <= 7",
                "hard": "total_correct_count == 0",
                "excluded_from_difficulty": (
                    "fewer than eight raw records, runtime/final-OOM error, parse error, "
                    "crop parse error, missing reward, or missing image confusion"
                ),
                "oom_recovered_policy": (
                    "eligible when retry inference and parsing succeeded and reward is valid"
                ),
                "complete_fields_are_not_difficulties": [
                    "m31_complete4",
                    "crop_complete4",
                    "cross_model_complete8",
                    "grpo_ready_m31",
                    "grpo_ready_crop",
                ],
                "grpo_ready_m31": (
                    "medium, m31_complete4, source/parse eligible, and the four "
                    "exact rewards contain both true and false"
                ),
                "grpo_ready_crop": (
                    "medium, crop_complete4, source/parse eligible, and the four "
                    "exact rewards contain both true and false"
                ),
            },
            "difficulty_counts": {
                difficulty: difficulty_counts[difficulty]
                for difficulty in DIFFICULTIES
            },
            "difficulty_ratios": {
                difficulty: difficulty_counts[difficulty] / complete8_total
                if complete8_total
                else 0.0
                for difficulty in DIFFICULTIES
            },
            "per_task_difficulty": [
                {
                    "task": task,
                    "difficulty": difficulty,
                    "samples": task_counts[(task, difficulty)],
                    "proportion": (
                        task_counts[(task, difficulty)] / task_totals[task]
                        if task_totals[task]
                        else 0.0
                    ),
                }
                for task in TASKS
                for difficulty in DIFFICULTIES
            ],
            "excluded_from_difficulty_counts": {
                "incomplete_eight_routes": exclusion_counts["incomplete_eight_routes"],
                "technical_error": exclusion_counts["technical_error"],
                "total": expected_total - complete8_total,
            },
            "grpo_ready_counts": {
                "m31": file_counts["grpo_m31_ready"],
                "crop": file_counts["grpo_crop_ready"],
            },
            "file_counts": file_counts,
            "error_counts": error_counts,
            "outcome_counts": {
                "all_correct": difficulty_counts["easy"],
                "all_wrong": difficulty_counts["hard"],
                "partially_correct": difficulty_counts["medium"],
            },
            "outcome_ratios": {
                "all_correct": difficulty_counts["easy"] / complete8_total
                if complete8_total
                else 0.0,
                "all_wrong": difficulty_counts["hard"] / complete8_total
                if complete8_total
                else 0.0,
                "partially_correct": difficulty_counts["medium"] / complete8_total
                if complete8_total
                else 0.0,
            },
            "correct_count_distribution_0to8": correct_count_distribution,
            **cumulative,
        }
        summary["visualizations"] = render_snapshot_gallery(
            temporary, bundle_root, complete_records
        )
        summary["statistics_workbook"] = "snapshot_statistics.xlsx"
        write_snapshot_workbook(temporary / "snapshot_statistics.xlsx", summary)
        atomic_json(temporary / "summary.json", summary)
        atomic_json(
            temporary / "manifest.json", build_snapshot_manifest(temporary, summary)
        )
        atomic_text(temporary / "_SUCCESS", summary["created_at"] + "\n")
        os.replace(temporary, destination)
    finally:
        if temporary.exists():
            shutil.rmtree(temporary)
    if export_selection_dir is not None:
        write_difficulty_exports(export_selection_dir, records)
    print(
        json.dumps(
            {
                "snapshot": str(destination),
                "kind": kind,
                "scheduled_hour": scheduled_hour,
                "counts": summary["difficulty_counts"],
                "delta": summary["file_counts"]["delta_since_previous"],
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    return destination, summary


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    create_snapshot(
        args.output_root.expanduser().resolve(strict=True),
        args.bundle_root.expanduser().resolve(strict=True),
        kind=args.kind,
        scheduled_hour=args.scheduled_hour,
        started_at_epoch=args.started_at_epoch,
        export_selection_dir=(
            args.export_selection_dir.expanduser().resolve(strict=False)
            if args.export_selection_dir is not None
            else None
        ),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
