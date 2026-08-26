#!/usr/bin/env python3
"""Task-aware, resumable UI5 detector/crop audit.

The GPU stages write immutable task-agnostic detections.  Every crop parameter
comparison after ``merge`` is CPU-only and associates proposals with each
``image x task`` annotation independently.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import time
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

from PIL import Image, ImageDraw, ImageOps

from analyze_ui5_source_overlap import (
    TASK_NAMES,
    analyze as analyze_overlap,
    assistant_answer,
    content_fingerprint,
    resolve_training_image,
)
from prepare_ui_defect_locany import TASKS


BASELINE_COMMIT = "c06f1479a11b0175579994b880466b57bba50a87"
PARSER_COMMIT = "06eaebf8eb4ea01e61b690f2ff972bf614915918"
BOX_PATTERN = re.compile(
    r"<box>\s*<(-?\d+(?:\.\d+)?)>\s*<(-?\d+(?:\.\d+)?)>\s*"
    r"<(-?\d+(?:\.\d+)?)>\s*<(-?\d+(?:\.\d+)?)>\s*</box>"
)
CONFIGS = {
    "A": {"horizontal_link_ratio": 0.015, "vertical_link_ratio": 0.015, "context_ratio": 0.10},
    "B": {"horizontal_link_ratio": 0.015, "vertical_link_ratio": 0.015, "context_ratio": 0.20},
    "C": {"horizontal_link_ratio": 0.025, "vertical_link_ratio": 0.025, "context_ratio": 0.20},
}
REGION_TASKS = (
    "ui_occlusion",
    "ui_cropping",
    "ui_text_overflow",
    "ui_text_ellipsis",
)
ELEMENT_TASKS = ("ui_occlusion", "ui_cropping")
TEXT_TASKS = ("ui_text_overflow", "ui_text_ellipsis")


def _task_aware_candidate(
    *,
    element_min_context_image_ratio: float,
    text_horizontal_link_ratio: float,
) -> dict[str, Any]:
    rules: dict[str, dict[str, float]] = {}
    for task in ELEMENT_TASKS:
        rules[task] = {
            "horizontal_link_ratio": 0.025,
            "vertical_link_ratio": 0.025,
            "context_ratio": 0.20,
            "min_context_image_ratio": element_min_context_image_ratio,
        }
    for task in TEXT_TASKS:
        rules[task] = {
            "horizontal_link_ratio": text_horizontal_link_ratio,
            "vertical_link_ratio": 0.025,
            "context_ratio": 0.20,
            "min_context_image_ratio": 0.0,
        }
    return {
        "task_aware": True,
        "element_min_context_image_ratio": element_min_context_image_ratio,
        "text_horizontal_link_ratio": text_horizontal_link_ratio,
        "task_rules": rules,
    }


TASK_AWARE_CANDIDATES = {
    "C": {
        "task_aware": False,
        "description": "uniform config C baseline",
        "task_rules": {
            task: {
                "horizontal_link_ratio": 0.025,
                "vertical_link_ratio": 0.025,
                "context_ratio": 0.20,
                "min_context_image_ratio": 0.0,
            }
            for task in REGION_TASKS
        },
    },
    "TA_CTX010_H035": _task_aware_candidate(
        element_min_context_image_ratio=0.010,
        text_horizontal_link_ratio=0.035,
    ),
    "TA_CTX015_H035": _task_aware_candidate(
        element_min_context_image_ratio=0.015,
        text_horizontal_link_ratio=0.035,
    ),
    "TA_CTX010_H050": _task_aware_candidate(
        element_min_context_image_ratio=0.010,
        text_horizontal_link_ratio=0.050,
    ),
    "TA_CTX015_H050": _task_aware_candidate(
        element_min_context_image_ratio=0.015,
        text_horizontal_link_ratio=0.050,
    ),
}
TASK_LABELS = {task["name"]: task["en"] for task in TASKS}
PIPELINE_STAGES = ("prepare", "text", "icon", "merge", "crop-audit")
CROP_AUDIT_FORMAT_VERSION = 3
ANOMALY_PRIORITY = (
    "gt_uncovered",
    "gt_partial_only",
    "basename_or_multitask_annotation_risk",
    "detector_empty_fallback",
    "near_full_image",
    "max_crops_or_forced_merge",
    "roundtrip_error",
)
METRIC_DEFINITIONS = {
    "gt_box_containment_recall": (
        "至少被一个 crop 完整包含的 GT bbox 数 / 所有正样本中的 GT bbox 总数；"
        "负样本 gt_count=0，不进入分母"
    ),
    "positive_sample_success_rate": (
        "全部 GT bbox 都被完整包含的正样本数 / 正样本总数；负样本不进入分母"
    ),
    "near_full_image_ratio": (
        "crop union area / original area > 0.8 的图片任务样本数 / 该范围全部图片任务样本数；"
        "包含正样本和负样本"
    ),
    "gt_gain_over_1_25_1_5_2_0_ratio": (
        "放大倍率超过阈值且已被完整包含的 GT bbox 数 / 已被完整包含且可计算放大倍率的 GT bbox 数"
    ),
    "negative_samples": (
        "只描述数据组成；不参与 GT bbox 完整覆盖率或正样本全部成功率的分母"
    ),
}
FINAL_TRAINING_GATE_CONDITIONS = frozenset(
    {
        "region_overall_recall_at_least_0_99",
        "each_region_task_recall_at_least_0_98",
        "detector_boundary_cut_count_zero",
        "region_roundtrip_error_over_1_count_zero",
        "partial_crop_training_eligible_count_zero",
        "hard_negative_max_one_per_image_task",
        "same_content_cross_train_val_count_zero",
        "content_missing_recall_equals_1",
        "content_missing_normalized_gt_mismatch_count_zero",
        "input_snapshot_unchanged",
        "all_reports_written_successfully",
    }
)


@dataclass(frozen=True)
class AuditPaths:
    output: Path
    crop_audit_name: str = "crop_audit_v3"

    @property
    def manifest(self) -> Path:
        return self.output / "manifest"

    @property
    def unique_images(self) -> Path:
        return self.manifest / "unique_images.jsonl"

    @property
    def task_samples(self) -> Path:
        return self.manifest / "task_samples.jsonl"

    @property
    def shards(self) -> Path:
        return self.manifest / "shards"

    @property
    def detections(self) -> Path:
        return self.output / "detections"

    def stage_dir(self, stage: str) -> Path:
        return self.detections / stage

    @property
    def merged(self) -> Path:
        return self.detections / "merged" / "detections.jsonl"

    @property
    def detector_config(self) -> Path:
        return self.detections / "detector_config.json"

    @property
    def crop_audit(self) -> Path:
        return self.output / self.crop_audit_name


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--stage",
        choices=("all", "prepare", "text", "icon", "merge", "crop-audit", "_worker"),
        default="all",
    )
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--locany-data-dir", type=Path, required=True)
    parser.add_argument("--parser-root", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument(
        "--crop-audit-name",
        default="crop_audit_v3",
        help="Named CPU audit directory under --output-dir; detector outputs remain shared/read-only",
    )
    parser.add_argument("--gpus", default="0,1,2,3")
    parser.add_argument("--workers-per-gpu", type=int, choices=(1, 2), default=1)
    parser.add_argument("--allow-two-processes-per-gpu", action="store_true")
    parser.add_argument(
        "--text-python",
        default=os.environ.get("TEXT_PYTHON"),
        help="Python executable for PP-OCRv5 workers; defaults to the launcher Python",
    )
    parser.add_argument(
        "--icon-python",
        default=os.environ.get("ICON_PYTHON"),
        help=(
            "Python executable for icon workers; may point to a separate Torch/"
            "LocateAnything environment"
        ),
    )
    parser.add_argument("--image-loader-threads", type=int, default=4)
    parser.add_argument("--shard-size", type=int, default=750)
    parser.add_argument(
        "--max-unique-images",
        type=int,
        default=0,
        help="0 means all; use 2000 with a separate output directory for process/GPU benchmarking",
    )
    parser.add_argument("--resume", action="store_true")
    parser.add_argument(
        "--progress-interval-seconds",
        type=float,
        default=10.0,
        help="Consolidated terminal/run_status.json update interval",
    )
    parser.add_argument(
        "--expected-unique-images",
        type=int,
        default=0,
        help="Fail crop-audit unless the prepared manifest has exactly this many images; 0 disables",
    )
    parser.add_argument(
        "--progress-every-images",
        type=int,
        default=25,
        help="Detector worker progress checkpoint frequency",
    )
    parser.add_argument(
        "--text-model-dir",
        type=Path,
        default=None,
        help="Optional local PP-OCRv5 directory; omitted means PaddleOCR auto-download/cache",
    )
    parser.add_argument("--icon-model", type=Path, default=None)
    parser.add_argument("--text-long-side", type=int, default=1920)
    parser.add_argument("--text-box-threshold", type=float, default=0.3)
    parser.add_argument(
        "--enable-mkldnn",
        action="store_true",
        help=(
            "Opt in to Paddle MKLDNN/OneDNN. Disabled by default because some "
            "Paddle PIR builds fail with ConvertPirAttribute2RuntimeAttribute."
        ),
    )
    parser.add_argument("--icon-long-side", type=int, default=1920)
    parser.add_argument("--icon-confidence", type=float, default=0.05)
    parser.add_argument("--max-crops", type=int, default=10)
    parser.add_argument("--boundary-margin-ratio", type=float, default=0.01)
    parser.add_argument(
        "--crop-workers",
        type=int,
        default=8,
        help="CPU workers for geometry/materialization; does not start detector models",
    )
    parser.add_argument(
        "--overview-samples-per-task",
        type=int,
        default=50,
        help="Deterministic selected-config overview samples per task",
    )
    parser.add_argument(
        "--overview-anomalies-per-category",
        type=int,
        default=50,
        help="Selected-config overview count for each anomaly category",
    )
    parser.add_argument("--worker-index", type=int, default=None, help=argparse.SUPPRESS)
    parser.add_argument("--worker-count", type=int, default=None, help=argparse.SUPPRESS)
    parser.add_argument("--detector-stage", choices=("text", "icon"), help=argparse.SUPPRESS)
    return parser.parse_args(argv)


def atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + f".tmp.{os.getpid()}")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(value, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary, path)


def atomic_write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + f".tmp.{os.getpid()}")
    count = 0
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
            count += 1
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary, path)
    return count


def format_duration(seconds: float | None) -> str:
    if seconds is None or not math.isfinite(seconds) or seconds < 0:
        return "--:--:--"
    value = int(round(seconds))
    hours, remainder = divmod(value, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


class ProgressReporter:
    """Write human-readable progress and an atomic machine-readable snapshot."""

    def __init__(
        self,
        *,
        stage: str,
        total: int,
        output_dir: Path,
        interval_seconds: float,
        initial_completed: int = 0,
        unit: str = "images",
    ) -> None:
        self.stage = stage
        self.total = max(0, int(total))
        self.output_dir = output_dir
        self.interval_seconds = max(1.0, float(interval_seconds))
        self.initial_completed = max(0, int(initial_completed))
        self.unit = unit
        self.started = time.monotonic()
        self.last_print = 0.0

    def update(
        self,
        completed: int,
        *,
        status: str = "running",
        detail: str | None = None,
        force: bool = False,
    ) -> dict[str, Any]:
        completed = max(0, min(int(completed), self.total)) if self.total else max(0, int(completed))
        elapsed = max(0.0, time.monotonic() - self.started)
        newly_completed = max(0, completed - self.initial_completed)
        rate = newly_completed / elapsed if elapsed > 0 and newly_completed else 0.0
        remaining = max(0, self.total - completed)
        eta_seconds = 0.0 if remaining == 0 else remaining / rate if rate > 0 else None
        percent = completed / self.total if self.total else (1.0 if status == "completed" else 0.0)
        stage_index = PIPELINE_STAGES.index(self.stage) + 1 if self.stage in PIPELINE_STAGES else None
        payload = {
            "stage": self.stage,
            "stage_index": stage_index,
            "stage_total": len(PIPELINE_STAGES),
            "status": status,
            "detail": detail,
            "completed": completed,
            "total": self.total,
            "unit": self.unit,
            "percent": round(percent, 6),
            "elapsed_seconds": round(elapsed, 3),
            "rate_per_second": round(rate, 6),
            "eta_seconds": round(eta_seconds, 3) if eta_seconds is not None else None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        atomic_write_json(self.output_dir / "run_status.json", payload)
        now = time.monotonic()
        if force or now - self.last_print >= self.interval_seconds:
            stage_label = (
                f"{self.stage} {stage_index}/{len(PIPELINE_STAGES)}"
                if stage_index is not None
                else self.stage
            )
            suffix = f" | {detail}" if detail else ""
            print(
                f"[进度 {stage_label}] {completed}/{self.total} {self.unit} "
                f"({percent:.1%}) | 已耗时 {format_duration(elapsed)} | "
                f"速度 {rate:.2f} {self.unit}/s | ETA {format_duration(eta_seconds)}{suffix}",
                flush=True,
            )
            self.last_print = now
        return payload


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, 1):
            if not line.strip():
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"invalid JSON at {path}:{line_no}: {exc}") from exc
    return rows


def digest_ids(ids: Iterable[str]) -> str:
    digest = hashlib.blake2b(digest_size=16)
    for image_id in sorted(ids):
        digest.update(image_id.encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def stable_id(prefix: str, value: str, length: int = 20) -> str:
    digest = hashlib.blake2b(value.encode("utf-8"), digest_size=16).hexdigest()
    return f"{prefix}_{digest[:length]}"


def normalized_1000_to_pixels(
    box: Sequence[float], width: int, height: int
) -> list[int]:
    values = [
        round(float(box[0]) * width / 1000),
        round(float(box[1]) * height / 1000),
        round(float(box[2]) * width / 1000),
        round(float(box[3]) * height / 1000),
    ]
    x1, x2 = sorted((max(0, min(width, values[0])), max(0, min(width, values[2]))))
    y1, y2 = sorted((max(0, min(height, values[1])), max(0, min(height, values[3]))))
    if x2 <= x1 or y2 <= y1:
        raise ValueError(f"zero-area GT after conversion: {box} at {width}x{height}")
    return [x1, y1, x2, y2]


def parse_gt_boxes(answer: str, width: int, height: int) -> tuple[list[list[int]], list[list[float]]]:
    norm = [[float(value) for value in match] for match in BOX_PATTERN.findall(answer)]
    pixels = [normalized_1000_to_pixels(box, width, height) for box in norm]
    # Exact duplicates carry no additional localization information.
    dedup: dict[tuple[int, int, int, int], list[float]] = {}
    for pixel_box, norm_box in zip(pixels, norm):
        dedup.setdefault(tuple(pixel_box), norm_box)
    return [list(box) for box in dedup], [dedup[box] for box in dedup]


def open_raw_image(path: Path) -> Image.Image:
    with Image.open(path) as image:
        return ImageOps.exif_transpose(image).convert("RGB")


def git_output(repo: Path, *args: str) -> str:
    result = subprocess.run(
        ["git", "-C", str(repo), *args],
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return result.stdout.strip()


def verify_revisions(project_root: Path, parser_root: Path) -> dict[str, str]:
    project_head = git_output(project_root, "rev-parse", "HEAD")
    ancestor = subprocess.run(
        ["git", "-C", str(project_root), "merge-base", "--is-ancestor", BASELINE_COMMIT, "HEAD"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    if ancestor.returncode != 0:
        raise RuntimeError(f"required baseline {BASELINE_COMMIT} is not an ancestor of {project_head}")
    parser_head = git_output(parser_root, "rev-parse", "HEAD")
    if parser_head != PARSER_COMMIT:
        raise RuntimeError(f"parser must be {PARSER_COMMIT}, found {parser_head}")
    return {"baseline": BASELINE_COMMIT, "project_head": project_head, "parser": parser_head}


def detector_config(args: argparse.Namespace) -> dict[str, Any]:
    text_model = args.text_model_dir.resolve(strict=False) if args.text_model_dir else None
    icon_model = args.icon_model or args.parser_root / "weights" / "icon_detect_v3" / "model.pt"
    return {
        "parser_commit": PARSER_COMMIT,
        "text": {
            "model_dir": str(text_model) if text_model else None,
            "auto_download": text_model is None,
            "model_name": "PP-OCRv5_server_det",
            "engine": "paddle_static",
            "long_side": args.text_long_side,
            "limit_type": "max",
            "pixel_threshold": 0.3,
            "box_threshold": args.text_box_threshold,
            "unclip_ratio": 1.5,
            "enable_mkldnn": bool(getattr(args, "enable_mkldnn", False)),
            "min_area": 0,
            "min_width": 0,
            "min_height": 0,
        },
        "icon": {
            "model": str(icon_model.resolve(strict=False)),
            "long_side": args.icon_long_side,
            "confidence": args.icon_confidence,
            "iou_threshold": 0.7,
            "max_detections": 1000,
            "min_area": 0,
            "min_width": 0,
            "min_height": 0,
        },
    }


def ensure_detector_config(path: Path, config: Mapping[str, Any]) -> None:
    if path.is_file():
        existing = json.loads(path.read_text(encoding="utf-8"))
        if existing != config:
            # Safe one-time migration for runs that prepared manifests before
            # MKLDNN became explicit but failed before completing any text shard.
            migrated = json.loads(json.dumps(config))
            migrated["text"].pop("enable_mkldnn", None)
            has_completed_text = any((path.parent / "text").glob("shard_*.done.json"))
            if existing == migrated and not has_completed_text:
                atomic_write_json(path, config)
                print(
                    "[config] migrated detector_config.json: text enable_mkldnn=false",
                    flush=True,
                )
                return
            raise RuntimeError(
                f"detector configuration is immutable once written: {path}; use a new output directory"
            )
    else:
        atomic_write_json(path, config)


def training_files(locany_data_dir: Path) -> list[Path]:
    files = [locany_data_dir / f"{task}_train.jsonl" for task in TASK_NAMES]
    missing = [path for path in files if not path.is_file()]
    if missing:
        raise FileNotFoundError("missing training JSONL: " + ", ".join(map(str, missing)))
    return files


def source_files(source_dir: Path) -> list[Path]:
    files = [source_dir / task["file"] for task in TASKS]
    missing = [path for path in files if not path.is_file()]
    if missing:
        raise FileNotFoundError("missing source JSONL: " + ", ".join(map(str, missing)))
    return files


def print_preflight(
    args: argparse.Namespace,
    *,
    unique_count: int | None = None,
    readable_count: int | None = None,
    detector_stage: str | None = None,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    revisions = verify_revisions(project_root, args.parser_root.resolve(strict=True))
    sources = source_files(args.source_dir)
    trains = training_files(args.locany_data_dir)
    config = detector_config(args)
    prepared_paths = AuditPaths(
        args.output_dir, str(getattr(args, "crop_audit_name", "crop_audit_v3"))
    )
    if unique_count is None and prepared_paths.unique_images.is_file():
        unique_count = len(read_jsonl(prepared_paths.unique_images))
    if (
        detector_stage in {"text", "all"}
        and config["text"]["model_dir"] is not None
        and not Path(config["text"]["model_dir"]).is_dir()
    ):
        raise FileNotFoundError(f"missing PP-OCRv5 model directory: {config['text']['model_dir']}")
    if detector_stage in {"icon", "all"} and not Path(config["icon"]["model"]).is_file():
        raise FileNotFoundError(f"missing icon detector weights: {config['icon']['model']}")
    lines = [
        "=== UI5 crop audit preflight ===",
        f"baseline_commit : {revisions['baseline'][:12]} (ancestor of {revisions['project_head'][:12]})",
        "CPT             : disabled (no CPT data or training entrypoint)",
        "input_files     : " + ", ".join(str(path) for path in (*sources, *trains)),
        f"images_readable : {readable_count if readable_count is not None else 'from prepared manifest'}",
        f"unique_images   : {unique_count if unique_count is not None else 'not prepared'}",
        f"expected_unique : {getattr(args, 'expected_unique_images', 0) or 'not enforced'}",
        f"parser_commit   : {revisions['parser']}",
        f"text_model      : {config['text']['model_dir'] or 'PaddleOCR automatic download/cache'}",
        f"icon_model      : {config['icon']['model']}",
        f"text_python     : {args.text_python}",
        f"icon_python     : {args.icon_python}",
        f"output_dir      : {args.output_dir.resolve(strict=False)}",
        f"crop_audit_name : {args.crop_audit_name}",
        f"GPUs            : {args.gpus}",
        f"workers/GPU     : {args.workers_per_gpu}",
        f"crop_workers    : {args.crop_workers}",
        f"parameters      : {json.dumps({'detector': config, 'crop_candidates': TASK_AWARE_CANDIDATES, 'boundary_margin_ratio': args.boundary_margin_ratio, 'max_crops': args.max_crops}, ensure_ascii=False)}",
    ]
    print("\n".join(lines), flush=True)


def prepared_manifest_valid(paths: AuditPaths) -> bool:
    """Return whether prepare outputs form a complete, self-consistent set."""
    summary_path = paths.manifest / "prepare_summary.json"
    if not paths.unique_images.is_file() or not paths.task_samples.is_file() or not summary_path.is_file():
        return False
    try:
        unique = read_jsonl(paths.unique_images)
        samples = read_jsonl(paths.task_samples)
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        shards = [row for shard in sorted(paths.shards.glob("shard_*.jsonl")) for row in read_jsonl(shard)]
    except (OSError, ValueError, json.JSONDecodeError):
        return False
    unique_ids = [row.get("image_id") for row in unique]
    shard_ids = [row.get("image_id") for row in shards]
    sample_ids = [row.get("sample_id") for row in samples]
    return (
        bool(unique_ids)
        and len(unique_ids) == len(set(unique_ids))
        and len(shard_ids) == len(set(shard_ids))
        and set(shard_ids) == set(unique_ids)
        and len(sample_ids) == len(set(sample_ids))
        and {row.get("image_id") for row in samples}.issubset(set(unique_ids))
        and int(summary.get("unique_images", -1)) == len(unique)
        and int(summary.get("task_samples", -1)) == len(samples)
        and int(summary.get("shards", -1)) == len(list(paths.shards.glob("shard_*.jsonl")))
    )


def build_task_aware_manifest(args: argparse.Namespace) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not 500 <= args.shard_size <= 1000:
        raise ValueError("--shard-size must be in [500, 1000]")
    paths = AuditPaths(args.output_dir)
    # Fail on missing JSONL/parser/icon weights before spending time hashing images.
    print_preflight(args, detector_stage="all")
    if args.resume and prepared_manifest_valid(paths):
        unique_images = read_jsonl(paths.unique_images)
        task_samples = read_jsonl(paths.task_samples)
        ensure_detector_config(paths.detector_config, detector_config(args))
        reporter = ProgressReporter(
            stage="prepare",
            total=len(unique_images),
            output_dir=args.output_dir,
            interval_seconds=args.progress_interval_seconds,
        )
        reporter.update(
            len(unique_images),
            status="completed",
            detail="--resume 验证通过，跳过 prepare，不重复扫描图片",
            force=True,
        )
        return unique_images, task_samples
    overlap_reporter = ProgressReporter(
        stage="prepare",
        total=1,
        output_dir=args.output_dir,
        interval_seconds=args.progress_interval_seconds,
        unit="substeps",
    )
    overlap_reporter.update(
        0,
        detail="子步骤 1/3：统计源数据与训练数据重叠",
        force=True,
    )
    overlap_phase_reporters: dict[str, ProgressReporter] = {}

    def overlap_progress(phase: str, completed: int, total: int) -> None:
        if phase not in overlap_phase_reporters:
            overlap_phase_reporters[phase] = ProgressReporter(
                stage="prepare",
                total=total,
                output_dir=args.output_dir,
                interval_seconds=args.progress_interval_seconds,
                unit="records",
            )
        overlap_phase_reporters[phase].update(
            completed,
            detail=f"子步骤 1/3：重叠统计 {phase}",
            force=completed in {0, total},
        )

    overlap = analyze_overlap(
        args.source_dir,
        args.locany_data_dir,
        paths.manifest / "overlap",
        progress_callback=overlap_progress,
    )
    overlap_reporter.update(
        1,
        status="completed",
        detail="子步骤 1/3：重叠统计完成",
        force=True,
    )
    fingerprints: dict[str, str] = {}
    aliases_by_content: dict[str, set[str]] = defaultdict(set)
    raw_samples: list[dict[str, Any]] = []
    readable = 0
    task_records: list[tuple[Mapping[str, str], Path, int, dict[str, Any]]] = []
    for task in TASKS:
        annotation = args.locany_data_dir / f"{task['name']}_train.jsonl"
        task_records.extend(
            (task, annotation, line_no, record)
            for line_no, record in enumerate(read_jsonl(annotation), 1)
        )
    manifest_reporter = ProgressReporter(
        stage="prepare",
        total=len(task_records),
        output_dir=args.output_dir,
        interval_seconds=args.progress_interval_seconds,
        unit="records",
    )
    manifest_reporter.update(
        0,
        detail="子步骤 2/3：解析图片、内容指纹与 task-aware GT",
        force=True,
    )
    for record_index, (task, annotation, line_no, record) in enumerate(task_records, 1):
            raw_image = str(record["image"])
            image_path = resolve_training_image(raw_image, args.source_dir, args.locany_data_dir)
            canonical = str(image_path)
            if canonical not in fingerprints:
                fingerprints[canonical] = content_fingerprint(image_path)
                readable += 1
            content_id = fingerprints[canonical]
            aliases_by_content[content_id].add(canonical)
            image_id = stable_id("img", content_id)
            with open_raw_image(image_path) as image:
                width, height = image.size
            gt_pixels, gt_1000 = parse_gt_boxes(assistant_answer(record), width, height)
            raw_samples.append(
                {
                    "task": task["name"],
                    "source_file": str(annotation.resolve()),
                    "line_no": line_no,
                    "image_path": raw_image,
                    "canonical_path": canonical,
                    "content_id": content_id,
                    "image_id": image_id,
                    "width": width,
                    "height": height,
                    "positive": bool(gt_pixels),
                    "gt_count": len(gt_pixels),
                    "gt_boxes": gt_pixels,
                    "gt_boxes_1000": gt_1000,
                    "split": "train",
                }
            )
            if (
                record_index % args.progress_every_images == 0
                or record_index == len(task_records)
            ):
                manifest_reporter.update(
                    record_index,
                    detail="子步骤 2/3：解析图片、内容指纹与 task-aware GT",
                )

    # Exactly one manifest sample per image x task.  Same-task duplicates retain
    # source provenance and their distinct GT is combined only within that task.
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for sample in raw_samples:
        grouped[(sample["image_id"], sample["task"])].append(sample)
    task_samples: list[dict[str, Any]] = []
    for (image_id, task), members in sorted(grouped.items()):
        statuses = {bool(member["positive"]) for member in members}
        gt_by_box: dict[tuple[int, int, int, int], list[float]] = {}
        for member in members:
            for pixel_box, norm_box in zip(member["gt_boxes"], member["gt_boxes_1000"]):
                gt_by_box.setdefault(tuple(pixel_box), norm_box)
        representative = min(members, key=lambda item: item["canonical_path"])
        sample = dict(representative)
        sample.update(
            {
                "sample_id": stable_id("sample", f"{image_id}\0{task}"),
                "canonical_path": min(
                    path for member in members for path in aliases_by_content[member["content_id"]]
                ),
                "positive": bool(gt_by_box),
                "gt_count": len(gt_by_box),
                "gt_boxes": [list(box) for box in gt_by_box],
                "gt_boxes_1000": [gt_by_box[box] for box in gt_by_box],
                "source_records": [
                    {"source_file": member["source_file"], "line_no": member["line_no"]}
                    for member in members
                ],
                "same_task_polarity_conflict": len(statuses) > 1,
            }
        )
        task_samples.append(sample)

    by_image: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for sample in task_samples:
        by_image[sample["image_id"]].append(sample)
    unique_images = []
    for image_id, samples in sorted(by_image.items()):
        representative = min(samples, key=lambda item: item["canonical_path"])
        aliases = sorted(aliases_by_content[representative["content_id"]])
        dimensions = {(sample["width"], sample["height"]) for sample in samples}
        if len(dimensions) != 1:
            raise ValueError(f"byte-identical image has inconsistent decoded dimensions: {image_id}")
        unique_images.append(
            {
                "image_id": image_id,
                "content_id": representative["content_id"],
                "image_path": representative["canonical_path"],
                "canonical_paths": aliases,
                "basename": Path(representative["canonical_path"]).name,
                "width": representative["width"],
                "height": representative["height"],
                "tasks": sorted({sample["task"] for sample in samples}),
            }
        )

    if args.max_unique_images:
        if args.max_unique_images < 1:
            raise ValueError("--max-unique-images must be 0 or positive")
        keep_ids = {row["image_id"] for row in unique_images[: args.max_unique_images]}
        unique_images = unique_images[: args.max_unique_images]
        task_samples = [sample for sample in task_samples if sample["image_id"] in keep_ids]

    print_preflight(
        args,
        unique_count=len(unique_images),
        readable_count=readable,
        detector_stage="all",
    )
    atomic_write_jsonl(paths.unique_images, unique_images)
    atomic_write_jsonl(paths.task_samples, task_samples)
    paths.shards.mkdir(parents=True, exist_ok=True)
    for stale in paths.shards.glob("shard_*.jsonl"):
        stale.unlink()
    for start in range(0, len(unique_images), args.shard_size):
        shard_index = start // args.shard_size
        atomic_write_jsonl(
            paths.shards / f"shard_{shard_index:05d}.jsonl",
            unique_images[start : start + args.shard_size],
        )
    ensure_detector_config(paths.detector_config, detector_config(args))
    prepare_summary = {
        "unique_images": len(unique_images),
        "task_samples": len(task_samples),
        "readable_canonical_paths": readable,
        "shards": math.ceil(len(unique_images) / args.shard_size),
        "same_content_cross_train_val": overlap["same_content_cross_train_val"],
        "cpt_enabled": False,
    }
    atomic_write_json(paths.manifest / "prepare_summary.json", prepare_summary)
    manifest_reporter.update(
        len(task_records),
        status="completed",
        detail=f"子步骤 3/3：已生成 {len(unique_images)} 张唯一图片、{prepare_summary['shards']} 个 shard",
        force=True,
    )
    return unique_images, task_samples


def load_parser_module(parser_root: Path, module_name: str):
    module_path = parser_root / f"{module_name}.py"
    if not module_path.is_file():
        raise FileNotFoundError(module_path)
    if module_name in sys.modules:
        return sys.modules[module_name]
    # ui_region_cropper imports ui_region_parser by its real module name.
    if module_name == "ui_region_cropper" and "ui_region_parser" not in sys.modules:
        load_parser_module(parser_root, "ui_region_parser")
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"cannot import {module_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def completed_shard_valid(input_path: Path, output_path: Path, done_path: Path, stage: str) -> bool:
    if not output_path.is_file() or not done_path.is_file():
        return False
    try:
        expected = read_jsonl(input_path)
        actual = read_jsonl(output_path)
        marker = json.loads(done_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, json.JSONDecodeError):
        return False
    expected_ids = [row.get("image_id") for row in expected]
    actual_ids = [row.get("image_id") for row in actual]
    return (
        len(actual_ids) == len(expected_ids)
        and len(set(actual_ids)) == len(actual_ids)
        and set(actual_ids) == set(expected_ids)
        and marker.get("stage") == stage
        and marker.get("count") == len(expected_ids)
        and marker.get("image_id_digest") == digest_ids(expected_ids)
    )


def _loaded_images(rows: Sequence[Mapping[str, Any]], threads: int) -> Iterator[tuple[Mapping[str, Any], Image.Image]]:
    def load(row: Mapping[str, Any]) -> tuple[Mapping[str, Any], Image.Image]:
        image = open_raw_image(Path(str(row["image_path"])))
        if image.size != (int(row["width"]), int(row["height"])):
            raise ValueError(
                f"image dimensions changed for {row['image_id']}: manifest "
                f"{row['width']}x{row['height']}, current {image.width}x{image.height}"
            )
        return row, image

    if threads <= 1:
        for row in rows:
            yield load(row)
        return
    with ThreadPoolExecutor(max_workers=threads) as executor:
        yield from executor.map(load, rows)


def normalize_detector_items(items: Iterable[Mapping[str, Any]], width: int, height: int) -> list[dict[str, Any]]:
    result = []
    for item in items:
        bbox = [int(round(float(value))) for value in item["bbox"]]
        bbox = [
            max(0, min(width, bbox[0])),
            max(0, min(height, bbox[1])),
            max(0, min(width, bbox[2])),
            max(0, min(height, bbox[3])),
        ]
        bbox[0], bbox[2] = sorted((bbox[0], bbox[2]))
        bbox[1], bbox[3] = sorted((bbox[1], bbox[3]))
        if bbox[2] <= bbox[0] or bbox[3] <= bbox[1]:
            continue
        normalized = dict(item)
        normalized["bbox"] = bbox
        if normalized.get("score") is not None:
            normalized["score"] = float(normalized["score"])
        result.append(normalized)
    return result


def run_detector_worker(args: argparse.Namespace) -> None:
    if args.worker_index is None or args.worker_count is None or not args.detector_stage:
        raise ValueError("internal detector worker arguments are incomplete")
    stage = args.detector_stage
    paths = AuditPaths(args.output_dir)
    config = detector_config(args)
    ensure_detector_config(paths.detector_config, config)
    module = load_parser_module(args.parser_root, "ui_region_parser")
    if stage == "text":
        settings = config["text"]
        detector = module.PaddleTextDetector(
            model_name=settings["model_name"],
            model_dir=Path(settings["model_dir"]) if settings["model_dir"] else None,
            device="cuda:0",
            engine=settings["engine"],
            limit_side_len=settings["long_side"],
            limit_type=settings["limit_type"],
            pixel_threshold=settings["pixel_threshold"],
            box_threshold=settings["box_threshold"],
            unclip_ratio=settings["unclip_ratio"],
            enable_mkldnn=settings["enable_mkldnn"],
            min_area=0,
            min_width=0,
            min_height=0,
        )
    else:
        settings = config["icon"]
        detector = module.OmniParserYOLOv9Detector(Path(settings["model"]), "cuda:0")

    shard_paths = sorted(paths.shards.glob("shard_*.jsonl"))
    assigned = [
        path for index, path in enumerate(shard_paths) if index % args.worker_count == args.worker_index
    ]
    output_dir = paths.stage_dir(stage)
    output_dir.mkdir(parents=True, exist_ok=True)
    worker_progress_path = output_dir / "progress" / f"worker_{args.worker_index:02d}.json"
    assigned_total = sum(len(read_jsonl(shard)) for shard in assigned)
    processed_this_run = 0

    def write_worker_progress(
        *,
        status: str,
        current_shard: str | None = None,
        current_completed: int = 0,
        current_total: int = 0,
    ) -> None:
        atomic_write_json(
            worker_progress_path,
            {
                "stage": stage,
                "worker_index": args.worker_index,
                "worker_count": args.worker_count,
                "status": status,
                "assigned_images": assigned_total,
                "processed_images_this_run": processed_this_run,
                "current_shard": current_shard,
                "current_shard_completed": current_completed,
                "current_shard_total": current_total,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            },
        )

    write_worker_progress(status="starting")
    worker_final_status = "failed"
    try:
        for shard in assigned:
            output_path = output_dir / shard.name
            done_path = output_dir / (shard.stem + ".done.json")
            if args.resume and completed_shard_valid(shard, output_path, done_path, stage):
                print(f"[{stage}] resume skip {shard.name}", flush=True)
                continue
            rows = read_jsonl(shard)
            outputs = []
            write_worker_progress(
                status="running",
                current_shard=shard.name,
                current_completed=0,
                current_total=len(rows),
            )
            for image_index, (row, image) in enumerate(
                _loaded_images(rows, args.image_loader_threads), 1
            ):
                try:
                    started = time.perf_counter()
                    if stage == "text":
                        detections = detector.predict(image)
                    else:
                        detections = detector.predict(
                            image=image,
                            confidence=settings["confidence"],
                            iou_threshold=settings["iou_threshold"],
                            long_side=settings["long_side"],
                            max_detections=settings["max_detections"],
                            min_area=0,
                            min_width=0,
                            min_height=0,
                        )
                    normalized = normalize_detector_items(detections, image.width, image.height)
                    elapsed_ms = round((time.perf_counter() - started) * 1000, 3)
                    outputs.append(
                        {
                            "image_id": row["image_id"],
                            "image": row["image_path"],
                            "width": image.width,
                            "height": image.height,
                            f"{stage}_detections": normalized,
                            "inference_ms": elapsed_ms,
                        }
                    )
                    processed_this_run += 1
                    if (
                        image_index % args.progress_every_images == 0
                        or image_index == len(rows)
                    ):
                        write_worker_progress(
                            status="running",
                            current_shard=shard.name,
                            current_completed=image_index,
                            current_total=len(rows),
                        )
                finally:
                    image.close()
            atomic_write_jsonl(output_path, outputs)
            atomic_write_json(
                done_path,
                {
                    "stage": stage,
                    "count": len(outputs),
                    "image_id_digest": digest_ids(row["image_id"] for row in outputs),
                    "input_shard": str(shard),
                },
            )
            write_worker_progress(status="running")
            print(f"[{stage}] completed {shard.name}: {len(outputs)} images", flush=True)
        worker_final_status = "completed"
    finally:
        close_model = getattr(getattr(detector, "model", None), "close", None)
        if callable(close_model):
            close_model()
        write_worker_progress(status=worker_final_status)


def detection_python(args: argparse.Namespace, stage: str) -> str:
    return str(args.text_python if stage == "text" else args.icon_python)


def detection_worker_command(args: argparse.Namespace, stage: str, worker_index: int, worker_count: int) -> list[str]:
    command = [
        detection_python(args, stage),
        str(Path(__file__).resolve()),
        "--stage", "_worker",
        "--detector-stage", stage,
        "--worker-index", str(worker_index),
        "--worker-count", str(worker_count),
        "--source-dir", str(args.source_dir),
        "--locany-data-dir", str(args.locany_data_dir),
        "--parser-root", str(args.parser_root),
        "--output-dir", str(args.output_dir),
        "--gpus", args.gpus,
        "--workers-per-gpu", str(args.workers_per_gpu),
        "--image-loader-threads", str(args.image_loader_threads),
        "--shard-size", str(args.shard_size),
        "--max-unique-images", str(args.max_unique_images),
        "--progress-interval-seconds", str(args.progress_interval_seconds),
        "--progress-every-images", str(args.progress_every_images),
        "--text-long-side", str(args.text_long_side),
        "--text-box-threshold", str(args.text_box_threshold),
        "--icon-long-side", str(args.icon_long_side),
        "--icon-confidence", str(args.icon_confidence),
    ]
    if args.text_model_dir:
        command.extend(("--text-model-dir", str(args.text_model_dir)))
    if args.icon_model:
        command.extend(("--icon-model", str(args.icon_model)))
    if args.resume:
        command.append("--resume")
    if args.enable_mkldnn:
        command.append("--enable-mkldnn")
    return command


def preflight_icon_runtime(
    python_executable: str,
    *,
    gpu: str,
    model_path: Path,
) -> dict[str, Any]:
    """Validate Torch, torchvision ops, CUDA and the TorchScript model once.

    The parser intentionally treats every import failure as one short error.  A
    subprocess probe preserves the real traceback and avoids starting four
    workers that are guaranteed to fail in the same environment.
    """
    probe = r"""
import json
import sys
import numpy
import PIL
import torch
import torchvision
from torchvision.ops import batched_nms

boxes = torch.tensor([[0.0, 0.0, 2.0, 2.0], [1.0, 1.0, 3.0, 3.0]])
scores = torch.tensor([0.9, 0.8])
classes = torch.tensor([0, 0])
batched_nms(boxes, scores, classes, 0.5)
model = torch.jit.load(sys.argv[1], map_location="cpu")
del model
payload = {
    "torch": torch.__version__,
    "torchvision": torchvision.__version__,
    "numpy": numpy.__version__,
    "pillow": PIL.__version__,
    "torch_cuda": torch.version.cuda,
    "cuda_available": torch.cuda.is_available(),
    "cuda_device_count": torch.cuda.device_count(),
    "cuda_device": torch.cuda.get_device_name(0) if torch.cuda.is_available() else None,
    "model_loaded": True,
}
print("UI5_ICON_RUNTIME=" + json.dumps(payload, ensure_ascii=False))
"""
    environment = os.environ.copy()
    environment["CUDA_VISIBLE_DEVICES"] = gpu
    result = subprocess.run(
        [python_executable, "-c", probe, str(model_path)],
        env=environment,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    output = "\n".join(part.strip() for part in (result.stdout, result.stderr) if part.strip())
    if result.returncode != 0:
        raise RuntimeError(
            "icon Python 环境预检失败；尚未启动 GPU worker。\n"
            f"icon_python: {python_executable}\n"
            f"CUDA_VISIBLE_DEVICES: {gpu}\n"
            "需要同一环境可导入 torch、torchvision.ops.batched_nms，且能加载 model.pt。\n"
            "原始异常如下：\n"
            f"{output or '(no subprocess output)'}"
        )
    prefix = "UI5_ICON_RUNTIME="
    payload_line = next(
        (line for line in result.stdout.splitlines() if line.startswith(prefix)),
        None,
    )
    if payload_line is None:
        raise RuntimeError(f"icon Python 预检没有返回版本信息：\n{output}")
    payload = json.loads(payload_line[len(prefix) :])
    if not payload.get("cuda_available") or int(payload.get("cuda_device_count", 0)) < 1:
        raise RuntimeError(
            "icon Python 可以导入 Torch，但该环境看不到请求的 CUDA GPU。\n"
            f"icon_python: {python_executable}\n"
            f"CUDA_VISIBLE_DEVICES: {gpu}\n"
            f"runtime: {json.dumps(payload, ensure_ascii=False)}"
        )
    print(
        "[icon 环境预检] "
        f"python={python_executable} | torch={payload['torch']} | "
        f"torchvision={payload['torchvision']} | torch CUDA={payload['torch_cuda']} | "
        f"GPU={payload['cuda_device']} | model.pt=OK",
        flush=True,
    )
    return payload


def run_detection_stage(args: argparse.Namespace, stage: str) -> None:
    paths = AuditPaths(args.output_dir)
    unique = read_jsonl(paths.unique_images)
    baseline_completed = 0
    for shard in sorted(paths.shards.glob("shard_*.jsonl")):
        output_path = paths.stage_dir(stage) / shard.name
        done_path = paths.stage_dir(stage) / (shard.stem + ".done.json")
        if args.resume and completed_shard_valid(shard, output_path, done_path, stage):
            baseline_completed += len(read_jsonl(shard))
    if args.resume and unique and baseline_completed == len(unique):
        # A fully completed stage must not import or initialize its model again.
        # This also lets merge/crop reuse immutable detections after GPU envs are released.
        print_preflight(
            args,
            unique_count=len(unique),
            readable_count=len(unique),
            detector_stage=None,
        )
        reporter = ProgressReporter(
            stage=stage,
            total=len(unique),
            output_dir=args.output_dir,
            interval_seconds=args.progress_interval_seconds,
            initial_completed=baseline_completed,
        )
        reporter.update(
            baseline_completed,
            status="completed",
            detail="--resume 已验证全部 shard，跳过模型加载和 GPU worker",
            force=True,
        )
        return
    print_preflight(args, unique_count=len(unique), readable_count=len(unique), detector_stage=stage)
    if args.workers_per_gpu == 2 and not args.allow_two_processes_per_gpu:
        raise ValueError(
            "2 processes/GPU requires --allow-two-processes-per-gpu after the 2,000-image "
            "benchmark confirms GPU <40%, memory <12GB, and higher throughput"
        )
    gpus = [gpu.strip() for gpu in args.gpus.split(",") if gpu.strip()]
    if not gpus:
        raise ValueError("--gpus must name at least one GPU")
    nvidia_smi = shutil.which("nvidia-smi")
    if nvidia_smi is None:
        raise FileNotFoundError("nvidia-smi is required to validate --gpus before detection")
    query = subprocess.run(
        [nvidia_smi, "--query-gpu=index", "--format=csv,noheader"],
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    available = {line.strip() for line in query.stdout.splitlines() if line.strip()}
    missing_gpus = [gpu for gpu in gpus if gpu not in available]
    if missing_gpus:
        raise ValueError(f"requested GPUs are unavailable: {missing_gpus}; available={sorted(available)}")
    runtime: dict[str, Any] = {"python": detection_python(args, stage)}
    if stage == "icon":
        config = detector_config(args)
        runtime.update(
            preflight_icon_runtime(
                detection_python(args, stage),
                gpu=gpus[0],
                model_path=Path(config["icon"]["model"]),
            )
        )
    slots = [(gpu, slot) for gpu in gpus for slot in range(args.workers_per_gpu)]
    processes = []
    wall_started = time.perf_counter()
    reporter = ProgressReporter(
        stage=stage,
        total=len(unique),
        output_dir=args.output_dir,
        interval_seconds=args.progress_interval_seconds,
        initial_completed=baseline_completed,
    )
    reporter.update(
        baseline_completed,
        detail=f"启动 {len(slots)} 个常驻 GPU worker",
        force=True,
    )
    for worker_index, (gpu, _slot) in enumerate(slots):
        env = os.environ.copy()
        env["CUDA_VISIBLE_DEVICES"] = gpu
        command = detection_worker_command(args, stage, worker_index, len(slots))
        processes.append((worker_index, gpu, subprocess.Popen(command, env=env)))
    def observed_completed() -> int:
        done_shards: set[str] = set()
        completed = 0
        for marker_path in paths.stage_dir(stage).glob("shard_*.done.json"):
            try:
                marker = json.loads(marker_path.read_text(encoding="utf-8"))
                if marker.get("stage") == stage:
                    completed += int(marker.get("count", 0))
                    done_shards.add(marker_path.name.replace(".done.json", ".jsonl"))
            except (OSError, ValueError, json.JSONDecodeError):
                continue
        for progress_path in (paths.stage_dir(stage) / "progress").glob("worker_*.json"):
            try:
                progress = json.loads(progress_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            current_shard = progress.get("current_shard")
            if current_shard and current_shard not in done_shards:
                completed += int(progress.get("current_shard_completed", 0))
        return min(len(unique), completed)

    while any(process.poll() is None for _, _, process in processes):
        reporter.update(observed_completed())
        time.sleep(min(1.0, args.progress_interval_seconds))
    failures = [
        (worker_index, gpu, int(process.returncode))
        for worker_index, gpu, process in processes
        if process.returncode
    ]
    if failures:
        reporter.update(
            observed_completed(),
            status="failed",
            detail=f"GPU worker 失败：{failures}",
            force=True,
        )
        raise RuntimeError(f"{stage} detector workers failed: {failures}")
    stage_rows = [
        row
        for path in sorted(paths.stage_dir(stage).glob("shard_*.jsonl"))
        for row in read_jsonl(path)
    ]
    total_inference_ms = sum(float(row.get("inference_ms", 0.0)) for row in stage_rows)
    wall_seconds = time.perf_counter() - wall_started
    atomic_write_json(
        paths.stage_dir(stage) / "stage_summary.json",
        {
            "stage": stage,
            "images": len(stage_rows),
            "workers": len(slots),
            "workers_per_gpu": args.workers_per_gpu,
            "runtime": runtime,
            "wall_seconds": round(wall_seconds, 3),
            "throughput_images_per_second": round(len(stage_rows) / wall_seconds, 6) if wall_seconds else 0.0,
            "sum_inference_ms": round(total_inference_ms, 3),
            "mean_inference_ms": round(total_inference_ms / len(stage_rows), 3) if stage_rows else 0.0,
        },
    )
    reporter.update(
        len(stage_rows),
        status="completed",
        detail=f"{len(slots)} 个 worker 已退出，检测结果已落盘",
        force=True,
    )


def merge_detections(args: argparse.Namespace) -> list[dict[str, Any]]:
    paths = AuditPaths(args.output_dir)
    if not paths.detector_config.is_file():
        raise FileNotFoundError(f"missing immutable detector config: {paths.detector_config}")
    ensure_detector_config(paths.detector_config, detector_config(args))
    expected_rows = read_jsonl(paths.unique_images)
    expected = {row["image_id"]: row for row in expected_rows}
    if len(expected) != len(expected_rows):
        raise ValueError("unique_images.jsonl contains duplicate image_id")
    reporter = ProgressReporter(
        stage="merge",
        total=len(expected),
        output_dir=args.output_dir,
        interval_seconds=args.progress_interval_seconds,
    )
    reporter.update(0, detail="验证 text/icon shard 完整性", force=True)
    by_stage: dict[str, dict[str, dict[str, Any]]] = {}
    for stage in ("text", "icon"):
        invalid_shards = []
        for shard in sorted(paths.shards.glob("shard_*.jsonl")):
            output_path = paths.stage_dir(stage) / shard.name
            done_path = paths.stage_dir(stage) / (shard.stem + ".done.json")
            if not completed_shard_valid(shard, output_path, done_path, stage):
                invalid_shards.append(shard.name)
        if invalid_shards:
            raise ValueError(
                f"{stage} has incomplete or invalid shards: {invalid_shards[:20]}"
            )
        rows = [row for path in sorted(paths.stage_dir(stage).glob("shard_*.jsonl")) for row in read_jsonl(path)]
        indexed: dict[str, dict[str, Any]] = {}
        duplicates = []
        for row in rows:
            image_id = row["image_id"]
            if image_id in indexed:
                duplicates.append(image_id)
            indexed[image_id] = row
        if duplicates:
            raise ValueError(f"duplicate {stage} detections: {sorted(set(duplicates))[:20]}")
        missing = sorted(set(expected) - set(indexed))
        extra = sorted(set(indexed) - set(expected))
        if len(rows) != len(expected) or missing or extra:
            raise ValueError(
                f"{stage} count mismatch: unique={len(expected)}, rows={len(rows)}, "
                f"missing={missing[:20]}, extra={extra[:20]}"
            )
        by_stage[stage] = indexed
    merged = []
    for merge_index, image_id in enumerate(sorted(expected), 1):
        manifest = expected[image_id]
        text_row = by_stage["text"][image_id]
        icon_row = by_stage["icon"][image_id]
        dimensions = {
            (int(manifest["width"]), int(manifest["height"])),
            (int(text_row["width"]), int(text_row["height"])),
            (int(icon_row["width"]), int(icon_row["height"])),
        }
        if len(dimensions) != 1:
            raise ValueError(f"dimension mismatch for {image_id}: {dimensions}")
        merged.append(
            {
                "image_id": image_id,
                "content_id": manifest["content_id"],
                "image": manifest["image_path"],
                "width": manifest["width"],
                "height": manifest["height"],
                "text_detections": text_row["text_detections"],
                "icon_detections": icon_row["icon_detections"],
            }
        )
        if merge_index % args.progress_every_images == 0 or merge_index == len(expected):
            reporter.update(merge_index, detail="按 image_id 合并 text/icon 检测")
    if len(merged) != len(expected):
        raise AssertionError("merged count changed unexpectedly")
    atomic_write_jsonl(paths.merged, merged)
    atomic_write_json(
        paths.merged.parent / "merge_summary.json",
        {
            "unique_images": len(expected),
            "text_rows": len(by_stage["text"]),
            "icon_rows": len(by_stage["icon"]),
            "merged_rows": len(merged),
            "image_id_digest": digest_ids(expected),
        },
    )
    reporter.update(
        len(merged),
        status="completed",
        detail="数量、重复、缺失和尺寸检查全部通过",
        force=True,
    )
    return merged


def box_area(box: Sequence[int]) -> int:
    return max(0, int(box[2]) - int(box[0])) * max(0, int(box[3]) - int(box[1]))


def rect_contains(outer: Sequence[int], inner: Sequence[int]) -> bool:
    return outer[0] <= inner[0] and outer[1] <= inner[1] and outer[2] >= inner[2] and outer[3] >= inner[3]


def rect_intersects(left: Sequence[int], right: Sequence[int]) -> bool:
    return not (left[2] <= right[0] or right[2] <= left[0] or left[3] <= right[1] or right[3] <= left[1])


def proposal_crops(
    cropper: Any,
    detection_record: Mapping[str, Any],
    config: Mapping[str, float],
    *,
    max_crops: int,
    boundary_margin_ratio: float,
    min_context_image_ratio: float = 0.0,
) -> dict[str, Any]:
    width, height = int(detection_record["width"]), int(detection_record["height"])
    detections = []
    for source in ("text", "icon"):
        for source_index, item in enumerate(detection_record[f"{source}_detections"]):
            detections.append(
                cropper.DetectionBox(
                    index=len(detections),
                    source=source,
                    source_index=source_index,
                    bbox=tuple(int(value) for value in item["bbox"]),
                    score=float(item["score"]) if item.get("score") is not None else None,
                )
            )
    boxes = [detection.bbox for detection in detections]
    components, edge_count = cropper.build_connected_components(
        detections,
        width,
        height,
        config["horizontal_link_ratio"],
        config["vertical_link_ratio"],
    )
    initial_count = len(components)
    empty_fallback = False
    if components:
        groups, merge_history = cropper.merge_groups_to_limit(components, detections, max_crops)
        groups, overlap_history = cropper.merge_overlapping_group_envelopes(groups)
        if min_context_image_ratio > 0:
            crops = []
            for group in groups:
                group_width = max(1, int(group.bbox[2]) - int(group.bbox[0]))
                group_height = max(1, int(group.bbox[3]) - int(group.bbox[1]))
                pad_x = max(
                    math.ceil(group_width * config["context_ratio"]),
                    math.ceil(width * min_context_image_ratio),
                )
                pad_y = max(
                    math.ceil(group_height * config["context_ratio"]),
                    math.ceil(height * min_context_image_ratio),
                )
                crops.append(
                    (
                        max(0, int(group.bbox[0]) - pad_x),
                        max(0, int(group.bbox[1]) - pad_y),
                        min(width, int(group.bbox[2]) + pad_x),
                        min(height, int(group.bbox[3]) + pad_y),
                    )
                )
            context_adjustments = [
                {
                    "reason": "task_min_context_image_ratio",
                    "min_context_image_ratio": min_context_image_ratio,
                }
            ]
            # A minimum context floor may make desired rectangles overlap.
            # Merge them losslessly; never trim through a detector box.
            while True:
                pair = next(
                    (
                        (left, right)
                        for left in range(len(crops))
                        for right in range(left + 1, len(crops))
                        if rect_intersects(crops[left], crops[right])
                    ),
                    None,
                )
                if pair is None:
                    break
                left, right = pair
                merged_context = cropper.union_bbox([crops[left], crops[right]])
                overlap_history.append(
                    {
                        "reason": "overlapping_task_min_context",
                        "left": list(crops[left]),
                        "right": list(crops[right]),
                        "merged": list(merged_context),
                    }
                )
                crops.pop(right)
                crops.pop(left)
                crops.append(merged_context)
        else:
            crops, context_adjustments = cropper.make_non_overlapping_context_crops(
                groups, width, height, config["context_ratio"]
            )
    else:
        groups, merge_history, overlap_history, context_adjustments = [], [], [], []
        crops = [(0, 0, width, height)]
        empty_fallback = True
    crop_boxes = [tuple(int(value) for value in crop) for crop in crops]
    boundary_merge_history = []
    if boxes:
        crop_boxes = [
            cropper.make_boundary_safe_crop(
                crop, boxes, width, height, boundary_margin_ratio
            )[0]
            for crop in crop_boxes
        ]
        # Boundary-safe expansion is lossless.  If two expanded crops overlap,
        # merge them instead of allowing overlap or trimming through a detector.
        while True:
            pair = next(
                (
                    (left, right)
                    for left in range(len(crop_boxes))
                    for right in range(left + 1, len(crop_boxes))
                    if rect_intersects(crop_boxes[left], crop_boxes[right])
                ),
                None,
            )
            if pair is None:
                break
            left, right = pair
            merged = cropper.union_bbox([crop_boxes[left], crop_boxes[right]])
            merged = cropper.make_boundary_safe_crop(
                merged, boxes, width, height, boundary_margin_ratio
            )[0]
            boundary_merge_history.append(
                {"left": list(crop_boxes[left]), "right": list(crop_boxes[right]), "merged": list(merged)}
            )
            crop_boxes.pop(right)
            crop_boxes.pop(left)
            crop_boxes.append(merged)
        crop_boxes.sort(key=lambda box: (box[1], box[0], box[3], box[2]))
    cut = [
        {"crop": list(crop), "detector_box": list(box)}
        for crop in crop_boxes
        for box in boxes
        if rect_intersects(crop, box) and not rect_contains(crop, box)
    ]
    if cut:
        raise AssertionError(f"crop boundary cuts {len(cut)} detector boxes")
    if not 1 <= len(crop_boxes) <= max_crops:
        raise AssertionError(f"invalid crop count: {len(crop_boxes)}")
    return {
        "crop_boxes": [list(crop) for crop in crop_boxes],
        "detection_boxes": [list(box) for box in boxes],
        "detection_count": len(boxes),
        "edge_count": edge_count,
        "component_count_before_merge": initial_count,
        "forced_merge": bool(merge_history or overlap_history or boundary_merge_history),
        "merge_history": merge_history,
        "boundary_merge_history": boundary_merge_history,
        "overlap_merge_history": overlap_history,
        "context_adjustments": context_adjustments,
        "empty_detection_fallback": empty_fallback,
        "detector_boundary_cut_count": len(cut),
    }


def normalize_gt_in_crop(gt: Sequence[int], crop: Sequence[int]) -> dict[str, Any]:
    crop_width = crop[2] - crop[0]
    crop_height = crop[3] - crop[1]
    if crop_width <= 0 or crop_height <= 0:
        raise ValueError(f"invalid crop: {crop}")
    local = [gt[0] - crop[0], gt[1] - crop[1], gt[2] - crop[0], gt[3] - crop[1]]
    norm1000 = [
        round(local[0] / crop_width * 1000),
        round(local[1] / crop_height * 1000),
        round(local[2] / crop_width * 1000),
        round(local[3] / crop_height * 1000),
    ]
    reconstructed_local = [
        round(norm1000[0] / 1000 * crop_width),
        round(norm1000[1] / 1000 * crop_height),
        round(norm1000[2] / 1000 * crop_width),
        round(norm1000[3] / 1000 * crop_height),
    ]
    reconstructed_original = [
        reconstructed_local[0] + crop[0],
        reconstructed_local[1] + crop[1],
        reconstructed_local[2] + crop[0],
        reconstructed_local[3] + crop[1],
    ]
    max_error = max(abs(left - right) for left, right in zip(gt, reconstructed_original))
    return {
        "original_bbox": list(gt),
        "local_bbox": local,
        "norm1000": norm1000,
        "roundtrip_original_bbox": reconstructed_original,
        "roundtrip_max_error_px": max_error,
    }


def build_answer(label: str, boxes: Sequence[Sequence[int]]) -> str:
    if not boxes:
        return "<box>none</box>"
    return f"<ref>{label}</ref>" + "".join(
        f"<box><{box[0]}><{box[1]}><{box[2]}><{box[3]}></box>" for box in boxes
    )


def build_preview_rows(
    sample: Mapping[str, Any],
    crop_boxes: Sequence[Sequence[int]],
    crop_paths: Sequence[Path],
    *,
    config_name: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    gt_boxes = sample["gt_boxes"]
    if uses_task_whole_image_policy(str(sample["task"])):
        full_box = [0, 0, int(sample["width"]), int(sample["height"])]
        if len(crop_boxes) != 1 or list(crop_boxes[0]) != full_box:
            raise AssertionError("ui_content_missing must use the exact full image")
        if len(crop_paths) != 1:
            raise AssertionError("ui_content_missing must reference exactly one full image")
        original_norm = [list(box) for box in sample["gt_boxes_1000"]]
        if len(original_norm) != len(gt_boxes):
            raise ValueError("content_missing pixel and normalized GT counts differ")
        transforms = [
            {
                "original_bbox": list(pixel_box),
                "original_norm1000": list(norm_box),
                "output_norm1000": list(norm_box),
                "label_transform_applied": False,
                "normalized_gt_identical": True,
                "roundtrip_max_error_px": 0,
                "roundtrip_gate_excluded": True,
            }
            for pixel_box, norm_box in zip(gt_boxes, original_norm)
        ]
        label = TASK_LABELS[str(sample["task"])]
        return [
            {
                "sample_id": sample["sample_id"],
                "image_id": sample["image_id"],
                "task": sample["task"],
                "config": config_name,
                "source_image": sample["canonical_path"],
                "image": str(crop_paths[0]),
                "crop_id": 1,
                "crop_bbox": full_box,
                "positive": bool(gt_boxes),
                "gt_count": len(gt_boxes),
                "contained_gt_indices": list(range(len(gt_boxes))),
                "partial_gt_indices": [],
                "training_eligible": True,
                "roundtrip_max_error_px": 0,
                "roundtrip_gate_excluded": True,
                "label_transform_applied": False,
                "normalized_gt_identical": True,
                "original_gt_boxes_1000": original_norm,
                "output_gt_boxes_1000": [list(box) for box in original_norm],
                "coordinate_transforms": transforms,
                "conversations": [
                    {
                        "from": "human",
                        "value": f"Locate all the instances that match the following description: {label}.",
                    },
                    {"from": "gpt", "value": build_answer(label, original_norm)},
                ],
            }
        ], []
    preview = []
    failures = []
    negative_kept = False
    for crop_index, (crop, crop_path) in enumerate(zip(crop_boxes, crop_paths), 1):
        contained = [index for index, gt in enumerate(gt_boxes) if rect_contains(crop, gt)]
        partial = [
            index
            for index, gt in enumerate(gt_boxes)
            if rect_intersects(crop, gt) and not rect_contains(crop, gt)
        ]
        transformed = [normalize_gt_in_crop(gt_boxes[index], crop) for index in contained]
        max_error = max((item["roundtrip_max_error_px"] for item in transformed), default=0)
        training_eligible = not partial and max_error <= 1
        if partial:
            failure_type = "partial_intersection"
        elif max_error > 1:
            failure_type = "roundtrip_error"
        else:
            failure_type = None
        if failure_type:
            failures.append(
                {
                    "sample_id": sample["sample_id"],
                    "image_id": sample["image_id"],
                    "task": sample["task"],
                    "crop_id": crop_index,
                    "crop_bbox": list(crop),
                    "failure_type": failure_type,
                    "training_eligible": training_eligible,
                    "partial_gt_indices": partial,
                    "roundtrip_max_error_px": max_error,
                }
            )
        if not contained and not partial:
            if negative_kept:
                continue
            negative_kept = True
        norm_boxes = [item["norm1000"] for item in transformed]
        label = TASK_LABELS[sample["task"]]
        preview.append(
            {
                "sample_id": sample["sample_id"],
                "image_id": sample["image_id"],
                "task": sample["task"],
                "config": config_name,
                "source_image": sample["canonical_path"],
                "image": str(crop_path),
                "crop_id": crop_index,
                "crop_bbox": list(crop),
                "positive": bool(contained) if training_eligible else None,
                "gt_count": len(contained),
                "contained_gt_indices": contained,
                "partial_gt_indices": partial,
                "training_eligible": training_eligible,
                "roundtrip_max_error_px": max_error,
                "roundtrip_gate_excluded": False,
                "label_transform_applied": True,
                "coordinate_transforms": transformed,
                "conversations": [
                    {
                        "from": "human",
                        "value": f"Locate all the instances that match the following description: {label}.",
                    },
                    {"from": "gpt", "value": build_answer(label, norm_boxes)},
                ],
            }
        )
    return preview, failures


def rectangle_union_area(rectangles: Sequence[Sequence[int]]) -> int:
    if not rectangles:
        return 0
    xs = sorted({coordinate for box in rectangles for coordinate in (box[0], box[2])})
    total = 0
    for left, right in zip(xs, xs[1:]):
        intervals = sorted(
            (box[1], box[3]) for box in rectangles if box[0] < right and box[2] > left
        )
        if not intervals:
            continue
        start, end = intervals[0]
        covered = 0
        for next_start, next_end in intervals[1:]:
            if next_start <= end:
                end = max(end, next_end)
            else:
                covered += end - start
                start, end = next_start, next_end
        covered += end - start
        total += (right - left) * covered
    return total


def atomic_save_png(image: Image.Image, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + f".tmp.{os.getpid()}")
    image.save(temporary, format="PNG", compress_level=1)
    os.replace(temporary, path)


def save_raw_crops(image: Image.Image, crop_boxes: Sequence[Sequence[int]], directory: Path, prefix: str = "crop") -> list[Path]:
    directory.mkdir(parents=True, exist_ok=True)
    paths = []
    for index, crop in enumerate(crop_boxes, 1):
        path = directory / f"{prefix}_{index:02d}.png"
        cropped = image.crop(tuple(crop))
        try:
            atomic_save_png(cropped, path)
        finally:
            cropped.close()
        paths.append(path.resolve())
    return paths


def uses_task_whole_image_policy(task: str) -> bool:
    """Global-defect policy is selected from task identity, never a filename or GT."""
    return task == "ui_content_missing"


def save_overview(
    image: Image.Image,
    crop_boxes: Sequence[Sequence[int]],
    gt_boxes: Sequence[Sequence[int]],
    path: Path,
) -> None:
    rendered = image.copy()
    draw = ImageDraw.Draw(rendered)
    line = max(2, round(min(image.size) / 400))
    for index, crop in enumerate(crop_boxes, 1):
        draw.rectangle(tuple(crop), outline=(0, 210, 255), width=line)
        draw.text((crop[0] + line, crop[1] + line), f"crop {index}", fill=(0, 110, 170))
    for index, gt in enumerate(gt_boxes, 1):
        contained = any(rect_contains(crop, gt) for crop in crop_boxes)
        color = (255, 205, 0) if contained else (255, 40, 40)
        draw.rectangle(tuple(gt), outline=color, width=line)
        draw.text((gt[0] + line, gt[1] + line), f"GT {index}", fill=color)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        atomic_save_png(rendered, path)
    finally:
        rendered.close()


def percentile(values: Sequence[float], percent: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    position = (len(ordered) - 1) * percent
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return float(ordered[lower])
    fraction = position - lower
    return float(ordered[lower] * (1 - fraction) + ordered[upper] * fraction)


def detection_density_bucket(detection_count: int) -> str:
    if detection_count <= 50:
        return "sparse"
    if detection_count <= 150:
        return "medium"
    return "dense"


def aggregate_scope(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    gt_total = sum(row["gt_count"] for row in rows)
    gt_contained = sum(row["gt_contained_count"] for row in rows)
    positive = [row for row in rows if row["positive"]]
    crop_counts = [row["crop_count"] for row in rows]
    area_ratios = [row["union_area_ratio"] for row in rows]
    gains = [gain for row in rows for gain in row["gt_enlargement_gains"]]
    total_original = sum(row["original_area"] for row in rows)
    total_union = sum(row["union_crop_area"] for row in rows)
    distribution = Counter(crop_counts)
    uncovered = gt_total - gt_contained
    partial_only = sum(row["partial_only_gt_count"] for row in rows)
    empty_fallback = len({row["image_id"] for row in rows if row["empty_detection_fallback"]})
    forced_merge = len({row["image_id"] for row in rows if row["forced_merge"]})
    boundary_cuts = sum(row["detector_boundary_cut_count"] for row in rows)
    roundtrip_failures = sum(row["roundtrip_error_over_1_count"] for row in rows)
    near_full = sum(row["union_area_ratio"] > 0.8 for row in rows)
    return {
        "samples": len(rows),
        "positive_samples": len(positive),
        "negative_samples": len(rows) - len(positive),
        "gt_count": gt_total,
        "gt_contained_count": gt_contained,
        "gt_box_containment_recall": gt_contained / gt_total if gt_total else 1.0,
        "positive_sample_success_count": sum(row["all_gt_contained"] for row in positive),
        "positive_sample_success_rate": (
            sum(row["all_gt_contained"] for row in positive) / len(positive) if positive else 1.0
        ),
        "uncovered_gt_count": uncovered,
        "partial_only_gt_count": partial_only,
        "partial_only_gt_ratio": (
            partial_only / gt_total if gt_total else 0.0
        ),
        "crop_count": {
            "mean": statistics.fmean(crop_counts) if crop_counts else 0.0,
            "p50": percentile(crop_counts, 0.50),
            "p90": percentile(crop_counts, 0.90),
            "max": max(crop_counts, default=0),
            "distribution_1_to_10": {str(value): distribution.get(value, 0) for value in range(1, 11)},
        },
        "union_area_ratio": {
            "mean": statistics.fmean(area_ratios) if area_ratios else 0.0,
            "p50": percentile(area_ratios, 0.50),
            "p90": percentile(area_ratios, 0.90),
        },
        "pixel_reduction_ratio": 1 - total_union / total_original if total_original else 0.0,
        "near_full_image_count": near_full,
        "near_full_image_ratio": near_full / len(rows) if rows else 0.0,
        "gt_gain_over_1_25_ratio": sum(gain > 1.25 for gain in gains) / len(gains) if gains else 0.0,
        "gt_gain_over_1_5_ratio": sum(gain > 1.5 for gain in gains) / len(gains) if gains else 0.0,
        "gt_gain_over_2_0_ratio": sum(gain > 2.0 for gain in gains) / len(gains) if gains else 0.0,
        "empty_detection_fallback_images": empty_fallback,
        "forced_merge_images": forced_merge,
        "detector_boundary_cut_count": boundary_cuts,
        "roundtrip_error_over_1_count": roundtrip_failures,
        "anomaly_event_count": (
            uncovered + partial_only + near_full + empty_fallback + forced_merge
            + boundary_cuts + roundtrip_failures
        ),
    }


def make_image_detail(
    sample: Mapping[str, Any],
    proposal: Mapping[str, Any],
    crop_boxes: Sequence[Sequence[int]],
    crop_paths: Sequence[Path],
    overview: Path,
    config_name: str,
    roundtrip_errors: Sequence[int],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    gt_coverage = []
    gains = []
    failures = []
    width, height = sample["width"], sample["height"]
    density = detection_density_bucket(int(proposal.get("detection_count", 0)))
    for gt_index, gt in enumerate(sample["gt_boxes"]):
        contained_by = [index + 1 for index, crop in enumerate(crop_boxes) if rect_contains(crop, gt)]
        partial_by = [
            index + 1
            for index, crop in enumerate(crop_boxes)
            if rect_intersects(crop, gt) and not rect_contains(crop, gt)
        ]
        if contained_by:
            gains.append(
                max(
                    min(width / (crop_boxes[index - 1][2] - crop_boxes[index - 1][0]), height / (crop_boxes[index - 1][3] - crop_boxes[index - 1][1]))
                    for index in contained_by
                )
            )
        else:
            failure_type = "partial_intersection" if partial_by else "uncovered"
            compensation: dict[str, int] | None = None
            compensation_max: int | None = None
            compensation_total: int | None = None
            compensation_bucket = "not_applicable"
            if partial_by:
                choices = []
                for crop_id in partial_by:
                    crop = crop_boxes[crop_id - 1]
                    required = {
                        "left": max(0, int(crop[0]) - int(gt[0])),
                        "top": max(0, int(crop[1]) - int(gt[1])),
                        "right": max(0, int(gt[2]) - int(crop[2])),
                        "bottom": max(0, int(gt[3]) - int(crop[3])),
                    }
                    maximum = max(required.values())
                    total = sum(required.values())
                    choices.append((maximum, total, crop_id, required))
                compensation_max, compensation_total, _, compensation = min(choices)
                if compensation_max <= 16:
                    compensation_bucket = "small_0_16px"
                elif compensation_max <= 64:
                    compensation_bucket = "medium_17_64px"
                else:
                    compensation_bucket = "large_over_64px"
            failures.append(
                {
                    "config": config_name,
                    "sample_id": sample["sample_id"],
                    "image_id": sample["image_id"],
                    "task": sample["task"],
                    "gt_index": gt_index,
                    "gt_bbox": gt,
                    "intersecting_crop_ids": partial_by,
                    "intersecting_crop_bboxes": [crop_boxes[index - 1] for index in partial_by],
                    "failure_type": failure_type,
                    "detection_density": density,
                    "required_compensation_px": compensation,
                    "required_max_single_side_px": compensation_max,
                    "required_total_px": compensation_total,
                    "compensation_bucket": compensation_bucket,
                    "visualization": str(overview.resolve()),
                }
            )
        gt_coverage.append({"contained_by": contained_by, "partial_by": partial_by})
    original_area = width * height
    union_area = rectangle_union_area(crop_boxes)
    whole_image = uses_task_whole_image_policy(str(sample["task"]))
    detection_count = int(proposal.get("detection_count", 0))
    detail = {
        "config": config_name,
        "sample_id": sample["sample_id"],
        "image_id": sample["image_id"],
        "task": sample["task"],
        "positive": sample["positive"],
        "gt_count": len(sample["gt_boxes"]),
        "gt_contained_count": sum(bool(item["contained_by"]) for item in gt_coverage),
        "partial_only_gt_count": sum(not item["contained_by"] and bool(item["partial_by"]) for item in gt_coverage),
        "all_gt_contained": all(bool(item["contained_by"]) for item in gt_coverage),
        "crop_count": len(crop_boxes),
        "original_area": original_area,
        "union_crop_area": union_area,
        "union_area_ratio": union_area / original_area if original_area else 0.0,
        "pixel_reduction_ratio": 1 - union_area / original_area if original_area else 0.0,
        "gt_enlargement_gains": gains,
        "empty_detection_fallback": proposal["empty_detection_fallback"],
        "forced_merge": proposal["forced_merge"],
        "component_count_before_merge": proposal["component_count_before_merge"],
        "detection_count": detection_count,
        "detection_density": density,
        "detector_boundary_cut_count": proposal["detector_boundary_cut_count"],
        "roundtrip_error_over_1_count": (
            0 if whole_image else sum(error > 1 for error in roundtrip_errors)
        ),
        "roundtrip_gate_excluded": whole_image,
        "label_transform_applied": not whole_image,
        "normalized_gt_identical": whole_image,
        "overview": str(overview.resolve()),
        "source_image": sample["canonical_path"],
        "crop_boxes": [list(crop) for crop in crop_boxes],
        "crop_paths": [str(path.resolve()) for path in crop_paths],
    }
    return detail, failures


def task_overlap_rows(overlap: Mapping[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for dataset_key, dataset_label in (("source_data", "original_source"), ("actual_training_data", "actual_training")):
        dataset = overlap[dataset_key]
        for identity_key, identity_label in (("path_overlap", "path"), ("content_overlap", "content")):
            for metric in ("counts", "jaccard"):
                rows.append([dataset_label, identity_label, metric, "task", *TASK_NAMES])
                matrix = dataset[identity_key][metric]
                for task in TASK_NAMES:
                    rows.append([dataset_label, identity_label, metric, task, *[matrix[task][other] for other in TASK_NAMES]])
                rows.append([])
    return rows


def write_statistics_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    columns = [
        "config", "sample_id", "image_id", "task", "positive", "gt_count",
        "gt_contained_count", "partial_only_gt_count", "all_gt_contained", "crop_count",
        "original_area", "union_crop_area", "union_area_ratio", "pixel_reduction_ratio",
        "detection_count", "detection_density", "empty_detection_fallback",
        "forced_merge", "detector_boundary_cut_count", "roundtrip_error_over_1_count",
        "roundtrip_gate_excluded", "label_transform_applied", "normalized_gt_identical",
        "partial_training_eligible_count", "hard_negative_count", "overview", "source_image",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + f".tmp.{os.getpid()}")
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column) for column in columns})
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temporary, path)


def write_excel_report(
    path: Path,
    summary: Mapping[str, Any],
    overlap: Mapping[str, Any],
    details: Sequence[Mapping[str, Any]],
    gt_failures: Sequence[Mapping[str, Any]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    candidate_summaries = summary.get("candidates", summary.get("configs", {}))

    summary_sheet = workbook.create_sheet("summary")
    summary_headers = [
        "config", "scope", "samples", "positive_samples", "negative_samples", "gt_count",
        "gt_contained_count", "gt_box_containment_recall", "positive_sample_success_rate",
        "uncovered_gt_count", "partial_only_gt_count", "partial_only_gt_ratio", "crop_mean",
        "crop_p50", "crop_p90", "crop_max", "union_area_mean", "union_area_p50",
        "union_area_p90", "pixel_reduction_ratio", "near_full_image_ratio",
        "gt_gain_over_1_25_ratio", "gt_gain_over_1_5_ratio", "gt_gain_over_2_0_ratio",
        "empty_detection_fallback_images", "forced_merge_images", "detector_boundary_cut_count",
        "roundtrip_error_over_1_count", "anomaly_event_count",
        "gt_recall_denominator_definition", "positive_success_denominator_definition",
        "near_full_denominator_definition", "gain_denominator_definition",
        "negative_samples_definition",
    ]
    summary_sheet.append(summary_headers)
    for config_name in candidate_summaries:
        scopes = ["ALL"]
        if "REGION_ALL" in candidate_summaries[config_name]["by_scope"]:
            scopes.append("REGION_ALL")
        scopes.extend(TASK_NAMES)
        for scope in scopes:
            metric = candidate_summaries[config_name]["by_scope"][scope]
            summary_sheet.append(
                [
                    config_name, scope, metric["samples"], metric["positive_samples"],
                    metric["negative_samples"], metric["gt_count"], metric["gt_contained_count"],
                    metric["gt_box_containment_recall"], metric["positive_sample_success_rate"],
                    metric["uncovered_gt_count"], metric["partial_only_gt_count"],
                    metric["partial_only_gt_ratio"], metric["crop_count"]["mean"],
                    metric["crop_count"]["p50"], metric["crop_count"]["p90"],
                    metric["crop_count"]["max"], metric["union_area_ratio"]["mean"],
                    metric["union_area_ratio"]["p50"], metric["union_area_ratio"]["p90"],
                    metric["pixel_reduction_ratio"], metric["near_full_image_ratio"],
                    metric["gt_gain_over_1_25_ratio"], metric["gt_gain_over_1_5_ratio"],
                    metric["gt_gain_over_2_0_ratio"], metric["empty_detection_fallback_images"],
                    metric["forced_merge_images"], metric["detector_boundary_cut_count"],
                    metric["roundtrip_error_over_1_count"], metric["anomaly_event_count"],
                    METRIC_DEFINITIONS["gt_box_containment_recall"],
                    METRIC_DEFINITIONS["positive_sample_success_rate"],
                    METRIC_DEFINITIONS["near_full_image_ratio"],
                    METRIC_DEFINITIONS["gt_gain_over_1_25_1_5_2_0_ratio"],
                    METRIC_DEFINITIONS["negative_samples"],
                ]
            )

    overlap_sheet = workbook.create_sheet("task_overlap")
    for row in task_overlap_rows(overlap):
        overlap_sheet.append(row)
    cross_task = summary.get("cross_task_supervision", {})
    if cross_task:
        overlap_sheet.append(["v3_audit", "metric", "key", "value"])
        for task, count in cross_task.get(
            "per_task_content_unique_images", {}
        ).items():
            overlap_sheet.append(
                ["v3_audit", "content_unique_images", task, count]
            )
        for cardinality, count in cross_task.get(
            "task_cardinality_by_content", {}
        ).items():
            overlap_sheet.append(
                ["v3_audit", "task_cardinality", cardinality, count]
            )
        for key in (
            "cross_task_positive_label_difference_images",
            "cross_task_gt_difference_images",
            "same_content_cross_train_val_count",
            "all_tasks_share_one_content_pool",
        ):
            overlap_sheet.append(["v3_audit", "summary", key, cross_task.get(key)])
        for key, value in summary.get("materialization", {}).items():
            if not isinstance(value, (list, dict)):
                overlap_sheet.append(["v3_audit", "physical_reuse", key, value])

    detail_sheet = workbook.create_sheet("image_detail")
    detail_headers = [
        "config", "sample_id", "image_id", "task", "positive", "gt_count",
        "gt_contained_count", "crop_count", "union_area_ratio", "pixel_reduction_ratio",
        "all_gt_contained", "detection_density", "empty_detection_fallback",
        "forced_merge", "roundtrip_error_over_1_count", "roundtrip_gate_excluded",
        "label_transform_applied", "normalized_gt_identical", "overview", "source_image",
        *[f"crop_{index:02d}" for index in range(1, 11)],
    ]
    detail_sheet.append(detail_headers)
    for row in details:
        values = dict(row)
        for index, crop_path in enumerate(row.get("crop_paths", []), 1):
            values[f"crop_{index:02d}"] = crop_path
        detail_sheet.append([values.get(column) for column in detail_headers])

    failure_sheet = workbook.create_sheet("gt_failures")
    failure_headers = [
        "config", "sample_id", "image_id", "task", "gt_index", "gt_bbox",
        "intersecting_crop_ids", "intersecting_crop_bboxes", "failure_type",
        "detection_density", "required_compensation_px",
        "required_max_single_side_px", "required_total_px", "compensation_bucket",
        "text_detection_count", "icon_detection_count", "crop_count_for_task",
        "manual_root_cause", "manual_note", "visualization", "visualization_4panel",
    ]
    failure_sheet.append(failure_headers)
    for row in gt_failures:
        failure_sheet.append(
            [
                row.get(column) if not isinstance(row.get(column), (list, dict)) else json.dumps(row.get(column), ensure_ascii=False)
                for column in failure_headers
            ]
        )

    compare_sheet = workbook.create_sheet("config_compare")
    compare_headers = [
        "config", "task", "density", "horizontal_link_ratio", "vertical_link_ratio",
        "context_ratio", "min_context_image_ratio", "gt_box_containment_recall",
        "positive_sample_success_rate", "union_area_mean", "pixel_reduction_ratio",
        "gt_gain_over_1_25_ratio", "gt_gain_over_1_5_ratio", "gt_gain_over_2_0_ratio",
        "uncovered_gt_count", "partial_only_gt_count", "detector_boundary_cut_count",
    ]
    compare_sheet.append(compare_headers)
    for config_name, config_summary in candidate_summaries.items():
        config = config_summary.get("parameters", {})
        scopes = ["ALL"]
        if "REGION_ALL" in config_summary["by_scope"]:
            scopes.append("REGION_ALL")
        scopes.extend(TASK_NAMES)
        for scope in scopes:
            metric = config_summary["by_scope"][scope]
            rule = config.get("task_rules", {}).get(scope, {})
            compare_sheet.append(
                [
                    config_name, scope, "ALL", rule.get("horizontal_link_ratio"),
                    rule.get("vertical_link_ratio"), rule.get("context_ratio"),
                    rule.get("min_context_image_ratio"),
                    metric["gt_box_containment_recall"], metric["positive_sample_success_rate"],
                    metric["union_area_ratio"]["mean"], metric["pixel_reduction_ratio"],
                    metric["gt_gain_over_1_25_ratio"], metric["gt_gain_over_1_5_ratio"],
                    metric["gt_gain_over_2_0_ratio"], metric["uncovered_gt_count"],
                    metric["partial_only_gt_count"], metric["detector_boundary_cut_count"],
                ]
            )
        for density, metric in config_summary.get(
            "region_by_detection_density", {}
        ).items():
            compare_sheet.append(
                [
                    config_name, "REGION_ALL", density, None, None, None, None,
                    metric["gt_box_containment_recall"],
                    metric["positive_sample_success_rate"],
                    metric["union_area_ratio"]["mean"],
                    metric["pixel_reduction_ratio"],
                    metric["gt_gain_over_1_25_ratio"],
                    metric["gt_gain_over_1_5_ratio"],
                    metric["gt_gain_over_2_0_ratio"],
                    metric["uncovered_gt_count"], metric["partial_only_gt_count"],
                    metric["detector_boundary_cut_count"],
                ]
            )
        for task, density_rows in config_summary.get(
            "region_by_task_and_detection_density", {}
        ).items():
            rule = config.get("task_rules", {}).get(task, {})
            for density, metric in density_rows.items():
                compare_sheet.append(
                    [
                        config_name, task, density,
                        rule.get("horizontal_link_ratio"),
                        rule.get("vertical_link_ratio"),
                        rule.get("context_ratio"),
                        rule.get("min_context_image_ratio"),
                        metric["gt_box_containment_recall"],
                        metric["positive_sample_success_rate"],
                        metric["union_area_ratio"]["mean"],
                        metric["pixel_reduction_ratio"],
                        metric["gt_gain_over_1_25_ratio"],
                        metric["gt_gain_over_1_5_ratio"],
                        metric["gt_gain_over_2_0_ratio"],
                        metric["uncovered_gt_count"],
                        metric["partial_only_gt_count"],
                        metric["detector_boundary_cut_count"],
                    ]
                )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 24
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            width = min(60, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            sheet.column_dimensions[letter].width = width

    percent_names = {
        "gt_box_containment_recall", "positive_sample_success_rate", "partial_only_gt_ratio",
        "union_area_mean", "union_area_p50", "union_area_p90", "pixel_reduction_ratio",
        "near_full_image_ratio", "gt_gain_over_1_25_ratio", "gt_gain_over_1_5_ratio",
        "gt_gain_over_2_0_ratio", "union_area_ratio",
    }
    for sheet in (summary_sheet, detail_sheet, compare_sheet):
        headers = {cell.value: cell.column for cell in sheet[1]}
        for name in percent_names & headers.keys():
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, headers[name]).number_format = "0.00%"
    decimal_names = {
        "crop_mean", "crop_p50", "crop_p90", "horizontal_link_ratio",
        "vertical_link_ratio", "context_ratio", "min_context_image_ratio",
        "required_max_single_side_px", "required_total_px",
    }
    for sheet in workbook.worksheets:
        headers = {cell.value: cell.column for cell in sheet[1]}
        for name in decimal_names & headers.keys():
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, headers[name]).number_format = "0.00"
    hyperlink_columns = [
        (detail_sheet, "overview"),
        (detail_sheet, "source_image"),
        (failure_sheet, "visualization"),
        (failure_sheet, "visualization_4panel"),
        *[(detail_sheet, f"crop_{index:02d}") for index in range(1, 11)],
    ]
    for sheet, header_name in hyperlink_columns:
        headers = {cell.value: cell.column for cell in sheet[1]}
        column = headers[header_name]
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            if cell.value:
                try:
                    cell.hyperlink = Path(str(cell.value)).resolve().as_uri()
                    cell.style = "Hyperlink"
                except ValueError:
                    pass
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.stem + f".tmp.{os.getpid()}" + path.suffix)
    workbook.save(temporary)
    os.replace(temporary, path)


def audit_input_snapshot(
    paths: AuditPaths,
    unique: Sequence[Mapping[str, Any]],
    detections: Sequence[Mapping[str, Any]],
    samples: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    unique_ids = [str(row["image_id"]) for row in unique]
    detection_ids = [str(row["image_id"]) for row in detections]
    if len(detection_ids) != len(set(detection_ids)) or set(detection_ids) != set(unique_ids):
        raise ValueError("merged detections do not exactly match unique manifest")
    return {
        "unique_images": len(unique_ids),
        "unique_image_id_digest": digest_ids(unique_ids),
        "merged_detections": len(detection_ids),
        "merged_image_id_digest": digest_ids(detection_ids),
        "task_samples": len(samples),
        "unique_images_file_digest": content_fingerprint(paths.unique_images),
        "task_samples_file_digest": content_fingerprint(paths.task_samples),
        "merged_detections_file_digest": content_fingerprint(paths.merged),
        "detector_config_file_digest": content_fingerprint(paths.detector_config),
    }


def audit_state_digest(state: Mapping[str, Any]) -> str:
    payload = json.dumps(state, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.blake2b(payload.encode("utf-8"), digest_size=16).hexdigest()


def invalidate_training_ready_marker(marker_path: Path) -> None:
    """Atomically make any previous success marker unusable before refresh."""
    marker_path.unlink(missing_ok=True)


def build_final_training_gate(
    candidate_gate: Mapping[str, Any],
    *,
    same_content_cross_train_val_count: int,
    content_missing_recall: float,
    content_missing_normalized_gt_mismatch_count: int,
    input_snapshot_unchanged: bool,
    all_reports_written_successfully: bool,
) -> dict[str, Any]:
    conditions = dict(candidate_gate["conditions"])
    conditions.update(
        {
            "same_content_cross_train_val_count_zero": (
                int(same_content_cross_train_val_count) == 0
            ),
            "content_missing_recall_equals_1": (
                math.isclose(float(content_missing_recall), 1.0, abs_tol=1e-12)
            ),
            "content_missing_normalized_gt_mismatch_count_zero": (
                int(content_missing_normalized_gt_mismatch_count) == 0
            ),
            "input_snapshot_unchanged": bool(input_snapshot_unchanged),
            "all_reports_written_successfully": bool(
                all_reports_written_successfully
            ),
        }
    )
    passes = all(bool(value) for value in conditions.values())
    return {
        **candidate_gate,
        "conditions": conditions,
        "passes": passes,
        "training_ready": passes,
        "training_started": False,
        "failed_conditions": [
            name for name, passed in conditions.items() if not passed
        ],
    }


def validate_training_ready_marker(audit_dir: Path) -> dict[str, Any]:
    """Validate marker digests against live audit inputs before crop training."""
    audit_dir = audit_dir.resolve(strict=True)
    marker_path = audit_dir / "training_ready.json"
    summary_path = audit_dir / "summary.json"
    state_path = audit_dir / "audit_state.json"
    if not marker_path.is_file():
        raise RuntimeError(f"training-ready marker is missing: {marker_path}")
    for required in (summary_path, state_path):
        if not required.is_file():
            raise RuntimeError(f"training-ready dependency is missing: {required}")
    marker = json.loads(marker_path.read_text(encoding="utf-8"))
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    state = json.loads(state_path.read_text(encoding="utf-8"))
    if marker.get("training_ready") is not True:
        raise RuntimeError("training-ready marker does not authorize training")
    if marker.get("training_started") is not False:
        raise RuntimeError("training-ready marker must record training_started=false")
    if marker.get("created_after_all_checks") is not True:
        raise RuntimeError("training-ready marker was not created after all checks")
    if marker.get("audit_state_digest") != audit_state_digest(state):
        raise RuntimeError("training-ready audit state digest mismatch")
    if marker.get("summary_file_digest") != content_fingerprint(summary_path):
        raise RuntimeError("training-ready summary digest mismatch")
    if summary.get("audit_state_digest") != marker.get("audit_state_digest"):
        raise RuntimeError("summary audit state digest mismatch")
    if summary.get("training_ready") is not True:
        raise RuntimeError("summary does not authorize training")
    if summary.get("training_started") is not False:
        raise RuntimeError("summary must record training_started=false")
    gate = summary.get("next_stage_gate", {})
    if gate.get("passes") is not True or gate.get("training_ready") is not True:
        raise RuntimeError("summary next-stage gate does not pass")
    conditions = gate.get("conditions", {})
    if set(conditions) != FINAL_TRAINING_GATE_CONDITIONS:
        raise RuntimeError("summary does not contain the complete final training gate")
    if not all(bool(value) for value in conditions.values()):
        raise RuntimeError("summary contains a failed or missing training gate condition")
    recommended = summary.get("recommended_config")
    if not recommended or marker.get("recommended_config") != recommended:
        raise RuntimeError("training-ready recommended config mismatch")
    snapshot = summary.get("input_snapshot_after")
    if not isinstance(snapshot, Mapping):
        raise RuntimeError("summary input_snapshot_after is missing")
    if marker.get("input_snapshot_digest") != audit_state_digest(snapshot):
        raise RuntimeError("training-ready input snapshot digest mismatch")
    if summary.get("input_snapshot_digest") != marker.get("input_snapshot_digest"):
        raise RuntimeError("summary input snapshot digest mismatch")
    output_dir = audit_dir.parent
    paths = AuditPaths(output_dir, audit_dir.name)
    live_snapshot = audit_input_snapshot(
        paths,
        read_jsonl(paths.unique_images),
        read_jsonl(paths.merged),
        read_jsonl(paths.task_samples),
    )
    if live_snapshot != snapshot:
        raise RuntimeError("live manifest/detections no longer match training-ready snapshot")
    return {
        "training_ready": True,
        "recommended_config": recommended,
        "audit_state_digest": marker["audit_state_digest"],
        "input_snapshot_digest": marker["input_snapshot_digest"],
        "summary_file_digest": marker["summary_file_digest"],
    }


def initialize_crop_audit_v3(
    args: argparse.Namespace,
    paths: AuditPaths,
    unique: Sequence[Mapping[str, Any]],
    input_snapshot: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Create immutable named v3 state without touching any other audit."""
    expected = {
        "format_version": CROP_AUDIT_FORMAT_VERSION,
        "crop_audit_name": paths.crop_audit_name,
        "unique_images": len(unique),
        "expected_unique_images": int(getattr(args, "expected_unique_images", 0)),
        "image_id_digest": digest_ids(str(row["image_id"]) for row in unique),
        "input_snapshot": dict(input_snapshot or {}),
        "candidates": TASK_AWARE_CANDIDATES,
        "max_crops": args.max_crops,
        "boundary_margin_ratio": args.boundary_margin_ratio,
        "content_missing_crop": "full_original_[0,0,W,H]",
        "content_missing_label_transform_applied": False,
        "overview_samples_per_task": int(
            getattr(args, "overview_samples_per_task", 50)
        ),
        "overview_anomalies_per_category": int(
            getattr(args, "overview_anomalies_per_category", 50)
        ),
    }
    root = paths.crop_audit
    state_path = root / "audit_state.json"
    if state_path.is_file():
        actual = json.loads(state_path.read_text(encoding="utf-8"))
        if actual != expected:
            raise RuntimeError(
                f"crop audit parameters are immutable for resume: {state_path}; "
                "use a new --output-dir for different geometry parameters"
            )
        return expected
    if root.is_dir() and any(root.iterdir()):
        raise RuntimeError(
            f"named crop audit directory already contains non-v3 output: {root}; "
            "choose a new --crop-audit-name. Existing audit directories are never moved or overwritten."
        )
    root.mkdir(parents=True, exist_ok=True)
    atomic_write_json(state_path, expected)
    return expected


# Backward-compatible import for older smoke tests and callers.  The state
# written by this alias is still format_version=3.
initialize_crop_audit_v2 = initialize_crop_audit_v3


def planned_crop_paths(
    config_root: Path,
    image_id: str,
    crop_boxes: Sequence[Sequence[int]],
    *,
    prefix: str,
) -> list[Path]:
    directory = config_root / "crops" / image_id
    paths = []
    for crop in crop_boxes:
        bbox_token = stable_id("bbox", ",".join(str(int(value)) for value in crop), 12)
        paths.append(directory / f"{prefix}_{bbox_token}.png")
    return paths


def planned_overview_path(config_root: Path, sample: Mapping[str, Any]) -> Path:
    return (
        config_root
        / "overviews"
        / str(sample["task"])
        / f"{sample['image_id']}_{sample['sample_id']}.png"
    )


def compute_geometry_record(
    cropper: Any,
    *,
    manifest: Mapping[str, Any],
    detection: Mapping[str, Any],
    image_samples: Sequence[Mapping[str, Any]],
    config_name: str,
    candidate: Mapping[str, Any],
    config_root: Path,
    max_crops: int,
    boundary_margin_ratio: float,
    shared_proposal_cache: dict[
        tuple[float, float, float, float], dict[str, Any]
    ] | None = None,
) -> dict[str, Any]:
    """Compute task-aware image/candidate geometry without opening the image."""
    image_id = str(manifest["image_id"])
    width, height = int(manifest["width"]), int(manifest["height"])
    whole_box = [0, 0, width, height]
    detection_count = sum(
        len(detection.get(f"{source}_detections", [])) for source in ("text", "icon")
    )
    whole_proposal = {
        "detection_count": detection_count,
        "edge_count": 0,
        "component_count_before_merge": 0,
        "forced_merge": False,
        "empty_detection_fallback": detection_count == 0,
        "detector_boundary_cut_count": 0,
    }
    proposal_cache = shared_proposal_cache if shared_proposal_cache is not None else {}
    sample_results: list[dict[str, Any]] = []
    for sample in image_samples:
        task = str(sample["task"])
        use_whole = uses_task_whole_image_policy(task)
        if use_whole:
            proposal = whole_proposal
            crop_boxes = [whole_box]
            crop_paths = [Path(str(manifest["image_path"])).resolve()]
            rule: Mapping[str, float] | None = None
        else:
            rule = candidate["task_rules"][task]
            rule_key = (
                float(rule["horizontal_link_ratio"]),
                float(rule["vertical_link_ratio"]),
                float(rule["context_ratio"]),
                float(rule.get("min_context_image_ratio", 0.0)),
            )
            if rule_key not in proposal_cache:
                proposal_cache[rule_key] = proposal_crops(
                    cropper,
                    detection,
                    rule,
                    max_crops=max_crops,
                    boundary_margin_ratio=boundary_margin_ratio,
                    min_context_image_ratio=float(
                        rule.get("min_context_image_ratio", 0.0)
                    ),
                )
            proposal = proposal_cache[rule_key]
            crop_boxes = proposal["crop_boxes"]
            crop_paths = planned_crop_paths(
                config_root, image_id, crop_boxes, prefix="region"
            )
        overview = planned_overview_path(config_root, sample)
        preview, preview_failures = build_preview_rows(
            sample, crop_boxes, crop_paths, config_name=config_name
        )
        errors = [row["roundtrip_max_error_px"] for row in preview]
        detail, failures = make_image_detail(
            sample,
            proposal,
            crop_boxes,
            crop_paths,
            overview,
            config_name,
            errors,
        )
        detail["partial_training_eligible_count"] = sum(
            row.get("failure_type") == "partial_intersection"
            and bool(row.get("training_eligible"))
            for row in preview_failures
        )
        detail["hard_negative_count"] = sum(
            row.get("training_eligible") and row.get("positive") is False
            for row in preview
        )
        proposal_summary = {
            key: proposal[key]
            for key in (
                "detection_count",
                "edge_count",
                "component_count_before_merge",
                "forced_merge",
                "empty_detection_fallback",
                "detector_boundary_cut_count",
            )
        }
        # Geometry reports never claim that a file was rendered.  The selected
        # config materialization pass fills these fields with verified paths.
        detail["overview"] = ""
        detail["crop_paths"] = []
        for failure in failures:
            failure["visualization"] = ""
        sample_results.append(
            {
                "sample_id": sample["sample_id"],
                "task": sample["task"],
                "gt_boxes": sample["gt_boxes"],
                "gt_boxes_1000": sample["gt_boxes_1000"],
                "crop_kind": "whole" if use_whole else "region",
                "task_geometry_rule": dict(rule) if rule is not None else None,
                "crop_boxes": [list(box) for box in crop_boxes],
                "proposal": proposal_summary,
                "detail": detail,
                "failures": failures,
                "preview_failures": preview_failures,
            }
        )
    unique_region_boxes = sorted(
        {
            tuple(int(value) for value in box)
            for result in sample_results
            if result["crop_kind"] == "region"
            for box in result["crop_boxes"]
        },
        key=lambda box: (box[1], box[0], box[3], box[2]),
    )
    return {
        "image_id": image_id,
        "config": config_name,
        "unique_region_boxes": [list(box) for box in unique_region_boxes],
        "whole_box": whole_box,
        "detection_count": detection_count,
        "sample_results": sample_results,
    }


_GEOMETRY_CROPPER: Any | None = None


def initialize_geometry_worker(parser_root: str) -> None:
    global _GEOMETRY_CROPPER
    _GEOMETRY_CROPPER = load_parser_module(Path(parser_root), "ui_region_cropper")


def geometry_worker(payload: Mapping[str, Any]) -> dict[str, Any]:
    if _GEOMETRY_CROPPER is None:
        raise RuntimeError("geometry worker cropper was not initialized")
    values = dict(payload)
    values["config_root"] = Path(str(values["config_root"]))
    return compute_geometry_record(_GEOMETRY_CROPPER, **values)


def geometry_bundle_worker(payload: Mapping[str, Any]) -> list[dict[str, Any]]:
    """Evaluate all missing candidates for one image while reusing rule geometry."""
    if _GEOMETRY_CROPPER is None:
        raise RuntimeError("geometry worker cropper was not initialized")
    values = dict(payload)
    candidates = values.pop("candidates")
    config_roots = values.pop("config_roots")
    shared_cache: dict[tuple[float, float, float, float], dict[str, Any]] = {}
    return [
        compute_geometry_record(
            _GEOMETRY_CROPPER,
            **values,
            config_name=name,
            candidate=candidate,
            config_root=Path(str(config_roots[name])),
            shared_proposal_cache=shared_cache,
        )
        for name, candidate in candidates.items()
    ]


def geometry_shard_valid(
    input_path: Path,
    output_path: Path,
    done_path: Path,
    config_name: str,
    expected_state_digest: str,
) -> bool:
    stage = f"crop-geometry-{config_name}"
    if not completed_shard_valid(input_path, output_path, done_path, stage):
        return False
    try:
        marker = json.loads(done_path.read_text(encoding="utf-8"))
        return (
            marker.get("audit_state_digest") == expected_state_digest
            and all(
                row.get("config") == config_name for row in read_jsonl(output_path)
            )
        )
    except (OSError, ValueError, json.JSONDecodeError):
        return False


def materialization_shard_valid(
    input_path: Path,
    output_path: Path,
    done_path: Path,
    config_name: str,
    expected_state_digest: str,
) -> bool:
    stage = f"crop-materialize-{config_name}"
    if not completed_shard_valid(input_path, output_path, done_path, stage):
        return False
    try:
        marker = json.loads(done_path.read_text(encoding="utf-8"))
        if marker.get("audit_state_digest") != expected_state_digest:
            return False
        rows = read_jsonl(output_path)
    except (OSError, ValueError, json.JSONDecodeError):
        return False
    for row in rows:
        paths = [*row.get("region_paths", []), *row.get("whole_paths", [])]
        paths.extend(row.get("whole_source_paths", []))
        paths.extend(row.get("overview_paths", {}).values())
        if any(not Path(str(path)).is_file() for path in paths):
            return False
    return True


def collect_config_audit(
    records: Sequence[Mapping[str, Any]],
    *,
    overlap: Mapping[str, Any],
    samples_by_image: Mapping[str, Sequence[Mapping[str, Any]]],
    max_crops: int,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    dict[str, list[dict[str, Any]]],
    list[dict[str, Any]],
]:
    details: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    anomaly_categories: dict[str, list[dict[str, Any]]] = defaultdict(list)
    preview_failures: list[dict[str, Any]] = []
    for record in records:
        for sample_result in record["sample_results"]:
            proposal = sample_result["proposal"]
            detail = dict(sample_result["detail"])
            details.append(detail)
            sample_failures = [dict(row) for row in sample_result["failures"]]
            failures.extend(sample_failures)
            for failure in sample_failures:
                category = (
                    "gt_partial_only"
                    if failure["failure_type"] == "partial_intersection"
                    else "gt_uncovered"
                )
                anomaly_categories[category].append(failure)
            if proposal["empty_detection_fallback"]:
                anomaly_categories["detector_empty_fallback"].append(detail)
            if detail["union_area_ratio"] > 0.8:
                anomaly_categories["near_full_image"].append(detail)
            if detail["crop_count"] == max_crops or proposal["forced_merge"]:
                anomaly_categories["max_crops_or_forced_merge"].append(detail)
            if detail["roundtrip_error_over_1_count"]:
                anomaly_categories["roundtrip_error"].append(detail)
            preview_failures.extend(dict(row) for row in sample_result["preview_failures"])
    for failure in preview_failures:
        if failure["failure_type"] == "roundtrip_error":
            failure["visualization"] = ""
            anomaly_categories["roundtrip_error"].append(failure)

    detail_by_sample = {str(row["sample_id"]): row for row in details}
    details_by_source: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in details:
        details_by_source[str(row["source_image"])].append(row)
    for conflict in overlap["actual_training_data"]["basename_conflicts"]["details"]:
        for canonical_path in conflict["canonical_paths"]:
            for row in details_by_source.get(str(canonical_path), []):
                anomaly_categories["basename_or_multitask_annotation_risk"].append(
                    {
                        "risk_type": "same_basename_different_identity",
                        "basename": conflict["basename"],
                        "sample_id": row["sample_id"],
                        "image_id": row["image_id"],
                        "task": row["task"],
                        "source_image": row["source_image"],
                        "visualization": "",
                    }
                )
    for image_id, image_samples in samples_by_image.items():
        if len({sample["task"] for sample in image_samples}) < 2:
            continue
        signatures = {
            str(sample["task"]): tuple(tuple(box) for box in sample["gt_boxes"])
            for sample in image_samples
        }
        if len(set(signatures.values())) <= 1:
            continue
        for sample in image_samples:
            detail = detail_by_sample[str(sample["sample_id"])]
            anomaly_categories["basename_or_multitask_annotation_risk"].append(
                {
                    "risk_type": "shared_image_distinct_task_gt",
                    "sample_id": sample["sample_id"],
                    "image_id": image_id,
                    "task": sample["task"],
                    "gt_signature": signatures[str(sample["task"])],
                    "visualization": detail["overview"],
                }
            )
    ordered = {name: anomaly_categories.get(name, []) for name in ANOMALY_PRIORITY}
    return details, failures, ordered, preview_failures


def candidate_selection_key(
    config_name: str,
    summary: Mapping[str, Any],
    all_details: Sequence[Mapping[str, Any]],
) -> tuple[float, float, float, float, float]:
    scopes = summary["candidates"][config_name]["by_scope"]
    gt_total = sum(scopes[task]["gt_count"] for task in REGION_TASKS)
    contained = sum(scopes[task]["gt_contained_count"] for task in REGION_TASKS)
    positives = sum(scopes[task]["positive_samples"] for task in REGION_TASKS)
    successful = sum(
        scopes[task]["positive_sample_success_count"] for task in REGION_TASKS
    )
    local_rows = [
        row
        for row in all_details
        if row["config"] == config_name and row["task"] in REGION_TASKS
    ]
    efficiency = aggregate_scope(local_rows)
    return (
        contained / gt_total if gt_total else 1.0,
        min(scopes[task]["gt_box_containment_recall"] for task in REGION_TASKS),
        successful / positives if positives else 1.0,
        efficiency["pixel_reduction_ratio"],
        efficiency["gt_gain_over_1_5_ratio"],
    )


def evaluate_candidate_gate(
    config_name: str,
    summary: Mapping[str, Any],
    all_details: Sequence[Mapping[str, Any]],
    preview_failures: Sequence[Mapping[str, Any]] = (),
) -> dict[str, Any]:
    scopes = summary["candidates"][config_name]["by_scope"]
    region_rows = [
        row
        for row in all_details
        if row["config"] == config_name and row["task"] in REGION_TASKS
    ]
    region = aggregate_scope(region_rows)
    overall_recall = region["gt_box_containment_recall"]
    per_task = {
        task: scopes[task]["gt_box_containment_recall"] for task in REGION_TASKS
    }
    partial_eligible_from_details = sum(
        int(row.get("partial_training_eligible_count", 0)) for row in region_rows
    )
    partial_eligible_from_preview = sum(
        row.get("failure_type") == "partial_intersection"
        and bool(row.get("training_eligible"))
        for row in preview_failures
    )
    if partial_eligible_from_details != partial_eligible_from_preview:
        raise AssertionError("partial eligibility accounting mismatch")
    partial_eligible = partial_eligible_from_details
    hard_negative_violations = sum(
        int(row.get("hard_negative_count", 0)) > 1 for row in region_rows
    )
    conditions = {
        "region_overall_recall_at_least_0_99": overall_recall >= 0.99,
        "each_region_task_recall_at_least_0_98": all(
            value >= 0.98 for value in per_task.values()
        ),
        "detector_boundary_cut_count_zero": (
            region["detector_boundary_cut_count"] == 0
        ),
        "region_roundtrip_error_over_1_count_zero": (
            region["roundtrip_error_over_1_count"] == 0
        ),
        "partial_crop_training_eligible_count_zero": partial_eligible == 0,
        "hard_negative_max_one_per_image_task": hard_negative_violations == 0,
    }
    return {
        "config": config_name,
        "region_overall_recall": overall_recall,
        "region_min_task_recall": min(per_task.values(), default=1.0),
        "region_task_recall": per_task,
        "detector_boundary_cut_count": region["detector_boundary_cut_count"],
        "region_roundtrip_error_over_1_count": region[
            "roundtrip_error_over_1_count"
        ],
        "partial_crop_training_eligible_count": partial_eligible,
        "hard_negative_limit_violation_count": hard_negative_violations,
        "conditions": conditions,
        "passes": all(conditions.values()),
    }


def select_overview_sample_ids(
    samples: Sequence[Mapping[str, Any]],
    anomalies: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    samples_per_task: int,
    anomalies_per_category: int,
) -> set[str]:
    selected: set[str] = set()
    by_task: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for sample in samples:
        by_task[str(sample["task"])].append(sample)
    for task in TASK_NAMES:
        for sample in sorted(
            by_task.get(task, []), key=lambda row: str(row["sample_id"])
        )[:samples_per_task]:
            selected.add(str(sample["sample_id"]))
    for category in ANOMALY_PRIORITY:
        category_rows = list(anomalies.get(category, []))
        if category == "gt_uncovered":
            for row in category_rows:
                if row.get("sample_id") is not None:
                    selected.add(str(row["sample_id"]))
            continue
        if category == "gt_partial_only":
            strata: dict[tuple[str, str, str], list[Mapping[str, Any]]] = defaultdict(list)
            for row in category_rows:
                strata[
                    (
                        str(row.get("task", "")),
                        str(row.get("detection_density", "")),
                        str(row.get("compensation_bucket", "")),
                    )
                ].append(row)
            queues = {
                key: sorted(
                    values, key=lambda item: str(item.get("sample_id", ""))
                )
                for key, values in strata.items()
            }
            category_rows = []
            offset = 0
            while any(offset < len(values) for values in queues.values()):
                for key in sorted(queues):
                    if offset < len(queues[key]):
                        category_rows.append(queues[key][offset])
                offset += 1
        added = 0
        for row in category_rows:
            sample_id = row.get("sample_id")
            if sample_id is None:
                continue
            selected.add(str(sample_id))
            added += 1
            if added >= anomalies_per_category:
                break
    return selected


def materialize_image_record(
    *,
    manifest: Mapping[str, Any],
    geometry: Mapping[str, Any],
    config_root: Path,
    overview_sample_ids: set[str],
) -> dict[str, Any]:
    """Decode once; write each unique region bbox once and reuse whole originals."""
    image_id = str(manifest["image_id"])
    sample_results = geometry["sample_results"]
    region_boxes = [list(box) for box in geometry["unique_region_boxes"]]
    region_paths = planned_crop_paths(
        config_root, image_id, region_boxes, prefix="region"
    )
    bbox_to_path = {
        tuple(box): path.resolve() for box, path in zip(region_boxes, region_paths)
    }
    source_path = Path(str(manifest["image_path"])).resolve()
    with open_raw_image(Path(str(manifest["image_path"]))) as image:
        for crop, path in zip(region_boxes, region_paths):
            cropped = image.crop(tuple(crop))
            try:
                atomic_save_png(cropped, path)
            finally:
                cropped.close()
        overview_paths: dict[str, str] = {}
        sample_paths: dict[str, list[str]] = {}
        for sample_result in sample_results:
            sample_id = str(sample_result["sample_id"])
            if sample_result["crop_kind"] == "whole":
                crop_paths = [source_path]
            else:
                crop_paths = [
                    bbox_to_path[tuple(box)] for box in sample_result["crop_boxes"]
                ]
            sample_paths[sample_id] = [str(path) for path in crop_paths]
            if sample_id not in overview_sample_ids:
                continue
            overview = (
                config_root
                / "overviews"
                / str(sample_result["task"])
                / f"{image_id}_{sample_id}.png"
            )
            save_overview(
                image,
                sample_result["crop_boxes"],
                sample_result["gt_boxes"],
                overview,
            )
            overview_paths[sample_id] = str(overview.resolve())
    return {
        "image_id": image_id,
        "region_paths": [str(path.resolve()) for path in region_paths],
        "whole_paths": [],
        "whole_source_paths": (
            [str(source_path)]
            if any(row["crop_kind"] == "whole" for row in sample_results)
            else []
        ),
        "region_reference_count": sum(
            len(row["crop_boxes"])
            for row in sample_results
            if row["crop_kind"] == "region"
        ),
        "whole_reference_count": sum(
            row["crop_kind"] == "whole" for row in sample_results
        ),
        "sample_paths": sample_paths,
        "overview_paths": overview_paths,
    }


def build_cross_task_supervision_audit(
    samples: Sequence[Mapping[str, Any]], overlap: Mapping[str, Any]
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    by_image: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for sample in samples:
        by_image[str(sample["image_id"])].append(sample)
    rows = []
    cardinality = Counter()
    positive_diff = 0
    gt_diff = 0
    for image_id, members in sorted(by_image.items()):
        supervision = {
            str(row["task"]): {
                "positive": bool(row["positive"]),
                "gt_count": int(row["gt_count"]),
                "gt_boxes_1000": row["gt_boxes_1000"],
            }
            for row in members
        }
        cardinality[len(supervision)] += 1
        has_positive_diff = len(
            {entry["positive"] for entry in supervision.values()}
        ) > 1
        has_gt_diff = len(
            {
                json.dumps(entry["gt_boxes_1000"], sort_keys=True)
                for entry in supervision.values()
            }
        ) > 1
        positive_diff += has_positive_diff
        gt_diff += has_gt_diff
        rows.append(
            {
                "image_id": image_id,
                "tasks": sorted(supervision),
                "task_count": len(supervision),
                "positive_labels_differ": has_positive_diff,
                "gt_labels_differ": has_gt_diff,
                "task_supervision": supervision,
            }
        )
    actual = overlap["actual_training_data"]
    split_overlap = int(
        overlap.get("same_content_cross_train_val", {}).get("count", 0)
    )
    return {
        "per_task_content_unique_images": {
            task: len(
                {
                    str(row["image_id"])
                    for row in samples
                    if str(row["task"]) == task
                }
            )
            for task in TASK_NAMES
        },
        "content_overlap": actual["content_overlap"],
        "task_cardinality_by_content": {
            str(size): cardinality.get(size, 0) for size in range(1, 6)
        },
        "cross_task_positive_label_difference_images": positive_diff,
        "cross_task_gt_difference_images": gt_diff,
        "same_content_cross_train_val_count": split_overlap,
        "all_tasks_share_one_content_pool": all(
            len(
                {
                    str(row["image_id"])
                    for row in samples
                    if str(row["task"]) == task
                }
            )
            == len(by_image)
            for task in TASK_NAMES
        ),
    }, rows


def materialization_reuse_metrics(
    rows: Sequence[Mapping[str, Any]], overview_count: int
) -> dict[str, Any]:
    region_paths = {
        str(path) for row in rows for path in row.get("region_paths", [])
    }
    whole_sources = {
        str(path) for row in rows for path in row.get("whole_source_paths", [])
    }
    region_references = sum(int(row.get("region_reference_count", 0)) for row in rows)
    whole_references = sum(int(row.get("whole_reference_count", 0)) for row in rows)
    return {
        "unique_images": len(rows),
        "region_physical_file_count": len(region_paths),
        "region_task_reference_count": region_references,
        "region_reference_reuse_ratio": (
            1 - len(region_paths) / region_references if region_references else 0.0
        ),
        "whole_original_physical_file_count": len(whole_sources),
        "whole_generated_file_count": 0,
        "whole_task_reference_count": whole_references,
        "whole_reference_reuse_ratio": (
            1 - len(whole_sources) / whole_references if whole_references else 0.0
        ),
        "overview_file_count": overview_count,
        "whole_image_policy": "reuse original path; no duplicate PNG",
    }


def run_crop_audit(args: argparse.Namespace) -> dict[str, Any]:
    """Run the named CPU-only v3 audit against immutable detector output."""
    paths = AuditPaths(
        args.output_dir, str(getattr(args, "crop_audit_name", "crop_audit_v3"))
    )
    marker_path = paths.crop_audit / "training_ready.json"
    invalidate_training_ready_marker(marker_path)
    detection_rows = read_jsonl(paths.merged)
    detections = {str(row["image_id"]): row for row in detection_rows}
    unique = read_jsonl(paths.unique_images)
    samples = read_jsonl(paths.task_samples)
    input_snapshot = audit_input_snapshot(paths, unique, detection_rows, samples)
    expected_unique = int(getattr(args, "expected_unique_images", 0))
    if expected_unique and len(unique) != expected_unique:
        raise RuntimeError(
            f"expected {expected_unique} unique images, found {len(unique)}; "
            "refusing an incomplete crop-only audit"
        )
    overlap = json.loads(
        (paths.manifest / "overlap" / "source_overlap.json").read_text(
            encoding="utf-8"
        )
    )
    supervision_audit, supervision_rows = build_cross_task_supervision_audit(
        samples, overlap
    )
    audit_state = initialize_crop_audit_v3(
        args, paths, unique, input_snapshot=input_snapshot
    )
    state_digest = audit_state_digest(audit_state)
    resume = bool(getattr(args, "resume", False))
    crop_workers = int(getattr(args, "crop_workers", 1))
    samples_by_image: dict[str, list[dict[str, Any]]] = defaultdict(list)
    samples_by_id: dict[str, dict[str, Any]] = {}
    for sample in samples:
        image_id = str(sample["image_id"])
        sample_id = str(sample["sample_id"])
        samples_by_image[image_id].append(sample)
        if sample_id in samples_by_id:
            raise ValueError(f"duplicate sample_id in task manifest: {sample_id}")
        samples_by_id[sample_id] = sample

    shard_paths = sorted(paths.shards.glob("shard_*.jsonl"))
    if not shard_paths:
        raise FileNotFoundError(f"no manifest shards found under {paths.shards}")
    expected_ids = {str(row["image_id"]) for row in unique}
    shard_ids = [
        str(row["image_id"])
        for shard in shard_paths
        for row in read_jsonl(shard)
    ]
    if len(shard_ids) != len(expected_ids) or set(shard_ids) != expected_ids:
        raise ValueError("manifest shards do not exactly match unique_images.jsonl")

    candidate_names = tuple(TASK_AWARE_CANDIDATES)
    geometry_total = len(unique) * len(candidate_names)
    geometry_completed = 0
    for candidate_name in candidate_names:
        root = paths.crop_audit / f"candidate_{candidate_name}" / "geometry"
        for shard in shard_paths:
            if resume and geometry_shard_valid(
                shard,
                root / shard.name,
                root / f"{shard.stem}.done.json",
                candidate_name,
                state_digest,
            ):
                geometry_completed += len(read_jsonl(shard))
    geometry_reporter = ProgressReporter(
        stage="crop-audit",
        total=geometry_total,
        output_dir=args.output_dir,
        interval_seconds=args.progress_interval_seconds,
        initial_completed=geometry_completed,
        unit="image-candidates",
    )
    geometry_reporter.update(
        geometry_completed,
        detail=(
            f"阶段 1/2：{crop_workers} 个 CPU worker，{len(candidate_names)} 个"
            " task-aware 候选；复用 merged detections，不运行 GPU"
        ),
        force=True,
    )
    geometry_remaining = geometry_completed < geometry_total
    local_cropper = (
        load_parser_module(args.parser_root, "ui_region_cropper")
        if geometry_remaining and crop_workers == 1
        else None
    )
    process_pool = (
        ProcessPoolExecutor(
            max_workers=crop_workers,
            initializer=initialize_geometry_worker,
            initargs=(str(args.parser_root),),
        )
        if geometry_remaining and crop_workers > 1
        else None
    )
    try:
        for shard in shard_paths:
            missing = []
            for name in candidate_names:
                root = paths.crop_audit / f"candidate_{name}" / "geometry"
                if not (
                    resume
                    and geometry_shard_valid(
                        shard,
                        root / shard.name,
                        root / f"{shard.stem}.done.json",
                        name,
                        state_digest,
                    )
                ):
                    missing.append(name)
            if not missing:
                continue
            shard_rows = read_jsonl(shard)
            payloads = [
                {
                    "manifest": manifest,
                    "detection": detections[str(manifest["image_id"])],
                    "image_samples": samples_by_image[str(manifest["image_id"])],
                    "candidates": {
                        name: TASK_AWARE_CANDIDATES[name] for name in missing
                    },
                    "config_roots": {
                        name: str(paths.crop_audit / f"candidate_{name}")
                        for name in missing
                    },
                    "max_crops": args.max_crops,
                    "boundary_margin_ratio": args.boundary_margin_ratio,
                }
                for manifest in shard_rows
            ]
            if process_pool is not None:
                iterator = process_pool.map(
                    geometry_bundle_worker, payloads, chunksize=4
                )
            else:
                assert local_cropper is not None

                def local_bundles() -> Iterator[list[dict[str, Any]]]:
                    for payload in payloads:
                        shared: dict[
                            tuple[float, float, float, float], dict[str, Any]
                        ] = {}
                        yield [
                            compute_geometry_record(
                                local_cropper,
                                manifest=payload["manifest"],
                                detection=payload["detection"],
                                image_samples=payload["image_samples"],
                                config_name=name,
                                candidate=TASK_AWARE_CANDIDATES[name],
                                config_root=Path(payload["config_roots"][name]),
                                max_crops=args.max_crops,
                                boundary_margin_ratio=args.boundary_margin_ratio,
                                shared_proposal_cache=shared,
                            )
                            for name in missing
                        ]

                iterator = local_bundles()
            rows_by_candidate: dict[str, list[dict[str, Any]]] = {
                name: [] for name in missing
            }
            for bundle in iterator:
                for row in bundle:
                    name = str(row["config"])
                    rows_by_candidate[name].append(row)
                    geometry_completed += 1
                    geometry_reporter.update(
                        geometry_completed,
                        detail=(
                            f"阶段 1/2：candidate {name}，{shard.name}，"
                            f"{row['detection_count']} boxes -> "
                            f"{len(row['unique_region_boxes'])} unique crops"
                        ),
                    )
            for name, output_rows in rows_by_candidate.items():
                root = paths.crop_audit / f"candidate_{name}" / "geometry"
                atomic_write_jsonl(root / shard.name, output_rows)
                atomic_write_json(
                    root / f"{shard.stem}.done.json",
                    {
                        "stage": f"crop-geometry-{name}",
                        "config": name,
                        "count": len(output_rows),
                        "image_id_digest": digest_ids(
                            str(row["image_id"]) for row in output_rows
                        ),
                        "input_shard": str(shard),
                        "audit_state_digest": state_digest,
                    },
                )
    finally:
        if process_pool is not None:
            process_pool.shutdown(wait=True, cancel_futures=True)
    geometry_reporter.update(
        geometry_total,
        detail="阶段 1/2 完成：task-aware 几何、GT 离线评价和坐标检查已落盘",
        force=True,
    )

    details_by_candidate: dict[str, list[dict[str, Any]]] = {}
    failures_by_candidate: dict[str, list[dict[str, Any]]] = {}
    anomalies_by_candidate: dict[str, dict[str, list[dict[str, Any]]]] = {}
    preview_failures_by_candidate: dict[str, list[dict[str, Any]]] = {}
    summary: dict[str, Any] = {
        "candidates": {},
        "cpt_enabled": False,
        "crop_only": True,
        "detector_stages_executed": [],
        "audit_format_version": CROP_AUDIT_FORMAT_VERSION,
        "crop_audit_name": paths.crop_audit_name,
        "input_snapshot_before": input_snapshot,
        "cross_task_supervision": supervision_audit,
        "detection_density_definition": {
            "sparse": "detector boxes <= 50",
            "medium": "51-150 detector boxes",
            "dense": "detector boxes > 150",
        },
        "training_started": False,
    }
    for name, candidate in TASK_AWARE_CANDIDATES.items():
        root = paths.crop_audit / f"candidate_{name}" / "geometry"
        records = [
            row for shard in shard_paths for row in read_jsonl(root / shard.name)
        ]
        record_ids = [str(row["image_id"]) for row in records]
        if len(record_ids) != len(expected_ids) or set(record_ids) != expected_ids:
            raise ValueError(f"candidate {name} geometry does not match unique images")
        details, failures, anomalies, preview_failures = collect_config_audit(
            records,
            overlap=overlap,
            samples_by_image=samples_by_image,
            max_crops=args.max_crops,
        )
        details_by_candidate[name] = details
        failures_by_candidate[name] = failures
        anomalies_by_candidate[name] = anomalies
        preview_failures_by_candidate[name] = preview_failures
        by_scope = {"ALL": aggregate_scope(details)}
        for task in TASK_NAMES:
            by_scope[task] = aggregate_scope(
                [row for row in details if row["task"] == task]
            )
        region_details = [row for row in details if row["task"] in REGION_TASKS]
        by_scope["REGION_ALL"] = aggregate_scope(region_details)
        by_density = {
            density: aggregate_scope(
                [row for row in region_details if row["detection_density"] == density]
            )
            for density in ("sparse", "medium", "dense")
        }
        by_task_density = {
            task: {
                density: aggregate_scope(
                    [
                        row
                        for row in region_details
                        if row["task"] == task
                        and row["detection_density"] == density
                    ]
                )
                for density in ("sparse", "medium", "dense")
            }
            for task in REGION_TASKS
        }
        partial_compensation = {}
        for task in REGION_TASKS:
            values = [
                float(row["required_max_single_side_px"])
                for row in failures
                if row["task"] == task
                and row["failure_type"] == "partial_intersection"
                and row.get("required_max_single_side_px") is not None
            ]
            partial_compensation[task] = {
                "count": len(values),
                "p50_px": percentile(values, 0.50),
                "p90_px": percentile(values, 0.90),
                "max_px": max(values, default=0.0),
            }
        summary["candidates"][name] = {
            "parameters": candidate,
            "by_scope": by_scope,
            "region_by_detection_density": by_density,
            "region_by_task_and_detection_density": by_task_density,
            "partial_required_single_side_compensation": partial_compensation,
            "gt_failure_count": len(failures),
            "anomaly_counts": {
                category: len(rows) for category, rows in anomalies.items()
            },
        }

    all_geometry_details = [
        row for name in candidate_names for row in details_by_candidate[name]
    ]
    candidate_gates = {
        name: evaluate_candidate_gate(
            name,
            summary,
            all_geometry_details,
            preview_failures_by_candidate[name],
        )
        for name in candidate_names
    }
    summary["candidate_gates"] = candidate_gates
    passing = [name for name in candidate_names if candidate_gates[name]["passes"]]
    selection_pool = passing or list(candidate_names)
    selected_name = max(
        selection_pool,
        key=lambda name: candidate_selection_key(
            name, summary, all_geometry_details
        ),
    )
    if passing:
        summary["recommended_config"] = selected_name
    else:
        summary["best_candidate_config"] = selected_name
    selected_root = paths.crop_audit / f"candidate_{selected_name}"
    summary["materialized_candidate"] = selected_name
    overview_sample_ids = select_overview_sample_ids(
        samples,
        anomalies_by_candidate[selected_name],
        samples_per_task=int(getattr(args, "overview_samples_per_task", 50)),
        anomalies_per_category=int(
            getattr(args, "overview_anomalies_per_category", 50)
        ),
    )
    selected_records = {
        str(row["image_id"]): row
        for shard in shard_paths
        for row in read_jsonl(selected_root / "geometry" / shard.name)
    }
    material_root = selected_root / "materialized"
    material_completed = 0
    for shard in shard_paths:
        if resume and materialization_shard_valid(
            shard,
            material_root / shard.name,
            material_root / f"{shard.stem}.done.json",
            selected_name,
            state_digest,
        ):
            material_completed += len(read_jsonl(shard))
    material_reporter = ProgressReporter(
        stage="crop-audit",
        total=len(unique),
        output_dir=args.output_dir,
        interval_seconds=args.progress_interval_seconds,
        initial_completed=material_completed,
        unit="images",
    )
    material_reporter.update(
        material_completed,
        detail=(
            f"阶段 2/2：落图 {selected_name}；同 bbox 物理去重，"
            f"content_missing 复用原图；{crop_workers} 个 CPU worker"
        ),
        force=True,
    )
    for shard in shard_paths:
        output = material_root / shard.name
        marker = material_root / f"{shard.stem}.done.json"
        if resume and materialization_shard_valid(
            shard, output, marker, selected_name, state_digest
        ):
            continue
        kwargs_rows = [
            {
                "manifest": manifest,
                "geometry": selected_records[str(manifest["image_id"])],
                "config_root": selected_root,
                "overview_sample_ids": overview_sample_ids,
            }
            for manifest in read_jsonl(shard)
        ]
        output_rows: list[dict[str, Any]] = []
        if crop_workers == 1:
            iterator: Iterable[dict[str, Any]] = (
                materialize_image_record(**kwargs) for kwargs in kwargs_rows
            )
            for row in iterator:
                output_rows.append(row)
                material_completed += 1
                material_reporter.update(
                    material_completed,
                    detail=f"阶段 2/2：candidate {selected_name}，{shard.name}",
                )
        else:
            with ThreadPoolExecutor(max_workers=crop_workers) as executor:
                futures = [
                    executor.submit(materialize_image_record, **kwargs)
                    for kwargs in kwargs_rows
                ]
                for future in futures:
                    output_rows.append(future.result())
                    material_completed += 1
                    material_reporter.update(
                        material_completed,
                        detail=f"阶段 2/2：candidate {selected_name}，{shard.name}",
                    )
        atomic_write_jsonl(output, output_rows)
        atomic_write_json(
            marker,
            {
                "stage": f"crop-materialize-{selected_name}",
                "config": selected_name,
                "count": len(output_rows),
                "image_id_digest": digest_ids(
                    str(row["image_id"]) for row in output_rows
                ),
                "input_shard": str(shard),
                "audit_state_digest": state_digest,
            },
        )

    material_rows = [
        row
        for shard in shard_paths
        for row in read_jsonl(material_root / shard.name)
    ]
    material_by_image = {str(row["image_id"]): row for row in material_rows}
    if len(material_by_image) != len(expected_ids) or set(material_by_image) != expected_ids:
        raise ValueError("selected candidate materialization does not match unique images")

    selected_details: list[dict[str, Any]] = []
    selected_failures: list[dict[str, Any]] = []
    preview_by_task: dict[str, list[dict[str, Any]]] = defaultdict(list)
    overview_by_sample = {
        str(key): str(value)
        for row in material_rows
        for key, value in row["overview_paths"].items()
    }
    for record in selected_records.values():
        material = material_by_image[str(record["image_id"])]
        for sample_result in record["sample_results"]:
            sample_id = str(sample_result["sample_id"])
            crop_paths = [
                Path(path) for path in material["sample_paths"][sample_id]
            ]
            overview_path = overview_by_sample.get(sample_id, "")
            detail = dict(sample_result["detail"])
            detail["crop_paths"] = [str(path.resolve()) for path in crop_paths]
            detail["overview"] = overview_path
            selected_details.append(detail)
            preview, _ = build_preview_rows(
                samples_by_id[sample_id],
                sample_result["crop_boxes"],
                crop_paths,
                config_name=selected_name,
            )
            preview_by_task[str(sample_result["task"])].extend(preview)
            for failure in sample_result["failures"]:
                updated = dict(failure)
                updated["visualization"] = overview_path
                selected_failures.append(updated)
    details_by_candidate[selected_name] = selected_details
    failures_by_candidate[selected_name] = selected_failures
    for category, rows in anomalies_by_candidate[selected_name].items():
        for row in rows:
            row["visualization"] = overview_by_sample.get(
                str(row.get("sample_id", "")), ""
            )

    for name in candidate_names:
        root = paths.crop_audit / f"candidate_{name}"
        atomic_write_jsonl(
            root / "task_aware_manifest.jsonl", details_by_candidate[name]
        )
        atomic_write_jsonl(
            root / "gt_failures.jsonl", failures_by_candidate[name]
        )
        atomic_write_json(root / "anomalies.json", anomalies_by_candidate[name])
    for task in TASK_NAMES:
        atomic_write_jsonl(
            selected_root / "preview" / f"{task}.jsonl", preview_by_task[task]
        )

    # The root machine-readable detail and Excel contain one row per image x
    # task for the materialized candidate.  Cross-candidate metrics stay in
    # summary/config_compare, and every candidate retains its own full JSONL.
    report_details = details_by_candidate[selected_name]
    report_failures = failures_by_candidate[selected_name]
    summary["materialization"] = {
        "candidate": selected_name,
        **materialization_reuse_metrics(material_rows, len(overview_by_sample)),
        "geometry_only_candidates": [
            name for name in candidate_names if name != selected_name
        ],
    }
    summary["root_detail_outputs"] = {
        "candidate": selected_name,
        "task_aware_manifest_rows": len(report_details),
        "gt_failure_rows": len(report_failures),
        "other_candidates": "full detail remains under candidate_*/",
    }
    content_scope = summary["candidates"][selected_name]["by_scope"][
        "ui_content_missing"
    ]
    content_preview = preview_by_task["ui_content_missing"]
    summary["content_missing_checks"] = {
        "full_image_bbox": "[0,0,W,H]",
        "gt_box_containment_recall": content_scope[
            "gt_box_containment_recall"
        ],
        "label_transform_applied": False,
        "normalized_gt_identical_count": sum(
            bool(row.get("normalized_gt_identical")) for row in content_preview
        ),
        "normalized_gt_box_count": sum(
            len(row.get("original_gt_boxes_1000", [])) for row in content_preview
        ),
        "normalized_gt_identical_box_count": sum(
            len(row.get("original_gt_boxes_1000", []))
            for row in content_preview
            if row.get("original_gt_boxes_1000")
            == row.get("output_gt_boxes_1000")
        ),
        "normalized_gt_mismatch_count": sum(
            row.get("original_gt_boxes_1000") != row.get("output_gt_boxes_1000")
            for row in content_preview
        ),
        "roundtrip_gate_excluded": True,
    }
    input_snapshot_after = audit_input_snapshot(
        paths, unique, detection_rows, samples
    )
    input_snapshot_unchanged = input_snapshot_after == input_snapshot
    summary["input_snapshot_after"] = input_snapshot_after
    summary["input_snapshot_unchanged"] = input_snapshot_unchanged
    summary["input_snapshot_digest"] = audit_state_digest(input_snapshot_after)
    summary["audit_state_digest"] = state_digest
    summary["metric_definitions"] = METRIC_DEFINITIONS
    provisional_gate = build_final_training_gate(
        candidate_gates[selected_name],
        same_content_cross_train_val_count=int(
            supervision_audit["same_content_cross_train_val_count"]
        ),
        content_missing_recall=float(
            content_scope["gt_box_containment_recall"]
        ),
        content_missing_normalized_gt_mismatch_count=int(
            summary["content_missing_checks"]["normalized_gt_mismatch_count"]
        ),
        input_snapshot_unchanged=input_snapshot_unchanged,
        all_reports_written_successfully=False,
    )
    summary["next_stage_gate"] = provisional_gate
    summary["training_ready"] = False
    # Leave a fail-closed summary on disk while report generation is in flight.
    # If any later writer raises, no valid marker exists and this provisional
    # summary names all currently failed conditions, including reports=false.
    atomic_write_json(paths.crop_audit / "summary.json", summary)

    atomic_write_jsonl(
        paths.crop_audit / "cross_task_supervision.jsonl", supervision_rows
    )
    write_statistics_csv(paths.crop_audit / "statistics.csv", report_details)
    atomic_write_jsonl(paths.crop_audit / "task_aware_manifest.jsonl", report_details)
    atomic_write_jsonl(paths.crop_audit / "gt_failures.jsonl", report_failures)
    atomic_write_json(
        paths.crop_audit / "materialization_summary.json", summary["materialization"]
    )
    from render_ui5_crop_failures import (
        EXPECTED_FAILURES_BY_DENSITY,
        EXPECTED_FAILURES_BY_TASK,
        render_failure_visualizations,
    )

    strict_current_audit = expected_unique == 17281
    if strict_current_audit and selected_name != "TA_CTX015_H050":
        raise RuntimeError(
            "the 17,281-image v3 refresh must select TA_CTX015_H050; "
            f"found {selected_name}"
        )
    summary["failure_visualizations"] = render_failure_visualizations(
        output_dir=paths.output,
        crop_audit_name=paths.crop_audit_name,
        config_name=selected_name,
        expected_failures=107 if strict_current_audit else len(report_failures),
        expected_partial=87 if strict_current_audit else sum(
            row["failure_type"] == "partial_intersection" for row in report_failures
        ),
        expected_uncovered=20 if strict_current_audit else sum(
            row["failure_type"] == "uncovered" for row in report_failures
        ),
        expected_by_task=EXPECTED_FAILURES_BY_TASK if strict_current_audit else None,
        expected_by_density=(
            EXPECTED_FAILURES_BY_DENSITY if strict_current_audit else None
        ),
        resume=resume,
    )
    visualized_failure_rows = read_jsonl(
        paths.crop_audit / "gt_failures_visualized.jsonl"
    )
    material_reporter.update(
        len(unique), detail="阶段 2/2 完成，正在写唯一 Excel 与机器可读报告", force=True
    )
    required_before_excel = [
        paths.crop_audit / "cross_task_supervision.jsonl",
        paths.crop_audit / "statistics.csv",
        paths.crop_audit / "task_aware_manifest.jsonl",
        paths.crop_audit / "gt_failures.jsonl",
        paths.crop_audit / "gt_failures_visualized.jsonl",
        paths.crop_audit / "failure_diagnosis_summary.json",
        paths.crop_audit / "materialization_summary.json",
        paths.crop_audit / "failure_visualizations" / "gallery" / "index.html",
        paths.crop_audit / "failure_visualizations" / "gallery" / "uncovered_all.html",
        paths.crop_audit / "failure_visualizations" / "gallery" / "representative_partial.html",
        paths.crop_audit / "failure_visualizations" / "gallery" / "diagnosis_summary.html",
    ]
    for name in candidate_names:
        candidate_root = paths.crop_audit / f"candidate_{name}"
        required_before_excel.extend(
            [
                candidate_root / "task_aware_manifest.jsonl",
                candidate_root / "gt_failures.jsonl",
                candidate_root / "anomalies.json",
            ]
        )
    required_before_excel.extend(
        selected_root / "preview" / f"{task}.jsonl" for task in TASK_NAMES
    )
    missing_reports = [str(path) for path in required_before_excel if not path.is_file()]
    reports_ready_for_excel = not missing_reports
    final_gate = build_final_training_gate(
        candidate_gates[selected_name],
        same_content_cross_train_val_count=int(
            supervision_audit["same_content_cross_train_val_count"]
        ),
        content_missing_recall=float(
            content_scope["gt_box_containment_recall"]
        ),
        content_missing_normalized_gt_mismatch_count=int(
            summary["content_missing_checks"]["normalized_gt_mismatch_count"]
        ),
        input_snapshot_unchanged=input_snapshot_unchanged,
        all_reports_written_successfully=reports_ready_for_excel,
    )
    summary["next_stage_gate"] = final_gate
    summary["training_ready"] = final_gate["passes"]
    summary["report_write_check"] = {
        "required_before_excel": [str(path.resolve()) for path in required_before_excel],
        "missing_before_excel": missing_reports,
    }
    write_excel_report(
        paths.crop_audit / "ui5_crop_audit.xlsx",
        summary,
        overlap,
        report_details,
        visualized_failure_rows,
    )
    excel_path = paths.crop_audit / "ui5_crop_audit.xlsx"
    if not excel_path.is_file():
        raise RuntimeError(f"Excel report was not written: {excel_path}")
    # Recompute with the Excel existence included, then atomically publish the
    # final summary.  training_ready.json is intentionally the final write.
    final_gate = build_final_training_gate(
        candidate_gates[selected_name],
        same_content_cross_train_val_count=int(
            supervision_audit["same_content_cross_train_val_count"]
        ),
        content_missing_recall=float(
            content_scope["gt_box_containment_recall"]
        ),
        content_missing_normalized_gt_mismatch_count=int(
            summary["content_missing_checks"]["normalized_gt_mismatch_count"]
        ),
        input_snapshot_unchanged=input_snapshot_unchanged,
        all_reports_written_successfully=(reports_ready_for_excel and excel_path.is_file()),
    )
    summary["next_stage_gate"] = final_gate
    summary["training_ready"] = final_gate["passes"]
    summary["report_write_check"]["excel"] = str(excel_path.resolve())
    summary["report_write_check"]["all_reports_written_successfully"] = bool(
        final_gate["conditions"]["all_reports_written_successfully"]
    )
    summary_path = paths.crop_audit / "summary.json"
    atomic_write_json(summary_path, summary)
    if final_gate["passes"]:
        atomic_write_json(
            marker_path,
            {
                "training_ready": True,
                "recommended_config": selected_name,
                "training_started": False,
                "audit_state_digest": state_digest,
                "input_snapshot_digest": audit_state_digest(input_snapshot_after),
                "summary_file_digest": content_fingerprint(summary_path),
                "created_after_all_checks": True,
            },
        )
    material_reporter.update(
        len(unique),
        status="completed",
        detail=(
            f"crop-only v3 完成；materialized={selected_name}，"
            f"training_ready={final_gate['passes']}，OCR/icon 未运行"
        ),
        force=True,
    )
    return summary


def resolve_required_directory(value: Path, option_name: str) -> Path:
    """Resolve a CLI directory with an actionable error for relative paths."""
    supplied = value.expanduser()
    candidate = supplied if supplied.is_absolute() else Path.cwd() / supplied
    resolved = candidate.resolve(strict=False)
    if not resolved.exists():
        relative_note = (
            f" Relative paths are resolved from current working directory {Path.cwd()}."
            if not supplied.is_absolute()
            else ""
        )
        raise FileNotFoundError(
            f"{option_name} directory does not exist: {resolved}. "
            f"Supplied value: {value!s}.{relative_note} "
            "Pass the existing dataset/tool directory; do not create an empty placeholder."
        )
    if not resolved.is_dir():
        raise NotADirectoryError(f"{option_name} is not a directory: {resolved}")
    return resolved.resolve(strict=True)


def resolve_python_executable(value: str | None, option_name: str) -> str:
    supplied = value or sys.executable
    expanded = str(Path(supplied).expanduser())
    if Path(expanded).is_file():
        # Do not resolve the final symlink: venv uses the executable location
        # (and adjacent pyvenv.cfg) to select its site-packages.
        return str(Path(expanded).absolute())
    located = shutil.which(expanded)
    if located:
        return str(Path(located).absolute())
    raise FileNotFoundError(
        f"{option_name} Python executable does not exist or is not on PATH: {supplied}"
    )


def normalize_args(args: argparse.Namespace) -> argparse.Namespace:
    args.source_dir = resolve_required_directory(args.source_dir, "--source-dir")
    args.locany_data_dir = resolve_required_directory(
        args.locany_data_dir, "--locany-data-dir"
    )
    args.parser_root = resolve_required_directory(args.parser_root, "--parser-root")
    args.text_python = resolve_python_executable(args.text_python, "--text-python")
    args.icon_python = resolve_python_executable(args.icon_python, "--icon-python")
    args.output_dir = args.output_dir.expanduser().resolve(strict=False)
    audit_name = str(args.crop_audit_name).strip()
    if (
        not audit_name
        or audit_name in {".", ".."}
        or Path(audit_name).name != audit_name
        or "/" in audit_name
        or "\\" in audit_name
    ):
        raise ValueError("--crop-audit-name must be one safe directory name")
    args.crop_audit_name = audit_name
    if args.text_model_dir:
        args.text_model_dir = args.text_model_dir.expanduser().resolve(strict=False)
    if args.icon_model:
        args.icon_model = args.icon_model.expanduser().resolve(strict=False)
    if not 1 <= args.max_crops <= 10:
        raise ValueError("--max-crops must be in [1, 10]")
    if not 2 <= args.image_loader_threads <= 4:
        raise ValueError("--image-loader-threads must be in [2, 4]")
    if args.progress_interval_seconds <= 0:
        raise ValueError("--progress-interval-seconds must be positive")
    if args.progress_every_images <= 0:
        raise ValueError("--progress-every-images must be positive")
    if not 1 <= args.crop_workers <= 32:
        raise ValueError("--crop-workers must be in [1, 32]")
    if args.overview_samples_per_task < 0:
        raise ValueError("--overview-samples-per-task must be non-negative")
    if args.overview_anomalies_per_category < 0:
        raise ValueError("--overview-anomalies-per-category must be non-negative")
    if args.expected_unique_images < 0:
        raise ValueError("--expected-unique-images must be non-negative")
    return args


def run(args: argparse.Namespace) -> Any:
    args = normalize_args(args)
    if args.stage == "_worker":
        return run_detector_worker(args)
    stages = PIPELINE_STAGES if args.stage == "all" else (args.stage,)
    result = None
    for stage in stages:
        stage_index = PIPELINE_STAGES.index(stage) + 1
        print(
            f"\n[流水线] 开始阶段 {stage_index}/{len(PIPELINE_STAGES)}：{stage}",
            flush=True,
        )
        stage_started = time.monotonic()
        try:
            if stage == "prepare":
                result = build_task_aware_manifest(args)
            elif stage in {"text", "icon"}:
                result = run_detection_stage(args, stage)
            elif stage == "merge":
                print_preflight(args)
                result = merge_detections(args)
            elif stage == "crop-audit":
                print_preflight(args)
                result = run_crop_audit(args)
        except Exception as exc:
            status_path = args.output_dir / "run_status.json"
            try:
                payload = (
                    json.loads(status_path.read_text(encoding="utf-8"))
                    if status_path.is_file()
                    else {}
                )
            except (OSError, json.JSONDecodeError):
                payload = {}
            payload.update(
                {
                    "stage": stage,
                    "stage_index": stage_index,
                    "stage_total": len(PIPELINE_STAGES),
                    "status": "failed",
                    "error": str(exc),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
            )
            atomic_write_json(status_path, payload)
            print(
                f"[流水线] 阶段 {stage_index}/{len(PIPELINE_STAGES)} 失败：{stage}：{exc}",
                file=sys.stderr,
                flush=True,
            )
            raise
        print(
            f"[流水线] 完成阶段 {stage_index}/{len(PIPELINE_STAGES)}：{stage}，"
            f"本阶段耗时 {format_duration(time.monotonic() - stage_started)}",
            flush=True,
        )
    return result


def main(argv: Sequence[str] | None = None) -> int:
    try:
        args = parse_args(argv)
        result = run(args)
        if isinstance(result, Mapping):
            print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
